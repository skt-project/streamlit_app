"""Upload-file parsing for BD Support's real templates.

Both templates put decoration above the header, so neither can be read with a
plain ``pd.read_excel``:

  NOO  ("Template" sheet)          SKU  ("SKU TEMPLATE FOR STREAMLIT" sheet)
    row 1  instruction banner        row 1  instruction banner
    row 2  HEADER                    row 2  (blank)
    row 3  "CONTOH" example          row 3  HEADER
    row 4+ data                      row 4  "CONTOH" marker
                                     row 5  example values
                                     row 6+ data

Parsing is deliberately tolerant about *where* the header is — admins add and
remove rows — so we locate it by signature instead of trusting a fixed offset,
then treat everything after it as data and drop the example rows.
"""
from __future__ import annotations

from dataclasses import dataclass, field

import openpyxl

from . import config
from .normalize import clean, norm_header

UPLOAD_NOO = "NOO"
UPLOAD_SKU = "SKU"


class ParseError(Exception):
    """Raised when the file cannot be read at all. Carries user-facing text."""


@dataclass
class ParsedFile:
    kind: str | None
    sheet_name: str
    headers: list
    rows: list                       # list[dict[header -> str]]
    header_row: int                  # 1-based sheet row where the header sits
    row_numbers: list = field(default_factory=list)  # sheet row per data row

    def __len__(self) -> int:
        return len(self.rows)


def _load_workbook(file_obj):
    try:
        return openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
    except Exception as exc:  # openpyxl raises a wide variety here
        raise ParseError(
            "File tidak bisa dibaca. Pastikan file berformat .xlsx dan tidak "
            "ter-password."
        ) from exc


def _score(headers) -> dict:
    """How strongly a header row matches each known template signature."""
    normalised = {norm_header(h) for h in headers if clean(h)}
    return {
        UPLOAD_NOO: len(normalised & config.NOO_SIGNATURE),
        UPLOAD_SKU: len(normalised & config.SKU_SIGNATURE),
    }


def _find_header_row(ws, max_scan: int = 12):
    """Return (row_index, headers, kind) for the best-matching header row."""
    best = (0, None, None, None)
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan,
                                           values_only=True), start=1):
        headers = [clean(c) for c in row]
        if not any(headers):
            continue
        scores = _score(headers)
        kind = max(scores, key=scores.get)
        if scores[kind] > best[0]:
            best = (scores[kind], idx, headers, kind)
    if not best[1]:
        return None, None, None
    return best[1], best[2], best[3]


def _is_example_row(values) -> bool:
    """The templates mark their sample with a literal CONTOH cell."""
    return any(clean(v).upper() == "CONTOH" for v in values)


def _is_marker_only_row(values) -> bool:
    """A row whose only content is the word CONTOH.

    The two templates differ here: the NOO template puts CONTOH in column A of
    the same row as the sample values, while the SKU template puts CONTOH alone
    on one row and the sample values on the *next* row. Without this check the
    SKU sample (TYY114002 ...) is read as real data and would be written to the
    pool on every single upload.
    """
    populated = [clean(v) for v in values if clean(v)]
    return len(populated) == 1 and populated[0].upper() == "CONTOH"


def parse_upload(file_obj) -> ParsedFile:
    """Read an uploaded workbook and detect which template it is.

    Detection is by header signature, not by sheet name or by which section the
    user uploaded into — that is what lets us tell an admin they have the two
    templates the wrong way round.
    """
    wb = _load_workbook(file_obj)
    best = None
    for ws in wb.worksheets:
        header_row, headers, kind = _find_header_row(ws)
        if not header_row:
            continue
        score = _score(headers)[kind]
        if best is None or score > best[0]:
            best = (score, ws, header_row, headers, kind)

    if best is None or best[0] == 0:
        raise ParseError(
            "Template tidak dikenali. Silakan gunakan template resmi yang "
            "bisa diunduh dari aplikasi ini."
        )

    _, ws, header_row, headers, kind = best
    expected = config.NOO_COLUMNS if kind == UPLOAD_NOO else config.SKU_COLUMNS
    # Trim trailing empties so a template with padding columns still lines up.
    width = max(len(expected), len([h for h in headers if h]))
    headers = headers[:width]

    # Re-key to canonical column names. The sheet's own header text is not
    # usable as a dict key: it carries a trailing space in 'Customer Code ' and
    # a double space in 'Customer Product Name  (...)', and clean() collapses
    # whitespace — so the raw text and the canonical constant never match.
    # Everything downstream addresses columns by the canonical name.
    canonical = {norm_header(c): c for c in expected}
    headers = [canonical.get(norm_header(h), h) for h in headers]

    rows, row_numbers = [], []
    skip_next_as_example = False
    for sheet_row, values in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        cells = [clean(v) for v in values[:width]]
        if not any(cells):
            continue
        if _is_example_row(values):
            # Only a bare CONTOH marker implies the sample sits on the next row.
            skip_next_as_example = _is_marker_only_row(values)
            continue
        if skip_next_as_example:
            skip_next_as_example = False
            continue
        rows.append({h: (cells[i] if i < len(cells) else "")
                     for i, h in enumerate(headers) if h})
        row_numbers.append(sheet_row)

    return ParsedFile(kind=kind, sheet_name=ws.title, headers=headers,
                      rows=rows, header_row=header_row,
                      row_numbers=row_numbers)


def check_template_kind(parsed: ParsedFile, expected_kind: str):
    """Return a user-facing error when the wrong template was uploaded."""
    if parsed.kind == expected_kind:
        return None
    if expected_kind == UPLOAD_NOO and parsed.kind == UPLOAD_SKU:
        return ("Ini template **SKU Mapping**, bukan NOO Mapping. "
                "Silakan upload file ini di section SKU Mapping.")
    if expected_kind == UPLOAD_SKU and parsed.kind == UPLOAD_NOO:
        return ("Ini template **NOO Mapping**, bukan SKU Mapping. "
                "Silakan upload file ini di section NOO / Store Mapping.")
    return ("Template tidak dikenali. Silakan unduh template terbaru dari "
            "aplikasi ini.")


def missing_columns(parsed: ParsedFile, expected_kind: str) -> list:
    """Expected headers absent from the file, compared after normalisation."""
    expected = (config.NOO_COLUMNS if expected_kind == UPLOAD_NOO
                else config.SKU_COLUMNS)
    present = {norm_header(h) for h in parsed.headers if h}
    return [c for c in expected if norm_header(c) not in present]


def column_lookup(parsed: ParsedFile, expected_kind: str) -> dict:
    """Map canonical column name -> the header string actually in the file."""
    expected = (config.NOO_COLUMNS if expected_kind == UPLOAD_NOO
                else config.SKU_COLUMNS)
    by_norm = {norm_header(h): h for h in parsed.headers if h}
    return {c: by_norm[norm_header(c)] for c in expected
            if norm_header(c) in by_norm}
