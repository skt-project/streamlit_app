"""I/O layer: Google Sheets, Google Drive and BigQuery readers.

Everything that touches the network lives here so the rest of the package stays
unit-testable. Reads are cached; the submission ledger deliberately is not,
because a stale ledger would let a duplicate through.
"""
from __future__ import annotations

import io
import json

from . import config, duplicates
from .normalize import clean, norm_key

# Sheets (read + append), Drive (fetch the .xlsx templates), and BigQuery (the
# product master and the PO-history fallback). Scoped credentials do NOT pick up
# BigQuery implicitly, so omitting it here fails at query time, not at auth time.
SCOPES_RW = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.readonly",
    "https://www.googleapis.com/auth/bigquery",
]


# ─── Credentials ──────────────────────────────────────────────────────────────
def load_credentials(secrets=None, scopes=None):
    """Service-account credentials from st.secrets, falling back to a key file.

    No key path is hardcoded for production use; the local fallback exists only
    so the app can be run offline during development.
    """
    from google.oauth2 import service_account

    scopes = scopes or SCOPES_RW
    reason = "st.secrets tidak tersedia"
    if secrets:
        # Distinguish "the block is absent" from "the block is present but
        # broken". Collapsing both into one message turns a malformed
        # private_key into a misleading 'credentials not found'.
        try:
            info = dict(secrets["connections"]["bigquery"])
        except Exception:
            reason = ("blok [connections.bigquery] tidak ditemukan di "
                      "st.secrets")
            info = None
        if info is not None:
            missing = [k for k in ("type", "project_id", "private_key",
                                   "client_email") if not info.get(k)]
            if missing:
                raise RuntimeError(
                    "Konfigurasi [connections.bigquery] tidak lengkap. "
                    f"Field yang hilang: {', '.join(missing)}.")
            if "private_key" in info:
                info["private_key"] = info["private_key"].replace("\\n", "\n")
            try:
                return service_account.Credentials.from_service_account_info(
                    info, scopes=scopes), info.get("project_id")
            except Exception as exc:
                raise RuntimeError(
                    "Kredensial [connections.bigquery] ada tetapi tidak valid "
                    f"({type(exc).__name__}). Periksa private_key — pastikan "
                    "disalin utuh termasuk baris BEGIN/END PRIVATE KEY."
                ) from exc

    import os

    path = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS")
    if not path:
        # Application Default Credentials: how a server or a developer running
        # `gcloud auth application-default login` is expected to authenticate.
        try:
            import google.auth

            adc, adc_project = google.auth.default(scopes=scopes)
            return adc, adc_project
        except Exception:
            pass
        raise RuntimeError(
            f"Kredensial Google tidak ditemukan ({reason}). Set "
            "st.secrets['connections']['bigquery'], "
            "GOOGLE_APPLICATION_CREDENTIALS, atau jalankan "
            "`gcloud auth application-default login`."
        )
    with open(path) as fh:
        project = json.load(fh).get("project_id")
    return service_account.Credentials.from_service_account_file(
        path, scopes=scopes), project


# ─── Sheets ───────────────────────────────────────────────────────────────────
class SheetsClient:
    """Thin wrapper over the Sheets v4 API. Append is the only write verb."""

    def __init__(self, credentials, spreadsheet_id):
        from googleapiclient.discovery import build

        self._svc = build("sheets", "v4", credentials=credentials,
                          cache_discovery=False)
        self.spreadsheet_id = spreadsheet_id

    def read_values(self, tab: str, a1: str = "A1:BZ") -> list:
        res = self._svc.spreadsheets().values().get(
            spreadsheetId=self.spreadsheet_id,
            range=f"'{tab}'!{a1}",
        ).execute()
        return res.get("values", [])

    def batch_read(self, ranges) -> list:
        res = self._svc.spreadsheets().values().batchGet(
            spreadsheetId=self.spreadsheet_id, ranges=ranges,
        ).execute()
        return [vr.get("values", []) for vr in res.get("valueRanges", [])]

    def append_values(self, tab: str, rows) -> dict:
        """Append rows to the end of a tab. Never overwrites existing cells."""
        return self._svc.spreadsheets().values().append(
            spreadsheetId=self.spreadsheet_id,
            range=f"'{tab}'!A1",
            valueInputOption="RAW",
            insertDataOption="INSERT_ROWS",
            body={"values": rows},
        ).execute()

    def tab_names(self) -> list:
        meta = self._svc.spreadsheets().get(
            spreadsheetId=self.spreadsheet_id,
            fields="sheets.properties.title",
        ).execute()
        return [s["properties"]["title"] for s in meta.get("sheets", [])]


# ─── Distributor master (DIST DATABASE) ───────────────────────────────────────
def load_distributors(client) -> dict:
    """Read DIST DATABASE into ``{code: {...}}``.

    This tab — not BigQuery — is the login population. BigQuery's
    master_distributor lists only 100 active distributors; DIST DATABASE lists
    215 and is a strict superset, so gating login on BigQuery would lock out
    more than half of the real admins.
    """
    values = client.read_values(config.TAB_DIST_DATABASE, "A2:AO2000")
    if not values:
        return {}
    out = {}
    for row in values[1:]:                      # row 1 of this range = header
        get = lambda i: clean(row[i]) if i < len(row) else ""  # noqa: E731
        code = norm_key(get(config.DIST_COL_CODE))
        if not code:
            continue
        out[code] = {
            "distributor_code": code,
            "distributor_name": get(config.DIST_COL_NAME)
                                or get(config.DIST_COL_COMPANY),
            "company": get(config.DIST_COL_COMPANY),
            "status": get(config.DIST_COL_STATUS),
            "region": get(config.DIST_COL_REGION),
            "branch_code": norm_key(get(config.DIST_COL_BRANCH_CODE)),
            "active": get(config.DIST_COL_STATUS).lower() == "active",
        }
    return out


def suffixes_from_dist_database(distributors) -> dict:
    return {c: d["branch_code"] for c, d in distributors.items()
            if d.get("branch_code")}


def suffixes_from_sku_history(client) -> dict:
    """Derive each distributor's abbreviation from what the tracker already used."""
    values = client.read_values(config.TAB_SKU_MAPPING, "F2:I20000")
    latest = {}
    for row in values:
        get = lambda i: clean(row[i]) if i < len(row) else ""  # noqa: E731
        cust, dist = norm_key(get(0)), norm_key(get(3))
        if cust and dist and len(cust) > 2:
            latest[dist] = cust[2:]      # later rows win — the sheet is append-ordered
    return latest


def suffixes_from_po_history(credentials, project) -> dict:
    """Widest-coverage source: the distributor's own PO history in BigQuery."""
    from google.cloud import bigquery

    client = bigquery.Client(credentials=credentials, project=project)
    query = """
        WITH ranked AS (
          SELECT distributor_code,
                 SUBSTR(customer_code, 3) AS suffix,
                 ROW_NUMBER() OVER (
                     PARTITION BY distributor_code ORDER BY MAX(order_date) DESC
                 ) AS rn
          FROM `skintific-data-warehouse.dms.gt_po_tracking_all_mv`
          WHERE customer_code IS NOT NULL AND customer_code != ''
            AND distributor_code IS NOT NULL
          GROUP BY distributor_code, suffix
        )
        SELECT distributor_code, suffix FROM ranked WHERE rn = 1
    """
    return {norm_key(r["distributor_code"]): norm_key(r["suffix"])
            for r in client.query(query).result()}


# ─── Product master ───────────────────────────────────────────────────────────
def load_products(credentials, project) -> dict:
    """``{normalised sku: {brand, product_name, pack_size}}``."""
    from google.cloud import bigquery

    client = bigquery.Client(credentials=credentials, project=project)
    query = """
        SELECT sku, brand, product_name, pack_size
        FROM `skintific-data-warehouse.gt_schema.master_product`
        WHERE sku IS NOT NULL AND sku != ''
    """
    return {
        norm_key(r["sku"]): {
            "brand": clean(r["brand"]).upper(),
            "product_name": clean(r["product_name"]),
            "pack_size": clean(r["pack_size"]),
        }
        for r in client.query(query).result()
    }


# ─── Login (shared credentials table) ─────────────────────────────────────────
# ─── Distributor accounts (BigQuery, bcrypt) ─────────────────────────────────
ACCOUNT_TABLE = "noo_sku_distributor_user"


def _account_table(project, dataset):
    return f"`{project}.{dataset}.{ACCOUNT_TABLE}`"


def load_account(credentials, project, dataset, distributor_code):
    """Fetch one account row. Returns None when absent or inactive.

    The password hash is returned so the caller can verify it in-process; it is
    never rendered, logged, or placed in session state.
    """
    from google.cloud import bigquery

    client = bigquery.Client(credentials=credentials, project=project)
    query = f"""
        SELECT distributor_code, distributor_name, password_hash,
               is_active, must_change_password, last_login_at
        FROM {_account_table(project, dataset)}
        WHERE UPPER(TRIM(distributor_code)) = @code
          AND is_active = TRUE
        LIMIT 1
    """
    job = client.query(query, job_config=bigquery.QueryJobConfig(
        query_parameters=[bigquery.ScalarQueryParameter(
            "code", "STRING", norm_key(distributor_code))]))
    rows = list(job.result())
    if not rows:
        return None
    row = rows[0]
    return {
        "distributor_code": clean(row["distributor_code"]),
        "distributor_name": clean(row["distributor_name"]),
        "password_hash": str(row["password_hash"] or ""),
        "is_active": bool(row["is_active"]),
        "must_change_password": bool(row["must_change_password"]),
        "last_login_at": row["last_login_at"],
    }


def set_password(credentials, project, dataset, distributor_code, password_hash):
    """Store a new bcrypt hash and clear the must-change flag."""
    from google.cloud import bigquery

    client = bigquery.Client(credentials=credentials, project=project)
    query = f"""
        UPDATE {_account_table(project, dataset)}
        SET password_hash = @hash,
            must_change_password = FALSE,
            updated_at = CURRENT_TIMESTAMP()
        WHERE UPPER(TRIM(distributor_code)) = @code
    """
    client.query(query, job_config=bigquery.QueryJobConfig(
        query_parameters=[
            bigquery.ScalarQueryParameter("hash", "STRING", password_hash),
            bigquery.ScalarQueryParameter("code", "STRING",
                                          norm_key(distributor_code))])).result()
    return True


def touch_last_login(credentials, project, dataset, distributor_code):
    """Best-effort login stamp. Never blocks a successful login."""
    from google.cloud import bigquery

    client = bigquery.Client(credentials=credentials, project=project)
    query = f"""
        UPDATE {_account_table(project, dataset)}
        SET last_login_at = CURRENT_TIMESTAMP()
        WHERE UPPER(TRIM(distributor_code)) = @code
    """
    try:
        client.query(query, job_config=bigquery.QueryJobConfig(
            query_parameters=[bigquery.ScalarQueryParameter(
                "code", "STRING", norm_key(distributor_code))])).result()
    except Exception:
        pass


def existing_account_codes(credentials, project, dataset):
    """Codes that already have an account, for idempotent seeding."""
    from google.cloud import bigquery

    client = bigquery.Client(credentials=credentials, project=project)
    query = (f"SELECT UPPER(TRIM(distributor_code)) AS code "
             f"FROM {_account_table(project, dataset)}")
    return {r["code"] for r in client.query(query).result()}


def insert_accounts(credentials, project, dataset, rows):
    """Append new account rows. Existing rows are never modified here."""
    from google.cloud import bigquery

    client = bigquery.Client(credentials=credentials, project=project)
    table = f"{project}.{dataset}.{ACCOUNT_TABLE}"
    errors = client.insert_rows_json(table, rows)
    if errors:
        raise RuntimeError(f"Gagal membuat akun: {errors[:3]}")
    return len(rows)


# ─── Distributor master (BigQuery) ────────────────────────────────────────────
def load_master_distributor(credentials, project) -> dict:
    """``{distributor_code: {...}}`` from gt_schema.master_distributor.

    Authoritative for distributor-level enrichment. Note this table lists only
    100 Active distributors against DIST DATABASE's 215, so it enriches but does
    NOT gate login — see load_distributors().
    """
    from google.cloud import bigquery

    client = bigquery.Client(credentials=credentials, project=project)
    query = """
        SELECT UPPER(TRIM(distributor_code)) AS distributor_code,
               distributor, distributor_company, region, region_g2g,
               asm, aom, pm, area_coverage, province, city, status,
               asm_skt, asm_tph, asm_fr,
               spv_skt, spv_tph, spv_fr,
               aom_skt, aom_tph, aom_fr
        FROM `skintific-data-warehouse.gt_schema.master_distributor`
        WHERE distributor_code IS NOT NULL AND distributor_code != ''
    """
    out = {}
    for r in client.query(query).result():
        out[r["distributor_code"]] = {k: clean(r[k]) for k in r.keys()}
    return out


# ─── Store master (BigQuery) ──────────────────────────────────────────────────
#: Fields pulled for store-level enrichment. Kept narrow on purpose: the pool
#: columns this feeds are the only ones BD Support actually uses.
_BASIS_FIELDS = (
    "cust_id", "store_name", "address", "city", "province", "area_coverage",
    "customer_type", "customer_category",
    "se_skt", "se_tph", "se_fcr",
    "spv_skt", "spv_tph",
    "aom_skt", "aom_tph",
    "reference_id_skt", "reference_id_tph",
)


def load_store_basis(credentials, project, distributor_codes=None) -> tuple:
    """Store master indexed two ways for the composite join.

    Returns ``(by_cust_id, by_reference)`` where ``by_reference`` is
    ``{brand_suffix: {reference_id: [records]}}``. A reference id with more than
    one record is kept as a list so the caller can detect ambiguity rather than
    silently taking the first.

    ``distributor_codes`` narrows the scan to the logged-in distributor, which
    keeps this cheap — the full table is 59k rows.
    """
    from google.cloud import bigquery

    client = bigquery.Client(credentials=credentials, project=project)
    where, params = "", []
    if distributor_codes:
        # 2026-09-03 fix: a store's reference_id can start with ANY of the
        # caller's authorised branch codes, not just the first one. Using a
        # single @prefix (codes[0]) here silently excluded matches for every
        # other branch and was the second half of the multi-branch NOO
        # Detector bug — see the EXISTS/UNNEST rewrite below.
        where = ("WHERE UPPER(TRIM(dst_id_skt)) IN UNNEST(@codes) "
                 "OR UPPER(TRIM(dst_id_tph)) IN UNNEST(@codes) "
                 "OR EXISTS(SELECT 1 FROM UNNEST(@codes) AS c "
                 "          WHERE STARTS_WITH(UPPER(TRIM(reference_id_skt)), c)) "
                 "OR EXISTS(SELECT 1 FROM UNNEST(@codes) AS c "
                 "          WHERE STARTS_WITH(UPPER(TRIM(reference_id_tph)), c))")
        codes = [norm_key(c) for c in distributor_codes]
        params = [bigquery.ArrayQueryParameter("codes", "STRING", codes)]
    query = f"SELECT {', '.join(_BASIS_FIELDS)} "             f"FROM `skintific-data-warehouse.gt_schema.master_store_database_basis` {where}"
    job = client.query(query, job_config=bigquery.QueryJobConfig(
        query_parameters=params)) if params else client.query(query)

    by_cust, by_ref = {}, {"skt": {}, "tph": {}, "fcr": {}}
    for r in job.result():
        record = {k: clean(r[k]) for k in r.keys()}
        cid = norm_key(record.get("cust_id"))
        if cid:
            by_cust[cid] = record
        for suffix, column in (("skt", "reference_id_skt"),
                               ("tph", "reference_id_tph"),
                               ("fcr", "reference_id_skt")):
            ref = norm_key(record.get(column))
            if ref:
                by_ref[suffix].setdefault(ref, []).append(record)
    return by_cust, by_ref


# ─── City reference (from the NOO template's own reference sheet) ─────────────
def load_city_reference(credentials, file_id=None) -> set:
    """Cities from the 'City & Store Type' sheet of BD Support's NOO template."""
    import openpyxl
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseDownload

    drive = build("drive", "v3", credentials=credentials, cache_discovery=False)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(
        buf, drive.files().get_media(fileId=file_id or config.NOO_TEMPLATE_FILE_ID,
                                     supportsAllDrives=True))
    done = False
    while not done:
        _, done = downloader.next_chunk()
    buf.seek(0)
    wb = openpyxl.load_workbook(buf, data_only=True, read_only=True)
    if "City & Store Type" not in wb.sheetnames:
        return set()
    ws = wb["City & Store Type"]
    return {clean(r[1]) for r in ws.iter_rows(min_row=2, values_only=True)
            if len(r) > 1 and clean(r[1])}


def load_local_sku_template() -> bytes:
    """The bundled SKU_MAPPING_TEMPLATE 2.0.xlsx — no network call.

    MoM 2026-09-03: BD Support handed over a fixed file rather than a live
    Drive link, and it is treated as the current source template, not a
    reference to inspect and reproduce. Raises FileNotFoundError with a clear
    message if the asset was not shipped with this deployment, rather than
    silently falling back to the outdated Drive-hosted template — serving the
    wrong SKU template is worse than failing loudly.
    """
    path = config.SKU_TEMPLATE_LOCAL_PATH
    if not path.is_file():
        raise FileNotFoundError(
            f"Template SKU tidak ditemukan di {path}. Pastikan file "
            "SKU_MAPPING_TEMPLATE 2.0.xlsx ikut ter-deploy bersama aplikasi."
        )
    return path.read_bytes()


def prepare_sku_template(raw: bytes) -> bytes:
    """Serve BD Support's SKU template without the removed gramasi column.

    MoM 31-Aug-2026 §6.1 removed "Product Size (ml/g)" from the SKU template,
    but the master file on Drive still carries it and belongs to BD Support, so
    we strip the column on the way out rather than editing their file. Keeps the
    downloadable template in step with the upload parser instead of asking
    admins for a field that is no longer used.

    Returns the original bytes unchanged if the column is already gone.
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        if config.SKU_SHEET_NAME not in wb.sheetnames:
            return raw
        ws = wb[config.SKU_SHEET_NAME]
        header = [clean(c.value) for c in ws[config.SKU_HEADER_ROW]]
        target = next((i for i, h in enumerate(header, start=1)
                       if h and "size" in h.lower()), None)
        if target is None:
            return raw
        ws.delete_cols(target)
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()
    except Exception:
        # A template we cannot rewrite is still better than no template.
        return raw


def download_template(credentials, file_id) -> bytes:
    """Fetch a template .xlsx from Drive so the app can serve it unchanged."""
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseDownload

    drive = build("drive", "v3", credentials=credentials, cache_discovery=False)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(
        buf, drive.files().get_media(fileId=file_id, supportsAllDrives=True))
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return buf.getvalue()


# ─── Submission ledger ────────────────────────────────────────────────────────
def _pool_rows_as_dicts(client, tab, headers):
    """Read a pool tab back as dicts keyed by its own header."""
    values = client.read_values(tab, "A2:BZ20000")
    out = []
    for row in values:
        if not any(clean(c) for c in row):
            continue
        out.append({h: (clean(row[i]) if i < len(row) else "")
                    for i, h in enumerate(headers)})
    return out


def _as_code_set(codes):
    """Accept a single code or an iterable of them."""
    if isinstance(codes, str):
        return {norm_key(codes)}
    return {norm_key(c) for c in (codes or []) if norm_key(c)}


def _ledger_from_pool(client, tab, headers, distributor_code, identity_fn,
                      content_fn):
    # A NOO upload may span several branches of one company, so the ledger has
    # to cover all of them - scoping to the login alone would miss a duplicate
    # submitted earlier under a sibling branch.
    want = _as_code_set(distributor_code)
    identities, contents = set(), set()
    for row in _pool_rows_as_dicts(client, tab, headers):
        if norm_key(row.get("customer_branch_code")) not in want:
            continue
        identities.add(identity_fn(row))
        contents.add(content_fn(row))
    return identities, contents


def load_noo_ledger(client, distributor_code) -> tuple:
    """Identities + content hashes this distributor has already submitted.

    The pool supplies both. The three brand trackers supply identities only — a
    store already promoted into the main tracker should read as a correction
    rather than as brand-new. They are never written to.
    """
    identities, contents = _ledger_from_pool(
        client, config.TAB_POOL_NOO, config.POOL_NOO_HEADERS, distributor_code,
        duplicates.noo_identity, duplicates.noo_content)

    want = _as_code_set(distributor_code)
    # Columns N and O of every brand tracker: Customer Branch Code, Customer
    # Store Code. Read-only.
    ranges = [f"'{t}'!N2:O10000" for t in config.TAB_NOO_MAIN.values()]
    for block in client.batch_read(ranges):
        for row in block:
            branch = norm_key(row[0]) if len(row) > 0 else ""
            store = norm_key(row[1]) if len(row) > 1 else ""
            if not store:
                continue
            if branch in want:
                identities.add(f"{branch}|{store}")
            else:
                for code in want:
                    if store.startswith(code):
                        identities.add(f"{code}|{store}")
                        break
    return identities, contents


def load_sku_ledger(client, distributor_code) -> tuple:
    """Identities + content hashes for SKU mappings.

    The main SKU MAPPING tab lacks some pool columns, so a comparable content
    hash cannot be rebuilt from it. History therefore contributes identities
    only: a mapping already in the main tracker classifies as a correction
    rather than an exact duplicate.
    """
    identities, contents = _ledger_from_pool(
        client, config.TAB_POOL_SKU, config.POOL_SKU_HEADERS, distributor_code,
        duplicates.sku_identity, duplicates.sku_content)

    want = _as_code_set(distributor_code)
    # F:L covers Customer Code .. Customer Product Name. Offsets within it:
    # 2 = H Product, 3 = I Distributor Code, 5 = K Customer Product Code.
    for row in client.read_values(config.TAB_SKU_MAPPING, "F2:L20000"):
        get = lambda i: clean(row[i]) if i < len(row) else ""  # noqa: E731
        code = norm_key(get(3))
        if code not in want:
            continue
        identities.add("|".join([code, norm_key(get(2)), norm_key(get(5))]))
    return identities, contents
