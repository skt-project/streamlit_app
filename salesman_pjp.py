import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
import re
import unicodedata
from google.oauth2 import service_account
from google.cloud import bigquery
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import FormulaRule

# ─── Page config (must be first) ──────────────────────────────────────────────

st.set_page_config(page_title="Salesman & PJP Template", page_icon="📋", layout="wide")

# ─── Distributor Passwords ────────────────────────────────────────────────────

DISTRIBUTOR_PASSWORDS = {
    "DST171": "5bcd0fc2",
    "DST157": "35b0e7bc",
    "DST152": "abf3d041",
    "DST109": "3a2e86c4",
    "DST160": "6f2fbafb",
    "DST057": "406ab114",
    "DST076": "7fc47ba9",
    "DST141": "b8e82723",
    "DST036": "6203ab21",
    "DST081": "dc0fc0b3",
    "DST196": "01f554e6",
    "DST197": "7d10ab68",
    "DST227": "3e5eec77",
    "DST108": "7202dd1d",
    "DST173": "de2168d4",
    "DST098": "3e2711e9",
    "DST250": "c41f74ce",
    "DST251": "474ebf78",
    "DST265": "f2ba48a4",
    "DST137": "a1869a0b",
    "DST192": "a81445ba",
    "DST193": "0759ffc9",
    "DST194": "c7410acf",
    "DST195": "3dea99f7",
    "DST204": "9ada7de4",
    "DST215": "bdeb210d",
    "DST216": "5bc07dbd",
    "DST217": "6fcae1a8",
    "DST218": "4921348e",
    "DST219": "a3bec96a",
    "DST220": "856f8f92",
    "DST221": "57688b82",
    "DST222": "131ea904",
    "DST223": "53904ab8",
    "DST224": "786dc184",
    "DST225": "041e1d37",
    "DST226": "1a6c6353",
    "DST101": "93b26366",
    "DST233": "dc291890",
    "DST236": "c1a10086",
    "DST237": "41ee140e",
    "DST238": "e7072f26",
    "DST239": "d4fe81e6",
    "DST240": "254d41a9",
    "DST241": "170d68cd",
    "DST242": "5dc64171",
    "DST243": "ad2c616e",
    "DST244": "6abe00f7",
    "DST245": "ce0e29df",
    "DST246": "6ff84751",
    "DST247": "92301c77",
    "DST248": "d8bc31bc",
    "DST249": "90019722",
    "DST143": "315ee2a1",
    "DST138": "4b51355b",
    "DST185": "797e3451",
    "DST252": "eddd74e5",
    "DST253": "9514af98",
    "DST254": "a25ceb29",
    "DST180": "6f414b90",
    "DST268": "63bf1aee",
    "DST269": "24dca8cf",
    "DST270": "ea146e45",
    "DST271": "416e4c21",
    "DST272": "bb461723",
    "DST181": "31f26be5",
    "DST182": "4772412e",
    "DST255": "6ec60126",
    "DST256": "a5b8497f",
    "DST257": "d08b956c",
    "DST258": "56a3516f",
    "DST259": "75d5bc79",
    "DST260": "3f8b3f24",
    "DST261": "c179ad32",
    "DST262": "e371419f",
    "DST263": "bffc53d7",
    "DST183": "e95a1d9f",
    "DST282": "997255a9",
    "DST283": "bb14d95b",
    "DST284": "da569e8f",
    "DST285": "5235c272",
    "DST286": "b1136924",
    "DST287": "b19a6be6",
    "DST288": "b868246a",
    "DST289": "649a174e",
    "DST290": "87c90fe4",
    "DST190": "d4bca582",
    "DST202": "c1711b29",
    "DST234": "29e438ab",
    "DST235": "e068abc5",
    "DST292": "0d2def3b",
    "DST305": "93443df2",
    "DST307": "b5cf6933",
    "DST308": "13500000",
    "DST310": "3b00aae1",
    "DST311": "a5b689d0",
    "DST312": "d06b143b",
    "DST313": "4008e3e0",
    "DST314": "69e270b7",
    "DST315": "88ab90fd",
    "DST324": "98f33e07",
    "DST326": "0453fc09",
    "DST327": "ca73a962",
    "DST328": "8d6c66c7",
    "DST329": "bee166e4",
    "DST330": "c66e5dda",
    "DST331": "2e1a0b81",
    "DST336": "b74ce041",
    "DST337": "c937e2c7",
    "DST338": "9eab97d2",
    "DST339": "1426bbbd",
    "DST340": "2094cfe8",
    "DST341": "62a19cf1",
    "DST342": "75b49aa5",
    "DST343": "2eb7f316",
    "DST344": "7303ada8",
    "DST345": "f3e22d41",
    "DST346": "0d2d6510",
    "DST347": "e5f29c41",
    "DST349": "b8e82723",
    "DST350": "98f33e07",
    "DST351": "12345678",
    "DST352": "12345678",
    "DST356": "1a2b3c4d",
    "DST363": "2b3c4d5e"
}

# ─── Input Period Deadline ─────────────────────────────────────────────────────
# Ubah tanggal ini setiap bulan sesuai jadwal. Format: datetime(YYYY, MM, DD)
INPUT_DEADLINE = datetime(2026, 9, 3).date()


def _get_password_for_distributor(dist_code: str) -> str | None:
    return DISTRIBUTOR_PASSWORDS.get(str(dist_code).strip().upper())


def _check_distributor_auth(dist_code: str) -> bool:
    key = f"auth_{dist_code}"
    return st.session_state.get(key, False)


def _render_password_gate(dist_code: str, dist_name: str) -> bool:
    key = f"auth_{dist_code}"
    if st.session_state.get(key, False):
        return True

    expected = _get_password_for_distributor(dist_code)

    st.markdown("---")
    st.markdown(
        f"""
        <div style='text-align:center; padding: 2rem 0 1rem 0;'>
            <span style='font-size:2.5rem'>🔒</span>
            <h3 style='margin:0.5rem 0 0.25rem 0;'>Akses Terkunci</h3>
            <p style='color:#888; margin:0;'>Masukkan password untuk distributor <b>{dist_name}</b></p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    col_l, col_c, col_r = st.columns([1, 2, 1])
    with col_c:
        entered = st.text_input(
            "Password Distributor",
            type="password",
            key=f"pw_input_{dist_code}",
            placeholder="••••••••",
        )
        if st.button(
            "🔓 Masuk",
            key=f"pw_btn_{dist_code}",
            type="primary",
            use_container_width=True,
        ):
            if expected is None:
                st.error(
                    "Password untuk distributor ini belum dikonfigurasi. Hubungi administrator."
                )
            elif entered == expected:
                st.session_state[key] = True
                st.rerun()
            else:
                st.error("Password salah. Silakan coba lagi.")

    return False


# ─── BigQuery credentials ─────────────────────────────────────────────────────


def get_credentials():
    try:
        gcp_secrets = st.secrets["connections"]["bigquery"]
        private_key = gcp_secrets["private_key"].replace("\\n", "\n")
        credentials = service_account.Credentials.from_service_account_info(
            {
                "type": gcp_secrets["type"],
                "project_id": gcp_secrets["project_id"],
                "private_key_id": gcp_secrets["private_key_id"],
                "private_key": private_key,
                "client_email": gcp_secrets["client_email"],
                "client_id": gcp_secrets["client_id"],
                "auth_uri": gcp_secrets["auth_uri"],
                "token_uri": gcp_secrets["token_uri"],
                "auth_provider_x509_cert_url": gcp_secrets[
                    "auth_provider_x509_cert_url"
                ],
                "client_x509_cert_url": gcp_secrets["client_x509_cert_url"],
            }
        )
        project_id = gcp_secrets["project_id"]
    except Exception:
        SERVICE_ACCOUNT_FILE = r"C:\Users\Bella Chelsea\Documents\skintific-data-warehouse-ea77119e2e7a.json"
        credentials = service_account.Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE
        )
        project_id = "skintific-data-warehouse"
    return credentials, project_id


# ─── BigQuery loaders ─────────────────────────────────────────────────────────


@st.cache_data(show_spinner="Memuat data distributor dari BigQuery...")
def load_distributor_data() -> pd.DataFrame:
    credentials, project_id = get_credentials()
    client = bigquery.Client(credentials=credentials, project=project_id)
    query = """
        SELECT
            UPPER(distributor)      AS distributor_name,
            UPPER(region_g2g)       AS region,
            UPPER(distributor_code) AS distributor_code,
            UPPER(asm_g2g)          AS asm
        FROM `gt_schema.master_distributor`
        WHERE region_g2g != '' AND status = 'Active'
    """
    df = client.query(query).to_dataframe()
    df["distributor_code"] = df["distributor_code"].astype(str).str.strip()
    df = df.drop_duplicates(subset=["distributor_code"]).reset_index(drop=True)
    return df


@st.cache_data(show_spinner="Memuat data toko dari Database...")
def load_store_master() -> pd.DataFrame:
    """
    Loads store master data directly from master_store_database_basis,
    sourcing distributor_code / region / asm straight off the store row
    (per the PJP redesign mapping table) instead of joining by distributor
    name. This is the single source of truth used to auto-populate
    Nama Toko / Region / ASM / Nama Distributor / Kode Distributor once a
    Kode Toko is chosen in the PJP flow.
    """
    credentials, project_id = get_credentials()
    client = bigquery.Client(credentials=credentials, project=project_id)
    query = """
        SELECT
            UPPER(cust_id)           AS store_code,
            UPPER(store_name)        AS store_name,
            UPPER(distributor_g2g)   AS distributor_name,
            UPPER(dst_id_g2g)  AS distributor_code,
            UPPER(region_g2g)        AS region,
            UPPER(asm_g2g)           AS asm
        FROM `gt_schema.master_store_database_basis`
        WHERE cust_id IS NOT NULL AND cust_id != ''
    """
    df = client.query(query).to_dataframe()
    df = df.dropna(subset=["store_code", "store_name", "distributor_code"])
    for c in ["store_code", "store_name", "distributor_name", "distributor_code", "region", "asm"]:
        df[c] = df[c].astype(str).str.strip()
    # Dropdown now shows the bare store code only (no "Kode - Nama" combo).
    df["store_label"] = df["store_code"]
    df = df.drop_duplicates(subset=["store_code"]).reset_index(drop=True)
    return df


# Backwards-compatible alias — older code referenced load_store_data().
load_store_data = load_store_master


@st.cache_data(show_spinner="Memuat daftar salesman aktif...")
def load_salesman_mapping(distributor_code: str) -> pd.DataFrame:
    """
    Loads the ACTIVE salesman roster for a given distributor straight from
    gt_salesman_mapping — the source of truth for salesman identity. Used
    to validate uploaded PJP files (salesman_id must belong to the selected
    distributor). NOTE: this distributor-scoped roster is no longer used to
    populate the Excel template's Salesman ID dropdown — see
    load_all_salesman_mapping() for that.
    """
    credentials, project_id = get_credentials()
    client = bigquery.Client(credentials=credentials, project=project_id)
    query = f"""
        SELECT
            salesman_id,
            salesman_type,
            UPPER(TRIM(distributor_code)) AS distributor_code,
            UPPER(TRIM(salesman))         AS salesman,
            is_active
        FROM `{MAPPING_TABLE}`
        WHERE UPPER(TRIM(distributor_code)) = UPPER(@kode)
          AND is_active = TRUE
        QUALIFY ROW_NUMBER() OVER (
            PARTITION BY salesman_id ORDER BY updated_at DESC
        ) = 1
        ORDER BY salesman
    """
    job_config = bigquery.QueryJobConfig(
        query_parameters=[
            bigquery.ScalarQueryParameter("kode", "STRING", distributor_code)
        ]
    )
    df = client.query(query, job_config=job_config).to_dataframe()
    df["salesman_id"] = df["salesman_id"].astype(str).str.strip()
    df["salesman"] = df["salesman"].astype(str).str.strip()
    # Dropdown now shows the bare Salesman ID only (no "ID - Nama" combo).
    df["salesman_label"] = df["salesman_id"]
    df = df.drop_duplicates(subset=["salesman_id"]).reset_index(drop=True)
    return df


@st.cache_data(show_spinner="Memuat seluruh Salesman ID aktif (semua distributor)...")
def load_all_salesman_mapping() -> pd.DataFrame:
    """
    Loads the ACTIVE salesman roster across ALL distributors from
    gt_salesman_mapping (no distributor filter). Used to populate the
    Salesman ID dropdown in the PJP Excel template so users can search and
    select any active Salesman ID regardless of which distributor it
    belongs to. Distributor ownership of the chosen Salesman ID is still
    enforced separately, at upload time, by validate_pjp_df() (using the
    distributor-scoped roster from load_salesman_mapping()).
    """
    credentials, project_id = get_credentials()
    client = bigquery.Client(credentials=credentials, project=project_id)
    query = f"""
        SELECT
            salesman_id,
            salesman_type,
            UPPER(TRIM(distributor_code)) AS distributor_code,
            UPPER(TRIM(salesman))         AS salesman,
            is_active
        FROM `{MAPPING_TABLE}`
        WHERE is_active = TRUE
        QUALIFY ROW_NUMBER() OVER (
            PARTITION BY salesman_id ORDER BY updated_at DESC
        ) = 1
        ORDER BY salesman
    """
    df = client.query(query).to_dataframe()
    df["salesman_id"] = df["salesman_id"].astype(str).str.strip()
    df["salesman"] = df["salesman"].astype(str).str.strip()
    # Dropdown now shows the bare Salesman ID only (no "ID - Nama" combo).
    df["salesman_label"] = df["salesman_id"]
    df = df.drop_duplicates(subset=["salesman_id"]).reset_index(drop=True)
    return df


def build_salesman_lookup(mapping_df: pd.DataFrame) -> dict:
    """{salesman_id: salesman_name}"""
    if mapping_df is None or mapping_df.empty:
        return {}
    return dict(zip(mapping_df["salesman_id"], mapping_df["salesman"]))


def build_store_lookup(store_df: pd.DataFrame) -> dict:
    """
    {store_code: {"store_name", "region", "asm", "nama_distributor", "kode_distributor"}}
    """
    if store_df is None or store_df.empty:
        return {}
    lookup = {}
    for _, r in store_df.iterrows():
        lookup[r["store_code"]] = {
            "store_name": r["store_name"],
            "region": r["region"],
            "asm": r["asm"],
            "nama_distributor": r["distributor_name"],
            "kode_distributor": r["distributor_code"],
        }
    return lookup


def build_lookup_tables(dist_df: pd.DataFrame):
    distributor_map = dict(zip(dist_df["distributor_code"], dist_df["distributor_name"]))
    asm_options = sorted(dist_df["asm"].dropna().unique().tolist())
    region_options = sorted(dist_df["region"].dropna().unique().tolist())
    return distributor_map, asm_options, region_options


# ─── Salesman Mapping table helpers ──────────────────────────────────────────

MAPPING_TABLE = "skintific-data-warehouse.gt_schema.gt_salesman_mapping"
SALESMAN_TABLE = "skintific-data-warehouse.gt_schema.gt_master_salesman"
PJP_TABLE = "skintific-data-warehouse.gt_schema.gt_master_salesman_pjp"

SALESMAN_TYPES = ["GTI", "MIX", "MTI"]


@st.cache_data(show_spinner=False)
def get_salesman_list(distributor_code: str) -> pd.DataFrame:
    try:
        credentials, project_id = get_credentials()
        client = bigquery.Client(credentials=credentials, project=project_id)
        query = f"""
            WITH ranked_salesman AS (
                SELECT
                    *,
                    ROW_NUMBER() OVER (
                        PARTITION BY
                            UPPER(TRIM(nama_salesman)),
                            UPPER(TRIM(kode_distributor))
                        ORDER BY uploaded_at DESC
                    ) AS rn
                FROM `{SALESMAN_TABLE}`
            )
            SELECT
                m.salesman_id,
                m.salesman_type,
                m.distributor_code,
                m.salesman,
                m.is_active,
                m.created_at,
                m.updated_at,
                s.nama_salesman,
                s.no_hp,
                s.status_salesman,
                s.region,
                s.asm
            FROM `{MAPPING_TABLE}` m
            LEFT JOIN ranked_salesman s
                ON  UPPER(TRIM(m.salesman))        = UPPER(TRIM(s.nama_salesman))
                AND UPPER(TRIM(m.distributor_code)) = UPPER(TRIM(s.kode_distributor))
                AND s.rn = 1
            WHERE UPPER(m.distributor_code) = UPPER(@kode)
            ORDER BY m.salesman_id, m.created_at DESC
        """
        job_config = bigquery.QueryJobConfig(
            query_parameters=[
                bigquery.ScalarQueryParameter("kode", "STRING", distributor_code)
            ]
        )
        df = client.query(query, job_config=job_config).to_dataframe()
        return df
    except Exception as e:
        st.error(f"Gagal memuat daftar salesman: {e}")
        return pd.DataFrame()


def get_pjp_list(
    distributor_code: str = None, salesman_name: str = None
) -> pd.DataFrame:
    try:
        credentials, project_id = get_credentials()
        client = bigquery.Client(credentials=credentials, project=project_id)

        conditions = []
        params = []

        if distributor_code:
            conditions.append("UPPER(kode_distributor) = UPPER(@kode)")
            params.append(
                bigquery.ScalarQueryParameter("kode", "STRING", distributor_code)
            )
        if salesman_name:
            conditions.append("UPPER(TRIM(nama_salesman)) = UPPER(TRIM(@salesman))")
            params.append(
                bigquery.ScalarQueryParameter("salesman", "STRING", salesman_name)
            )

        where_clause = "WHERE " + " AND ".join(conditions) if conditions else ""

        query = f"""
            SELECT *
            FROM `{PJP_TABLE}`
            {where_clause}
            ORDER BY nama_salesman, hari
        """
        job_config = bigquery.QueryJobConfig(query_parameters=params)
        df = client.query(query, job_config=job_config).to_dataframe()
        return df
    except Exception as e:
        st.error(f"Gagal memuat data PJP: {e}")
        return pd.DataFrame()


def get_latest_running_number(distributor_code: str, salesman_type: str) -> int:
    try:
        credentials, project_id = get_credentials()
        client = bigquery.Client(credentials=credentials, project=project_id)
        prefix = f"{salesman_type}{distributor_code}"
        query = f"""
            SELECT MAX(CAST(SUBSTR(salesman_id, {len(prefix) + 1}) AS INT64)) AS max_num
            FROM `{MAPPING_TABLE}`
            WHERE UPPER(distributor_code) = UPPER(@kode)
              AND UPPER(salesman_type)    = UPPER(@stype)
              AND STARTS_WITH(salesman_id, @prefix)
        """
        job_config = bigquery.QueryJobConfig(
            query_parameters=[
                bigquery.ScalarQueryParameter("kode", "STRING", distributor_code),
                bigquery.ScalarQueryParameter("stype", "STRING", salesman_type),
                bigquery.ScalarQueryParameter("prefix", "STRING", prefix),
            ]
        )
        result = client.query(query, job_config=job_config).to_dataframe()
        max_num = result["max_num"].iloc[0]
        return int(max_num) if pd.notna(max_num) else 0
    except Exception:
        return 0


def generate_salesman_id(distributor_code: str, salesman_type: str) -> str:
    latest = get_latest_running_number(distributor_code, salesman_type)
    next_num = latest + 1
    return f"{salesman_type}{distributor_code}{str(next_num).zfill(3)}"


def insert_salesman_record(salesman_data: dict) -> tuple[bool, str]:
    try:
        credentials, project_id = get_credentials()
        client = bigquery.Client(credentials=credentials, project=project_id)

        row = {**salesman_data, "uploaded_at": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")}
        bq_df = pd.DataFrame([row])

        for col in ["tanggal_lahir", "tanggal_join_g2g"]:
            if col in bq_df.columns:
                bq_df[col] = pd.to_datetime(bq_df[col], errors="coerce").dt.tz_localize("UTC")

        job_config = bigquery.LoadJobConfig(write_disposition="WRITE_APPEND", autodetect=True)
        client.load_table_from_dataframe(bq_df, SALESMAN_TABLE, job_config=job_config).result()
        return True, ""
    except Exception as e:
        return False, str(e)


def insert_mapping_record(
    salesman_id: str,
    distributor_code: str,
    salesman_type: str,
    nama_salesman: str = "",
) -> tuple[bool, str]:
    try:
        credentials, project_id = get_credentials()
        client = bigquery.Client(credentials=credentials, project=project_id)
        now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
        row = {
            "salesman_id": salesman_id,
            "salesman_type": salesman_type,
            "distributor_code": distributor_code,
            "salesman": sanitize_salesman_name(nama_salesman) if nama_salesman else "",
            "is_active": True,
            "created_at": now,
            "updated_at": now,
        }
        job_config = bigquery.LoadJobConfig(write_disposition="WRITE_APPEND", autodetect=True)
        client.load_table_from_dataframe(
            pd.DataFrame([row]), MAPPING_TABLE, job_config=job_config
        ).result()
        return True, ""
    except Exception as e:
        return False, str(e)


def update_salesman_record(
    nama_salesman: str, distributor_code: str, updated_fields: dict
) -> tuple[bool, str]:
    try:
        credentials, project_id = get_credentials()
        client = bigquery.Client(credentials=credentials, project=project_id)

        allowed = {
            "nama_salesman": "STRING",
            "nama_spv_external": "STRING",
            "nama_spv_internal": "STRING",
            "nama_spv_internal_2": "STRING",
            "status_salesman": "STRING",
            "total_outlet_coverage_pjp": "INT64",
            "gaji_pokok": "INT64",
            "tunjangan_dan_insentif": "INT64",
            "tanggal_lahir": "TIMESTAMP",
            "jenis_kelamin": "STRING",
            "pendidikan_terakhir": "STRING",
            "pengalaman_bulan": "INT64",
            "principal_lain": "STRING",
            "no_hp": "STRING",
            "tanggal_join_g2g": "TIMESTAMP",
        }

        set_clauses = []
        params = [
            bigquery.ScalarQueryParameter("target_nama", "STRING", nama_salesman),
            bigquery.ScalarQueryParameter("target_dist", "STRING", distributor_code),
        ]

        for field, value in updated_fields.items():
            if field not in allowed:
                continue
            param_name = f"p_{field}"
            bq_type = allowed[field]

            if bq_type == "INT64":
                try:
                    value = int(value) if value is not None and not (isinstance(value, float) and (pd.isna(value) or value != value)) else 0
                except (TypeError, ValueError):
                    value = 0

            elif bq_type == "FLOAT64":
                try:
                    value = float(value) if value is not None and not (isinstance(value, float) and (pd.isna(value) or value != value)) else 0.0
                except (TypeError, ValueError):
                    value = 0.0

            elif bq_type == "TIMESTAMP" and value is not None:
                try:
                    value = pd.to_datetime(value).strftime("%Y-%m-%dT00:00:00")
                except Exception:
                    pass

            elif bq_type == "STRING":
                if value is None:
                    pass
                else:
                    try:
                        if pd.isna(value):
                            value = None
                    except (TypeError, ValueError):
                        pass

            set_clauses.append(f"{field} = @{param_name}")
            params.append(bigquery.ScalarQueryParameter(param_name, bq_type, value))

        if not set_clauses:
            return False, "Tidak ada field yang diupdate."

        now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
        params.append(bigquery.ScalarQueryParameter("updated_at", "STRING", now))

        query = f"""
            UPDATE `{SALESMAN_TABLE}`
            SET {", ".join(set_clauses)},
                uploaded_at = @updated_at
            WHERE UPPER(TRIM(nama_salesman))     = UPPER(TRIM(@target_nama))
              AND UPPER(TRIM(kode_distributor))  = UPPER(TRIM(@target_dist))
        """
        job_config = bigquery.QueryJobConfig(query_parameters=params)
        client.query(query, job_config=job_config).result()
        return True, ""
    except Exception as e:
        return False, str(e)


def deactivate_previous_mapping(salesman_id: str) -> tuple[bool, str]:
    try:
        credentials, project_id = get_credentials()
        client = bigquery.Client(credentials=credentials, project=project_id)
        now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
        query = f"""
            UPDATE `{MAPPING_TABLE}`
            SET is_active  = FALSE,
                updated_at = @updated_at
            WHERE salesman_id = @sid
              AND is_active   = TRUE
        """
        job_config = bigquery.QueryJobConfig(
            query_parameters=[
                bigquery.ScalarQueryParameter("sid", "STRING", salesman_id),
                bigquery.ScalarQueryParameter("updated_at", "STRING", now),
            ]
        )
        client.query(query, job_config=job_config).result()
        return True, ""
    except Exception as e:
        return False, str(e)


def deactivate_salesman_mapping(salesman_id: str) -> tuple[bool, str]:
    return deactivate_previous_mapping(salesman_id)


# ─── Static option lists ──────────────────────────────────────────────────────

STATUS_OPTIONS = ["Mix", "Eksklusif"]
GENDER_OPTIONS = ["Male", "Female"]
EDUCATION_OPTIONS = ["SD", "SMP", "SMA", "S1", "S2"]

# ─── Frekuensi (Column I) / Hari (Column J) / Hari Minggu (Column K) /
# Nomor Minggu (Column L, -> DB `callcycle`) ────────────────────────────────
# Pure logic lives in pjp_hari_minggu.py (no Streamlit/BigQuery dependency)
# so it can be unit tested directly — see tests/test_pjp_hari_minggu.py.
# Frekuensi DRIVES both Hari's allowed day-count and callcycle's allowed
# values — see that module's docstring for the full rule table. Hari is
# SENIN..SABTU only ("MINGGU"/Sunday is never valid for this template).
from pjp_hari_minggu import (  # noqa: E402
    FREKUENSI_OPTIONS,
    FREKUENSI_RANGE_SUFFIX,
    HARI_CANONICAL_ORDER,
    HARI_COMBOS_BY_FREKUENSI,
    HARI_SEPARATOR,
    MINGGU_COL,
    MINGGU_COL_ALIASES,
    MINGGU_GANJIL,
    MINGGU_GANJIL_GENAP,
    MINGGU_GENAP,
    MINGGU_OPTIONS_BY_FREKUENSI,
    MINGGU_RANGE_SUFFIX,
    KET_MINGGU_COL,
    KET_MINGGU_COL_ALIASES,
    KET_OPTIONS_BY_FREKUENSI_MINGGU,
    LEGACY_MINGGU_MAP,
    auto_callcycle,
    derive_minggu_from_callcycle,
    hari_options_for_frekuensi,
    ket_minggu_options,
    migrate_legacy_minggu,
    minggu_options_for_frekuensi,
    normalize_callcycle,
    normalize_hari,
    normalize_minggu,
)

DAY_OPTIONS = HARI_CANONICAL_ORDER   # name kept for backward-compat call sites
FREQUENCY_OPTIONS = FREKUENSI_OPTIONS  # name kept for backward-compat call sites

SALESMAN_COLS = [
    ("Nama Salesman", True, "text"),
    ("Nama SPV External", False, "text"),
    ("Nama SPV Internal", True, "text"),
    ("Nama SPV Internal 2", False, "text"),
    ("ASM", True, "cascade"),
    ("Region", True, "cascade"),
    ("Nama Distributor", True, "cascade"),
    ("Kode Distributor", True, "auto"),
    ("Status Salesman", True, "dropdown"),
    ("Total Outlet Coverage PJP", True, "numeric"),
    ("Gaji Pokok", True, "numeric"),
    ("Tunjangan dan insentif", True, "numeric"),
    ("Tanggal Lahir", True, "date"),
    ("Jenis Kelamin", True, "dropdown"),
    ("Pendidikan Terakhir", True, "dropdown"),
    ("Pengalaman di Perusahaan Sebelumnya (Dalam Bulan)", True, "numeric"),
    ("Principal Lain yang Ditanggungjawabi", False, "text"),
    ("No. HP", True, "text"),
    ("Tanggal Join di G2G", True, "date"),
]

# ─── PJP columns: Salesman-ID-first flow ──────────────────────────────────────
# ASM / Region / Nama Distributor / Kode Distributor are constant for the
# whole file (1 file = 1 distributor, enforced by scoping the Salesman ID /
# Kode Toko dropdowns to `selected_dist_code`), so they're pre-filled
# directly from the selected distributor rather than looked up per row.
# Step 1: user picks Salesman ID -> Step 2: Nama Salesman auto-fills (read-only)
# Step 3: user picks Kode Toko   -> Step 4: Nama Toko auto-fills (read-only)
# Step 5: user picks Frekuensi (I) -> drives which Hari (J) / Nomor Minggu (L)
#         options are offered; Hari Minggu (K) is derived/display-only from L.
PJP_COLS = [
    ("ASM", False, "auto"),
    ("Region", False, "auto"),
    ("Nama Distributor", False, "auto"),
    ("Kode Distributor", False, "auto"),
    ("Salesman ID", True, "salesman_dropdown"),
    ("Nama Salesman", False, "auto"),
    ("Kode Toko", True, "store_dropdown"),
    ("Nama Toko", False, "auto"),
    ("Frekuensi", True, "dropdown"),
    ("Hari", True, "dependent_dropdown"),
    (MINGGU_COL, True, "dependent_dropdown"),
    (KET_MINGGU_COL, True, "dependent_dropdown"),
]

SALESMAN_REQUIRED = [c for c, r, _ in SALESMAN_COLS if r]
PJP_REQUIRED = [c for c, r, _ in PJP_COLS if r]


# ─── Named-range key sanitiser ────────────────────────────────────────────────


def _safe_name(text: str) -> str:
    nfkd = unicodedata.normalize("NFKD", text)
    ascii_s = nfkd.encode("ascii", "ignore").decode("ascii")
    cleaned = re.sub(r"[^A-Za-z0-9]", "_", ascii_s)
    if not cleaned or cleaned[0].isdigit():
        cleaned = "_" + cleaned
    return ("NR_" + cleaned)[:255]


def _indirect_clean(cell_ref: str) -> str:
    special = [" ", "-", "/", "(", ")", "+", "&", ".", "'"]
    expr = cell_ref
    for ch in special:
        expr = f'SUBSTITUTE({expr},"{ch}","_")'
    return expr


# ─── Style helpers ────────────────────────────────────────────────────────────


def _thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _header_font():
    return Font(bold=True, color="FFFFFF", size=10, name="Calibri")


def _note_font():
    return Font(italic=True, color="808080", size=9, name="Calibri")


def _req_font():
    return Font(bold=True, color="C00000", size=9, name="Calibri")


def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _vcenter(wrap=False):
    return Alignment(vertical="center", wrap_text=wrap)


# ─── Build Lookup sheet + named ranges (PJP: Salesman ID + Kode Toko driven) ──

# Legacy separator kept only so _extract_combo_key() can still gracefully
# parse older uploaded files that used the "KEY - Label" combo format. New
# templates write the bare key (Salesman ID / Kode Toko) directly, with no
# separator, so this is a fallback path only.
_COMBO_SEP = " - "


def _build_lookup_and_named_ranges(wb, salesman_df, store_df):
    """
    Builds a hidden 'Lookup' sheet containing:
      - NR_SALESMAN_COMBO : list of bare salesman_id values for the
                             Salesman ID dropdown. `salesman_df` is expected
                             to be the FULL active roster across ALL
                             distributors (see load_all_salesman_mapping()) —
                             this is no longer scoped to a single
                             distributor, so users can search/select any
                             active Salesman ID from any distributor.
      - NR_SALESMAN_LOOKUP: 2-col table (salesman_id, salesman) for VLOOKUP.
      - NR_STORE_COMBO     : list of bare cust_id (store code) values for the
                              Kode Toko dropdown. `store_df` is expected to
                              be the FULL store master across ALL
                              distributors, likewise unscoped.
      - NR_STORE_LOOKUP    : 6-col table (cust_id, store_name, region, asm,
                              nama_distributor, kode_distributor) for VLOOKUP.

    NOTE: Distributor ownership of whatever Salesman ID / Kode Toko the user
    ends up choosing is enforced separately at upload time by
    validate_pjp_df(), not here.
    """
    LK = "Lookup"
    lk = wb.create_sheet(LK)
    lk.sheet_state = "hidden"

    # ── Salesman combo list ────────────────────────────────────────────────
    sal_df = salesman_df.sort_values("salesman") if salesman_df is not None and not salesman_df.empty else pd.DataFrame(columns=["salesman_id", "salesman", "salesman_label"])

    combo_col = 1
    lk.cell(row=1, column=combo_col, value="__SALESMAN_COMBO__")
    for i, label in enumerate(sal_df["salesman_label"].tolist(), start=2):
        lk.cell(row=i, column=combo_col, value=label)
    c = get_column_letter(combo_col)
    last_row = max(2, 1 + len(sal_df))
    wb.defined_names["NR_SALESMAN_COMBO"] = DefinedName(
        "NR_SALESMAN_COMBO", attr_text=f"'{LK}'!${c}$2:${c}${last_row}"
    )

    id_col = combo_col + 1
    name_col = combo_col + 2
    lk.cell(row=1, column=id_col, value="__SALESMAN_ID__")
    lk.cell(row=1, column=name_col, value="__SALESMAN_NAME__")
    for i, r in enumerate(sal_df.itertuples(index=False), start=2):
        lk.cell(row=i, column=id_col, value=r.salesman_id)
        lk.cell(row=i, column=name_col, value=r.salesman)
    ic = get_column_letter(id_col)
    nc = get_column_letter(name_col)
    last_row = max(2, 1 + len(sal_df))
    wb.defined_names["NR_SALESMAN_LOOKUP"] = DefinedName(
        "NR_SALESMAN_LOOKUP", attr_text=f"'{LK}'!${ic}$2:${nc}${last_row}"
    )

    # ── Store combo list + multi-column lookup table ───────────────────────
    st_df = store_df.sort_values("store_code") if store_df is not None and not store_df.empty else pd.DataFrame(
        columns=["store_code", "store_name", "region", "asm", "distributor_name", "distributor_code", "store_label"]
    )

    store_combo_col = name_col + 1
    lk.cell(row=1, column=store_combo_col, value="__STORE_COMBO__")
    for i, label in enumerate(st_df["store_label"].tolist(), start=2):
        lk.cell(row=i, column=store_combo_col, value=label)
    sc = get_column_letter(store_combo_col)
    last_row = max(2, 1 + len(st_df))
    wb.defined_names["NR_STORE_COMBO"] = DefinedName(
        "NR_STORE_COMBO", attr_text=f"'{LK}'!${sc}$2:${sc}${last_row}"
    )

    lut_start = store_combo_col + 1  # cust_id
    cols_order = ["store_code", "store_name", "region", "asm", "distributor_name", "distributor_code"]
    headers = ["__STORE_ID__", "__STORE_NAME__", "__STORE_REGION__", "__STORE_ASM__", "__STORE_DIST__", "__STORE_DISTCODE__"]
    for offset, (col_key, header) in enumerate(zip(cols_order, headers)):
        col_idx = lut_start + offset
        lk.cell(row=1, column=col_idx, value=header)
        for i, val in enumerate(st_df[col_key].tolist(), start=2):
            lk.cell(row=i, column=col_idx, value=val)
    lut_first_letter = get_column_letter(lut_start)
    lut_last_letter = get_column_letter(lut_start + len(cols_order) - 1)
    last_row = max(2, 1 + len(st_df))
    wb.defined_names["NR_STORE_LOOKUP"] = DefinedName(
        "NR_STORE_LOOKUP", attr_text=f"'{LK}'!${lut_first_letter}$2:${lut_last_letter}${last_row}"
    )

    # ── Frekuensi-dependent Hari / Nomor Minggu (callcycle) dropdowns ───────
    # Column J (Hari) and Column L (Nomor Minggu) are DEPENDENT dropdowns
    # keyed off Column I (Frekuensi) — see FREKUENSI_RANGE_SUFFIX below and
    # _attach_pjp_dvs()'s INDIRECT("NR_HARI_"&...) / INDIRECT("NR_CALLCYCLE_"
    # &...) formulas. Every valid combination for each Frekuensi is
    # pre-rendered here in canonical form, so picking from the dropdown
    # always yields an already-correctly-formatted, Frekuensi-consistent
    # value — for F4/F4+ the "dropdown" has exactly one option (the fixed
    # callcycle), which is the native-Excel (no VBA) way to make that field
    # effectively non-editable while keeping every row's DV mechanism
    # uniform (see module docstring in pjp_hari_minggu.py for the F4/F4+
    # "no manual choice" rule).
    next_col = lut_start + len(cols_order)

    def _write_combo_column(values, tag):
        nonlocal next_col
        col_idx = next_col
        next_col += 1
        lk.cell(row=1, column=col_idx, value=tag)
        for i, val in enumerate(values, start=2):
            lk.cell(row=i, column=col_idx, value=val)
        letter = get_column_letter(col_idx)
        last = max(2, 1 + len(values))
        return f"'{LK}'!${letter}$2:${letter}${last}"

    wb.defined_names["NR_FREKUENSI_COMBO"] = DefinedName(
        "NR_FREKUENSI_COMBO", attr_text=_write_combo_column(FREKUENSI_OPTIONS, "__FREKUENSI_COMBO__")
    )
    # Column J — one day-combination list per Frekuensi.
    for freq, suffix in FREKUENSI_RANGE_SUFFIX.items():
        name = f"NR_HARI_{suffix}"
        wb.defined_names[name] = DefinedName(
            name, attr_text=_write_combo_column(HARI_COMBOS_BY_FREKUENSI[freq], f"__HARI_{suffix}__")
        )
    # Column K — one Minggu (week-parity) list per Frekuensi. F1/F2 offer
    # Ganjil/Genap; F4/F4+ offer only "Minggu Ganjil + Genap" (a one-entry
    # list is how "automatic + locked" is expressed without VBA).
    for freq, suffix in FREKUENSI_RANGE_SUFFIX.items():
        name = f"NR_MINGGU_{suffix}"
        wb.defined_names[name] = DefinedName(
            name, attr_text=_write_combo_column(MINGGU_OPTIONS_BY_FREKUENSI[freq], f"__MINGGU_{suffix}__")
        )
    # Column L — one Ket. Minggu list per (Frekuensi, Minggu) pair. F1 gets
    # a real 2-option choice (Ganjil -> 1|3, Genap -> 2|4); F2/F4/F4+ get a
    # single automatic value.
    for (freq, minggu), options in KET_OPTIONS_BY_FREKUENSI_MINGGU.items():
        suffix = f"{FREKUENSI_RANGE_SUFFIX[freq]}_{MINGGU_RANGE_SUFFIX[minggu]}"
        name = f"NR_KET_{suffix}"
        wb.defined_names[name] = DefinedName(
            name, attr_text=_write_combo_column(options, f"__KET_{suffix}__")
        )


# ─── Attach Salesman ID / Kode Toko dropdown DVs ──────────────────────────────


def _attach_pjp_dvs(ws, col_names, first_data, last_data):
    def cl(name):
        return get_column_letter(col_names.index(name) + 1)

    def sqref(name):
        c = cl(name)
        return f"{c}{first_data}:{c}{last_data}"

    dv_sal = DataValidation(
        type="list",
        formula1="NR_SALESMAN_COMBO",
        allow_blank=True,
        showInputMessage=True,
        promptTitle="Langkah 1 - Salesman ID",
        prompt="Pilih Salesman ID dari daftar (seluruh distributor). Nama Salesman akan terisi otomatis. Kepemilikan distributor akan divalidasi saat upload.",
        showErrorMessage=True,
        errorTitle="Input Tidak Valid",
        error="Pilih Salesman ID dari daftar dropdown.",
    )
    ws.add_data_validation(dv_sal)
    dv_sal.sqref = sqref("Salesman ID")

    dv_store = DataValidation(
        type="list",
        formula1="NR_STORE_COMBO",
        allow_blank=True,
        showInputMessage=True,
        promptTitle="Langkah 2 - Kode Toko",
        prompt="Pilih Kode Toko dari daftar (seluruh distributor). Nama Toko, Region, ASM, dan Distributor akan terisi otomatis. Kepemilikan distributor akan divalidasi saat upload.",
        showErrorMessage=True,
        errorTitle="Input Tidak Valid",
        error="Pilih Kode Toko dari daftar dropdown.",
    )
    ws.add_data_validation(dv_store)
    dv_store.sqref = sqref("Kode Toko")


# ─── PJP Excel ────────────────────────────────────────────────────────────────

# Auto (formula-driven, protected, read-only) columns and which NR_..._LOOKUP
# column index (1-based, within the lookup table) each one pulls from, keyed
# by which driving/input column it depends on.
_AUTO_FROM_SALESMAN = {
    "Nama Salesman": 2,  # NR_SALESMAN_LOOKUP col 2 = salesman name
}
_AUTO_FROM_STORE = {
    "Nama Toko": 2,  # NR_STORE_LOOKUP col 2 = store_name
    "Region": 3,  # NR_STORE_LOOKUP col 3 = region
    "ASM": 4,  # NR_STORE_LOOKUP col 4 = asm
    "Nama Distributor": 5,  # NR_STORE_LOOKUP col 5 = distributor_name
    "Kode Distributor": 6,  # NR_STORE_LOOKUP col 6 = distributor_code
}
# Kept for backwards-compat readability in a couple of comments/docstrings
# below; ASM / Region / Nama Distributor / Kode Distributor are now looked
# up per-row via VLOOKUP off Kode Toko (see _AUTO_FROM_STORE above) instead
# of being written once as static values.
_DIST_CONST_COLS = []


def _extract_key_formula(cell_ref: str) -> str:
    """
    The Salesman ID / Kode Toko cell now holds the bare key value directly
    (e.g. "GTIDST171001" or "ST00123") — no "KEY - Label" combo string — so
    the VLOOKUP key is just the trimmed cell value itself.
    """
    return f"TRIM({cell_ref})"


def create_pjp_excel(
    df,
    salesman_df,
    store_df,
    selected_dist_code,
    selected_dist_name,
    selected_dist_asm,
    selected_dist_region,
) -> BytesIO:
    """
    Builds the PJP Template workbook.

    Column order:
    ASM
    Region
    Nama Distributor
    Kode Distributor
    Salesman ID
    Nama Salesman
    Kode Toko
    Nama Toko
    Frekuensi       (I) drives J, K, L
    Hari            (J) day count set by Frekuensi
    Minggu          (K) week parity; options set by Frekuensi
    Ket. Minggu     (L) -> DB `callcycle`; options set by (I, K)

    Flow implemented in the sheet:
      Step 1 — user picks "Salesman ID" from a dropdown of bare salesman_id
               values sourced from the FULL active gt_salesman_mapping
               roster across ALL distributors (`salesman_df` is expected to
               be unscoped — see load_all_salesman_mapping()). The chosen
               Salesman ID is NOT restricted to `selected_dist_code` here;
               ownership is enforced later at upload-validation time.
      Step 2 — "Nama Salesman" auto-fills (read-only) via VLOOKUP.
      Step 3 — user picks "Kode Toko" from a dropdown of bare cust_id
               (store code) values sourced from the FULL store master
               across ALL distributors (`store_df` is expected to be
               unscoped). Likewise not restricted to `selected_dist_code`
               here.
      Step 4 — "Nama Toko" (and Region / ASM / Nama Distributor / Kode
               Distributor) auto-fill (read-only) via VLOOKUP based on
               whichever store was picked.

    `selected_dist_code` / `selected_dist_name` / `selected_dist_asm` /
    `selected_dist_region` are kept as parameters for context/labeling
    purposes only — they no longer filter `salesman_df` / `store_df`, which
    the caller is expected to pass in as the FULL, unfiltered datasets.
    Distributor ownership of the row's actual Salesman ID / Kode Toko is
    validated separately at upload time (see validate_pjp_df()).
    """
    wb = Workbook()
    wb.remove(wb.active)
    _build_lookup_and_named_ranges(wb, salesman_df, store_df)

    col_names = [c for c, _, _ in PJP_COLS]
    col_req = {c: r for c, r, _ in PJP_COLS}

    FIRST_DATA = 4
    LAST_DATA = 30003

    AUTO_COLS = set(_AUTO_FROM_SALESMAN) | set(_AUTO_FROM_STORE) | set(_DIST_CONST_COLS)
    DROPDOWN_COLS = {"Salesman ID", "Kode Toko", "Frekuensi", "Hari", MINGGU_COL, KET_MINGGU_COL}

    notes_pjp = {
        "ASM": "Otomatis terisi dari Kode Toko",
        "Region": "Otomatis terisi dari Kode Toko",
        "Nama Distributor": "Otomatis terisi dari Kode Toko",
        "Kode Distributor": "Otomatis terisi dari Kode Toko",
        "Salesman ID": "Langkah 1 - Pilih Salesman ID (dari seluruh distributor). Kepemilikan distributor divalidasi saat upload.",
        "Nama Salesman": "Otomatis terisi dari Salesman ID",
        "Kode Toko": "Langkah 2 - Pilih Kode Toko (dari seluruh distributor). Kepemilikan distributor divalidasi saat upload.",
        "Nama Toko": "Otomatis terisi dari Kode Toko",
        "Frekuensi": (
            "Langkah 3 - Pilih Frekuensi terlebih dahulu: F1 = 1x/bulan | "
            "F2 = 2x/bulan | F4 = setiap minggu (tanpa minggu ke-5) | "
            "F4+ = setiap minggu (+ minggu ke-5 bila ada). Menentukan pilihan "
            "Hari dan Nomor Minggu di kolom berikutnya."
        ),
        "Hari": (
            "Langkah 4 - Pilih Frekuensi (kolom I) dahulu. Jumlah hari yang "
            "boleh dipilih: F1=1 hari, F2=2 hari, F4=1-4 hari, F4+=1-5 hari. "
            "SENIN-SABTU saja (hari Minggu/Sunday tidak berlaku)."
        ),
        MINGGU_COL: (
            "Langkah 5 - Pilih Frekuensi (kolom I) dahulu. F1/F2: pilih "
            "Minggu Ganjil atau Minggu Genap. F4/F4+: hanya ada 1 pilihan, "
            "Minggu Ganjil + Genap."
        ),
        KET_MINGGU_COL: (
            "Langkah 6 - Terisi berdasarkan Frekuensi + Minggu. "
            "F1 Ganjil: pilih 1 atau 3. F1 Genap: pilih 2 atau 4. "
            "F2/F4/F4+: hanya ada 1 pilihan (otomatis). Nilai ini yang "
            "disimpan sebagai callcycle."
        ),
    }

    ws = wb.create_sheet("PJP Template")

    for ci, cn in enumerate(col_names, 1):
        cell = ws.cell(row=1, column=ci, value=notes_pjp.get(cn, ""))
        cell.font = _note_font()
        cell.alignment = _vcenter(wrap=True)

    for ci, cn in enumerate(col_names, 1):
        if col_req.get(cn):
            cell = ws.cell(row=2, column=ci, value="Wajib Diisi")
            cell.font = _req_font()
            cell.alignment = _center()

    for ci, cn in enumerate(col_names, 1):
        cell = ws.cell(row=3, column=ci, value=cn)
        cell.font = _header_font()
        cell.fill = _fill("1A7A6E" if (cn in DROPDOWN_COLS or cn in AUTO_COLS) else "ED7D31")
        cell.alignment = _center()
        cell.border = _thin_border()

    ws.row_dimensions[1].height = 42
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 44
    ws.freeze_panes = "A4"

    widths = [18, 18, 28, 18, 20, 30, 18, 30, 12, 28, 22, 16]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    def col_letter(name):
        return get_column_letter(col_names.index(name) + 1)

    def dr(name):
        return f"{col_letter(name)}{FIRST_DATA}:{col_letter(name)}{LAST_DATA}"

    _attach_pjp_dvs(ws, col_names, FIRST_DATA, LAST_DATA)

    freq_cl = col_letter("Frekuensi")
    freq_anchor = f"${freq_cl}{FIRST_DATA}"  # column fixed, row relative — see below

    # Frekuensi (Column I) is the DRIVER: a plain, closed dropdown with
    # exactly the 4 allowed values. Hari (J) and Nomor Minggu (L) are
    # DEPENDENT dropdowns keyed off whatever Frekuensi the row's own I cell
    # holds — errorStyle="stop" (blocking) here, unlike Hari/Minggu in the
    # rest of this sheet, because Frekuensi has a small closed vocabulary
    # with no legitimate free-text variant to tolerate.
    frekuensi_dv = DataValidation(
        type="list",
        formula1="NR_FREKUENSI_COMBO",
        allow_blank=True,
        showInputMessage=True,
        promptTitle="Langkah 3 - Frekuensi",
        prompt=(
            "Pilih Frekuensi terlebih dahulu — menentukan pilihan Hari dan "
            "Nomor Minggu di kolom berikutnya.\n"
            "  F1  = 1x per bulan\n"
            "  F2  = 2x per bulan\n"
            "  F4  = setiap minggu (minggu 1-4, tanpa minggu ke-5)\n"
            "  F4+ = setiap minggu (+ minggu ke-5 bila bulan tsb memilikinya)"
        ),
        showErrorMessage=True,
        errorTitle="Input Tidak Valid",
        error="Pilih F1, F2, F4, atau F4+ dari daftar dropdown.",
    )
    ws.add_data_validation(frekuensi_dv)
    frekuensi_dv.sqref = dr("Frekuensi")

    # Hari (Column J): the dropdown source range switches with Frekuensi via
    # INDIRECT("NR_HARI_"&...) — NR_HARI_F1/F2/F4/F4PLUS each hold every
    # valid day-combination for that Frekuensi (see _build_lookup_and_named_
    # ranges()), so picking from the list always yields a Frekuensi-
    # consistent value. "+" isn't valid inside an Excel name, hence the
    # SUBSTITUTE("+","PLUS") to match the FREKUENSI_RANGE_SUFFIX mapping.
    # errorStyle="warning" (not "stop") so a manually-typed variant isn't
    # hard-blocked client-side; validate_pjp_df() on upload is the
    # authoritative gate.
    hari_dv = DataValidation(
        type="list",
        formula1=f'INDIRECT("NR_HARI_"&SUBSTITUTE({freq_anchor},"+","PLUS"))',
        allow_blank=True,
        showInputMessage=True,
        promptTitle="Langkah 4 - Hari (tergantung Frekuensi)",
        prompt=(
            "Pilih Frekuensi (kolom I) terlebih dahulu. Jumlah hari: "
            "F1=1, F2=2, F4=1-4, F4+=1-5. Hanya SENIN-SABTU, dipisah slash "
            "atau koma — contoh: SENIN/SELASA. Tidak boleh duplikat."
        ),
        showErrorMessage=True,
        errorStyle="warning",
        errorTitle="Format Hari Tidak Standar",
        error=(
            "Invalid Hari untuk Frekuensi ini. Gunakan SENIN-SABTU (bukan "
            "MINGGU) sesuai jumlah hari yang diizinkan Frekuensi. Nilai akan "
            "divalidasi ulang saat upload."
        ),
    )
    ws.add_data_validation(hari_dv)
    hari_dv.sqref = dr("Hari")

    # Minggu (Column K): dependent on Frekuensi via the same INDIRECT
    # cascade. NR_MINGGU_F1/F2 hold {Ganjil, Genap}; NR_MINGGU_F4/F4PLUS
    # hold only {Ganjil + Genap} — a one-entry list, which is how
    # "automatically set + locked" is expressed with native Excel Data
    # Validation and no VBA (the spec's §24 explicitly prefers this over
    # introducing VBA). errorStyle="stop": the vocabulary is closed, there
    # is no legitimate free-text variant to tolerate.
    minggu_cl = col_letter(MINGGU_COL)
    minggu_anchor = f"${minggu_cl}{FIRST_DATA}"
    minggu_dv = DataValidation(
        type="list",
        formula1=f'INDIRECT("NR_MINGGU_"&SUBSTITUTE({freq_anchor},"+","PLUS"))',
        allow_blank=True,
        showInputMessage=True,
        promptTitle="Langkah 5 - Minggu (tergantung Frekuensi)",
        prompt=(
            "Pilih Frekuensi (kolom I) terlebih dahulu.\n"
            "  F1 / F2  -> Minggu Ganjil atau Minggu Genap\n"
            "  F4 / F4+ -> hanya Minggu Ganjil + Genap (otomatis)"
        ),
        showErrorMessage=True,
        errorStyle="stop",
        errorTitle="Input Tidak Valid",
        error="Pilih Minggu dari daftar dropdown yang sesuai dengan Frekuensi.",
    )
    ws.add_data_validation(minggu_dv)
    minggu_dv.sqref = dr(MINGGU_COL)

    # Ket. Minggu (Column L) -> DB `callcycle`. Depends on BOTH Frekuensi
    # and Minggu: the range name is NR_KET_<FREQ>_<MINGGU>, e.g.
    # NR_KET_F1_GANJIL = {1,3}, NR_KET_F1_GENAP = {2,4},
    # NR_KET_F2_GANJIL = {1,3} (single, automatic),
    # NR_KET_F4_GANJILGENAP = {1,2,3,4}, NR_KET_F4PLUS_GANJILGENAP =
    # {1,2,3,4,5}. The nested SUBSTITUTEs turn the Minggu cell's text into
    # that suffix: "Minggu Ganjil + Genap" -> strip "Minggu " -> drop "+"
    # -> drop spaces -> upper -> "GANJILGENAP". This is what makes the
    # invalid F1 combinations (Ganjil+2, Ganjil+4, Genap+1, Genap+3) not
    # merely rejected but literally unselectable.
    ket_source = (
        f'INDIRECT("NR_KET_"&SUBSTITUTE({freq_anchor},"+","PLUS")&"_"'
        f'&UPPER(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE({minggu_anchor},"Minggu ",""),"+","")," ","")))'
    )
    ket_dv = DataValidation(
        type="list",
        formula1=ket_source,
        allow_blank=True,
        showInputMessage=True,
        promptTitle="Langkah 6 - Ket. Minggu (tergantung Frekuensi + Minggu)",
        prompt=(
            "Isi Frekuensi (kolom I) dan Minggu (kolom K) terlebih dahulu.\n"
            "  F1 + Minggu Ganjil -> pilih 1 atau 3\n"
            "  F1 + Minggu Genap  -> pilih 2 atau 4\n"
            "  F2  -> otomatis 1,3 (Ganjil) / 2,4 (Genap)\n"
            "  F4  -> otomatis 1,2,3,4\n"
            "  F4+ -> otomatis 1,2,3,4,5"
        ),
        showErrorMessage=True,
        errorStyle="stop",
        errorTitle="Input Tidak Valid",
        error="Pilih Ket. Minggu dari daftar dropdown (tergantung Frekuensi dan Minggu).",
    )
    ws.add_data_validation(ket_dv)
    ket_dv.sqref = dr(KET_MINGGU_COL)

    # ── Conditional formatting ────────────────────────────────────────────
    # Reuses the template's existing palette: FFF3CD = warning/required
    # (same amber already used elsewhere), D6E4F0 = the informational blue
    # used by the auto-filled columns. No new colours introduced.
    amber_fill = _fill("FFF3CD")
    info_fill = _fill("D6E4F0")
    hari_cl = col_letter("Hari")
    ket_cl = col_letter(KET_MINGGU_COL)

    # Advisory: a filled Hari cell that isn't an exact match for one of its
    # own Frekuensi's canonical combinations (e.g. "selasa/senin" — still
    # accepted and auto-normalized on upload, just not canonical yet).
    ws.conditional_formatting.add(
        dr("Hari"),
        FormulaRule(
            formula=[
                f'AND({hari_cl}{FIRST_DATA}<>"",'
                f'COUNTIF(INDIRECT("NR_HARI_"&SUBSTITUTE({freq_anchor},"+","PLUS")),{hari_cl}{FIRST_DATA})=0)'
            ],
            fill=amber_fill,
        ),
    )

    # Ket. Minggu status model (spec §16):
    #   State 1 REQUIRED — Frekuensi chosen, Minggu still empty, L empty.
    #   State 2 READY    — Frekuensi + Minggu chosen, L still empty.
    #   State 3 VALID    — L filled: no rule fires, normal cell styling.
    ws.conditional_formatting.add(
        dr(KET_MINGGU_COL),
        FormulaRule(
            formula=[
                f'AND({ket_cl}{FIRST_DATA}="",{freq_anchor}<>"",{minggu_anchor}="")'
            ],
            fill=amber_fill,
        ),
    )
    ws.conditional_formatting.add(
        dr(KET_MINGGU_COL),
        FormulaRule(
            formula=[
                f'AND({ket_cl}{FIRST_DATA}="",{freq_anchor}<>"",{minggu_anchor}<>"")'
            ],
            fill=info_fill,
        ),
    )
    # Minggu itself gets the same REQUIRED cue while Frekuensi is set but
    # Minggu is still blank, so the chain reads left-to-right.
    ws.conditional_formatting.add(
        dr(MINGGU_COL),
        FormulaRule(
            formula=[f'AND({minggu_cl}{FIRST_DATA}="",{freq_anchor}<>"")'],
            fill=amber_fill,
        ),
    )

    # STALE-VALUE cue (spec §21/§6): changing Frekuensi cannot clear an
    # already-filled Minggu / Ket. Minggu — native Excel Data Validation
    # only constrains NEW entries, it never rewrites existing cells, and
    # without VBA nothing can. So e.g. F2 + "Minggu Ganjil" + "1,3" that is
    # switched to F4 leaves both cells holding values that are no longer
    # selectable. Those rows ARE rejected on upload by validate_pjp_df(),
    # but they would otherwise look fine on screen. These two rules flag
    # them immediately, in the same amber warning colour, so the user fixes
    # them in the sheet rather than discovering it at upload time.
    #   - The Ket. Minggu rule wraps COUNTIF in IFERROR because an invalid
    #     (Frekuensi, Minggu) pair makes INDIRECT resolve to a named range
    #     that does not exist (e.g. NR_KET_F4_GANJIL) -> #REF!; IFERROR
    #     turns that into 0 so the cell is still flagged rather than the
    #     rule silently not firing.
    ws.conditional_formatting.add(
        dr(MINGGU_COL),
        FormulaRule(
            formula=[
                f'AND({minggu_cl}{FIRST_DATA}<>"",{freq_anchor}<>"",'
                f'IFERROR(COUNTIF(INDIRECT("NR_MINGGU_"&SUBSTITUTE({freq_anchor},"+","PLUS")),{minggu_cl}{FIRST_DATA}),0)=0)'
            ],
            fill=amber_fill,
        ),
    )
    ws.conditional_formatting.add(
        dr(KET_MINGGU_COL),
        FormulaRule(
            formula=[
                f'AND({ket_cl}{FIRST_DATA}<>"",{freq_anchor}<>"",{minggu_anchor}<>"",'
                f'IFERROR(COUNTIF({ket_source},{ket_cl}{FIRST_DATA}),0)=0)'
            ],
            fill=amber_fill,
        ),
    )

    sal_id_cl = col_letter("Salesman ID")
    kode_toko_cl = col_letter("Kode Toko")
    df_reindexed = df.reindex(columns=col_names)

    for excel_row in range(FIRST_DATA, LAST_DATA + 1):
        dfi = excel_row - FIRST_DATA
        has_data = dfi < len(df_reindexed)

        for ci, cn in enumerate(col_names, 1):
            cell = ws.cell(row=excel_row, column=ci)

            if cn in _AUTO_FROM_SALESMAN:
                key_formula = _extract_key_formula(f"{sal_id_cl}{excel_row}")
                lut_col = _AUTO_FROM_SALESMAN[cn]
                cell.value = f'=IFERROR(VLOOKUP({key_formula},NR_SALESMAN_LOOKUP,{lut_col},0),"")'
                cell.fill = _fill("D6E4F0")
                cell.font = Font(italic=True, color="1A7A6E", size=10, name="Calibri")
                cell.alignment = _vcenter()
                cell.border = _thin_border()
                cell.number_format = "@"
                cell.protection = Protection(locked=True)
                continue

            if cn in _AUTO_FROM_STORE:
                key_formula = _extract_key_formula(f"{kode_toko_cl}{excel_row}")
                lut_col = _AUTO_FROM_STORE[cn]
                cell.value = f'=IFERROR(VLOOKUP({key_formula},NR_STORE_LOOKUP,{lut_col},0),"")'
                cell.fill = _fill("D6E4F0")
                cell.font = Font(italic=True, color="1A7A6E", size=10, name="Calibri")
                cell.alignment = _vcenter()
                cell.border = _thin_border()
                cell.number_format = "@"
                cell.protection = Protection(locked=True)
                continue

            cell.number_format = "@"
            if has_data:
                val = df_reindexed.iloc[dfi].get(cn, "")
                cell.value = "" if pd.isna(val) else (str(val) if val != "" else "")

            cell.alignment = _vcenter()
            cell.border = _thin_border()
            cell.protection = Protection(locked=False)

    ws.protection.sheet = True
    ws.protection.password = "skintific"
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ─── Helpers ──────────────────────────────────────────────────────────────────


def normalize_phone_id(phone) -> str:
    if phone is None:
        return ""
    hp = str(phone).strip().replace(" ", "").replace("-", "").rstrip(".0")
    if hp.startswith("+62+62"):
        hp = hp[3:]
    if hp.startswith("+62"):
        digits_after = hp[3:]
        if digits_after.isdigit() and 8 <= len(digits_after) <= 13:
            return hp
        return hp
    if hp.startswith("62"):
        if hp[2:].isdigit() and 8 <= len(hp[2:]) <= 13:
            return "+" + hp
        return hp
    if hp.startswith("0"):
        if hp[1:].isdigit() and 8 <= len(hp[1:]) <= 13:
            return "+62" + hp[1:]
        return hp
    if hp.isdigit():
        if 8 <= len(hp) <= 13:
            return "+62" + hp
        return hp
    return hp


def sanitize_salesman_name(name: str) -> str:
    return re.sub(r"\s+", " ", str(name).strip().upper())


def _is_empty(val) -> bool:
    return pd.isna(val) or str(val).strip() == ""


def _get_unique_distributors(df, col="Kode Distributor") -> list:
    if col not in df.columns:
        return []
    return (
        df[col]
        .dropna()
        .astype(str)
        .str.strip()
        .replace("", pd.NA)
        .dropna()
        .unique()
        .tolist()
    )


def validate_row_completeness(df, required_cols, sheet_label) -> list:
    errors = []
    for i, row in df.iterrows():
        n = i + 4
        values = {c: row.get(c, "") for c in required_cols}
        non_empty = [c for c, v in values.items() if not _is_empty(v)]
        empty = [c for c, v in values.items() if _is_empty(v)]
        if non_empty and empty:
            errors.append(f"Baris {n}: kolom wajib belum terisi — {', '.join(empty)}")
    return errors


def validate_pjp_df(df, distributor_map, store_df=None, salesman_df=None, selected_dist_code=None):
    """
    Validates an uploaded/parsed PJP dataframe (after read_template_sheet has
    already extracted 'salesman_id' and 'kode_toko' from the combo strings).

    Rules enforced:
      - salesman_id must exist in gt_salesman_mapping AND be active
        (checked against the FULL active roster across all distributors —
        salesman_df should be load_all_salesman_mapping()'s output, NOT the
        distributor-scoped load_salesman_mapping()).
      - salesman_id's distributor_code must match selected_dist_code,
        otherwise: "Salesman ID '<id>' bukan milik distributor yang dipilih."
      - kode_toko must exist in master_store_database_basis (checked
        against the FULL store master — store_df should be the full,
        unscoped dataset, not pre-filtered to one distributor).
      - kode_toko's distributor_code must equal selected_dist_code,
        otherwise: "Store does not belong to selected distributor."
      - Frekuensi must be one of F1/F2/F4/F4+.
      - Hari (SENIN..SABTU, never MINGGU) must satisfy the row's Frekuensi
        day-count rule (F1=1, F2=2, F4=1-4, F4+=1-5) — see normalize_hari().
        Skipped if Frekuensi itself is invalid/missing (nothing to check
        the count against).
      - Nomor Minggu / callcycle must satisfy the row's Frekuensi rule
        (F1: 1 of 1-4; F2: 2 unique of 1-4; F4: fixed "1,2,3,4"; F4+: fixed
        "1,2,3,4,5") — see normalize_callcycle(). Legacy values ("Minggu
        Ganjil" etc.) are rejected here with a hint pointing at their new
        equivalent; use migrate_legacy_minggu() to fix already-stored
        historical rows. Skipped if Frekuensi itself is invalid/missing.

    IMPORTANT: callers MUST pass the FULL, unscoped salesman_df/store_df
    here (covering all distributors), NOT a subset pre-filtered to
    selected_dist_code. Pre-filtering before this function runs would
    conflate "doesn't exist at all" with "exists, but under a different
    distributor" — a store/salesman that's real but has any distributor_code
    formatting mismatch would then wrongly disappear from the pre-filtered
    subset and get reported as "not found" instead of the more accurate
    "belongs to a different distributor". Passing the full dataset here and
    letting this function do the selected_dist_code comparison itself keeps
    the two failure modes distinguishable.
    """
    errors, warnings = [], []
    missing = [c for c in PJP_REQUIRED if c not in df.columns]
    if missing:
        errors.append(f"Kolom wajib tidak ditemukan: {', '.join(missing)}")
        return errors, warnings

    # ── Outdated-template detection ────────────────────────────────────────
    # The column layout changed (Frekuensi moved K->I, Hari I->J, and the
    # single old "Minggu Ganjil/Minggu Genap/Minggu Ganjil + Genap" column
    # became Minggu (K) + Ket. Minggu (L)). A workbook saved before that
    # change still imports field-for-field correctly — read_template_sheet()
    # matches on HEADER TEXT, never on column position, so the swapped
    # Frekuensi/Hari positions do not corrupt anything — but it cannot
    # supply Column K, so every row would otherwise fail with several
    # confusing per-row errors that never say what actually went wrong.
    # Detect it once and give one actionable instruction instead.
    if KET_MINGGU_COL in df.columns:
        filled_ket = [v for v in df[KET_MINGGU_COL] if pd.notna(v) and str(v).strip()]
        legacy_ket = [v for v in filled_ket if migrate_legacy_minggu(v)]
        if filled_ket and len(legacy_ket) == len(filled_ket):
            errors.append(
                "Template lama terdeteksi — kolom 'Ket. Minggu' masih berisi nilai "
                "format lama (mis. 'Minggu Ganjil'). Struktur template sudah berubah: "
                "sekarang ada kolom I Frekuensi, J Hari, K Minggu, dan L Ket. Minggu. "
                "Silakan download ulang PJP Template terbaru pada tab "
                "'Download Template', isi ulang, lalu upload kembali."
            )
            return errors, warnings

    errors += validate_row_completeness(df, PJP_REQUIRED, "PJP")

    unique_dist = _get_unique_distributors(df, col="kode_distributor")
    if len(unique_dist) > 1:
        errors.append(
            f"Sheet PJP Template hanya boleh berisi 1 kode distributor per file. "
            f"Ditemukan {len(unique_dist)} kode: {', '.join(unique_dist)}"
        )
        return errors, warnings

    valid_salesman_ids = set()
    if salesman_df is not None and not salesman_df.empty:
        valid_salesman_ids = set(salesman_df["salesman_id"].dropna().astype(str).str.strip().tolist())

    store_lookup = build_store_lookup(store_df) if store_df is not None else {}

    for i, row in df.iterrows():
        n = i + 4

        # ── Salesman ID validation ──────────────────────────────────────────
        sal_id = str(row.get("salesman_id", "")).strip()
        if sal_id:
            if valid_salesman_ids and sal_id not in valid_salesman_ids:
                errors.append(
                    f"Baris {n}: Salesman ID '{sal_id}' tidak ditemukan atau tidak aktif."
                )
            elif selected_dist_code and salesman_df is not None and not salesman_df.empty:
                match = salesman_df.loc[salesman_df["salesman_id"] == sal_id]
                if not match.empty and str(match.iloc[0]["distributor_code"]).strip().upper() != str(selected_dist_code).strip().upper():
                    actual_dist = match.iloc[0]["distributor_code"]
                    errors.append(
                        f"Baris {n}: Salesman ID '{sal_id}' bukan milik distributor yang dipilih "
                        f"(terdaftar untuk distributor '{actual_dist}', bukan '{selected_dist_code}')."
                    )

        # ── Kode Toko validation ────────────────────────────────────────────
        kode_toko = str(row.get("kode_toko", "")).strip()
        if kode_toko:
            store_info = store_lookup.get(kode_toko)
            if store_info is None:
                errors.append(f"Baris {n}: Kode Toko '{kode_toko}' tidak ditemukan di master store.")
            elif selected_dist_code and str(store_info["kode_distributor"]).strip().upper() != str(selected_dist_code).strip().upper():
                actual_dist = store_info["kode_distributor"]
                errors.append(
                    f"Baris {n}: Store does not belong to selected distributor "
                    f"(Kode Toko '{kode_toko}' terdaftar untuk distributor '{actual_dist}', "
                    f"bukan '{selected_dist_code}')."
                )

        # ── Frekuensi validation (Column I — drives Hari & callcycle) ──────
        freq_val = row.get("Frekuensi", "")
        freq_str = str(freq_val).strip().upper() if pd.notna(freq_val) else ""
        freq_present_and_valid = bool(freq_str) and freq_str in FREKUENSI_OPTIONS
        if freq_str and not freq_present_and_valid:
            errors.append(f"Baris {n}: 'Frekuensi' nilai tidak valid (harus F1, F2, F4, atau F4+)")

        # ── Hari validation (Column J — count depends on Frekuensi) ────────
        hari_val = row.get("Hari", "")
        if pd.notna(hari_val) and str(hari_val).strip():
            _, hari_err = normalize_hari(hari_val, frekuensi=freq_str if freq_present_and_valid else None)
            if hari_err:
                errors.append(
                    f"Baris {n}: Invalid Hari. Gunakan SENIN-SABTU (bukan MINGGU) "
                    f"dipisahkan dengan slash atau koma, sesuai jumlah hari yang "
                    f"diizinkan Frekuensi (F1=1, F2=2, F4=1-4, F4+=1-5). "
                    f"(nilai: '{hari_val}', Frekuensi: '{freq_val}')"
                )

        # ── Minggu validation (Column K — options depend on Frekuensi) ─────
        minggu_val = row.get(MINGGU_COL, "")
        minggu_ok = None
        if pd.notna(minggu_val) and str(minggu_val).strip():
            if not freq_present_and_valid:
                errors.append(
                    f"Baris {n}: Minggu tidak dapat divalidasi tanpa Frekuensi yang valid "
                    f"(nilai Frekuensi: '{freq_val}')."
                )
            else:
                minggu_ok, mg_err = normalize_minggu(minggu_val, freq_str)
                if mg_err:
                    allowed = ", ".join(minggu_options_for_frekuensi(freq_str))
                    errors.append(
                        f"Baris {n}: Invalid Minggu untuk Frekuensi {freq_str}. "
                        f"Pilihan yang valid: {allowed}. (nilai: '{minggu_val}')"
                    )

        # ── Ket. Minggu / callcycle validation (Column L) ───────────────────
        # Depends on BOTH Frekuensi and Minggu — this is what rejects the
        # F1 mismatches (Ganjil+2, Ganjil+4, Genap+1, Genap+3) that the
        # Excel dependent dropdown already makes unselectable.
        ket_val = row.get(KET_MINGGU_COL, "")
        if pd.notna(ket_val) and str(ket_val).strip():
            if not freq_present_and_valid:
                errors.append(
                    f"Baris {n}: Ket. Minggu tidak dapat divalidasi tanpa Frekuensi yang valid "
                    f"(nilai Frekuensi: '{freq_val}')."
                )
            elif minggu_ok is None:
                # Show "(kosong)" rather than letting pandas' NaN reach the
                # user as the literal string 'nan'.
                shown = "(kosong)" if (pd.isna(minggu_val) or not str(minggu_val).strip()) else f"'{minggu_val}'"
                errors.append(
                    f"Baris {n}: Ket. Minggu tidak dapat divalidasi tanpa Minggu yang valid "
                    f"(kolom K Minggu: {shown}). Isi kolom Minggu terlebih dahulu."
                )
            else:
                _, ket_err = normalize_callcycle(ket_val, freq_str, minggu_ok)
                if ket_err:
                    allowed = ", ".join(ket_minggu_options(freq_str, minggu_ok))
                    legacy_equiv = migrate_legacy_minggu(ket_val)
                    hint = f" Nilai lama terdeteksi — gunakan '{legacy_equiv}'." if legacy_equiv else ""
                    auto = auto_callcycle(freq_str, minggu_ok)
                    auto_note = (
                        f" Untuk {freq_str} nilai ini terisi otomatis dan tidak boleh diubah manual."
                        if auto else ""
                    )
                    errors.append(
                        f"Baris {n}: Invalid Ket. Minggu untuk {freq_str} + {minggu_ok}. "
                        f"Pilihan yang valid: {allowed}.{auto_note}{hint} (nilai: '{ket_val}')"
                    )

    return errors, warnings


def _extract_combo_key(val) -> str:
    """Returns the trimmed key value chosen via the Excel dropdown
    (Salesman ID / Kode Toko). Templates now write the bare key directly,
    so this is normally just a trim; the old 'KEY - Label' combo format is
    still handled as a fallback for backward-compat with older uploads."""
    if pd.isna(val):
        return ""
    s = str(val).strip()
    if _COMBO_SEP in s:
        return s.split(_COMBO_SEP, 1)[0].strip()
    return s


def read_template_sheet(
    uploaded_file, sheet_name, header_row, salesman_df=None, store_df=None
):
    """
    Reads an uploaded PJP Template sheet and re-derives every auto column
    in Python (rather than trusting Excel's cached formula values, which are
    not guaranteed to be present since the template is generated by
    openpyxl, not by Excel itself).

    `salesman_df` MUST be a DataFrame with 'salesman_id' / 'salesman'
    columns (e.g. the output of load_salesman_mapping(distributor_code)) —
    not a {code: name} distributor map — otherwise the Nama Salesman
    auto-fill and downstream validation will be silently wrong/broken.

    Adds extra lowercase helper columns used by validate_pjp_df /
    push_to_bigquery:
      - salesman_id      : extracted from the "Salesman ID" combo column
      - kode_toko         : extracted from the "Kode Toko" combo column
      - kode_distributor  : mirror of "Kode Distributor" for convenience

    Frekuensi is normalized first (trimmed/uppercased) since it DRIVES the
    Hari day-count rule, the Minggu options, and (with Minggu) the Ket.
    Minggu rule. Hari / Minggu / Ket. Minggu are each re-derived through
    their normalize_*() function rather than a blunt .str.title() — this
    canonicalises well-formed cells (order, separator, spacing, casing) and
    leaves malformed cells untouched (raw) so validate_pjp_df() re-derives
    the same error and reports it with the exact offending text.

    Header back-compat: Minggu is matched by MINGGU_COL_ALIASES and Ket.
    Minggu by KET_MINGGU_COL_ALIASES, so workbooks downloaded under any
    earlier revision's headers still import; both are renamed to the
    current headers on the way out.

    Ket. Minggu is AUTO-FILLED where it is deterministic — for F2/F4/F4+ a
    valid (Frekuensi, Minggu) pair has exactly one legal value, so a blank
    cell is populated rather than reported as missing. F1 is never
    auto-filled: its two candidate weeks are a genuine user choice.
    Conversely, if Minggu is blank but Ket. Minggu is present, Minggu is
    back-derived from it so legacy/partial workbooks still validate.
    """
    df = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=header_row)
    df = df.dropna(how="all")

    if "Frekuensi" in df.columns:
        df["Frekuensi"] = df["Frekuensi"].astype(str).str.strip().str.upper()
        df.loc[df["Frekuensi"] == "NAN", "Frekuensi"] = ""

    def _row_freq(row):
        f = row.get("Frekuensi") if hasattr(row, "get") else None
        return f if f in FREKUENSI_OPTIONS else None

    if "Hari" in df.columns:
        def _norm_hari_row(row):
            v = row["Hari"]
            normalized, _err = normalize_hari(v, frekuensi=_row_freq(row))
            return normalized if normalized is not None else v
        df["Hari"] = df.apply(_norm_hari_row, axis=1)

    minggu_col_found = next((c for c in MINGGU_COL_ALIASES if c in df.columns), None)
    if minggu_col_found and minggu_col_found != MINGGU_COL:
        df = df.rename(columns={minggu_col_found: MINGGU_COL})
        minggu_col_found = MINGGU_COL

    ket_col_found = next((c for c in KET_MINGGU_COL_ALIASES if c in df.columns), None)
    if ket_col_found and ket_col_found != KET_MINGGU_COL:
        df = df.rename(columns={ket_col_found: KET_MINGGU_COL})
        ket_col_found = KET_MINGGU_COL

    # Back-derive a missing Minggu from an existing Ket. Minggu so partial
    # or older workbooks (which had no Minggu column at all) still resolve.
    if ket_col_found and MINGGU_COL not in df.columns:
        df[MINGGU_COL] = df[KET_MINGGU_COL].apply(derive_minggu_from_callcycle)
    elif ket_col_found and MINGGU_COL in df.columns:
        def _fill_minggu(row):
            cur = row.get(MINGGU_COL)
            if pd.notna(cur) and str(cur).strip():
                return cur
            return derive_minggu_from_callcycle(row.get(KET_MINGGU_COL))
        df[MINGGU_COL] = df.apply(_fill_minggu, axis=1)

    if MINGGU_COL in df.columns:
        def _norm_minggu_row(row):
            v = row[MINGGU_COL]
            freq = _row_freq(row)
            if freq is None:
                return v
            normalized, _err = normalize_minggu(v, freq)
            return normalized if normalized is not None else v
        df[MINGGU_COL] = df.apply(_norm_minggu_row, axis=1)

    if MINGGU_COL in df.columns:
        if KET_MINGGU_COL not in df.columns:
            df[KET_MINGGU_COL] = ""

        def _norm_ket_row(row):
            v = row[KET_MINGGU_COL]
            freq = _row_freq(row)
            minggu = row.get(MINGGU_COL)
            if freq is None or pd.isna(minggu) or not str(minggu).strip():
                return v
            blank = pd.isna(v) or str(v).strip() == ""
            if blank:
                # Deterministic for F2/F4/F4+; None for F1 (user must pick).
                return auto_callcycle(freq, minggu) or v
            normalized, _err = normalize_callcycle(v, freq, minggu)
            return normalized if normalized is not None else v
        df[KET_MINGGU_COL] = df.apply(_norm_ket_row, axis=1)

    salesman_lookup = build_salesman_lookup(salesman_df) if salesman_df is not None else {}
    store_lookup = build_store_lookup(store_df) if store_df is not None else {}

    if "Salesman ID" in df.columns:
        df["salesman_id"] = df["Salesman ID"].apply(_extract_combo_key)
        df["Nama Salesman"] = df["salesman_id"].apply(lambda k: salesman_lookup.get(k, ""))

    if "Kode Toko" in df.columns:
        df["kode_toko"] = df["Kode Toko"].apply(_extract_combo_key)
        df["Nama Toko"] = df["kode_toko"].apply(lambda k: store_lookup.get(k, {}).get("store_name", ""))
        df["Region"] = df["kode_toko"].apply(lambda k: store_lookup.get(k, {}).get("region", ""))
        df["ASM"] = df["kode_toko"].apply(lambda k: store_lookup.get(k, {}).get("asm", ""))
        df["Nama Distributor"] = df["kode_toko"].apply(lambda k: store_lookup.get(k, {}).get("nama_distributor", ""))
        df["Kode Distributor"] = df["kode_toko"].apply(lambda k: store_lookup.get(k, {}).get("kode_distributor", ""))
        df["kode_distributor"] = df["Kode Distributor"]

    if "No. HP" in df.columns:
        df["No. HP"] = df["No. HP"].apply(normalize_phone_id)

    return df


# ─── BigQuery writer ──────────────────────────────────────────────────────────

_SAL_COL_MAP = {
    "Nama Salesman": "nama_salesman",
    "Nama SPV External": "nama_spv_external",
    "Nama SPV Internal": "nama_spv_internal",
    "Nama SPV Internal 2": "nama_spv_internal_2",
    "ASM": "asm",
    "Region": "region",
    "Nama Distributor": "nama_distributor",
    "Kode Distributor": "kode_distributor",
    "Status Salesman": "status_salesman",
    "Total Outlet Coverage PJP": "total_outlet_coverage_pjp",
    "Gaji Pokok": "gaji_pokok",
    "Tunjangan dan insentif": "tunjangan_dan_insentif",
    "Tanggal Lahir": "tanggal_lahir",
    "Jenis Kelamin": "jenis_kelamin",
    "Pendidikan Terakhir": "pendidikan_terakhir",
    "Pengalaman di Perusahaan Sebelumnya (Dalam Bulan)": "pengalaman_bulan",
    "Principal Lain yang Ditanggungjawabi": "principal_lain",
    "No. HP": "no_hp",
    "Tanggal Join di G2G": "tanggal_join_g2g",
}

# NOTE: "salesman_id" and "snapshot_month" are derived lowercase helper
# columns added by read_template_sheet() / the upload flow (not literal
# Excel headers), included here so they get persisted going forward.
# Requires the BigQuery migrations:
#   ALTER TABLE gt_master_salesman_pjp ADD COLUMN salesman_id STRING;
#   ALTER TABLE gt_master_salesman_pjp ADD COLUMN snapshot_month STRING;
#   ALTER TABLE gt_master_salesman_pjp ADD COLUMN callcycle STRING;
# `callcycle` is the ONE new column this template's Frekuensi/Hari/Minggu/
# Ket. Minggu redesign adds — see backend/scripts/migrations/
# migrate_pjp_hari_minggu_format.py in the sfa-step repo. The legacy
# `minggu` column is intentionally NOT written here any more (frozen for
# historical rows only). Column K "Minggu" is a template/UI input used only
# to drive Column L — it is deliberately NOT persisted (no DB column for
# it, per spec §2/§34: the DB gains `callcycle` and nothing else).
_PJP_COL_MAP = {
    "salesman_id": "salesman_id",
    "ASM": "asm",
    "Region": "region",
    "Nama Distributor": "nama_distributor",
    "Kode Distributor": "kode_distributor",
    "Nama Salesman": "nama_salesman",
    "Kode Toko": "kode_toko",
    "Nama Toko": "nama_toko",
    "Frekuensi": "frekuensi",
    "Hari": "hari",
    KET_MINGGU_COL: "callcycle",
    "snapshot_month": "snapshot_month",
}


def push_to_bigquery(df, col_map, table_id) -> tuple[bool, str]:
    try:
        credentials, project_id = get_credentials()
        client = bigquery.Client(credentials=credentials, project=project_id)
        existing_cols = {c: col_map[c] for c in col_map if c in df.columns}
        bq_df = df[list(existing_cols.keys())].rename(columns=existing_cols).copy()
        bq_df["uploaded_at"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
        job_config = bigquery.LoadJobConfig(
            write_disposition="WRITE_APPEND",
            autodetect=True,
            # `callcycle` may not exist yet on the destination table (see
            # the ALTER TABLE note above) — without this, BQ load jobs
            # reject any DataFrame column absent from the existing schema
            # instead of adding it.
            schema_update_options=[bigquery.SchemaUpdateOption.ALLOW_FIELD_ADDITION],
        )
        job = client.load_table_from_dataframe(bq_df, table_id, job_config=job_config)
        job.result()
        return True, f"Berhasil menyimpan {len(bq_df)} baris ke Database."
    except Exception as e:
        return False, f"Gagal menyimpan ke Database: {e}"


def delete_pjp_records(
    distributor_code: str = None, salesman_name: str = None
) -> tuple[bool, str]:
    """
    Deletes PJP records for the given scope (distributor and/or salesman),
    but ONLY within the CURRENT snapshot_month (based on system date at
    call time). Historical months (previous snapshot_month values) are
    never touched here — this is what lets an upload replace this month's
    data while automatically preserving prior months, with no manual
    month/year selection required from the user.
    """
    try:
        credentials, project_id = get_credentials()
        client = bigquery.Client(credentials=credentials, project=project_id)

        current_snapshot_month = datetime.now().strftime("%Y-%m")

        conditions = ["snapshot_month = @snapshot_month"]
        params = [
            bigquery.ScalarQueryParameter(
                "snapshot_month", "STRING", current_snapshot_month
            )
        ]

        if distributor_code:
            conditions.append("UPPER(kode_distributor) = UPPER(@kode)")
            params.append(
                bigquery.ScalarQueryParameter("kode", "STRING", distributor_code)
            )
        if salesman_name:
            conditions.append("UPPER(TRIM(nama_salesman)) = UPPER(TRIM(@salesman))")
            params.append(
                bigquery.ScalarQueryParameter("salesman", "STRING", salesman_name)
            )

        if not distributor_code and not salesman_name:
            return False, "Tidak ada filter yang ditetapkan — operasi dibatalkan untuk keamanan."

        where_clause = "WHERE " + " AND ".join(conditions)
        query = f"DELETE FROM `{PJP_TABLE}` {where_clause}"
        job_config = bigquery.QueryJobConfig(query_parameters=params)
        client.query(query, job_config=job_config).result()
        return True, ""
    except Exception as e:
        return False, str(e)


# ─── Shared salesman form fields ──────────────────────────────────────────────


def _render_salesman_form_fields(key_prefix: str):
    c1, c2 = st.columns(2)
    with c1:
        nama = st.text_input("Nama Salesman *", key=f"{key_prefix}_nama")
        spv_ext = st.text_input("Nama SPV External", key=f"{key_prefix}_spv_ext")
        spv_int = st.text_input("Nama SPV Internal *", key=f"{key_prefix}_spv_int")
        spv_int2 = st.text_input("Nama SPV Internal 2", key=f"{key_prefix}_spv_int2")
        status_sal = st.selectbox(
            "Status Salesman *", STATUS_OPTIONS, key=f"{key_prefix}_status"
        )
        outlet_cov = st.number_input(
            "Total Outlet Coverage PJP *", min_value=0, step=1, key=f"{key_prefix}_outlet"
        )
        gaji = st.number_input(
            "Gaji Pokok (Rp) *", min_value=0, step=1000, key=f"{key_prefix}_gaji"
        )
        tunjangan = st.number_input(
            "Tunjangan dan Insentif (Rp) *",
            min_value=0,
            step=1000,
            key=f"{key_prefix}_tunj",
        )
    with c2:
        tgl_lahir = st.date_input(
            "Tanggal Lahir *",
            min_value=datetime(1945, 1, 1).date(),
            key=f"{key_prefix}_lahir",
        )
        gender = st.selectbox(
            "Jenis Kelamin *", GENDER_OPTIONS, key=f"{key_prefix}_gender"
        )
        pendidikan = st.selectbox(
            "Pendidikan Terakhir *", EDUCATION_OPTIONS, key=f"{key_prefix}_pendidikan"
        )
        pengalaman = st.number_input(
            "Pengalaman Sebelumnya (bulan) *",
            min_value=0,
            step=1,
            key=f"{key_prefix}_exp",
        )
        principal = st.text_input(
            "Principal Lain (opsional)", key=f"{key_prefix}_principal"
        )
        no_hp = st.text_input(
            "No. HP *", placeholder="08123456789", key=f"{key_prefix}_hp"
        )
        tgl_join = st.date_input("Tanggal Join di G2G *", key=f"{key_prefix}_join")
    return {
        "nama": nama,
        "spv_ext": spv_ext,
        "spv_int": spv_int,
        "spv_int2": spv_int2,
        "status_sal": status_sal,
        "outlet_cov": outlet_cov,
        "gaji": gaji,
        "tunjangan": tunjangan,
        "tgl_lahir": tgl_lahir,
        "gender": gender,
        "pendidikan": pendidikan,
        "pengalaman": pengalaman,
        "principal": principal,
        "no_hp": no_hp,
        "tgl_join": tgl_join,
    }


def _build_salesman_data(
    fields, dist_df, selected_dist_code, selected_dist_name
) -> dict:
    hp_norm = normalize_phone_id(fields["no_hp"])
    spv_int2_val = fields.get("spv_int2", "")
    return {
        "nama_salesman": sanitize_salesman_name(fields["nama"]),
        "nama_spv_external": fields["spv_ext"].strip().upper() if fields["spv_ext"].strip() else None,
        "nama_spv_internal": fields["spv_int"].strip().upper(),
        "nama_spv_internal_2": spv_int2_val.strip().upper() if spv_int2_val.strip() else None,
        "asm": dist_df.loc[dist_df["distributor_code"] == selected_dist_code, "asm"].iloc[0],
        "region": dist_df.loc[dist_df["distributor_code"] == selected_dist_code, "region"].iloc[0],
        "nama_distributor": selected_dist_name,
        "kode_distributor": selected_dist_code,
        "status_salesman": fields["status_sal"],
        "total_outlet_coverage_pjp": int(fields["outlet_cov"]),
        "gaji_pokok": float(fields["gaji"]),
        "tunjangan_dan_insentif": float(fields["tunjangan"]),
        "tanggal_lahir": fields["tgl_lahir"],
        "jenis_kelamin": fields["gender"],
        "pendidikan_terakhir": fields["pendidikan"],
        "pengalaman_bulan": int(fields["pengalaman"]),
        "principal_lain": fields["principal"].strip() if fields["principal"].strip() else None,
        "no_hp": hp_norm,
        "tanggal_join_g2g": fields["tgl_join"],
    }


def _validate_salesman_fields(fields) -> list:
    errors = []
    if not fields["nama"].strip():
        errors.append("Nama Salesman wajib diisi.")
    if not fields["spv_int"].strip():
        errors.append("Nama SPV Internal wajib diisi.")
    if not fields["no_hp"].strip():
        errors.append("No. HP wajib diisi.")
    return errors


# ══════════════════════════════════════════════════════════════════════════════
# NAVIGATION
# ══════════════════════════════════════════════════════════════════════════════

PAGES = {
    "👥 Kelola Salesman": "salesman",
    "🗓️ PJP Template": "pjp_template",
}

with st.sidebar:
    st.title("📋 G2G Template Manager")
    st.markdown("---")
    selected_page = st.radio(
        "Navigasi",
        list(PAGES.keys()),
        label_visibility="collapsed",
    )
    st.markdown("---")
    st.caption("Salesman & PJP Template Manager · G2G")

# ─── Load shared data ─────────────────────────────────────────────────────────

try:
    dist_df = load_distributor_data()
    store_df = load_store_data()
    distributor_map, asm_options, region_options = build_lookup_tables(dist_df)
except Exception as e:
    st.error(f"Gagal memuat data dari Database: {e}")
    st.stop()

# ─── Distributor selector ────────────────────────────────────────────────────

dist_labels = [
    f"{row['distributor_code']} — {row['distributor_name']}"
    for _, row in dist_df.sort_values("distributor_name").iterrows()
]
dist_code_from_label = {
    f"{row['distributor_code']} — {row['distributor_name']}": row["distributor_code"]
    for _, row in dist_df.iterrows()
}

with st.sidebar:
    st.markdown("### 🏢 Pilih Distributor")
    selected_label = st.selectbox(
        "Distributor",
        ["— Pilih distributor —"] + dist_labels,
        key="dist_selector",
        label_visibility="collapsed",
    )

if selected_label == "— Pilih distributor —":
    st.title("📋 Salesman & PJP Template Manager")
    st.info("👈 Pilih distributor di sidebar untuk melanjutkan.")
    st.stop()

selected_dist_code = dist_code_from_label[selected_label]
selected_dist_name = dist_df.loc[
    dist_df["distributor_code"] == selected_dist_code, "distributor_name"
].iloc[0]
selected_dist_asm = dist_df.loc[
    dist_df["distributor_code"] == selected_dist_code, "asm"
].iloc[0]
selected_dist_region = dist_df.loc[
    dist_df["distributor_code"] == selected_dist_code, "region"
].iloc[0]

# ── Detect distributor switch and reset auth + salesman cache ──────────────
_prev = st.session_state.get("_prev_dist_code")
if _prev is not None and _prev != selected_dist_code:
    auth_key = f"auth_{selected_dist_code}"
    st.session_state.pop(auth_key, None)
    st.session_state.pop("salesman_df", None)
    st.session_state.pop("_cached_dist", None)
    st.session_state.pop("pjp_salesman_df", None)
    st.session_state.pop("_pjp_cached_dist", None)
st.session_state["_prev_dist_code"] = selected_dist_code

with st.sidebar:
    st.success(f"**{selected_dist_name}**\n\n`{selected_dist_code}`")

# ─── PASSWORD GATE ────────────────────────────────────────────────────────────

if not _render_password_gate(selected_dist_code, selected_dist_name):
    st.stop()

# ─── PERIOD LOCK GATE ─────────────────────────────────────────────────────────

_today = datetime.now().date()
if _today > INPUT_DEADLINE:
    st.markdown(
        """
        <div style='
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            padding: 5rem 2rem;
            text-align: center;
        '>
            <div style='font-size: 4rem; line-height: 1; margin-bottom: 1.5rem;'>🔒</div>
            <h2 style='margin: 0 0 0.75rem 0; font-size: 1.75rem; font-weight: 600;'>
                Periode Input Sudah Ditutup
            </h2>
            <p style='color: #888; margin: 0 0 0.5rem 0; font-size: 1rem; max-width: 420px;'>
                Batas akhir pengisian adalah <b>{deadline}</b>.
            </p>
            <p style='color: #aaa; margin: 0; font-size: 0.9rem;'>
                Hubungi tim G2G jika ada pertanyaan.
            </p>
        </div>
        """.format(deadline=INPUT_DEADLINE.strftime("%d %B %Y")),
        unsafe_allow_html=True,
    )
    st.stop()

# ─── Load salesman data globally (shared across ALL pages) ────────────────────
# This is the "Kelola Salesman" roster (joined with gt_master_salesman for
# extra display fields like no_hp / status_salesman), used by the "Kelola
# Salesman" page and as the scope-selection dropdown in "Update PJP".

if (
    "salesman_df" not in st.session_state
    or st.session_state.get("_cached_dist") != selected_dist_code
):
    with st.spinner("Memuat daftar salesman..."):
        st.session_state.salesman_df = get_salesman_list(selected_dist_code)
        st.session_state._cached_dist = selected_dist_code

# ─── Load ACTIVE salesman-mapping roster for the new PJP flow ─────────────────
# This is the source-of-truth roster (gt_salesman_mapping, is_active=TRUE)
# used to VALIDATE uploaded PJP files (salesman_id must exist & be active &
# belong to selected_dist_code). It is distributor-scoped on purpose,
# because it drives ownership enforcement — it is NOT used anymore to build
# the Excel template's Salesman ID dropdown (that now uses
# pjp_salesman_all_df, loaded below).

if (
    "pjp_salesman_df" not in st.session_state
    or st.session_state.get("_pjp_cached_dist") != selected_dist_code
):
    with st.spinner("Memuat daftar Salesman ID aktif..."):
        st.session_state.pjp_salesman_df = load_salesman_mapping(selected_dist_code)
        st.session_state._pjp_cached_dist = selected_dist_code

# ─── Load FULL (all-distributor) active salesman roster for Excel dropdown ────
# Not distributor-scoped — this is what populates the Salesman ID dropdown
# in the downloaded Excel template, so users can search/select any active
# salesman from any distributor. Distributor ownership is still enforced
# at upload time using pjp_salesman_df (above), not this dataset.

if "pjp_salesman_all_df" not in st.session_state:
    with st.spinner("Memuat seluruh Salesman ID aktif..."):
        st.session_state.pjp_salesman_all_df = load_all_salesman_mapping()


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: KELOLA SALESMAN
# ══════════════════════════════════════════════════════════════════════════════

if PAGES[selected_page] == "salesman":
    st.title("👥 Kelola Salesman")
    st.caption(f"Distributor: **{selected_dist_name}** ({selected_dist_code})")

    if "action_mode" not in st.session_state:
        st.session_state.action_mode = None

    salesman_df = st.session_state.salesman_df

    col_search, col_filter, col_refresh = st.columns([3, 2, 1])
    with col_search:
        search_query = st.text_input(
            "🔍 Cari salesman:",
            key="search_salesman",
            placeholder="Nama atau ID salesman...",
        )
    with col_filter:
        filter_status = st.selectbox(
            "Filter status:", ["Semua", "Aktif", "Tidak Aktif"], key="filter_status"
        )
    with col_refresh:
        st.write("")
        if st.button("🔄 Refresh", use_container_width=True):
            st.cache_data.clear()
            st.session_state.pop("salesman_df", None)
            st.session_state.pop("_cached_dist", None)
            st.session_state.pop("pjp_salesman_df", None)
            st.session_state.pop("_pjp_cached_dist", None)
            st.session_state.pop("pjp_salesman_all_df", None)
            st.session_state.action_mode = None
            st.rerun()

    if not salesman_df.empty:
        display_df = salesman_df.copy()

        if search_query.strip():
            q = search_query.strip().upper()
            mask = display_df["salesman_id"].astype(str).str.upper().str.contains(
                q, na=False
            ) | display_df["nama_salesman"].astype(str).str.upper().str.contains(
                q, na=False
            )
            display_df = display_df[mask]

        if filter_status == "Aktif" and "is_active" in display_df.columns:
            display_df = display_df[display_df["is_active"] == True]
        elif filter_status == "Tidak Aktif" and "is_active" in display_df.columns:
            display_df = display_df[display_df["is_active"] == False]

        st.caption(f"Menampilkan **{len(display_df)}** salesman.")

        hcols = st.columns([1.2, 2.2, 1.2, 1.4, 1.2, 1.2, 0.8, 0.8, 0.8])
        headers = ["ID Salesman", "Nama", "Tipe", "No. HP", "Region", "ASM", "", "", ""]
        for hc, ht in zip(hcols, headers):
            hc.markdown(f"**{ht}**")
        st.divider()

        for _, row in display_df.iterrows():
            sal_id = row["salesman_id"]
            is_active = row.get("is_active", True)

            rcols = st.columns([1.2, 2.2, 1.2, 1.4, 1.2, 1.2, 0.8, 0.8, 0.8])
            id_label = f"🟢 {sal_id}" if is_active else f"🔴 {sal_id}"
            rcols[0].markdown(id_label)
            rcols[1].markdown(row.get("nama_salesman", "-"))
            rcols[2].markdown(f"`{row.get('salesman_type', '-')}`")
            rcols[3].markdown(row.get("no_hp", "-") or "-")
            rcols[4].markdown(row.get("region", "-") or "-")
            rcols[5].markdown(row.get("asm", "-") or "-")

            if is_active:
                if rcols[6].button("✏️ Edit", key=f"edit_{sal_id}", use_container_width=True):
                    st.session_state.action_mode = (
                        None
                        if st.session_state.action_mode == ("edit", sal_id)
                        else ("edit", sal_id)
                    )
                    st.rerun()
                if rcols[7].button("🔄 Ganti", key=f"rep_{sal_id}", use_container_width=True):
                    st.session_state.action_mode = (
                        None
                        if st.session_state.action_mode == ("replace", sal_id)
                        else ("replace", sal_id)
                    )
                    st.rerun()
                if rcols[8].button(
                    "❌ Nonaktif", key=f"deact_{sal_id}", use_container_width=True
                ):
                    st.session_state.action_mode = (
                        None
                        if st.session_state.action_mode == ("deactivate", sal_id)
                        else ("deactivate", sal_id)
                    )
                    st.rerun()
            else:
                rcols[6].markdown("—")
                rcols[7].markdown("—")
                rcols[8].markdown("—")

            # ── Inline Edit Panel ─────────────────────────────────────────────
            if st.session_state.action_mode == ("edit", sal_id):
                with st.container(border=True):
                    st.markdown(f"#### ✏️ Edit Info Salesman — `{sal_id}`")
                    st.info(
                        "Edit informasi salesman ini. Hanya field yang diubah yang akan diperbarui di database. "
                        "Untuk mengganti salesman sepenuhnya, gunakan tombol **🔄 Ganti**."
                    )

                    cur_nama = row.get("nama_salesman", "") or ""
                    cur_hp = row.get("no_hp", "") or ""
                    cur_dist_code = row.get("distributor_code", selected_dist_code) or selected_dist_code

                    @st.cache_data(show_spinner=False)
                    def _fetch_salesman_detail(nama: str, dist_code: str) -> dict:
                        try:
                            creds, proj = get_credentials()
                            c = bigquery.Client(credentials=creds, project=proj)
                            q = f"""
                                SELECT * FROM `{SALESMAN_TABLE}`
                                WHERE UPPER(TRIM(nama_salesman))    = UPPER(TRIM(@nama))
                                  AND UPPER(TRIM(kode_distributor)) = UPPER(TRIM(@dist_code))
                                ORDER BY uploaded_at DESC LIMIT 1
                            """
                            jc = bigquery.QueryJobConfig(
                                query_parameters=[
                                    bigquery.ScalarQueryParameter("nama", "STRING", nama),
                                    bigquery.ScalarQueryParameter("dist_code", "STRING", dist_code),
                                ]
                            )
                            rows = list(c.query(q, job_config=jc).result())
                            return dict(rows[0]) if rows else {}
                        except Exception:
                            return {}

                    detail = _fetch_salesman_detail(cur_nama, cur_dist_code)

                    def _s(key, fallback=""):
                        v = detail.get(key, fallback)
                        return "" if v is None else str(v)

                    def _d(key):
                        v = detail.get(key)
                        if v is None:
                            return datetime.today().date()
                        try:
                            return pd.to_datetime(v).date()
                        except Exception:
                            return datetime.today().date()

                    def _n(key, fallback=0):
                        v = detail.get(key, fallback)
                        try:
                            return float(v) if v is not None else fallback
                        except Exception:
                            return fallback

                    with st.form(f"form_edit_{sal_id}"):
                        st.markdown("**Informasi Dasar**")
                        ec1, ec2 = st.columns(2)
                        with ec1:
                            e_nama = st.text_input(
                                "Nama Salesman *", value=cur_nama, key=f"e_nama_{sal_id}"
                            )
                            e_spv_ext = st.text_input(
                                "Nama SPV External",
                                value=_s("nama_spv_external"),
                                key=f"e_spvext_{sal_id}",
                            )
                            e_spv_int = st.text_input(
                                "Nama SPV Internal *",
                                value=_s("nama_spv_internal"),
                                key=f"e_spvint_{sal_id}",
                            )
                            e_spv_int2 = st.text_input(
                                "Nama SPV Internal 2",
                                value=_s("nama_spv_internal_2"),
                                key=f"e_spvint2_{sal_id}",
                            )
                            e_hp = st.text_input(
                                "No. HP *",
                                value=cur_hp,
                                key=f"e_hp_{sal_id}",
                                placeholder="08123456789",
                            )
                            e_status = st.selectbox(
                                "Status Salesman *",
                                STATUS_OPTIONS,
                                index=(
                                    STATUS_OPTIONS.index(_s("status_salesman"))
                                    if _s("status_salesman") in STATUS_OPTIONS
                                    else 0
                                ),
                                key=f"e_status_{sal_id}",
                            )
                            e_outlet = st.number_input(
                                "Total Outlet Coverage PJP *",
                                min_value=0,
                                step=1,
                                value=int(_n("total_outlet_coverage_pjp")),
                                key=f"e_outlet_{sal_id}",
                            )
                            e_gaji = st.number_input(
                                "Gaji Pokok (Rp) *",
                                min_value=0,
                                step=1000,
                                value=int(_n("gaji_pokok")),
                                key=f"e_gaji_{sal_id}",
                            )
                            e_tunj = st.number_input(
                                "Tunjangan dan Insentif (Rp) *",
                                min_value=0,
                                step=1000,
                                value=int(_n("tunjangan_dan_insentif")),
                                key=f"e_tunj_{sal_id}",
                            )
                        with ec2:
                            e_lahir = st.date_input(
                                "Tanggal Lahir *",
                                value=_d("tanggal_lahir"),
                                min_value=datetime(1950, 1, 1).date(),
                                key=f"e_lahir_{sal_id}",
                            )
                            e_gender = st.selectbox(
                                "Jenis Kelamin *",
                                GENDER_OPTIONS,
                                index=(
                                    GENDER_OPTIONS.index(_s("jenis_kelamin"))
                                    if _s("jenis_kelamin") in GENDER_OPTIONS
                                    else 0
                                ),
                                key=f"e_gender_{sal_id}",
                            )
                            e_pendidikan = st.selectbox(
                                "Pendidikan Terakhir *",
                                EDUCATION_OPTIONS,
                                index=(
                                    EDUCATION_OPTIONS.index(_s("pendidikan_terakhir"))
                                    if _s("pendidikan_terakhir") in EDUCATION_OPTIONS
                                    else 0
                                ),
                                key=f"e_pend_{sal_id}",
                            )
                            e_exp = st.number_input(
                                "Pengalaman Sebelumnya (bulan) *",
                                min_value=0,
                                step=1,
                                value=int(_n("pengalaman_bulan")),
                                key=f"e_exp_{sal_id}",
                            )
                            e_principal = st.text_input(
                                "Principal Lain (opsional)",
                                value=_s("principal_lain"),
                                key=f"e_principal_{sal_id}",
                            )
                            e_join = st.date_input(
                                "Tanggal Join di G2G *",
                                value=_d("tanggal_join_g2g"),
                                key=f"e_join_{sal_id}",
                            )

                        submitted_edit = st.form_submit_button(
                            "💾 Simpan Perubahan", type="primary"
                        )

                    if submitted_edit:
                        edit_errors = []
                        if not e_nama.strip():
                            edit_errors.append("Nama Salesman wajib diisi.")
                        if not e_spv_int.strip():
                            edit_errors.append("Nama SPV Internal wajib diisi.")
                        if not e_hp.strip():
                            edit_errors.append("No. HP wajib diisi.")

                        if edit_errors:
                            for err in edit_errors:
                                st.error(err)
                        else:
                            updated = {
                                "nama_salesman": sanitize_salesman_name(e_nama),
                                "nama_spv_external": e_spv_ext.strip().upper() if e_spv_ext.strip() else None,
                                "nama_spv_internal": e_spv_int.strip().upper(),
                                "nama_spv_internal_2": e_spv_int2.strip().upper() if e_spv_int2.strip() else None,
                                "no_hp": normalize_phone_id(e_hp),
                                "status_salesman": e_status,
                                "total_outlet_coverage_pjp": int(e_outlet),
                                "gaji_pokok": float(e_gaji),
                                "tunjangan_dan_insentif": float(e_tunj),
                                "tanggal_lahir": e_lahir,
                                "jenis_kelamin": e_gender,
                                "pendidikan_terakhir": e_pendidikan,
                                "pengalaman_bulan": int(e_exp),
                                "principal_lain": e_principal.strip() if e_principal.strip() else None,
                                "tanggal_join_g2g": e_join,
                            }

                            with st.spinner("Menyimpan perubahan..."):
                                ok_e, err_e = update_salesman_record(
                                    cur_nama, cur_dist_code, updated
                                )

                            if not ok_e:
                                st.error(f"Gagal menyimpan perubahan: {err_e}")
                            else:
                                new_nama = sanitize_salesman_name(e_nama)
                                if new_nama != cur_nama:
                                    try:
                                        creds2, proj2 = get_credentials()
                                        c2 = bigquery.Client(
                                            credentials=creds2, project=proj2
                                        )
                                        upd_map_q = f"""
                                            UPDATE `{MAPPING_TABLE}`
                                            SET salesman   = @new_nama,
                                                updated_at = @ts
                                            WHERE salesman_id      = @sid
                                              AND is_active         = TRUE
                                              AND UPPER(TRIM(distributor_code)) = UPPER(TRIM(@dist_code))
                                        """
                                        jc2 = bigquery.QueryJobConfig(
                                            query_parameters=[
                                                bigquery.ScalarQueryParameter("new_nama", "STRING", new_nama),
                                                bigquery.ScalarQueryParameter("ts", "STRING", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")),
                                                bigquery.ScalarQueryParameter("sid", "STRING", sal_id),
                                                bigquery.ScalarQueryParameter("dist_code", "STRING", cur_dist_code),
                                            ]
                                        )
                                        c2.query(upd_map_q, job_config=jc2).result()
                                    except Exception:
                                        pass

                                st.success(
                                    f"✅ Info salesman **{new_nama}** (`{sal_id}`) berhasil diperbarui."
                                )
                                st.session_state.action_mode = None
                                _fetch_salesman_detail.clear()
                                st.cache_data.clear()
                                st.session_state.pop("salesman_df", None)
                                st.session_state.pop("_cached_dist", None)
                                st.session_state.pop("pjp_salesman_df", None)
                                st.session_state.pop("_pjp_cached_dist", None)
                                st.session_state.pop("pjp_salesman_all_df", None)
                                st.rerun()

            # ── Inline Replace Panel ──────────────────────────────────────────
            if st.session_state.action_mode == ("replace", sal_id):
                with st.container(border=True):
                    st.markdown(f"#### 🔄 Ganti Salesman — `{sal_id}`")
                    st.info(
                        f"Mengganti: **{row.get('nama_salesman', '-')}** | Tipe: `{row.get('salesman_type', '-')}`\n\n"
                        "Mapping lama akan dinonaktifkan, lalu mapping baru dibuat dengan kode yang sama."
                    )
                    with st.form(f"form_replace_{sal_id}"):
                        st.markdown("**Data Salesman Pengganti**")
                        fields_r = _render_salesman_form_fields(f"rep_{sal_id}")
                        submitted_rep = st.form_submit_button(
                            "🔄 Simpan Penggantian", type="primary"
                        )

                    if submitted_rep:
                        errs = _validate_salesman_fields(fields_r)
                        if errs:
                            for e in errs:
                                st.error(e)
                        else:
                            sal_data_r = _build_salesman_data(
                                fields_r, dist_df, selected_dist_code, selected_dist_name
                            )
                            with st.spinner("Menyimpan..."):
                                ok1, err1 = insert_salesman_record(sal_data_r)
                            if not ok1:
                                st.error(f"Gagal menyimpan data salesman: {err1}")
                            else:
                                with st.spinner("Menonaktifkan mapping lama..."):
                                    ok2, err2 = deactivate_previous_mapping(sal_id)
                                if not ok2:
                                    st.error(
                                        f"Data baru tersimpan, tapi gagal menonaktifkan mapping lama: {err2}"
                                    )
                                else:
                                    with st.spinner("Membuat mapping baru..."):
                                        ok3, err3 = insert_mapping_record(
                                            sal_id,
                                            selected_dist_code,
                                            str(row.get("salesman_type", "")),
                                            nama_salesman=sanitize_salesman_name(
                                                fields_r["nama"]
                                            ),
                                        )
                                    if not ok3:
                                        st.error(
                                            f"Mapping lama dinonaktifkan, tapi gagal membuat mapping baru: {err3}"
                                        )
                                    else:
                                        st.success(
                                            f"✅ Salesman berhasil diganti! Kode `{sal_id}` kini dipegang oleh **{fields_r['nama'].strip().upper()}**."
                                        )
                                        st.session_state.action_mode = None
                                        st.cache_data.clear()
                                        st.session_state.pop("salesman_df", None)
                                        st.session_state.pop("_cached_dist", None)
                                        st.session_state.pop("pjp_salesman_df", None)
                                        st.session_state.pop("_pjp_cached_dist", None)
                                        st.session_state.pop("pjp_salesman_all_df", None)
                                        st.rerun()

            # ── Inline Deactivate Panel ───────────────────────────────────────
            if st.session_state.action_mode == ("deactivate", sal_id):
                with st.container(border=True):
                    st.markdown(f"#### ❌ Non-Aktifkan — `{sal_id}`")
                    st.warning(
                        f"Anda akan menonaktifkan **{row.get('nama_salesman', sal_id)}**. "
                        "Tindakan ini akan menandai mapping sebagai tidak aktif."
                    )
                    dcols = st.columns([3, 1])
                    confirm = dcols[0].checkbox(
                        f"Saya konfirmasi ingin menonaktifkan salesman ini",
                        key=f"confirm_deact_{sal_id}",
                    )
                    if dcols[1].button(
                        "❌ Non-Aktifkan",
                        key=f"do_deact_{sal_id}",
                        type="primary",
                        disabled=not confirm,
                        use_container_width=True,
                    ):
                        with st.spinner("Menonaktifkan salesman..."):
                            ok_d, err_d = deactivate_salesman_mapping(sal_id)
                        if ok_d:
                            st.success(
                                f"✅ Salesman **{row.get('nama_salesman', sal_id)}** (`{sal_id}`) berhasil dinonaktifkan."
                            )
                            st.session_state.action_mode = None
                            st.cache_data.clear()
                            st.session_state.pop("salesman_df", None)
                            st.session_state.pop("_cached_dist", None)
                            st.session_state.pop("pjp_salesman_df", None)
                            st.session_state.pop("_pjp_cached_dist", None)
                            st.session_state.pop("pjp_salesman_all_df", None)
                            st.rerun()
                        else:
                            st.error(f"Gagal menonaktifkan salesman: {err_d}")

            st.divider()

    else:
        st.info("Belum ada data salesman untuk distributor ini.")

    # ── Add New Salesman ───────────────────────────────────────────────────────
    st.markdown("---")
    if "show_add_form" not in st.session_state:
        st.session_state.show_add_form = False

    add_btn_label = (
        "➕ Tambah Salesman Baru"
        if not st.session_state.show_add_form
        else "✖ Tutup Form Tambah"
    )
    if st.button(add_btn_label, type="primary", use_container_width=True):
        st.session_state.show_add_form = not st.session_state.show_add_form
        st.session_state.action_mode = None
        st.rerun()

    if st.session_state.show_add_form:
        with st.container(border=True):
            st.subheader("➕ Tambah Salesman Baru")
            st.info(
                "Sistem akan otomatis membuat **Salesman ID** baru berdasarkan:\n"
                "`Tipe Salesman + Kode Distributor + Nomor Urut`"
            )
            salesman_type_add = st.selectbox(
                "Tipe Salesman *", SALESMAN_TYPES, key="add_type_bottom"
            )
            preview_id = generate_salesman_id(selected_dist_code, salesman_type_add)
            st.info(f"ID yang akan dibuat: **`{preview_id}`**")

            with st.form("form_add_salesman_bottom"):
                st.markdown("**Data Salesman**")
                fields_add = _render_salesman_form_fields("add_bottom")
                submitted_add = st.form_submit_button(
                    "✅ Simpan Salesman Baru", type="primary"
                )

            if submitted_add:
                errs = _validate_salesman_fields(fields_add)
                if errs:
                    for e in errs:
                        st.error(e)
                else:
                    sal_data_add = _build_salesman_data(
                        fields_add, dist_df, selected_dist_code, selected_dist_name
                    )
                    salesman_id_new = generate_salesman_id(
                        selected_dist_code, salesman_type_add
                    )
                    with st.spinner("Menyimpan data salesman..."):
                        ok1, err1 = insert_salesman_record(sal_data_add)
                    if not ok1:
                        st.error(f"Gagal menyimpan ke tabel salesman: {err1}")
                    else:
                        with st.spinner("Membuat mapping salesman..."):
                            ok2, err2 = insert_mapping_record(
                                salesman_id_new,
                                selected_dist_code,
                                salesman_type_add,
                                nama_salesman=sanitize_salesman_name(fields_add["nama"]),
                            )
                        if not ok2:
                            st.error(
                                f"Data salesman tersimpan, tapi gagal membuat mapping: {err2}"
                            )
                        else:
                            st.success(
                                f"✅ Salesman baru berhasil ditambahkan!\n\n**ID Salesman: `{salesman_id_new}`**"
                            )
                            st.session_state.show_add_form = False
                            for key in [
                                "add_bottom_status",
                                "add_bottom_nama",
                                "add_bottom_hp",
                            ]:
                                st.session_state.pop(key, None)
                            st.cache_data.clear()
                            st.session_state.pop("salesman_df", None)
                            st.session_state.pop("_cached_dist", None)
                            st.session_state.pop("pjp_salesman_df", None)
                            st.session_state.pop("_pjp_cached_dist", None)
                            st.session_state.pop("pjp_salesman_all_df", None)
                            st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: PJP TEMPLATE
# ══════════════════════════════════════════════════════════════════════════════

elif PAGES[selected_page] == "pjp_template":
    st.title("🗓️ PJP Template")
    st.caption(f"Distributor: **{selected_dist_name}** ({selected_dist_code})")

    tab_download, tab_update, tab_view = st.tabs(
        [
            "📥 Download Template",
            "🔄 Update PJP",
            "📋 Lihat PJP",
        ]
    )

    # ── Tab 1: Download ───────────────────────────────────────────────────────
    with tab_download:
        st.subheader("📥 Download PJP Template")

        with st.expander("📖 Panduan Pengisian", expanded=False):
            st.markdown("""
            ### ⚡ ATURAN DASAR:
            - **1 file = 1 distributor** (tidak boleh campur)
            - **Semua kolom "Wajib Diisi" harus terisi**
            - **Jangan edit kolom hasil auto-fill** (Nama Salesman, Nama Toko, Region, ASM, Nama Distributor, Kode Distributor)

            ### 🔄 URUTAN PENGISIAN (WAJIB!):
            1. **Salesman ID** (pilih dari dropdown — daftar berisi seluruh salesman aktif dari semua distributor, cari dengan mengetik ID-nya) → Nama Salesman otomatis terisi
            2. **Kode Toko** (pilih dari dropdown — daftar berisi seluruh toko dari semua distributor, cari dengan mengetik kode tokonya) → Nama Toko, Region, ASM, Nama Distributor, Kode Distributor otomatis terisi

            ### ⚠️ VALIDASI KEPEMILIKAN DISTRIBUTOR:
            Walaupun dropdown menampilkan seluruh Salesman ID dan Kode Toko dari semua distributor,
            saat file diupload sistem akan tetap memvalidasi bahwa Salesman ID dan Kode Toko yang
            dipilih benar-benar **milik distributor `{selected_dist_code}`**. Baris yang tidak sesuai
            akan ditolak saat upload.

            ### ✅ FORMAT DATA YANG BENAR:
            - **Frekuensi PJP**: F4+ / F4 / F2 / F1
            - **Hari**: Pilih dari dropdown
            - **Minggu**: Pilih Ganjil / Genap / Ganjil+Genap

            ### 📞 BUTUH BANTUAN?
            Hubungi tim support G2G
            """)

        # ── Full (all-distributor) datasets used to populate the Excel
        # dropdowns — per the new requirement, Salesman ID and Kode Toko
        # are no longer scoped to `selected_dist_code` here. Distributor
        # ownership is still enforced separately at upload-validation time
        # (Tab 2, using the distributor-scoped datasets there).
        pjp_salesman_df_excel = st.session_state.pjp_salesman_all_df
        pjp_store_df_excel = store_df

        # Distributor-scoped subsets, kept only to warn if THIS distributor
        # currently has no active salesman / stores of its own — purely
        # informational, does not affect what appears in the Excel dropdown.
        _dist_salesman_df = st.session_state.pjp_salesman_df
        _dist_store_df = store_df[store_df["distributor_code"] == selected_dist_code]

        if _dist_salesman_df.empty:
            st.warning(
                "⚠️ Tidak ada Salesman ID aktif yang tercatat untuk distributor ini. "
                "Dropdown Salesman ID pada template tetap menampilkan seluruh salesman aktif dari semua "
                "distributor, namun baris dengan Salesman ID yang bukan milik distributor ini akan ditolak saat upload."
            )
        if _dist_store_df.empty:
            st.warning(
                "⚠️ Tidak ada data toko yang tercatat untuk distributor ini. "
                "Dropdown Kode Toko pada template tetap menampilkan seluruh toko dari semua distributor, "
                "namun baris dengan Kode Toko yang bukan milik distributor ini akan ditolak saat upload."
            )

        @st.cache_data(show_spinner="Menyiapkan template Excel...")
        def _cached_pjp_excel(
            _salesman_df, _store_df, _dist_code, _dist_name, _dist_asm, _dist_region
        ):
            return create_pjp_excel(
                pd.DataFrame(columns=[c for c, _, _ in PJP_COLS]),
                _salesman_df,
                _store_df,
                _dist_code,
                _dist_name,
                _dist_asm,
                _dist_region,
            )

        pjp_excel = _cached_pjp_excel(
            pjp_salesman_df_excel,
            pjp_store_df_excel,
            selected_dist_code,
            selected_dist_name,
            selected_dist_asm,
            selected_dist_region,
        )

        st.download_button(
            "⬇️ Download PJP Template",
            data=pjp_excel.getvalue(),
            file_name=f"PJP_Template_{selected_dist_code}_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )

    # ── Tab 2: Update PJP ─────────────────────────────────────────────────────
    with tab_update:
        st.subheader("🔄 Update Daftar PJP")
        st.info(
            "Fitur ini memungkinkan **User** untuk memperbarui data PJP dengan memilih "
            "berdasarkan **Distributor** atau **Salesman** tertentu. "
            "Data PJP **bulan berjalan** (ditentukan otomatis dari tanggal sistem) untuk filter yang "
            "dipilih akan **dihapus** dan diganti dengan data baru dari file yang diupload. "
            "**Data PJP bulan-bulan sebelumnya tidak pernah dihapus** dan tetap tersimpan sebagai riwayat — "
            "tidak perlu memilih bulan/tahun secara manual."
        )

        st.markdown("### 1️⃣ Pilih Scope Update")
        update_scope = st.radio(
            "Update berdasarkan:",
            ["Distributor (semua salesman)", "Salesman tertentu"],
            key="pjp_update_scope",
            horizontal=True,
        )

        scope_dist_code = None
        scope_salesman = None

        if update_scope == "Distributor (semua salesman)":
            st.markdown(
                f"Scope: semua PJP milik distributor **{selected_dist_name}** (`{selected_dist_code}`) "
                "akan diganti."
            )
            scope_dist_code = selected_dist_code

        else:
            st.markdown("Pilih salesman yang PJP-nya ingin diperbarui:")

            sal_list_df = st.session_state.get("salesman_df", pd.DataFrame())

            if sal_list_df.empty:
                st.warning("Tidak ada salesman aktif untuk distributor ini.")
                st.stop()

            if "is_active" in sal_list_df.columns:
                active_sal = sal_list_df[sal_list_df["is_active"] == True]
            else:
                active_sal = sal_list_df

            sal_name_options = sorted(
                active_sal["nama_salesman"].dropna().astype(str).unique().tolist()
            )
            if not sal_name_options:
                st.warning("Tidak ada salesman aktif yang ditemukan.")
                st.stop()

            chosen_salesman = st.selectbox(
                "Pilih Salesman:",
                sal_name_options,
                key="pjp_update_salesman_select",
            )
            scope_salesman = chosen_salesman
            scope_dist_code = selected_dist_code

        st.markdown("### 2️⃣ Upload File PJP Baru")
        st.warning(
            "File yang diupload **harus menggunakan template resmi**. "
            "Sistem akan memvalidasi bahwa setiap Salesman ID dan Kode Toko yang dipilih "
            "benar-benar milik distributor yang dipilih di atas — baris yang bukan milik "
            "distributor ini akan ditolak."
        )

        uploaded_update = st.file_uploader(
            "Pilih file Excel PJP baru (.xlsx)", type=["xlsx"], key="pjp_update_uploader"
        )

        if uploaded_update:
            try:
                xl_u = pd.ExcelFile(uploaded_update)
            except Exception as e:
                st.error(f"Gagal membaca file: {e}")
                st.stop()

            if "PJP Template" not in xl_u.sheet_names:
                st.error(
                    "Sheet 'PJP Template' tidak ditemukan. Gunakan template resmi."
                )
                st.stop()

            # Use the FULL (unscoped, all-distributor) salesman/store
            # datasets for the actual existence lookup — NOT a
            # distributor-pre-filtered subset. Pre-filtering before the
            # lookup conflates two different checks ("does this
            # store/salesman exist at all?" vs "does it belong to the
            # selected distributor?"): if a store's distributor_code field
            # doesn't line up exactly with selected_dist_code for any
            # reason, it silently disappears from a pre-filtered subset and
            # gets wrongly reported as "not found" even though it's a
            # perfectly real store — just (apparently) under a different
            # distributor. Looking it up in the full dataset first, and
            # then checking ownership separately in validate_pjp_df() by
            # comparing the row's own distributor_code to
            # selected_dist_code, gives accurate diagnostics either way.
            pjp_salesman_lookup_df = st.session_state.pjp_salesman_all_df
            pjp_store_lookup_df = store_df

            try:
                pjp_new_df = read_template_sheet(
                    uploaded_update, "PJP Template", 2, pjp_salesman_lookup_df, pjp_store_lookup_df
                )
                # Determine "row has data" from the RAW Kode Toko input,
                # not from the derived/looked-up "Nama Distributor" column.
                # Nama Distributor is re-computed via a VLOOKUP-equivalent
                # against pjp_store_df (scoped to selected_dist_code) — if
                # that lookup fails to match a row's Kode Toko for any
                # reason, Nama Distributor silently becomes "", which used
                # to cause the row (and potentially ALL rows) to be dropped
                # here with no explanation. Using the raw "kode_toko" column
                # instead means genuinely filled rows always survive to the
                # per-row validation below, where a real lookup miss is
                # reported explicitly (e.g. "Kode Toko 'X' tidak ditemukan
                # di master store.") instead of vanishing silently.
                _row_key_col = "kode_toko" if "kode_toko" in pjp_new_df.columns else "Kode Toko"
                pjp_new_df = pjp_new_df[
                    pjp_new_df[_row_key_col].notna()
                    & (pjp_new_df[_row_key_col].astype(str).str.strip() != "")
                ]
                pjp_new_df = pjp_new_df.reset_index(drop=True)
                # Tag every uploaded row with the CURRENT snapshot month
                # (YYYY-MM, from system date at upload time) — this is what
                # drives the automatic replace-current/preserve-history
                # logic below. No manual month/year selection needed.
                pjp_new_df["snapshot_month"] = datetime.now().strftime("%Y-%m")
            except Exception as e:
                st.error(f"Gagal membaca sheet: {e}")
                st.stop()

            if pjp_new_df.empty:
                st.warning("Tidak ada data di file yang diupload.")
                st.stop()

            pjp_u_errors, pjp_u_warnings = validate_pjp_df(
                pjp_new_df,
                distributor_map,
                store_df=pjp_store_lookup_df,
                salesman_df=pjp_salesman_lookup_df,
                selected_dist_code=selected_dist_code,
            )

            if pjp_u_errors or pjp_u_warnings:
                if pjp_u_errors:
                    st.error(f"**❌ {len(pjp_u_errors)} ERROR:**")
                    for e in pjp_u_errors:
                        st.markdown(f"- {e}")
                if pjp_u_warnings:
                    st.warning(f"**⚠️ {len(pjp_u_warnings)} PERINGATAN:**")
                    for w in pjp_u_warnings:
                        st.markdown(f"- {w}")
                st.error("Perbaiki semua masalah di atas sebelum melanjutkan update.")
            else:
                st.success("✅ Validasi data baru berhasil!")

                st.markdown("### 3️⃣ Konfirmasi & Eksekusi Update")

                _current_snapshot_month = datetime.now().strftime("%Y-%m")
                scope_label = (
                    f"semua PJP distributor **{selected_dist_name}**"
                    if scope_salesman is None
                    else f"PJP salesman **{scope_salesman}** di distributor **{selected_dist_name}**"
                )

                st.info(
                    f"📅 Snapshot bulan berjalan: **{_current_snapshot_month}** "
                    "(ditentukan otomatis dari tanggal sistem, tidak perlu dipilih manual)."
                )
                st.warning(
                    f"⚠️ Proses ini akan **menghapus data PJP bulan {_current_snapshot_month}** "
                    f"untuk {scope_label}, kemudian memasukkan **{len(pjp_new_df)} baris baru** "
                    f"sebagai snapshot bulan {_current_snapshot_month}. "
                    "**Data PJP bulan-bulan sebelumnya tetap aman dan tidak akan terhapus.** "
                    "Tindakan ini **tidak dapat dibatalkan**."
                )

                confirm_update = st.checkbox(
                    "Saya memahami bahwa data PJP bulan berjalan akan dihapus dan diganti "
                    "(data bulan sebelumnya tetap tersimpan).",
                    key="confirm_pjp_update",
                )

                if st.button(
                    "🔄 Update PJP",
                    key="exec_pjp_update",
                    type="primary",
                    disabled=not confirm_update,
                ):
                    with st.spinner(f"Menghapus data PJP bulan {_current_snapshot_month}..."):
                        ok_del, err_del = delete_pjp_records(
                            distributor_code=scope_dist_code,
                            salesman_name=scope_salesman,
                        )

                    if not ok_del:
                        st.error(f"Gagal menghapus data PJP bulan berjalan: {err_del}")
                    else:
                        with st.spinner("Memasukkan data PJP baru..."):
                            ok_ins, msg_ins = push_to_bigquery(
                                pjp_new_df, _PJP_COL_MAP, PJP_TABLE
                            )

                        if ok_ins:
                            st.success(
                                f"✅ Update PJP berhasil! Snapshot bulan **{_current_snapshot_month}** "
                                f"diganti dengan **{len(pjp_new_df)} baris baru**. "
                                "Data bulan-bulan sebelumnya tetap tersimpan."
                            )
                            st.cache_data.clear()
                        else:
                            st.error(
                                f"Data bulan {_current_snapshot_month} sudah dihapus, tetapi gagal "
                                f"memasukkan data baru: {msg_ins}\n\n"
                                "Segera hubungi administrator untuk memulihkan data."
                            )

    # ── Tab 3: Lihat PJP ─────────────────────────────────────────────────────
    with tab_view:
        st.subheader("📋 Lihat Data PJP")

        col_month, col_refresh = st.columns([3, 1])
        with col_month:
            current_ym = datetime.now().strftime("%Y-%m")
            # Build last 6 months as options
            month_options = [
                (datetime.now().replace(day=1) - pd.DateOffset(months=i)).strftime("%Y-%m")
                for i in range(6)
            ]
            selected_view_month = st.selectbox(
                "Snapshot Bulan",
                month_options,
                index=0,
                key="pjp_view_month",
            )
        with col_refresh:
            st.markdown("<br>", unsafe_allow_html=True)
            refresh_view = st.button("🔄 Refresh", key="pjp_view_refresh", use_container_width=True)

        pjp_view_df = get_pjp_list(distributor_code=selected_dist_code)

        if not pjp_view_df.empty and "snapshot_month" in pjp_view_df.columns:
            pjp_view_df = pjp_view_df[pjp_view_df["snapshot_month"] == selected_view_month]

        if pjp_view_df.empty:
            st.info(f"Tidak ada data PJP untuk distributor **{selected_dist_name}** pada bulan **{selected_view_month}**.")
        else:
            st.caption(f"Menampilkan **{len(pjp_view_df)} baris** PJP · {selected_dist_name} · {selected_view_month}")

            # Column order mirrors the template's own Frekuensi -> Hari ->
            # Minggu -> Ket. Minggu flow. `callcycle` is the authoritative
            # week pattern for anything uploaded since the redesign;
            # `minggu` is the frozen legacy column, still shown (labelled)
            # so rows predating that change remain readable instead of
            # appearing to have no week pattern at all. The Minggu (week
            # parity) shown here is back-derived from callcycle for
            # display only — it is never a stored column.
            if "callcycle" in pjp_view_df.columns:
                pjp_view_df = pjp_view_df.copy()
                pjp_view_df["minggu_kategori"] = pjp_view_df["callcycle"].apply(
                    derive_minggu_from_callcycle
                )

            display_cols = [c for c in [
                "salesman_id", "nama_salesman", "kode_toko", "nama_toko",
                "frekuensi", "hari", "minggu_kategori", "callcycle",
                "minggu", "kode_distributor", "snapshot_month"
            ] if c in pjp_view_df.columns]
            st.dataframe(
                pjp_view_df[display_cols],
                use_container_width=True,
                hide_index=True,
                column_config={
                    "minggu_kategori": st.column_config.TextColumn(
                        "Minggu (kolom K)",
                        help="Kategori minggu — diturunkan dari callcycle untuk tampilan saja, tidak disimpan sebagai kolom database.",
                    ),
                    "callcycle": st.column_config.TextColumn(
                        "callcycle (Ket. Minggu)",
                        help="Kolom L template. F1: satu angka (1/3 bila Ganjil, 2/4 bila Genap). F2: 1,3 atau 2,4. F4: 1,2,3,4. F4+: 1,2,3,4,5.",
                    ),
                    "minggu": st.column_config.TextColumn(
                        "minggu (lama/legacy)",
                        help="Kolom lama, tidak lagi diisi oleh upload baru. Hanya untuk baris sebelum perubahan format.",
                    ),
                },
            )

            csv_bytes = pjp_view_df[display_cols].to_csv(index=False).encode("utf-8")
            st.download_button(
                "⬇️ Download CSV",
                data=csv_bytes,
                file_name=f"PJP_{selected_dist_code}_{selected_view_month}.csv",
                mime="text/csv",
            )
