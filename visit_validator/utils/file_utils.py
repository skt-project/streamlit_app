"""
file_utils.py — File I/O helpers.

Handles reading uploaded files (Excel / CSV), exporting DataFrames to Excel,
and generating the downloadable template workbook.
"""

from io import BytesIO
from typing import List, Optional, Tuple

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Read uploaded file ─────────────────────────────────────────────────────────

def read_file(
    file_obj: BytesIO,
    filename: str,
    sheet_name: Optional[str] = None,
) -> Tuple[pd.DataFrame, Optional[str]]:
    """
    Parse an uploaded file into a DataFrame.
    Returns (df, None) on success or (empty_df, error_message) on failure.
    """
    try:
        ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
        if ext == "csv":
            df = pd.read_csv(file_obj)
        elif ext == "xlsx":
            df = pd.read_excel(file_obj, sheet_name=sheet_name, engine="openpyxl")
        elif ext == "xls":
            df = pd.read_excel(file_obj, sheet_name=sheet_name)
        else:
            return pd.DataFrame(), f"Unsupported file format: .{ext}"
        return df.reset_index(drop=True), None
    except Exception as exc:
        return pd.DataFrame(), str(exc)


def get_sheet_names(file_obj: BytesIO) -> List[str]:
    """Return all sheet names from an Excel file without consuming the buffer."""
    try:
        xl = pd.ExcelFile(file_obj)
        return xl.sheet_names
    except Exception:
        return []


# ── Export helpers ─────────────────────────────────────────────────────────────

def to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Results") -> bytes:
    """Serialise a DataFrame to Excel and return raw bytes for st.download_button."""
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name[:31])
    return buf.getvalue()


# ── Template generation ────────────────────────────────────────────────────────

def generate_template() -> bytes:
    """
    Build a two-sheet Excel template:
      Sheet 1 — Instructions with usage guidelines and an example row.
      Sheet 2 — Empty Template with the required headers pre-formatted.

    Returns raw bytes so it can be passed directly to st.download_button.
    """
    wb = openpyxl.Workbook()

    # ── Shared style objects ──────────────────────────────────────────────────
    BLUE = "1E6B8A"
    DARK = "1E3A4A"
    WHITE = "FFFFFF"
    ACCENT = "E8F4FD"

    header_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
    accent_fill = PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type="solid")
    header_font = Font(color=WHITE, bold=True, size=11)
    title_font = Font(bold=True, size=14, color=DARK)
    section_font = Font(bold=True, size=12, color=BLUE)
    body_font = Font(size=11, color=DARK)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="D0D8E0")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ═══════════════════════════════════════════════════════════════════
    # SHEET 1 — Instructions
    # ═══════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 56
    for col in ["B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 22

    # Title banner
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = "📍 Visit Validator — Usage Instructions"
    title_cell.font = title_font
    title_cell.alignment = center
    title_cell.fill = accent_fill
    ws.row_dimensions[1].height = 36

    # Guidelines section header
    ws.cell(row=3, column=1, value="GUIDELINES").font = section_font
    ws.row_dimensions[3].height = 24

    guidelines = [
        "1. Do not modify the header names.",
        "2. Latitude and longitude must be in decimal format.",
        "3. Use a period (.) as the decimal separator, not a comma.",
        "4. Ensure both salesman and store coordinates are valid.",
        "5. The validation radius is 1 kilometre.",
        "6. Records where the salesman GPS is (0, 0) are classified as MISSING GPS.",
        "7. The completed file can be uploaded directly into the application.",
    ]
    for i, text in enumerate(guidelines, start=4):
        c = ws.cell(row=i, column=1, value=text)
        c.font = body_font
        c.alignment = left_align
        ws.row_dimensions[i].height = 18

    # Example section
    ws.cell(row=12, column=1, value="EXAMPLE DATA").font = section_font
    ws.row_dimensions[12].height = 24

    example_headers = [
        "Store_Latitude",
        "Store_Longitude",
        "Salesman_Latitude",
        "Salesman_Longitude",
    ]
    for col_idx, header in enumerate(example_headers, start=1):
        c = ws.cell(row=13, column=col_idx, value=header)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = cell_border
    ws.row_dimensions[13].height = 22

    example_values = [-6.208763, 106.845599, -6.208500, 106.845700]
    for col_idx, val in enumerate(example_values, start=1):
        c = ws.cell(row=14, column=col_idx, value=val)
        c.alignment = center
        c.border = cell_border
        c.font = body_font
    ws.row_dimensions[14].height = 18

    # Result columns note
    ws.cell(row=16, column=1, value="OUTPUT COLUMNS APPENDED BY THE APP").font = section_font
    ws.row_dimensions[16].height = 24

    output_headers = ["Distance_KM", "Visit_Status", "Validation_Remark"]
    output_examples = [
        (0.35, "VALID VISIT", "Salesman is within the store radius"),
        (1.87, "INVALID VISIT", "Salesman is outside the store radius"),
        ("", "MISSING GPS", "Salesman GPS not captured (coordinates are 0, 0)"),
    ]
    for col_idx, header in enumerate(output_headers, start=1):
        c = ws.cell(row=17, column=col_idx, value=header)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = cell_border

    for row_offset, row_values in enumerate(output_examples, start=18):
        for col_idx, val in enumerate(row_values, start=1):
            c = ws.cell(row=row_offset, column=col_idx, value=val)
            c.font = body_font
            c.border = cell_border

    # ═══════════════════════════════════════════════════════════════════
    # SHEET 2 — Template
    # ═══════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Template")
    ws2.sheet_view.showGridLines = False

    template_headers = [
        "ID",
        "Activity_Start",
        "Date_and_Time",
        "Place_ID",
        "Place",
        "Address",
        "Representative_ID",
        "Representative",
        "Salesman_Longitude",   # col I — matches source file layout
        "Salesman_Latitude",    # col J
        "Store_Longitude",      # col K
        "Store_Latitude",       # col L
        "Check_In_Type",
    ]
    col_widths = [38, 22, 22, 18, 36, 40, 20, 30, 22, 22, 22, 22, 18]

    for col_idx, (header, width) in enumerate(zip(template_headers, col_widths), start=1):
        c = ws2.cell(row=1, column=col_idx, value=header)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = cell_border
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    ws2.row_dimensions[1].height = 26
    ws2.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
