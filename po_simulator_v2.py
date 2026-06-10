import streamlit as st
import pandas as pd
from datetime import datetime
import io
import zipfile
import math
import re
import base64
import urllib.request
from pathlib import Path
import openpyxl
import numpy as np
from typing import List
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from google.cloud import bigquery
from google.oauth2 import service_account
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
import os

BASE_DIR = os.path.dirname(__file__)

pdfmetrics.registerFont(TTFont('Trebuchet', os.path.join(BASE_DIR, 'trebuc.ttf')))
pdfmetrics.registerFont(TTFont('Trebuchet-Bold', os.path.join(BASE_DIR, 'trebucbd.ttf')))

try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    MATPLOTLIB_OK = True
except ImportError:
    MATPLOTLIB_OK = False

st.set_page_config(
    page_title="PO Simulator - G2G",
    layout="wide",
    page_icon='📁',
    initial_sidebar_state="expanded",
)

DASHBOARD_URL_DEFAULT = "https://po-simulator.streamlit.app/"
TEMPLATE_DRIVE_URL = "https://docs.google.com/spreadsheets/d/1_4SFn2_SvGm1on0EJkntYjC2cLvNZyDjX54zcQAWRtQ/edit?gid=0#gid=0"
TEMPLATE_PO_URL = "https://docs.google.com/spreadsheets/d/1_4SFn2_SvGm1on0EJkntYjC2cLvNZyDjX54zcQAWRtQ/edit?gid=0#gid=0"
BQ_DATASET = "rsa"
BQ_TABLE = "stock_analysis"

try:
    gcp_secrets = st.secrets["connections"]["bigquery"]
    private_key = gcp_secrets["private_key"].replace("\\n", "\n")
    _bq_credentials = service_account.Credentials.from_service_account_info({
        "type": gcp_secrets["type"],
        "project_id": gcp_secrets["project_id"],
        "private_key_id": gcp_secrets["private_key_id"],
        "private_key": private_key,
        "client_email": gcp_secrets["client_email"],
        "client_id": gcp_secrets["client_id"],
        "auth_uri": gcp_secrets["auth_uri"],
        "token_uri": gcp_secrets["token_uri"],
        "auth_provider_x509_cert_url": gcp_secrets["auth_provider_x509_cert_url"],
        "client_x509_cert_url": gcp_secrets["client_x509_cert_url"],
    })
    GCP_PROJECT_ID = st.secrets["bigquery"]["project"]
except Exception:
    GCP_CREDENTIALS_PATH = r"C:\Users\Shaltsa Nadya\Documents\try python\streamlit\skintific-data-warehouse-ea77119e2e7a.json"
    GCP_PROJECT_ID = "skintific-data-warehouse"
    _bq_credentials = service_account.Credentials.from_service_account_file(GCP_CREDENTIALS_PATH)


# ─── BigQuery ─────────────────────────────────────────────────────────────────

@st.cache_resource(show_spinner=False)
def get_bq_client() -> bigquery.Client:
    return bigquery.Client(credentials=_bq_credentials, project=GCP_PROJECT_ID)


@st.cache_data(ttl=21600, show_spinner=False)
def get_zero_price_skus() -> set:
    try:
        client = get_bq_client()
        query = "SELECT UPPER(sku) as sku FROM `skintific-data-warehouse.gt_schema.master_product` WHERE price_for_distri = 0 AND brand = 'G2G'"
        rows = client.query(query).result()
        return {r.sku for r in rows if r.sku}
    except Exception:
        return set()


@st.cache_data(ttl=3600, show_spinner=False)
def fetch_customer_names() -> list:
    try:
        client = get_bq_client()
        table_id = f"{GCP_PROJECT_ID}.{BQ_DATASET}.{BQ_TABLE}"
        rows = client.query(f"SELECT DISTINCT distributor FROM `{table_id}` ORDER BY distributor").result()
        return [r.distributor for r in rows if r.distributor]
    except Exception as e:
        st.warning(f"⚠️ Gagal memuat daftar distributor: {e}")
        return []


CUSTOMER_NAMES = fetch_customer_names()


@st.cache_data(ttl=21600, show_spinner="Fetching SKU data from BigQuery...")
def get_sku_data(sku_list) -> pd.DataFrame:
    if not sku_list:
        return pd.DataFrame()
    client = get_bq_client()
    sku_list_str = ", ".join([f"'{sku}'" for sku in sku_list])
    query = f"SELECT sku, product_name, price_for_distri FROM `{GCP_PROJECT_ID}.gt_schema.master_product` WHERE UPPER(sku) IN ({sku_list_str})"
    try:
        return client.query(query).to_dataframe()
    except Exception as e:
        st.error(f"Error fetching SKU data: {e}")
        return pd.DataFrame()


@st.cache_data(ttl=21600, show_spinner="Fetching NPD data from BigQuery...")
def get_npd_data(sku_list) -> pd.DataFrame:
    if not sku_list:
        return pd.DataFrame()
    client = get_bq_client()
    query = f"SELECT calendar_date, region, sku FROM `{GCP_PROJECT_ID}.gt_schema.npd_allocation` WHERE calendar_date = '2026-06-01'"
    try:
        return client.query(query).to_dataframe()
    except Exception as e:
        st.error(f"Error fetching NPD data: {e}")
        return pd.DataFrame()


@st.cache_data(ttl=21600, show_spinner="Fetching suggestions from BigQuery...")
def get_distributor_suggestions(distributor_names, brand_name: str = "All") -> pd.DataFrame:
    if isinstance(distributor_names, str):
        distributor_names = [distributor_names]
    distributor_names = [d for d in distributor_names if d and d != "(Pilih Distributor)"]
    if not distributor_names:
        return pd.DataFrame()
    client = get_bq_client()
    table_id = f"{GCP_PROJECT_ID}.{BQ_DATASET}.{BQ_TABLE}"
    _brand_filter = f"AND UPPER(brand) = UPPER('{brand_name}')" if brand_name and brand_name != "All" else ""
    _dist_str = ", ".join([f"'{d.upper()}'" for d in distributor_names])
    _REJECT_APPROVAL = ["G2G-840","G2G-844","G2G-841","G2G-800","G2G-213","G2G-217","G2G-30701","G2G-30702","G2G-30703","G2G-30704","G2G-243"]
    _NO_TOL = ["G2G-2721","G2G-224","G2G-226","G2G-228","G2G-74"]
    _steve_str = ", ".join([f"'{s.upper()}'" for s in _REJECT_APPROVAL])
    _no_tol_str = ", ".join([f"'{s.upper()}'" for s in _NO_TOL])
    query = f"""
    SELECT UPPER(region) AS REGION, UPPER(distributor) AS DISTRIBUTOR, UPPER(sku) AS SKU,
        ROUND(current_woi_by_lm,2) AS CURRENT_WOI,
        buffer_plan_by_lm_qty_adj AS SUGGESTION_QTY,
        ROUND(SAFE_DIVIDE(COALESCE(total_stock,0)+COALESCE(buffer_plan_by_lm_qty_adj,0),NULLIF(avg_weekly_st_lm_qty,0)),2) AS WOI_AFTER_PO,
        remaining_allocation_qty_region AS REMAINING_ALLOCATION,
        CASE
            WHEN UPPER(sku) IN ({_steve_str}) AND remaining_allocation_qty_region > 0 THEN 'Reject by Steve'
            WHEN remaining_allocation_qty_region > 0 THEN 'Terdapat Alokasi'
            WHEN remaining_allocation_qty_region <= 0 THEN 'Alokasi Habis'
            ELSE NULL
        END AS STATUS_ALOKASI
    FROM `{table_id}`
    WHERE UPPER(distributor) IN ({_dist_str}) AND buffer_plan_by_lm_qty_adj > 0
    {_brand_filter} AND SKU NOT IN ({_no_tol_str})
    ORDER BY DISTRIBUTOR, SUGGESTION_QTY DESC
    """
    try:
        return client.query(query).to_dataframe()
    except Exception as e:
        st.error(f"Error fetching suggestions: {e}")
        return pd.DataFrame()


@st.cache_data(ttl=21600, show_spinner="Fetching Stock Analysis from BigQuery...")
def get_stock_data(distributor_name: str, sku_list) -> pd.DataFrame:
    if not sku_list:
        return pd.DataFrame()
    client = get_bq_client()
    table_id = f"{GCP_PROJECT_ID}.{BQ_DATASET}.{BQ_TABLE}"
    sku_list_str = ", ".join([f"'{sku}'" for sku in sku_list])
    query = f"""
    SELECT UPPER(region) AS region, UPPER(distributor) AS distributor, sku,
        assortment, supply_control_status_gt, total_stock,
        buffer_plan_by_lm_qty_adj, avg_weekly_st_lm_qty,
        buffer_plan_by_lm_val_adj, remaining_allocation_qty_region, woi_end_of_month_by_lm
    FROM `{table_id}`
    WHERE UPPER(distributor) = '{distributor_name}'
    AND (UPPER(sku) IN ({sku_list_str}) OR buffer_plan_by_lm_qty_adj > 0)
    """
    try:
        return client.query(query).to_dataframe()
    except Exception as e:
        st.error(f"Error fetching stock data: {e}")
        return pd.DataFrame()


@st.cache_data(ttl=21600, show_spinner=False)
def get_brand_list() -> list:
    try:
        client = get_bq_client()
        rows = client.query(f"SELECT DISTINCT brand FROM `{GCP_PROJECT_ID}.gt_schema.master_product` WHERE brand IS NOT NULL ORDER BY brand").result()
        return [r.brand for r in rows if r.brand]
    except Exception:
        return []


# ─── Helpers ──────────────────────────────────────────────────────────────────

def calculate_woi(stock, po_qty, avg_weekly_sales):
    return np.where(avg_weekly_sales > 0, (stock + po_qty) / avg_weekly_sales, 0)


def apply_sku_rejection_rules(sku_list, df, regions, is_in):
    regions_upper = [r.upper() for r in regions]
    if "SKU" not in df.columns or "region" not in df.columns:
        return df
    if not is_in:
        condition = df["SKU"].isin(sku_list) & ~df["region"].str.upper().isin(regions_upper)
    else:
        condition = df["SKU"].isin(sku_list) & df["region"].str.upper().isin(regions_upper)
    df.loc[condition, "Remark"] = "Reject (Stop by Steve)"
    return df


def _drive_to_direct(url: str) -> str:
    m = re.search(r'/spreadsheets/d/([a-zA-Z0-9_-]+)', url)
    if m:
        return f"https://docs.google.com/spreadsheets/d/{m.group(1)}/export?format=xlsx"
    m = re.search(r'/file/d/([a-zA-Z0-9_-]+)', url)
    if m:
        return f"https://drive.usercontent.google.com/download?id={m.group(1)}&export=download&authuser=0"
    return url


@st.cache_data(show_spinner=False)
def _fetch_template_bytes(url: str) -> bytes:
    if not url or not url.strip():
        raise ValueError("TEMPLATE_DRIVE_URL belum diset.")
    direct = _drive_to_direct(url)
    req = urllib.request.Request(direct, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        data = resp.read()
    if not data or data[:4] != b'PK\x03\x04':
        raise ValueError("File yang didownload bukan xlsx yang valid.")
    return data


def _logo_src() -> str:
    local = Path(__file__).parent / "logo.png"
    if local.exists():
        b64 = base64.b64encode(local.read_bytes()).decode()
        return f"data:image/png;base64,{b64}"
    return "https://glad2glow.com/cdn/shop/files/logo.png?height=628&v=1745724802&width=1200"


LOGO_URL = _logo_src()


@st.cache_data(show_spinner=False)
def create_po_template_excel() -> bytes:
    try:
        return _fetch_template_bytes(TEMPLATE_PO_URL)
    except Exception:
        wb = Workbook()
        ws = wb.active
        ws.title = "PO Template"
        hdr_fill = PatternFill(start_color="BF3979", end_color="BF3979", fill_type="solid")
        hdr_font = Font(bold=True, color="FFFFFF")
        for col_idx, col_name in enumerate(["DISTRIBUTOR", "PRODUCT CODE", "DESCRIPTION", "QTY", "DPP"], 1):
            cell = ws.cell(row=8, column=col_idx, value=col_name)
            cell.fill = hdr_fill
            cell.font = hdr_font
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

PO_TEMPLATE_COLS = [
    'Distributor','SKU','Product Name','Assortment','Supply Control',
    'Avg Weekly Sales LM (Qty)','Total Stock (Qty)','Current WOI',
    'PO Qty','PO Value','WOI (Stock + PO Ori)','Remark',
    'Suggested PO Qty','Suggested PO Value',
    'WOI After Buffer (Stock + Suggested Qty)',
    'Stock + Suggested Qty WOI (Projection at EOM)',
    'Remaining Allocation (By Region)','RSA Notes',
]
PO_IMG_COLS = [PO_TEMPLATE_COLS[1], PO_TEMPLATE_COLS[2]] + PO_TEMPLATE_COLS[6:14]
PO_COLS_copy = PO_TEMPLATE_COLS[:13]


def _sanitize_xlsx_bytes(xlsx_bytes: bytes) -> bytes:
    _src = io.BytesIO(xlsx_bytes)
    _dst = io.BytesIO()
    try:
        with zipfile.ZipFile(_src, "r") as _zin:
            names = _zin.namelist()
            rels_name = "xl/_rels/workbook.xml.rels"
            drop_rids: set = set()
            if rels_name in names:
                _rels_txt = _zin.read(rels_name).decode("utf-8", "ignore")
                for _m in re.finditer(r'<Relationship\b[^>]*?Id="([^"]+)"[^>]*?Type="[^"]*externalLink[^"]*"[^>]*/>', _rels_txt):
                    drop_rids.add(_m.group(1))
            with zipfile.ZipFile(_dst, "w", zipfile.ZIP_DEFLATED) as _zout:
                for name in names:
                    if name.startswith("xl/externalLinks/"):
                        continue
                    data = _zin.read(name)
                    if name == "xl/workbook.xml":
                        _txt = data.decode("utf-8", "ignore")
                        _txt = re.sub(r"<externalReferences>.*?</externalReferences>", "", _txt, flags=re.DOTALL)
                        data = _txt.encode("utf-8")
                    elif name == rels_name and drop_rids:
                        _txt = data.decode("utf-8", "ignore")
                        for _rid in drop_rids:
                            _txt = re.sub(rf'<Relationship\b[^>]*?Id="{re.escape(_rid)}"[^>]*/>', "", _txt)
                        data = _txt.encode("utf-8")
                    _zout.writestr(name, data)
        return _dst.getvalue()
    except Exception:
        return xlsx_bytes


def _excel_engine(fname: str) -> str:
    if fname.lower().endswith('.xls'):
        import importlib.util
        if importlib.util.find_spec('xlrd') is None:
            raise ImportError("Library `xlrd` tidak ditemukan.")
        return 'xlrd'
    return 'openpyxl'


def detect_header_row(file_bytes: bytes, fname: str = "", max_scan: int = 15, sheet_name=0) -> int:
    engine = _excel_engine(fname)
    df_raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, header=None, engine=engine, dtype=str, nrows=max_scan)
    best_row, best_score = 0, -1
    for i in range(len(df_raw)):
        vals = [str(v).strip() for v in df_raw.iloc[i].values if pd.notna(v) and str(v).strip()]
        text_count = sum(1 for v in vals if not v.replace('.','',1).replace(',','',1).lstrip('-').isdigit())
        score = text_count * 10 + len(vals)
        if score > best_score:
            best_score = score; best_row = i
    return best_row


def _get_sheet_names(file_bytes: bytes, engine: str) -> list:
    try:
        if engine == 'xlrd':
            import xlrd
            book = xlrd.open_workbook(file_contents=file_bytes)
            if hasattr(book, 'sheet_visibility'):
                return [book.sheet_name(i) for i in range(book.nsheets) if book.sheet_visibility[i] == 0]
            return [book.sheet_name(i) for i in range(book.nsheets)]
        else:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            sheets = [ws.title for ws in wb.worksheets if ws.sheet_state == 'visible']
            wb.close()
            return sheets
    except Exception:
        return []


def _convert_to_xlsx(fname: str, fbytes: bytes):
    ext = fname.rsplit(".", 1)[-1].lower()
    base = fname.rsplit(".", 1)[0]
    if ext == "xlsx":
        return fname, fbytes
    elif ext == "xls":
        try:
            import xlrd as _xlrd
            _book = _xlrd.open_workbook(file_contents=fbytes)
            _wb = openpyxl.Workbook()
            _wb.remove(_wb.active)
            for _si in range(_book.nsheets):
                _ws_old = _book.sheet_by_index(_si)
                _ws_new = _wb.create_sheet(title=_ws_old.name)
                for _r in range(_ws_old.nrows):
                    for _c in range(_ws_old.ncols):
                        _ws_new.cell(row=_r+1, column=_c+1, value=_ws_old.cell_value(_r, _c))
            _buf = io.BytesIO(); _wb.save(_buf)
            return base + ".xlsx", _buf.getvalue()
        except Exception:
            pass
        _all_sheets = pd.read_excel(io.BytesIO(fbytes), sheet_name=None, dtype=object)
        _buf = io.BytesIO()
        with pd.ExcelWriter(_buf, engine="openpyxl") as _wr:
            for _nm, _df in _all_sheets.items():
                _df.to_excel(_wr, sheet_name=str(_nm)[:31], index=False)
        return base + ".xlsx", _buf.getvalue()
    elif ext == "csv":
        _df = pd.read_csv(io.BytesIO(fbytes), dtype=str, encoding_errors="replace")
        _buf = io.BytesIO()
        _df.to_excel(_buf, index=False, engine="openpyxl")
        return base + ".xlsx", _buf.getvalue()
    return fname, fbytes


def _edit_qty_via_excel_com(xlsx_bytes, sheet_name, hdr_row_0, sku_col_name, qty_col_name, cell_writer):
    try:
        import tempfile
        import win32com.client as _wc
        import pythoncom
    except Exception:
        return None
    _fin_path = None; _xl = None
    try:
        pythoncom.CoInitialize()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as _fin:
            _fin.write(xlsx_bytes); _fin_path = _fin.name
        _xl = _wc.DispatchEx("Excel.Application")
        _xl.Visible = False; _xl.DisplayAlerts = False
        _wb = _xl.Workbooks.Open(_fin_path, UpdateLinks=0)
        _ws = None
        for _s in _wb.Worksheets:
            if _s.Name == sheet_name: _ws = _s; break
        if _ws is None: _ws = _wb.Worksheets(1)
        _hdr_row = hdr_row_0 + 1
        _used = _ws.UsedRange
        _max_col = _used.Columns.Count + _used.Column - 1
        _max_row = _used.Rows.Count + _used.Row - 1
        _sku_ci = _qty_ci = None
        for _c in range(1, _max_col+1):
            _v = _ws.Cells(_hdr_row, _c).Value
            if _v == sku_col_name: _sku_ci = _c
            elif _v == qty_col_name: _qty_ci = _c
        changed = 0
        if _sku_ci and _qty_ci:
            for _r in range(_hdr_row+1, _max_row+1):
                _sv = str(_ws.Cells(_r, _sku_ci).Value or "").strip()
                _qv = _ws.Cells(_r, _qty_ci).Value
                _new = cell_writer(_sv, _qv)
                if _new is not None:
                    _ws.Cells(_r, _qty_ci).Value = _new; changed += 1
        _wb.Save(); _wb.Close(SaveChanges=False)
        with open(_fin_path, "rb") as _f:
            return _sanitize_xlsx_bytes(_f.read()), changed
    except Exception:
        return None
    finally:
        try:
            if _xl: _xl.Quit()
        except Exception: pass
        if _fin_path:
            try: os.remove(_fin_path)
            except Exception: pass
        try: pythoncom.CoUninitialize()
        except Exception: pass


def numeric_coerce(df: pd.DataFrame) -> pd.DataFrame:
    for col in df.columns:
        try:
            numeric_vals = pd.to_numeric(df[col], errors='coerce')
            if numeric_vals.notna().sum() / max(df[col].notna().sum(), 1) > 0.8:
                df[col] = numeric_vals
        except Exception:
            pass
    return df


def gsheet_to_csv_url(url: str):
    match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', url)
    if not match:
        return None, None
    sheet_id = match.group(1)
    gid_match = re.search(r'gid=(\d+)', url)
    gid = gid_match.group(1) if gid_match else '0'
    csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
    return csv_url, sheet_id


def create_zip_of_files(file_dict: dict) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for fname, data in file_dict.items():
            zf.writestr(fname, data)
    return buf.getvalue()

# ─── Excel export with styling ────────────────────────────────────────────────

def _write_po_rows(ws, df_no_flag, is_po_sku_series, npd_sku_list=None):
    rows = list(dataframe_to_rows(df_no_flag, index=False, header=True))
    headers = list(df_no_flag.columns)
    col_map = {col: i for i, col in enumerate(headers)}
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    po_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    suggestion_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    npd_fill = PatternFill(start_color="B1DBF0", end_color="B1DBF0", fill_type="solid")
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    header_alignment = Alignment(horizontal='left', vertical='center')
    proceed_font = Font(bold=True, color="54CE54")
    reject_font = Font(bold=True, color="D73E3E")
    suggest_font = Font(bold=True, color="F3C94C")
    currency_cols = ["PO Value", "Suggested PO Value"]
    integer_cols = ["Remaining Allocation (By Region)", "Avg Weekly Sales LM (Qty)"]
    decimal_cols = ["WOI (Stock + PO Ori)", "Current WOI",
                    "WOI After Buffer (Stock + Suggested Qty)",
                    "Stock + Suggested Qty WOI (Projection at EOM)"]

    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            if r_idx == 1:
                cell.font = header_font; cell.fill = header_fill; cell.alignment = header_alignment
            else:
                orig_idx = r_idx - 2
                col_name = headers[c_idx - 1] if c_idx - 1 < len(headers) else ""
                if c_idx <= 11:
                    is_po_row = is_po_sku_series.iloc[orig_idx]
                    cell.fill = po_fill if is_po_row else suggestion_fill
                if col_name in currency_cols: cell.number_format = "#,##0.00"
                elif col_name in integer_cols: cell.number_format = "#,##0"
                elif col_name in decimal_cols: cell.number_format = "0.00"
                if col_name == "Remark":
                    val = str(cell.value or "")
                    if "Proceed" in val: cell.font = proceed_font
                    elif "Reject" in val: cell.font = reject_font
                    elif "Additional" in val: cell.font = suggest_font
                if col_name in ["Remaining Allocation (By Region)", "Suggested PO Qty", "Suggested PO Value"] and npd_sku_list:
                    sku_ci = col_map.get("SKU")
                    if sku_ci is not None:
                        sku_val = ws.cell(row=r_idx, column=sku_ci + 1).value
                        if sku_val in npd_sku_list:
                            cell.fill = npd_fill


def to_excel_with_styling(dfs: dict, npd_sku_list=None) -> bytes:
    output = io.BytesIO()
    wb = Workbook()
    del wb["Sheet"]
    for sheet_name, df in dfs.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        is_po_sku_series = df["is_po_sku"]
        df_no_flag = df.drop("is_po_sku", axis=1)
        _write_po_rows(ws, df_no_flag, is_po_sku_series, npd_sku_list)
    wb.save(output); output.seek(0)
    return output.getvalue()


def to_excel_single_sheet(df: pd.DataFrame, npd_sku_list=None) -> bytes:
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active; ws.title = "PO Simulator"
    is_po_sku_series = df["is_po_sku"]
    df_no_flag = df.drop("is_po_sku", axis=1)
    _write_po_rows(ws, df_no_flag, is_po_sku_series, npd_sku_list)
    wb.save(output); output.seek(0)
    return output.getvalue()


def to_excel_single_sheet_with_sku(df: pd.DataFrame, npd_sku_list=None, sku_master_df=None) -> bytes:
    return to_excel_single_sheet(df, npd_sku_list)


# ─── Image export ─────────────────────────────────────────────────────────────

_REMARK_STYLES = [
    ('reject with suggestion', '#FFF3CD', '#856404'),
    ('reject (stop by steve',  '#F8D7DA', '#721C24'),
    ('reject',                 '#F8D7DA', '#721C24'),
    ('proceed with suggestion','#D4EDDA', '#155724'),
    ('proceed',                '#D4EDDA', '#155724'),
    ('additional suggestion',  '#D1ECF1', '#0C5460'),
]


def df_to_image_bytes(df: pd.DataFrame, title: str = "") -> bytes:
    if not MATPLOTLIB_OK:
        raise RuntimeError("matplotlib tidak tersedia.")
    n_rows, n_cols = df.shape
    fig_w = max(16, n_cols * 2.0)
    fig_h = max(2.5, n_rows * 0.38 + 1.6)
    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.axis('off')
    if title:
        ax.set_title(title, fontsize=11, fontweight='bold', color='#1a1a2e', pad=10)
    cell_text = [[str(v) if pd.notna(v) else '' for v in row] for row in df.values]
    tbl = ax.table(cellText=cell_text, colLabels=df.columns.tolist(), loc='center', cellLoc='left')
    tbl.auto_set_font_size(False); tbl.set_fontsize(7.5)
    tbl.auto_set_column_width(col=list(range(n_cols)))
    for j in range(n_cols):
        tbl[0, j].set_facecolor('#1a1a2e')
        tbl[0, j].set_text_props(color='white', fontweight='bold')
    remark_idx = df.columns.tolist().index('Remark') if 'Remark' in df.columns else None
    for i in range(1, n_rows + 1):
        row_bg = '#EBF5FB' if i % 2 == 0 else '#FFFFFF'
        for j in range(n_cols):
            tbl[i, j].set_facecolor(row_bg)
            tbl[i, j].set_text_props(color='#1a1a2e')
        if remark_idx is not None:
            val = str(df.iloc[i-1, remark_idx]).lower().strip()
            for keyword, bg, fg in _REMARK_STYLES:
                if keyword in val:
                    tbl[i, remark_idx].set_facecolor(bg)
                    tbl[i, remark_idx].set_text_props(color=fg, fontweight='bold')
                    break
    fig.patch.set_facecolor('white')
    buf = io.BytesIO()
    plt.savefig(buf, format='jpg', bbox_inches='tight', dpi=150, facecolor='white')
    buf.seek(0); plt.close(fig)
    return buf.getvalue()


# ─── Auth ─────────────────────────────────────────────────────────────────────

def check_password():
    def login_form():
        st.markdown("""<style>
        html,body,[data-testid="stAppViewContainer"]{background:linear-gradient(145deg,#F13E93 40%,#FFA6A6 50%,#F26076 60%)!important;}
        [data-testid="stSidebar"]{display:none!important;}
        </style>""", unsafe_allow_html=True)
        _, col, _ = st.columns([1, 1.5, 1])
        with col:
            st.markdown(f'<div style="text-align:center;padding:.5rem 0 .2rem;"><img src="{LOGO_URL}" style="max-width:200px;" /></div>', unsafe_allow_html=True)
            st.markdown("""<div style="text-align:center;margin:1rem 0 1.5rem;">
                <div style="font-size:1.5rem;font-weight:700;color:#6E253A;">DataFlow Automator</div>
                <div style="color:#000;font-size:.85rem;margin-top:.3rem;">Masukkan passwordmu</div>
            </div>""", unsafe_allow_html=True)
            with st.form("login_form"):
                password = st.text_input("🔒 Password", type="password", placeholder="Masukkan password...")
                submitted = st.form_submit_button("Login", use_container_width=True)
                if submitted:
                    if password == st.secrets["glowithyou"]:
                        st.session_state["authenticated"] = True
                        st.rerun()
                    else:
                        st.error("❌ Password salah.")
            st.markdown('<div style="text-align:center;margin-top:1.5rem;color:#6E253A;font-size:.72rem;">Glad2Glow</div>', unsafe_allow_html=True)

    if st.session_state.get("authenticated"):
        return True
    login_form()
    return False


#if not check_password():
#    st.stop()


# ─── CSS ──────────────────────────────────────────────────────────────────────

st.markdown("""<style>
body,p,div,span,label,input,textarea,select,button,h1,h2,h3,h4,h5,h6,li,td,th,caption,small,strong,em{font-family:'Trebuchet MS',sans-serif;box-sizing:border-box;}
[data-testid="stSidebarCollapseButton"]{display:none!important;}
:root{--rose:#CA6180;--rose-dark:#A84D6A;--blush:#FCB7C7;--dark:#BF3979;--g-rose:linear-gradient(135deg,#751F58,#A84D6A);}
h1,h2,h3,h4,h5,h6{color:#FFFFFF!important;}
[data-testid="stSidebar"]{background:var(--dark)!important;box-shadow:4px 0 24px rgba(0,0,0,.35);}
[data-testid="stSidebar"],[data-testid="stSidebar"] *{color:#FFFFFF!important;}
[data-testid="stSidebar"] input{background:rgba(255,255,255,.1)!important;border:1px solid rgba(255,255,255,.2)!important;color:#2D1B26!important;border-radius:8px!important;}
[data-testid="stSidebar"] [data-baseweb="select"]>div{background:rgba(255,255,255,.1)!important;border:1px solid rgba(255,255,255,.2)!important;color:#FFFFFF!important;}
[data-testid="stMain"] input,[data-testid="stMain"] textarea{color:#FFFFFF!important;background:rgba(255,255,255,.15)!important;border:1.5px solid rgba(255,255,255,.35)!important;border-radius:10px!important;}
[data-testid="stMain"] .stSelectbox>div>div{color:#FFFFFF!important;background:rgba(255,255,255,.15)!important;border:1.5px solid rgba(255,255,255,.35)!important;border-radius:10px!important;}
[data-testid="stSidebar"] [data-testid="stButton-nav_po"] button {background: #FF89AA !important;  /* PO Simulator - RSA */
}
[data-testid="stMain"] [data-baseweb="select"] span,[data-testid="stMain"] [data-baseweb="select"] div{color:#FFFFFF!important;}
.hero-wrap{padding:.5rem 0 1.5rem;} .hero-tag{display:inline-block;background:rgba(168,77,106,.6);color:#FFFFFF!important;font-size:.72rem;font-weight:700;letter-spacing:1.5px;text-transform:uppercase;padding:.25rem .8rem;border-radius:20px;border:1px solid rgba(255,255,255,.25);margin-bottom:.75rem;}
.hero-title{font-size:2rem;font-weight:700;color:#FFFFFF!important;margin-bottom:.5rem;}
.pipeline-step{background:#A84D6A;border:none;border-radius:12px;padding:.9rem 1.3rem;margin-bottom:1rem;box-shadow:0 2px 12px rgba(0,0,0,.15);}
.pipeline-step *,.pipeline-step strong,.pipeline-step code{color:#FFFFFF!important;}
.pipeline-step.active{background:#8B2040;}
.step-number{display:inline-flex;align-items:center;justify-content:center;width:24px;height:24px;border-radius:50%;background:var(--g-rose);color:#FFF!important;font-weight:700;font-size:.75rem;margin-right:.6rem;vertical-align:middle;}
.metric-card{background:rgba(168,77,106,.35);border:1px solid rgba(255,255,255,.2);border-radius:12px;padding:1.1rem 1.3rem;position:relative;overflow:hidden;}
.metric-card::before{content:'';position:absolute;top:0;left:0;width:3px;height:100%;background:rgba(255,255,255,.5);}
.badge{display:inline-block;padding:.18rem .65rem;border-radius:20px;font-size:.65rem;font-weight:700;text-transform:uppercase;}
.badge-info{background:rgba(255,255,255,.2);color:#FFFFFF!important;border:1px solid rgba(255,255,255,.3);}
.stDownloadButton>button,.stButton>button{background:var(--g-rose)!important;color:#FFF!important;border:none!important;border-radius:10px!important;font-weight:700!important;padding:.55rem 1.4rem!important;}
hr{border-color:#E8EAED!important;margin:1.2rem 0!important;}
</style>""", unsafe_allow_html=True)

st.markdown("""<style>
html,body,[data-testid="stAppViewContainer"],[data-testid="stMain"],.main,.block-container{background:#FFFFFF!important;}
[data-testid="stHeader"]{background:#FFFFFF!important;}
[data-testid="stMain"] *:not(svg):not(path):not(button){color:#1F1F1F!important;}
[data-testid="stMain"] input,[data-testid="stMain"] textarea{color:#1F1F1F!important;background:#FFFFFF!important;border:1.5px solid #C9C9C9!important;}
[data-testid="stMain"] .stSelectbox>div>div{color:#1F1F1F!important;background:#FFFFFF!important;border:1.5px solid #C9C9C9!important;}
[data-testid="stMain"] [data-baseweb="select"] span,[data-testid="stMain"] [data-baseweb="select"] div{color:#1F1F1F!important;}
[data-testid="stMain"] .stDownloadButton>button,[data-testid="stMain"] .stButton>button{background:#D9D9D9!important;color:#111111!important;border:1px solid #B0B0B0!important;font-weight:700!important;}
[data-testid="stMain"] .pipeline-step{background:#F5EEF2!important;border:1px solid #E6D7DF!important;box-shadow:none!important;}
[data-testid="stMain"] .pipeline-step.active{background:#F0D9E2!important;}
[data-testid="stMain"] .pipeline-step,[data-testid="stMain"] .pipeline-step *,[data-testid="stMain"] .pipeline-step strong,[data-testid="stMain"] .pipeline-step code{color:#5A1E38!important;}
[data-testid="stMain"] .pipeline-step .step-number{background:#8B2040!important;color:#FFFFFF!important;}
[data-testid="stMain"] hr{margin:.6rem 0!important;}
[data-testid="stMain"] .stAlert,[data-testid="stMain"] .stAlert *{color:#1F1F1F!important;}
</style>""", unsafe_allow_html=True)


# ─── Sidebar ──────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown(f'<div style="text-align:center;padding:.5rem 0 .2rem;"><img src="{LOGO_URL}" style="max-width:200px;height:auto;" /></div>', unsafe_allow_html=True)

    if 'page' not in st.session_state:
        st.session_state['page'] = 'spv'

    st.markdown("<div style='padding:0 .4rem .3rem;text-align:center;font-size:1rem;font-weight:700;letter-spacing:2.5px;color:rgba(255,255,255,.35);text-transform:uppercase;'>MENU</div>", unsafe_allow_html=True)

    if st.button("📍 Request PO - SPV", use_container_width=True, key="nav_spv"):
        st.session_state['page'] = 'spv'; st.rerun()
    if st.button("📍 PO Simulator - SPV", use_container_width=True, key="nav_po_spv"):
        st.session_state['page'] = 'po_spv'; st.rerun()

    st.markdown("<div style='height:50px;'></div>", unsafe_allow_html=True)
    if st.button("🔒 PO Simulator - RSA", use_container_width=True, key="nav_po"):
        st.session_state['page'] = 'po_changer_login'
        st.session_state.pop('rsa_authenticated', None)
        st.rerun()

    # Shared constants used in PO simulation pages
    _MANUAL_REJECT_APPROVAL = ["G2G-840","G2G-844","G2G-841","G2G-800","G2G-213","G2G-217","G2G-30701","G2G-30702","G2G-30703","G2G-30704","G2G-243"]
    _MANUAL_REJECT_NO_TOL = ["G2G-2721","G2G-224","G2G-226","G2G-228","G2G-74"]
    _MANUAL_REJECT_ALL = _MANUAL_REJECT_APPROVAL + _MANUAL_REJECT_NO_TOL
    _LIMITED_SKUS_QTY = []
    ___MAX_QTY_LIMIT = 500
    _REJECTED_SKUS_1 = ["G2G-29700","G2G-27300"]
    __REGION_LIST_1 = ["Central Sumatera","Northern Sumatera","Jakarta (Csa)","West Kalimantan","South Kalimantan","East Kalimantan"]
    __REJECTED_SKUS_2 = []
    _REGION_LIST_2 = []

    st.markdown("<div style='height:100px;'></div>", unsafe_allow_html=True)
    st.divider()

    if st.button("Logout", key="logout_btn", use_container_width=True):
        st.session_state["authenticated"] = False
        st.session_state.pop('rsa_authenticated', None)
        for k in ("raw_df","data_source","source_type"):
            st.session_state.pop(k, None)
        st.rerun()
    st.markdown("<div style='text-align:center;color:rgba(255,255,255,.25);font-size:.62rem;margin-top:1rem;'>DataFlow v1.0 · Glad2Glow</div>", unsafe_allow_html=True)


# ─── Shared PO simulation logic ───────────────────────────────────────────────

def _run_po_simulation(sim_df, sku_col, qty_col, dist_col,
                       manual_reject_approval, manual_reject_no_tol,
                       rejected_skus_1, _region_list_1,
                       _rejected_skus_2, region_list_2,
                       limited_skus_qty, __MAX_QTY_LIMIT):
    manual_reject_all = manual_reject_approval + manual_reject_no_tol
    sim_df[qty_col] = pd.to_numeric(sim_df[qty_col], errors="coerce")
    sim_df = sim_df.dropna(subset=[qty_col])
    sim_df = sim_df[sim_df[qty_col] > 0].copy()
    sim_df[dist_col] = sim_df[dist_col].astype(str).str.strip().str.upper()
    sim_df[sku_col] = sim_df[sku_col].astype(str).str.strip().str.upper()
    sim_df = sim_df.rename(columns={dist_col:"Distributor", sku_col:"Customer SKU Code", qty_col:"PO Qty"})
    sim_df["is_po_sku"] = True
    sim_df = sim_df[["Distributor","Customer SKU Code","PO Qty","is_po_sku"]]

    all_npd = []; excel_dfs = {}
    zero_price_skus = get_zero_price_skus()
    prog = st.progress(0)
    distributors = sim_df["Distributor"].unique().tolist()

    for di, dist_name in enumerate(distributors):
        prog.progress((di+1)/len(distributors), f"Processing {dist_name}...")
        cur_po = sim_df[sim_df["Distributor"] == dist_name].copy()
        sku_list = cur_po["Customer SKU Code"].dropna().astype(str).str.strip().loc[lambda s: s.ne("") & s.str.lower().ne("nan")].unique().tolist()

        sku_df = get_sku_data(tuple(sku_list))
        stock_df = get_stock_data(dist_name, tuple(sku_list))

        if sku_df.empty and stock_df.empty:
            st.warning(f"Tidak ada data untuk distributor: {dist_name}"); continue

        sku_df = sku_df.rename(columns={"sku":"Customer SKU Code","price_for_distri":"SIP","product_name":"Product Name"})
        if "Customer SKU Code" in sku_df.columns:
            sku_df["Customer SKU Code"] = sku_df["Customer SKU Code"].astype(str).str.strip().str.upper()
        if "sku" in stock_df.columns:
            stock_df = stock_df.rename(columns={"sku":"Customer SKU Code"})
        if "Customer SKU Code" in stock_df.columns:
            stock_df["Customer SKU Code"] = stock_df["Customer SKU Code"].astype(str).str.strip().str.upper()
        stock_df = stock_df.drop(columns=["distributor","Distributor","product_name"], errors="ignore")

        skus_in_sku = set(sku_df["Customer SKU Code"].tolist()) if not sku_df.empty else set()
        skus_in_stock = set(stock_df["Customer SKU Code"].tolist()) if not stock_df.empty else set()
        skus_not_found = set(sku_list) - (skus_in_sku | skus_in_stock)
        if skus_not_found:
            skus_clean = [s for s in skus_not_found if s and pd.notna(s) and str(s).strip()]
            if skus_clean:
                st.warning(f"SKU tidak ditemukan ({dist_name}): {', '.join(sorted(skus_clean))}")

        res_df = pd.merge(cur_po, sku_df, on="Customer SKU Code", how="left")
        res_df["SIP"] = pd.to_numeric(res_df.get("SIP", 0), errors="coerce").fillna(0)
        res_df["PO Value"] = res_df["SIP"] * res_df["PO Qty"]
        res_df = pd.merge(res_df, stock_df, on="Customer SKU Code", how="outer")
        res_df["Distributor"] = dist_name

        miss_pn = res_df["Product Name"].isna()
        if miss_pn.any():
            extra_skus = res_df.loc[miss_pn, "Customer SKU Code"].unique().tolist()
            extra_df = get_sku_data(tuple(extra_skus))
            if not extra_df.empty:
                extra_df = extra_df.rename(columns={"sku":"Customer SKU Code","product_name":"Product Name","price_for_distri":"SIP"})
                extra_df["Customer SKU Code"] = extra_df["Customer SKU Code"].astype(str).str.strip().str.upper()
                name_map = extra_df.set_index("Customer SKU Code")["Product Name"].to_dict()
                res_df.loc[miss_pn, "Product Name"] = res_df.loc[miss_pn, "Customer SKU Code"].map(name_map)
                if "SIP" in extra_df.columns:
                    sip_map = extra_df.set_index("Customer SKU Code")["SIP"].to_dict()
                    miss_sip = res_df["SIP"].isna() | (res_df["SIP"] == 0)
                    res_df.loc[miss_sip, "SIP"] = res_df.loc[miss_sip, "Customer SKU Code"].map(sip_map)

        npd_df = get_npd_data(tuple(res_df["Customer SKU Code"].unique().tolist()))
        cur_npd = npd_df["sku"].unique().tolist() if not npd_df.empty else []
        all_npd = list(set(all_npd + cur_npd))
        npd_sku_upper = [s.upper() for s in cur_npd]

        res_df["is_po_sku"] = res_df["is_po_sku"].astype("boolean").fillna(False)
        for fc in ["PO Qty","PO Value","total_stock","buffer_plan_by_lm_qty_adj","avg_weekly_st_lm_qty","buffer_plan_by_lm_val_adj","remaining_allocation_qty_region","woi_end_of_month_by_lm"]:
            if fc in res_df.columns:
                res_df[fc] = pd.to_numeric(res_df[fc], errors="coerce").fillna(0)

        res_df["SIP"] = pd.to_numeric(res_df["SIP"], errors="coerce").fillna(0)
        zv = (res_df["buffer_plan_by_lm_val_adj"] == 0) & (res_df["buffer_plan_by_lm_qty_adj"] > 0) & (res_df["SIP"] > 0)
        res_df.loc[zv, "buffer_plan_by_lm_val_adj"] = res_df.loc[zv, "SIP"] * res_df.loc[zv, "buffer_plan_by_lm_qty_adj"]

        bp = res_df.get("buffer_plan_by_lm_qty_adj", pd.Series([0]*len(res_df), index=res_df.index))
        res_df = res_df[(res_df["PO Qty"] > 0) | (bp > 0)].copy()

        sugg_mask = res_df["is_po_sku"] == False
        sc_s = res_df.get("supply_control_status_gt", pd.Series([""]*len(res_df), index=res_df.index))
        ra_s = res_df.get("remaining_allocation_qty_region", pd.Series([0]*len(res_df), index=res_df.index))
        bp_s = res_df.get("buffer_plan_by_lm_qty_adj", pd.Series([0]*len(res_df), index=res_df.index))

        excl = (res_df["Customer SKU Code"].isin(skus_not_found) |
                res_df["Customer SKU Code"].isin(manual_reject_all) |
                (ra_s < 0) |
                sc_s.str.upper().isin(["STOP PO","DISCONTINUEDD","OOS","UNAVAILABLE"]) |
                (res_df["Customer SKU Code"].isin(limited_skus_qty) & (bp_s > __MAX_QTY_LIMIT)) |
                (bp_s == 0))
        if rejected_skus_1:
            reg_up = [r.upper() for r in __REGION_LIST_1]
            reg_s = res_df.get("region", pd.Series([""]*len(res_df), index=res_df.index))
            excl = excl | (res_df["Customer SKU Code"].isin(rejected_skus_1) & ~reg_s.str.upper().isin(reg_up))
        res_df = res_df[~(sugg_mask & excl)].copy()

        avg = res_df.get("avg_weekly_st_lm_qty", pd.Series([0]*len(res_df), index=res_df.index))
        stk = res_df.get("total_stock", pd.Series([0]*len(res_df), index=res_df.index))
        bp2 = res_df.get("buffer_plan_by_lm_qty_adj", pd.Series([0]*len(res_df), index=res_df.index))
        res_df["WOI PO Original"] = calculate_woi(stk, res_df["PO Qty"], avg)
        res_df["WOI Suggest"] = calculate_woi(stk, bp2, avg)
        res_df["Current WOI"] = calculate_woi(stk, 0, avg)

        ra2 = res_df.get("remaining_allocation_qty_region", pd.Series([0]*len(res_df), index=res_df.index))
        sc2 = res_df.get("supply_control_status_gt", pd.Series([""]*len(res_df), index=res_df.index))
        bp3 = res_df.get("buffer_plan_by_lm_qty_adj", pd.Series([0]*len(res_df), index=res_df.index))
        avg2 = res_df.get("avg_weekly_st_lm_qty", pd.Series([0]*len(res_df), index=res_df.index))

        conds = [
            res_df["Customer SKU Code"].isin(zero_price_skus),
            res_df["Customer SKU Code"].isin(skus_not_found),
            res_df["Customer SKU Code"].isin(limited_skus_qty) & (res_df["PO Qty"] > __MAX_QTY_LIMIT),
            ra2 < 0,
            res_df["is_po_sku"] == False,
            res_df["Customer SKU Code"].isin(manual_reject_approval),
            res_df["Customer SKU Code"].isin(manual_reject_no_tol),
            sc2.str.upper().isin(["STOP PO","DISCONTINUEDD","OOS","UNAVAILABLE"]),
            ((avg2 == 0) & (bp3 == 0) & ~res_df["Customer SKU Code"].str.upper().isin(npd_sku_upper) & ~sc2.str.upper().isin(["STOP PO","DISCONTINUEDD","OOS"])),
            bp3 == 0,
            res_df["PO Qty"] > bp3,
            res_df["PO Qty"] < bp3,
            res_df["PO Qty"] == bp3,
        ]
        choices = [
            "Price Not Available Yet",
            "Reject (SKU Not Found in System)",
            f"Reject (Exceeds Qty Limit of {__MAX_QTY_LIMIT})",
            "Reject (Negative Allocation)",
            "Additional Suggestion",
            "Reject (Stop by Steve - Need approval email)",
            "Reject (Stop by Steve - No tolerance to open)",
            "Reject",
            "Proceed",
            "Reject",
            "Reject with suggestion",
            "Proceed with suggestion",
            "Proceed",
        ]
        res_df["Remark"] = np.select(conds, choices, default="N/A (Missing Data)")
        res_df = res_df.rename(columns={
            "Customer SKU Code":"SKU","assortment":"Assortment","supply_control_status_gt":"Supply Control",
            "total_stock":"Total Stock (Qty)","avg_weekly_st_lm_qty":"Avg Weekly Sales LM (Qty)",
            "buffer_plan_by_lm_qty_adj":"Suggested PO Qty","buffer_plan_by_lm_val_adj":"Suggested PO Value",
            "WOI PO Original":"WOI (Stock + PO Ori)","WOI Suggest":"WOI After Buffer (Stock + Suggested Qty)",
            "woi_end_of_month_by_lm":"Stock + Suggested Qty WOI (Projection at EOM)",
            "remaining_allocation_qty_region":"Remaining Allocation (By Region)",
        })
        if rejected_skus_1:
            res_df = apply_sku_rejection_rules(rejected_skus_1, res_df, __REGION_LIST_1, is_in=False)
        if __REJECTED_SKUS_2:
            res_df = apply_sku_rejection_rules(__REJECTED_SKUS_2, res_df, region_list_2, is_in=False)

        res_df["RSA Notes"] = ""
        out_cols = ["Distributor","SKU","Product Name","Assortment","Supply Control",
                    "Avg Weekly Sales LM (Qty)","Total Stock (Qty)","Current WOI",
                    "PO Qty","PO Value","WOI (Stock + PO Ori)","Remark",
                    "Suggested PO Qty","Suggested PO Value",
                    "WOI After Buffer (Stock + Suggested Qty)",
                    "Stock + Suggested Qty WOI (Projection at EOM)",
                    "Remaining Allocation (By Region)","is_po_sku","RSA Notes"]
        res_df = res_df.reindex(columns=out_cols)
        res_df.sort_values(by=["is_po_sku","SKU"], ascending=[False,True], inplace=True)

        zero_sugg = res_df[(res_df["Suggested PO Qty"] == 0) & (res_df["is_po_sku"] == True)][["SKU","Product Name","PO Qty","Suggested PO Qty","Remark"]]
        if not zero_sugg.empty:
            st.warning(f"⚠️ **{dist_name}** — {len(zero_sugg)} SKU memiliki Suggested PO Qty = 0")
            st.dataframe(zero_sugg, use_container_width=True, hide_index=True)

        excel_dfs[dist_name] = res_df.copy()

    prog.progress(1.0, "Selesai")
    return excel_dfs, all_npd


def _render_sim_results(e_dfs, e_npd, folder_res, sku_col_sim, qty_col_sim, dist_col_sim):
    final_disp = pd.concat(e_dfs.values(), ignore_index=True)

    combined_raw = folder_res["df"].copy()
    combined_raw[sku_col_sim] = combined_raw[sku_col_sim].astype(str).str.strip().str.upper()
    combined_raw[dist_col_sim] = combined_raw[dist_col_sim].astype(str).str.strip().str.upper()
    combined_raw = combined_raw.rename(columns={sku_col_sim:"SKU", dist_col_sim:"Distributor"})
    sim_cols_upper = {c.upper() for c in final_disp.columns}
    exclude = {"SKU","DISTRIBUTOR", qty_col_sim.upper()}
    extra_cols = [c for c in combined_raw.columns if c.upper() not in exclude and c.upper() not in sim_cols_upper]
    if extra_cols:
        combined_agg = combined_raw.groupby(["SKU","Distributor"], as_index=False)[extra_cols].first()
        final_disp = final_disp.merge(combined_agg, on=["SKU","Distributor"], how="left")

    woi_col = next((c for c in final_disp.columns if "woi" in c.lower() and "stock" in c.lower()), None)
    if woi_col is None:
        woi_col = next((c for c in final_disp.columns if "woi" in c.lower()), "Current WOI")

    st.markdown("""<div class="pipeline-step active"><span class="step-number">2</span><strong>Preview Data — Top 10 WOI</strong></div>""", unsafe_allow_html=True)
    prev_df = final_disp[final_disp["Remark"].str.contains("Reject", na=False)].copy()
    prev_df[woi_col] = pd.to_numeric(prev_df[woi_col], errors="coerce")
    top10 = prev_df.nlargest(10, woi_col)[PO_IMG_COLS].reset_index(drop=True)
    st.dataframe(top10.style.set_properties(**{"background-color":"#D6EAF8","color":"#1a1a2e","border":"1px solid #AED6F1"}).format(na_rep="-"), use_container_width=True, hide_index=True)

    alloc_col = next((c for c in final_disp.columns if "allocation" in c.lower()), None)
    if alloc_col:
        st.markdown("""<div class="pipeline-step active"><span class="step-number">3</span><strong>Remaining Allocation (per Distributor)</strong></div>""", unsafe_allow_html=True)
        alloc_df = final_disp.copy()
        alloc_df[alloc_col] = pd.to_numeric(alloc_df[alloc_col], errors="coerce")
        alloc_df = alloc_df[alloc_df[alloc_col].notna() & (alloc_df[alloc_col] != 0)]
        show_cols = list(dict.fromkeys([c for c in PO_IMG_COLS + [alloc_col] if c in alloc_df.columns]))
        if alloc_df.empty:
            st.info("Tidak ada baris dengan allocation tersedia.")
        else:
            for dist_alloc, grp in alloc_df.groupby("Distributor"):
                with st.expander(f"📦 {dist_alloc} — {len(grp)} baris", expanded=True):
                    st.dataframe(grp[show_cols].reset_index(drop=True), use_container_width=True, hide_index=True)

    final_step = 3 + (1 if alloc_col else 0)

    dl_data = to_excel_single_sheet_with_sku(final_disp, e_npd)
    st.download_button(label=f"Download PO Result.xlsx ({len(final_disp)} baris)", data=dl_data,
                        file_name=f"PO Result {datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

    st.markdown(f"""<div class="pipeline-step active"><span class="step-number">{final_step+1}</span><strong>Auto Kategorisasi &amp; Download Gambar</strong></div>""", unsafe_allow_html=True)

    if not MATPLOTLIB_OK:
        st.error("❌ Library matplotlib tidak tersedia."); st.stop()

    img_df = final_disp[PO_COLS_copy].copy()
    for nc in img_df.select_dtypes(include=['float','float64','float32']).columns:
        img_df[nc] = img_df[nc].round(2)
    sc_col = next((c for c in img_df.columns if "supply" in c.lower() and "control" in c.lower()), None)
    stop_mask = img_df[sc_col].str.strip().str.upper().isin(["STOP PO","OOS","DISCONTINUED","UNAVAILABLE"])
    stop_df = img_df[stop_mask].reset_index(drop=True)
    non_stop_df = img_df[~stop_mask].copy()
    steve_mask = (non_stop_df["Remark"].str.lower().str.contains("reject (stop by steve", na=False, regex=False) |
                  non_stop_df["Remark"].str.lower().str.contains("reject (negative allocation)", na=False, regex=False))
    steve_df = non_stop_df[steve_mask].reset_index(drop=True)
    approval_mask = (non_stop_df["Remark"].str.strip().str.lower().isin(["reject","reject with suggestion"]) & ~steve_mask)
    approval_df = non_stop_df[approval_mask].reset_index(drop=True)

    cat1, cat2, cat3 = st.columns(3)
    for col_ui, df_cat, title, label, icon in [
        (cat1, stop_df, "Product Stop PO", "product_stop_po", "🚫"),
        (cat2, steve_df, "Reject by Steve", "reject_by_steve", "❌"),
        (cat3, approval_df, "Products Need Approval", "products_need_approval", "⚠️"),
    ]:
        with col_ui:
            st.markdown(f"""<div class="metric-card" style="text-align:center;"><div style="font-size:1.6rem;">{icon}</div><div style="font-weight:700;color:#CA6180;">{title}</div></div>""", unsafe_allow_html=True)
            if df_cat.empty:
                st.info("Tidak ada data kategori ini.")
            else:
                st.caption(f"{len(df_cat)} baris ditemukan")
                try:
                    img_bytes = df_to_image_bytes(df_cat, title=title)
                    st.download_button(label=f"Download {label} (.png)", data=img_bytes,
                                       file_name=f"{label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png",
                                       mime="image/png", use_container_width=True)
                except Exception as e:
                    st.error(f"Gagal generate gambar: {e}")

    st.markdown(f"""<div class="pipeline-step active"><span class="step-number">{final_step+2}</span><strong>Product Code yang Harus Dihapus</strong></div>""", unsafe_allow_html=True)
    sku_series = [d["SKU"].dropna().astype(str).str.strip() for d in (stop_df, steve_df) if "SKU" in d.columns]
    sku_all = (pd.concat(sku_series).pipe(lambda s: s[s != ""]).drop_duplicates().sort_values().reset_index(drop=True) if sku_series else pd.Series(dtype=str))
    remark_series = []
    if "Supply Control" in stop_df.columns: remark_series.append(stop_df["Supply Control"].dropna().astype(str).str.strip())
    if "Remark" in steve_df.columns: remark_series.append(steve_df["Remark"].dropna().astype(str).str.strip())
    remark_all = (pd.concat(remark_series, ignore_index=True).astype(str).str.strip().loc[lambda s: s.ne("")].drop_duplicates().sort_values().reset_index(drop=True) if remark_series else pd.Series(dtype="string"))

    if sku_all.empty:
        st.info("Tidak ada product code yang perlu dihapus.")
    else:
        pc1, pc2 = st.columns(2)
        with pc1:
            st.dataframe(pd.DataFrame({"Remark/Supply Control": remark_all}), use_container_width=True, hide_index=True)
        with pc2:
            with st.expander(f"📋 Copy SKU ({len(sku_all)} SKU)", expanded=False):
                st.code("\n".join(sku_all.tolist()), language=None)

    st.markdown(f"""<div class="pipeline-step active"><span class="step-number">{final_step+3}</span><strong>Summary PO</strong></div>""", unsafe_allow_html=True)
    summary_df = final_disp.copy()
    summary_df["PO Value"] = pd.to_numeric(summary_df["PO Value"], errors="coerce").fillna(0)
    summary_df["Supply Control"] = summary_df["Supply Control"].fillna("").astype(str)
    summary_df["Remark"] = summary_df["Remark"].fillna("").astype(str)
    po_date_str = datetime.now().strftime("%d %B %Y")
    stop_kw = ["STOP PO","OOS","DISCONTINUED","UNAVAILABLE"]

    def _rp(v):
        try: return f"Rp {v:,.0f}".replace(",",".")
        except: return "Rp 0"

    for dist_sum in sorted(summary_df["Distributor"].dropna().unique()):
        grp = summary_df[summary_df["Distributor"] == dist_sum]
        grp_po = grp[grp["is_po_sku"] == True] if "is_po_sku" in grp.columns else grp
        stop_mask_s = grp_po["Supply Control"].str.strip().str.upper().isin([k.upper() for k in stop_kw])
        stop_grp = grp_po[stop_mask_s]
        steve_mask_s = (grp_po["Remark"].str.lower().str.contains("reject (stop by steve", na=False, regex=False) |
                        grp_po["Remark"].str.lower().str.contains("reject (negative allocation)", na=False, regex=False))
        steve_grp = grp_po[steve_mask_s]
        total_reduction = stop_grp["PO Value"].sum() + steve_grp["PO Value"].sum()
        grand_total_po = grp_po["PO Value"].sum()
        grand_total_after = grand_total_po - total_reduction

        stop_cats = (stop_grp["Supply Control"].dropna().astype(str).str.strip().str.title().replace({"Stop Po":"Stop PO","Oos":"OOS"}).unique().tolist())
        stop_cats = [c for c in stop_cats if c]
        stop_label = ", ".join(stop_cats) if stop_cats else "Discontinued / Stop PO"
        steve_cats = []
        if len(steve_grp):
            if grp_po["Remark"].str.lower().str.contains("reject (stop by steve", na=False, regex=False).any():
                steve_cats.append("Stop by Steve")
            if grp_po["Remark"].str.lower().str.contains("reject (negative allocation)", na=False, regex=False).any():
                steve_cats.append("Negative Allocation")
        steve_label = ", ".join(steve_cats) if steve_cats else "Reject by Steve"

        st.markdown(f"""
        <div style="background:#FFF5F8;border:1px solid #F0C8D6;border-radius:12px;padding:1rem 1.2rem;margin:.6rem 0 .4rem;">
            <div style="font-size:.95rem;font-weight:700;color:#8B2040;margin-bottom:.5rem;">
                Summary PO dari Distributor <strong>{dist_sum}</strong> — PO Date: <strong>{po_date_str}</strong>
            </div>
            <ul style="margin:0;padding-left:1.2rem;color:#1F1F1F;font-size:.88rem;line-height:1.7;">
                <li>Total SKU: <strong>{grp_po["SKU"].nunique():,}</strong></li>
                <li>Grand Total PO (sebelum pengurangan): <strong>{_rp(grand_total_po)}</strong></li>
                <li>{stop_label}: <strong>{stop_grp["SKU"].nunique():,}</strong> SKU — <strong>{_rp(stop_grp["PO Value"].sum())}</strong></li>
                <li>{steve_label}: <strong>{steve_grp["SKU"].nunique():,}</strong> SKU — <strong>{_rp(steve_grp["PO Value"].sum())}</strong></li>
                <li>Total pengurangan: <strong>{_rp(total_reduction)}</strong></li>
                <li>Grand Total setelah pengurangan: <strong>{_rp(grand_total_after)}</strong></li>
            </ul>
        </div>""", unsafe_allow_html=True)

    return final_disp

# ─── Shared file upload section ───────────────────────────────────────────────

def _file_upload_section(page_key: str):
    _INVALID_QTY = {"-","null","none","","0","0.0"}

    def _read_one(fname, fbytes, sheet_name=0):
        ext = fname.rsplit(".",1)[-1].lower()
        if ext == "csv":
            raw_preview = pd.read_csv(io.BytesIO(fbytes), header=None, dtype=str, nrows=15, encoding_errors="replace")
            best_row, best_score = 0, -1
            for i in range(len(raw_preview)):
                vals = [str(v).strip() for v in raw_preview.iloc[i].values if pd.notna(v) and str(v).strip()]
                text_count = sum(1 for v in vals if not v.replace('.','',1).replace(',','',1).lstrip('-').isdigit())
                score = text_count * 10 + len(vals)
                if score > best_score:
                    best_score, best_row = score, i
            return pd.read_csv(io.BytesIO(fbytes), header=best_row, dtype=str, encoding_errors="replace"), best_row
        else:
            engine = _excel_engine(fname)
            hrow = detect_header_row(fbytes, fname, sheet_name=sheet_name)
            return pd.read_excel(io.BytesIO(fbytes), sheet_name=sheet_name, header=hrow, engine=engine, dtype=str), hrow

    def _parse_idx(rng):
        rng = rng.strip()
        if not rng or rng == ":": return None, None
        if ":" not in rng:
            n = int(rng); return n, n+1
        left, right = rng.split(":",1)
        start = int(left.strip()) if left.strip() else None
        end = int(right.strip()) if right.strip() else None
        return start, (end+1) if end is not None else None

    def _apply_range(df, rng_r, rng_c):
        if rng_r.strip():
            try:
                rs, re = _parse_idx(rng_r); df = df.iloc[rs:re]
            except Exception: pass
        if rng_c.strip():
            try:
                cs, ce = _parse_idx(rng_c); df = df.iloc[:, cs:ce]
            except Exception: pass
        return df

    folder_files = st.file_uploader("📁 Upload File PO (.xlsx / .xls / .csv / .zip)",
                                     type=["xlsx","xls","csv","zip"],
                                     accept_multiple_files=True, key=f"po_folder_{page_key}")
    if not folder_files:
        return None, None

    raw_entries = []; converted_names = []
    for uf in folder_files:
        fb = uf.read()
        if uf.name.lower().endswith(".zip"):
            with zipfile.ZipFile(io.BytesIO(fb)) as zf:
                for zname in zf.namelist():
                    ext = zname.rsplit(".",1)[-1].lower() if "." in zname else ""
                    if ext in ("xlsx","xls","csv"):
                        new_name, new_bytes = _convert_to_xlsx(zname, zf.read(zname))
                        raw_entries.append((new_name, new_bytes))
                        if ext != "xlsx": converted_names.append(f"{zname} → {new_name}")
        else:
            new_name, new_bytes = _convert_to_xlsx(uf.name, fb)
            raw_entries.append((new_name, new_bytes))
            if uf.name.rsplit(".",1)[-1].lower() != "xlsx":
                converted_names.append(f"{uf.name} → {new_name}")

    if converted_names:
        st.caption("🔄 Auto-convert: " + "  ·  ".join(converted_names))

    st.markdown("""<div class="pipeline-step active"><span class="step-number">1</span>
    <strong>Konfigurasi per File</strong></div>""", unsafe_allow_html=True)

    parsed = []
    for idx, (fname, fbytes) in enumerate(raw_entries):
        with st.container(border=True):
            hc1, hc2 = st.columns([2,1])
            with hc1:
                st.markdown(f"**#{idx+1} &nbsp; {fname}**")
            with hc2:
                try:
                    wb_tmp = openpyxl.load_workbook(io.BytesIO(fbytes), data_only=True)
                    sheets = [ws.title for ws in wb_tmp.worksheets if ws.sheet_state == 'visible']
                    wb_tmp.close()
                except Exception:
                    sheets = []
                if len(sheets) > 1:
                    sheet_sel = st.selectbox("Sheet:", options=sheets, key=f"fs_{page_key}_{idx}_{fname}", label_visibility="collapsed")
                elif sheets:
                    sheet_sel = sheets[0]; st.caption(f"📄 `{sheets[0]}`")
                else:
                    sheet_sel = 0

            try:
                df_f, hrow = _read_one(fname, fbytes, sheet_sel)
                df_f = df_f.loc[:, ~df_f.columns.astype(str).str.startswith('Unnamed')].dropna(how='all')
                df_f = df_f[df_f.apply(lambda r: r.astype(str).str.strip().ne('').any(), axis=1)]
                parse_err = None
            except Exception as e:
                df_f, hrow, parse_err = None, "-", str(e)

            if df_f is not None:
                df_f.columns = [str(c).strip().upper() for c in df_f.columns]
                qty_col = next((c for c in df_f.columns if c.strip().upper() in ("QTY","QUANTITY")), None)
                if qty_col:
                    def _qty_valid(v):
                        s = str(v).strip().lower()
                        if s in _INVALID_QTY: return False
                        try: return float(s.replace(",",".")) > 0
                        except: return False
                    before = len(df_f)
                    df_f = df_f[df_f[qty_col].apply(_qty_valid)].reset_index(drop=True)
                    removed = before - len(df_f)
                    if removed:
                        st.caption(f"🗑 {removed:,} baris dibuang (QTY tidak valid — kolom **{qty_col}**)")

                has_dist = any("DISTRIBUTOR" in c.upper() for c in df_f.columns)
                dc1, dc2 = st.columns([1,2])
                with dc1:
                    st.caption("Distributor" + (" *(sudah ada)*" if has_dist else ""))
                with dc2:
                    dist_val = st.selectbox("Distributor", options=["(Pilih)"] + CUSTOMER_NAMES,
                                            key=f"dist_{page_key}_{idx}_{fname}", label_visibility="collapsed")

                with st.expander(f"👁 Lihat isi · header row {hrow} · {len(df_f)} baris", expanded=False):
                    st.dataframe(df_f.iloc[:,:6].reset_index(drop=True), use_container_width=True)

                rc1, rc2 = st.columns(2)
                with rc1:
                    row_rng = st.text_input("Row Range", value="", key=f"row_{page_key}_{idx}_{fname}", placeholder="5:10")
                with rc2:
                    col_rng = st.text_input("Column Range", value="", key=f"col_{page_key}_{idx}_{fname}", placeholder="0:3")

                parsed.append({"name":fname,"df":df_f,"row_rng":row_rng,"col_rng":col_rng,
                                "dist_val":dist_val,"has_dist":has_dist,"error":None})
            else:
                st.error(f"❌ {parse_err}")
                parsed.append({"name":fname,"df":None,"row_rng":"","col_rng":"","dist_val":"","has_dist":False,"error":parse_err})

    ready = [p for p in parsed if p["df"] is not None]
    st.divider()

    if st.button("🔎 Cek Semua File", disabled=not ready, use_container_width=True, key=f"concat_btn_{page_key}"):
        frames = []
        for p in ready:
            df_tmp = _apply_range(p["df"].copy(), p["row_rng"], p["col_rng"])
            if p["dist_val"] not in ("","(Pilih)"):
                if p["has_dist"]:
                    df_tmp["DISTRIBUTOR"] = p["dist_val"]
                else:
                    df_tmp.insert(0, "DISTRIBUTOR", p["dist_val"])
            df_tmp["_source_file"] = p["name"]
            frames.append(df_tmp)

        try:
            tpl_bytes = _fetch_template_bytes(TEMPLATE_DRIVE_URL)
        except Exception:
            tpl_bytes = None

        seen = set(); ref_cols = []
        for f in frames:
            for c in f.columns:
                c_str = str(c).strip()
                if c_str and c_str not in seen:
                    seen.add(c_str); ref_cols.append(c_str)
        frames = [f.reindex(columns=ref_cols) for f in frames]
        combined_df = pd.concat(frames, ignore_index=True)
        combined_df = combined_df.loc[:, ~combined_df.columns.duplicated()]
        st.session_state[f"folder_result_{page_key}"] = {"df": combined_df, "tpl_bytes": tpl_bytes}
        st.session_state.pop(f"sim_result_{page_key}", None)
        st.rerun()

    res = st.session_state.get(f"folder_result_{page_key}")
    if res is not None:
        combined_df = res["df"]
        st.success(f"✅ **{len(combined_df):,}** baris · {combined_df.shape[1]} kolom dari {len(ready)} file")
        for col in ['QTY','DPP','TOTAL PRICE']:
            if col in combined_df.columns:
                combined_df[col] = pd.to_numeric(combined_df[col], errors='coerce')
        st.subheader("📊 Hasil Gabungan")
        st.dataframe(combined_df, use_container_width=True, hide_index=True)

    return raw_entries, res


# ─── Modify QTY section ───────────────────────────────────────────────────────

def _modify_qty_section(raw_entries, page_key: str):
    if not raw_entries:
        st.info("ℹ️ Upload file PO di section atas terlebih dahulu.")
        return

    st.markdown("""<div class="pipeline-step active"><span class="step-number">1</span>
    <strong>Pilih File untuk Modifikasi</strong></div>""", unsafe_allow_html=True)

    for fi, (tpl_fname, tpl_orig_bytes) in enumerate(raw_entries):
        with st.container(border=True):
            st.markdown(f"**#{fi+1} &nbsp; {tpl_fname}**")
            tpl_name, tpl_bytes = _convert_to_xlsx(tpl_fname, tpl_orig_bytes)
            tpl_sheets = _get_sheet_names(tpl_bytes, "openpyxl")
            if not tpl_sheets:
                st.warning("⚠️ Tidak ada sheet visible."); continue

            sc1, sc2 = st.columns([2,1])
            with sc1:
                if len(tpl_sheets) > 1:
                    tpl_selected_sheet = st.selectbox("Sheet:", options=tpl_sheets, key=f"tpl_sheet_{page_key}_{fi}")
                else:
                    tpl_selected_sheet = tpl_sheets[0]; st.caption(f"📄 Sheet: **{tpl_selected_sheet}**")
            with sc2:
                auto_hrow = detect_header_row(tpl_bytes, tpl_name, sheet_name=tpl_selected_sheet)
                hrow_input = st.number_input("Header row", min_value=1, value=int(auto_hrow)+1, step=1, key=f"tpl_hrow_{page_key}_{fi}")
            tpl_hrow = int(hrow_input) - 1

            try:
                tpl_df = pd.read_excel(io.BytesIO(tpl_bytes), sheet_name=tpl_selected_sheet,
                                        header=tpl_hrow, engine="openpyxl", dtype=str)
                tpl_df = tpl_df.loc[:, ~tpl_df.columns.str.startswith('Unnamed')].dropna(how='all').reset_index(drop=True)
            except Exception as e:
                st.error(f"❌ Gagal membaca file: {e}"); continue

            st.caption(f"**{len(tpl_df):,} baris · {len(tpl_df.columns)} kolom**")
            with st.expander("👁 Preview data", expanded=False):
                st.dataframe(tpl_df, use_container_width=True, hide_index=True)

            qty_col_t = next((c for c in tpl_df.columns if any(k in c.lower() for k in ['qty','quantity'])), None)
            sku_col_t = next((c for c in tpl_df.columns if any(k in c.lower() for k in ['sku','product code','kode','code'])), None)
            if not qty_col_t or not sku_col_t:
                st.info("ℹ️ Kolom SKU / QTY tidak terdeteksi."); continue

            st.markdown("""<div class="pipeline-step active"><span class="step-number">2</span>
            <strong>Modifikasi Quantity per Product Code</strong></div>""", unsafe_allow_html=True)

            def _save_tpl_file(cell_writer):
                com_res = _edit_qty_via_excel_com(tpl_bytes, tpl_selected_sheet, tpl_hrow, sku_col_t, qty_col_t, cell_writer)
                if com_res is not None:
                    return com_res
                out_buf = io.BytesIO(); changed = 0
                wb_tmp = openpyxl.load_workbook(io.BytesIO(tpl_bytes), data_only=False)
                ws_tmp = next((s for s in wb_tmp.worksheets if s.title == tpl_selected_sheet), wb_tmp.active)
                hdr_row = tpl_hrow + 1
                hdrs = {ws_tmp.cell(row=hdr_row, column=c).value: c for c in range(1, ws_tmp.max_column+1)}
                sku_ci = hdrs.get(sku_col_t); qty_ci = hdrs.get(qty_col_t)
                if sku_ci and qty_ci:
                    for row in ws_tmp.iter_rows(min_row=hdr_row+1, max_row=ws_tmp.max_row):
                        sv = str(row[sku_ci-1].value or "").strip()
                        qcell = row[qty_ci-1]
                        new = cell_writer(sv, qcell.value)
                        if new is not None:
                            qcell.value = new; changed += 1
                wb_tmp.save(out_buf)
                return out_buf.getvalue(), changed

            with st.container(border=True):
                st.caption(f"SKU: **{sku_col_t}** · Quantity: **{qty_col_t}**")
                reduce_codes = st.text_area("Daftar Product Code (satu per baris)", placeholder="SKU001\nSKU-ABC", height=150, key=f"reduce_codes_{page_key}_{fi}")

                btn1, btn2 = st.columns(2)
                with btn1:
                    if st.button("Modifikasi QTY", use_container_width=True, key=f"btn_qty_{page_key}_{fi}"):
                        parsed_skus = [c.strip() for c in reduce_codes.strip().splitlines() if c.strip()]
                        if parsed_skus:
                            st.session_state[f"reduce_skus_{page_key}_{fi}"] = parsed_skus
                            st.session_state[f"mod_mode_{page_key}_{fi}"] = "qty"
                        else:
                            st.warning("⚠️ Tidak ada SKU yang valid")

                with btn2:
                    if st.button("Auto Hapus SKU", use_container_width=True, key=f"btn_del_{page_key}_{fi}"):
                        parsed_skus = [c.strip() for c in reduce_codes.strip().splitlines() if c.strip()]
                        if not parsed_skus:
                            st.warning("⚠️ Tidak ada SKU yang valid")
                        else:
                            del_set = set(parsed_skus)
                            wb_del = openpyxl.load_workbook(io.BytesIO(tpl_bytes), data_only=False)
                            ws_del = next((s for s in wb_del.worksheets if s.title == tpl_selected_sheet), wb_del.active)
                            hdr_row_del = tpl_hrow + 1
                            hdrs_del = {ws_del.cell(row=hdr_row_del, column=c).value: c for c in range(1, ws_del.max_column+1)}
                            sku_ci_del = hdrs_del.get(sku_col_t); qty_ci_del = hdrs_del.get(qty_col_t)
                            zeroed = 0
                            if sku_ci_del and qty_ci_del:
                                for r in range(hdr_row_del+1, ws_del.max_row+1):
                                    sv = str(ws_del.cell(row=r, column=sku_ci_del).value or "").strip()
                                    if sv in del_set:
                                        ws_del.cell(row=r, column=qty_ci_del).value = None; zeroed += 1
                            buf_del = io.BytesIO(); wb_del.save(buf_del)
                            st.session_state[f"tpl_out_{page_key}_{fi}"] = {
                                "buf": _sanitize_xlsx_bytes(buf_del.getvalue()), "cleared": zeroed, "mode": "delete",
                                "ext": "xlsx", "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            }
                            st.success(f"✅ {zeroed} baris dihapus.")

            skus_r = st.session_state.get(f"reduce_skus_{page_key}_{fi}", [])
            if skus_r and st.session_state.get(f"mod_mode_{page_key}_{fi}") == "qty":
                st.markdown("**Atur quantity baru per Product Code:**")
                sku_qty_map = (tpl_df[[sku_col_t, qty_col_t]].dropna(subset=[sku_col_t])
                               .assign(**{sku_col_t: lambda d: d[sku_col_t].astype(str).str.strip()})
                               .set_index(sku_col_t)[qty_col_t].to_dict())
                for sku_r in skus_r:
                    cur_q = sku_qty_map.get(sku_r, None)
                    try: cur_q_int = int(float(cur_q)) if cur_q not in (None,"") else 0
                    except: cur_q_int = 0
                    def_key = f"edit_val_{page_key}_{fi}_{sku_r}"
                    if def_key not in st.session_state:
                        st.session_state[def_key] = cur_q_int
                    with st.container(border=True):
                        rc1, rc2, rc3 = st.columns([3,2,3])
                        with rc1: st.markdown(f"**{sku_r}**")
                        with rc2:
                            st.caption("QTY saat ini")
                            st.markdown(f"**{cur_q if cur_q is not None else '-'}**")
                        with rc3:
                            st.number_input("Quantity baru", min_value=0, step=1, key=def_key)

                if st.button("Change QTY", use_container_width=True, key=f"apply_qty_{page_key}_{fi}"):
                    edit_map = {sku_r: st.session_state.get(f"edit_val_{page_key}_{fi}_{sku_r}", 0) for sku_r in skus_r}
                    def _edit_writer(sku_val, qty_val):
                        if sku_val not in edit_map: return None
                        new = float(edit_map[sku_val])
                        return int(new) if new == int(new) else new
                    buf, cnt = _save_tpl_file(_edit_writer)
                    st.session_state[f"tpl_out_{page_key}_{fi}"] = {
                        "buf": buf, "cleared": cnt,
                        "ext": "xlsx", "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    }

            res_t = st.session_state.get(f"tpl_out_{page_key}_{fi}")
            if res_t:
                mode_label = "SKU di-set QTY null" if res_t.get("mode") == "delete" else "baris diubah QTY-nya"
                st.success(f"✅ **{res_t['cleared']}** {mode_label}.")
                customer_name = st.selectbox("Distributor", options=["(Pilih)"] + CUSTOMER_NAMES,
                                              key=f"tpl_cust_{page_key}_{fi}", label_visibility="collapsed")
                file_label = re.sub(r'[\\/*?:"<>|]', "", (customer_name or "").strip()) or "Unnamed_Customer"
                st.download_button(
                    label=f"⬇ Download Hasil Modifikasi (.{res_t['ext']})",
                    data=res_t["buf"],
                    file_name=f"Form PO {file_label}.{res_t['ext']}",
                    mime=res_t["mime"],
                    use_container_width=True, key=f"tpl_dl_{page_key}_{fi}",
                )

# === HALAMAN BARU: Login RSA ===
if st.session_state.get('page') == 'po_changer_login':
    st.markdown("""<div class="hero-wrap">
        <div class="hero-tag">✦ PO Management</div>
        <div class="hero-title">PO Simulator - RSA</div></div>""", unsafe_allow_html=True)
    st.divider()
    _, col, _ = st.columns([1, 1.5, 1])
    with col:
        st.markdown("""<div style="text-align:center;margin:1rem 0 1.5rem;">
            <div style="font-size:1.2rem;font-weight:700;color:#6E253A;">🔒 Masukkan Password RSA</div>
            <div style="color:#555;font-size:.85rem;margin-top:.3rem;">Fitur ini hanya untuk RSA</div>
        </div>""", unsafe_allow_html=True)
        with st.form("rsa_login_form"):
            rsa_password = st.text_input("Password", type="password", placeholder="Masukkan password RSA...")
            submitted = st.form_submit_button("Masuk", use_container_width=True)
            if submitted:
                if rsa_password == st.secrets["glowithyou"]:
                    st.session_state['rsa_authenticated'] = True
                    st.session_state['page'] = 'po_changer'
                    st.rerun()
                else:
                    st.error("❌ Password salah.")
    st.stop()

# === HALAMAN po_changer: tambah guard di baris pertama ===
if st.session_state.get('page') == 'po_changer':
    if not st.session_state.get('rsa_authenticated'):  # <-- TAMBAHAN INI SAJA
        st.session_state['page'] = 'po_changer_login'
        st.rerun()

    st.markdown("""<div class="hero-wrap">
        <div class="hero-tag">✦ PO Management</div>
        <div class="hero-title">PO Simulator - RSA</div></div>""", unsafe_allow_html=True)
    st.divider()
# ─── Page: PO Simulator (For RSA) ────────────────────────────────────────────

    raw_entries, folder_res = _file_upload_section("rsa")

    if folder_res is not None:
        sim_df = folder_res["df"].copy()
        sku_col_sim = next((c for c in sim_df.columns if c.upper() in ("SKU","PRODUCT CODE")), None)
        qty_col_sim = next((c for c in sim_df.columns if c.upper() in ("QTY","QUANTITY")), None)
        dist_col_sim = next((c for c in sim_df.columns if "DISTRIBUTOR" in c.upper()), None)

        if sku_col_sim and qty_col_sim and dist_col_sim:
            st.divider()
            if st.session_state.get("sim_result_rsa") is None:
                excel_dfs, all_npd = _run_po_simulation(
                    sim_df.copy(), sku_col_sim, qty_col_sim, dist_col_sim,
                    _MANUAL_REJECT_APPROVAL, _MANUAL_REJECT_NO_TOL,
                    _REJECTED_SKUS_1, __REGION_LIST_1,
                    __REJECTED_SKUS_2, _REGION_LIST_2,
                    _LIMITED_SKUS_QTY, ___MAX_QTY_LIMIT,
                )
                st.session_state["sim_result_rsa"] = {"dfs": excel_dfs, "npd": all_npd}
                st.rerun()

            sim_out = st.session_state.get("sim_result_rsa")
            if sim_out and sim_out["dfs"]:
                st.success(f"Simulasi selesai — {len(sim_out['dfs'])} distributor")
                _render_sim_results(sim_out["dfs"], sim_out["npd"], folder_res, sku_col_sim, qty_col_sim, dist_col_sim)

    _modify_qty_section(raw_entries, "rsa")
    st.stop()


# ─── Page: PO Simulator (For SPV) ────────────────────────────────────────────

if st.session_state.get('page') == 'po_spv':
    st.markdown("""<div class="hero-wrap">
        <div class="hero-tag">✦ PO Management</div>
        <div class="hero-title">PO Simulator (SPV)</div></div>""", unsafe_allow_html=True)
    st.divider()

    tab1, tab2 = st.tabs(["📖 How to Use", "📂 Upload & Simulate"])

    with tab1:
        st.header("How to Use the PO Simulator")
        with st.expander("📋 Step-by-Step Guide"):
            st.markdown("""
1. **Upload PO Data**: Upload file Excel atau CSV dengan kolom: `DISTRIBUTOR`, `PRODUCT CODE`, `DESCRIPTION`, `QTY`.
2. **Review Rejection Lists**: Cek manual rejection SKUs dan regional rejection jika ada.
3. **Simulate**: App akan fetch data stock & sales dari BigQuery, hitung WOI, dan apply approval/rejection rules.
4. **View Results**: Review simulated data termasuk Remark (Proceed / Reject / Suggest).
5. **Download Excel**: Pilih format output — Separate Sheets (per distributor) atau Single Sheet.
""")

        st.header("Rules & Calculations Logic")
        with st.expander("⚖️ Rules & Calculations Logic"):
            st.markdown("""
**Remark ditentukan berurutan:**

1. **Reject** jika:
   - SKU **tidak ditemukan** di sistem (BigQuery)
   - Ada di **regional rejection list** (Stop by Steve), kecuali region diizinkan
   - Ada di **manual rejection list**:
     - *Need approval email*: G2G-840, G2G-844, G2G-841, G2G-800, dll
     - *No tolerance to open*: G2G-2721, G2G-224, dll
   - **Remaining Allocation < 0** (Negative Allocation)
   - Supply Control = STOP PO / DISCONTINUED / OOS
   - PO Qty **>** Suggested Qty → **Reject with Suggestion**

2. **Proceed** jika:
   - PO Qty **<** Suggested Qty → **Proceed with Suggestion**
   - PO Qty **=** Suggested Qty → **Proceed**
   - NPD dengan allocation > 0 dan PO Qty ≤ remaining allocation
   - Tidak ada historical data, supply control bukan STOP/DISCONTINUED/OOS

3. **Additional Suggestion**: SKU tidak ada di PO tapi disarankan sistem (auto-filtered jika akan di-reject)

**WOI** = (Total Stock + PO/Suggested Qty) / Avg Weekly Sales LM
""")

        st.header("Manual Rejection SKUs")
        with st.expander("🚫 SKU yang di-reject manual (Steve)"):
            reject_data = [{"SKU": s, "Remark": "Need approval email"} for s in _MANUAL_REJECT_APPROVAL]
            reject_data += [{"SKU": s, "Remark": "No tolerance to open"} for s in _MANUAL_REJECT_NO_TOL]
            st.dataframe(pd.DataFrame(reject_data).sort_values("SKU").reset_index(drop=True),
                         use_container_width=True, hide_index=True)

        if _REJECTED_SKUS_1:
            st.header("Regional Rejection Rules")
            with st.expander("🌍 SKU dengan Pembatasan Regional"):
                st.markdown(f"**SKU: {', '.join(_REJECTED_SKUS_1)}**\n\nHanya diizinkan di region berikut:")
                for r in __REGION_LIST_1:
                    st.markdown(f"- {r}")
                st.markdown("**Region lain akan otomatis di-reject.**")

    with tab2:
        st.header("1. Download PO Template")

        st.download_button(
            label="📥 Download PO Template",
            data=create_po_template_excel(),
            file_name="po_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.header("2. Upload PO Data")
        uploaded_file = st.file_uploader(
            "Upload a PO file (.xlsx/.csv) containing: 'PRODUCT CODE', 'DESCRIPTION', 'QTY', and 'DISTRIBUTOR'.",
            type=["xlsx", "xls", "csv"],
        )

        if uploaded_file:
            try:
                required_cols = ["PRODUCT CODE", "DESCRIPTION", "QTY", "DISTRIBUTOR"]

                if uploaded_file.name.endswith(".xlsx"):
                    po_df = pd.read_excel(uploaded_file, header=7, engine="openpyxl")
                else:
                    po_df = pd.read_csv(uploaded_file, header=7)
                st.success("File uploaded successfully!")

                po_df.dropna(axis=1, how='all', inplace=True)
                po_df.drop(columns=[col for col in po_df.columns if 'Unnamed' in str(col)], inplace=True)

                if not all(col in po_df.columns for col in required_cols):
                    st.error("The uploaded file is missing one or more required columns.")
                    st.write("Please check for these columns:", required_cols)
                    st.stop()

                po_df["QTY"] = pd.to_numeric(po_df["QTY"], errors="coerce")
                po_df.dropna(subset=["QTY"], inplace=True)

                po_df["DISTRIBUTOR"] = po_df["DISTRIBUTOR"].str.upper()
                po_df["PRODUCT CODE"] = po_df["PRODUCT CODE"].astype(str).str.strip().str.upper()

                st.write("Preview of uploaded PO data:")
                st.dataframe(po_df.head())

                po_df = po_df[po_df["QTY"] > 0]
                po_df.dropna(subset=["PRODUCT CODE", "QTY", "DISTRIBUTOR"], inplace=True)

                po_df["is_po_sku"] = True

                po_df.rename(
                    columns={
                        "PRODUCT CODE": "Customer SKU Code",
                        "QTY": "PO Qty",
                        "DISTRIBUTOR": "Distributor"
                    },
                    inplace=True,
                )

                po_df = po_df[["Distributor", "Customer SKU Code", "PO Qty", "is_po_sku"]]

                st.header("3. PO Simulation")

                st.subheader("4. Select Output Format and Download")
                output_format = st.radio(
                    "Choose Excel Output Format:",
                    ["Separate Sheets (One per Distributor)", "Single Sheet (All Distributors Stacked)"],
                    index=0
                )

                progress = st.progress(0)
                progress_step = 1.0 / len(po_df["Distributor"].unique())

                all_npd_sku_list = []
                excel_dfs = {}
                display_dfs = []

                uploaded_distributors = po_df["Distributor"].unique().tolist()

                for i, distributor_name in enumerate(uploaded_distributors):
                    st.subheader(f"Processing data for: **{distributor_name}**")
                    progress.progress((i * progress_step), f"Processing {distributor_name}...")

                    current_po_df = po_df[po_df["Distributor"] == distributor_name].copy()
                    sku_list = current_po_df["Customer SKU Code"].unique().tolist()

                    # --- Fetch data from BigQuery ---
                    sku_df = get_sku_data(sku_list)
                    sku_data_df = get_stock_data(distributor_name, sku_list)

                    # Rename sku_df columns — now includes product_name
                    sku_df.rename(
                        columns={
                            "sku": "Customer SKU Code",
                            "price_for_distri": "SIP",
                            "product_name": "Product Name",
                        },
                        inplace=True,
                    )

                    if sku_data_df.empty:
                        st.warning(
                            "Could not find stock and sales data for the uploaded Distributor/SKUs in BigQuery. Please check the Distributor Name or SKU codes."
                        )
                        st.stop()

                    # Rename stock data columns
                    if "sku" in sku_data_df.columns:
                        sku_data_df.rename(
                            columns={"sku": "Customer SKU Code", "distributor": "Distributor"},
                            inplace=True,
                        )

                    # Drop product_name from stock data — we use master_product's version
                    if "product_name" in sku_data_df.columns:
                        sku_data_df.drop(columns=["product_name"], inplace=True)

                    # ===== TRACK SKUs NOT FOUND IN BIGQUERY =====
                    skus_in_sku_df = set(sku_df["Customer SKU Code"].tolist()) if not sku_df.empty else set()
                    skus_in_stock_df = set(sku_data_df["Customer SKU Code"].tolist()) if not sku_data_df.empty else set()
                    skus_not_in_bq = set(sku_list) - (skus_in_sku_df | skus_in_stock_df)

                    if skus_not_in_bq:
                        st.warning(
                            f"The following SKUs were **not found** in the system and will be rejected: "
                            f"`{', '.join(sorted(skus_not_in_bq))}`"
                        )
                    # ===== END TRACK SKUs NOT FOUND =====

                    # Merge PO data with SKU price + product_name data
                    result_df = pd.merge(current_po_df, sku_df, on="Customer SKU Code", how="left")

                    result_df["SIP"] = pd.to_numeric(result_df["SIP"], errors="coerce").fillna(0)
                    result_df["PO Value"] = result_df["SIP"] * result_df["PO Qty"]

                    # Merge with stock data (outer to include suggested SKUs)
                    result_df = pd.merge(result_df, sku_data_df, on="Customer SKU Code", how="outer")

                    # After outer merge, fill Product Name for suggested SKUs from master_product
                    # by re-merging just the product names for any new SKUs introduced via outer join
                    missing_product_names = result_df["Product Name"].isna()
                    if missing_product_names.any():
                        suggested_skus = result_df.loc[missing_product_names, "Customer SKU Code"].unique().tolist()
                        if suggested_skus:
                            extra_sku_df = get_sku_data(suggested_skus)
                            if not extra_sku_df.empty:
                                extra_sku_df.rename(
                                    columns={
                                        "sku": "Customer SKU Code",
                                        "price_for_distri": "SIP",
                                        "product_name": "Product Name",
                                    },
                                    inplace=True,
                                )
                                extra_name_map = extra_sku_df.set_index("Customer SKU Code")["Product Name"].to_dict()
                                result_df.loc[missing_product_names, "Product Name"] = (
                                    result_df.loc[missing_product_names, "Customer SKU Code"].map(extra_name_map)
                                )

                    result_sku_list = result_df["Customer SKU Code"].unique().tolist()

                    npd_df = get_npd_data(result_sku_list)
                    current_npd_sku_list = npd_df['sku'].unique().tolist() if not npd_df.empty else []

                    all_npd_sku_list.extend(current_npd_sku_list)
                    all_npd_sku_list = list(set(all_npd_sku_list))

                    result_df["is_po_sku"] = result_df["is_po_sku"].astype("boolean")
                    result_df["is_po_sku"] = result_df["is_po_sku"].fillna(False)

                    result_df[
                        [
                            "PO Qty",
                            "PO Value",
                            "total_stock",
                            "buffer_plan_by_lm_qty_adj",
                            "avg_weekly_st_lm_qty",
                            "buffer_plan_by_lm_val_adj",
                            "remaining_allocation_qty_region",
                            "woi_end_of_month_by_lm",
                        ]
                    ] = result_df[
                        [
                            "PO Qty",
                            "PO Value",
                            "total_stock",
                            "buffer_plan_by_lm_qty_adj",
                            "avg_weekly_st_lm_qty",
                            "buffer_plan_by_lm_val_adj",
                            "remaining_allocation_qty_region",
                            "woi_end_of_month_by_lm",
                        ]
                    ].fillna(0)

                    result_df = result_df[
                        (result_df["PO Qty"] > 0) | (result_df["buffer_plan_by_lm_qty_adj"] > 0)
                    ]

                    # ===== ENHANCED REJECTION LOGIC FOR SUGGESTED SKUs =====
                    suggested_skus_mask = result_df["is_po_sku"] == False

                    exclude_suggested = (
                        (result_df["Customer SKU Code"].isin(skus_not_in_bq)) |
                        (result_df["Customer SKU Code"].isin(_MANUAL_REJECT_APPROVAL)) |
                        (result_df["Customer SKU Code"].isin(_MANUAL_REJECT_NO_TOL)) |
                        (result_df["remaining_allocation_qty_region"] < 0) |
                        (result_df["supply_control_status_gt"].str.upper().isin(["STOP PO", "DISCONTINUED", "OOS", "UNAVAILABLE"])) |
                        (
                            (result_df["Customer SKU Code"].isin(_LIMITED_SKUS_QTY)) &
                            (result_df["buffer_plan_by_lm_qty_adj"] > ___MAX_QTY_LIMIT)
                        ) |
                        (result_df["buffer_plan_by_lm_qty_adj"] == 0)
                    )

                    if _REJECTED_SKUS_1:
                        regions_upper_1 = [r.upper() for r in _REJECTED_SKUS_1]
                        regional_reject_1 = (
                            (result_df["Customer SKU Code"].isin(_REJECTED_SKUS_1)) &
                            (~result_df["region"].str.upper().isin(regions_upper_1))
                        )
                        exclude_suggested = exclude_suggested | regional_reject_1

                    result_df = result_df[~(suggested_skus_mask & exclude_suggested)]
                    # ===== END ENHANCED REJECTION LOGIC =====

                    result_df["distributor_name"] = distributor_name

                    result_df["WOI PO Original"] = calculate_woi(
                        result_df["total_stock"],
                        result_df["PO Qty"],
                        result_df["avg_weekly_st_lm_qty"],
                    )

                    result_df["WOI Suggest"] = calculate_woi(
                        result_df["total_stock"],
                        result_df["buffer_plan_by_lm_qty_adj"],
                        result_df["avg_weekly_st_lm_qty"],
                    )

                    result_df["Current WOI"] = calculate_woi(
                        result_df["total_stock"],
                        0,
                        result_df["avg_weekly_st_lm_qty"],
                    )

                    conditions = [
                        result_df["Customer SKU Code"].isin(skus_not_in_bq),
                        (result_df["Customer SKU Code"].isin(_LIMITED_SKUS_QTY)) & (result_df["PO Qty"] > ___MAX_QTY_LIMIT),
                        (result_df["remaining_allocation_qty_region"] < 0),
                        (result_df["is_po_sku"] == False),
                        result_df["Customer SKU Code"].isin(_MANUAL_REJECT_APPROVAL),
                        result_df["Customer SKU Code"].isin(_MANUAL_REJECT_NO_TOL),
                        (result_df["supply_control_status_gt"].str.upper().isin(["STOP PO", "DISCONTINUED", "OOS", "UNAVAILABLE"])),
                        (
                            (result_df["avg_weekly_st_lm_qty"] == 0) &
                            (result_df["buffer_plan_by_lm_qty_adj"] == 0) &
                            (~result_df["Customer SKU Code"].str.upper().isin(all_npd_sku_list)) &
                            (~result_df["supply_control_status_gt"].str.upper().isin(["STOP PO", "DISCONTINUED", "OOS"]))
                        ),
                        (result_df["buffer_plan_by_lm_qty_adj"] == 0),
                        (result_df["PO Qty"] > result_df["buffer_plan_by_lm_qty_adj"]),
                        (result_df["PO Qty"] < result_df["buffer_plan_by_lm_qty_adj"]),
                        (result_df["PO Qty"] == result_df["buffer_plan_by_lm_qty_adj"]),
                    ]

                    choices = [
                        "Reject (SKU Not Found in System)",
                        f"Reject (Exceeds Qty Limit of {___MAX_QTY_LIMIT})",
                        "Reject (Negative Allocation)",
                        "Additional Suggestion",
                        "Reject (Stop by Steve - Need approval email)",
                        "Reject (Stop by Steve - No tolerance to open)",
                        "Reject",
                        "Proceed",
                        "Reject",
                        "Reject with suggestion",
                        "Proceed with suggestion",
                        "Proceed",
                    ]

                    result_df["Remark"] = np.select(conditions, choices, default="N/A (Missing Data)")

                    # Rename columns — note: product_name already renamed to "Product Name" via sku_df merge
                    new_column_names = {
                        "distributor_name": "Distributor",
                        "Customer SKU Code": "SKU",
                        "assortment": "Assortment",
                        "supply_control_status_gt": "Supply Control",
                        "PO Qty": "PO Qty",
                        "PO Value": "PO Value",
                        "total_stock": "Total Stock (Qty)",
                        "avg_weekly_st_lm_qty": "Avg Weekly Sales LM (Qty)",
                        "buffer_plan_by_lm_qty_adj": "Suggested PO Qty",
                        "buffer_plan_by_lm_val_adj": "Suggested PO Value",
                        "WOI PO Original": "WOI (Stock + PO Ori)",
                        "WOI Suggest": "WOI After Buffer (Stock + Suggested Qty)",
                        "woi_end_of_month_by_lm": "Stock + Suggested Qty WOI (Projection at EOM)",
                        "remaining_allocation_qty_region": "Remaining Allocation (By Region)",
                    }

                    result_df.rename(columns=new_column_names, inplace=True)

                    if _REJECTED_SKUS_1:
                        result_df = apply_sku_rejection_rules(
                            _REJECTED_SKUS_1,
                            result_df,
                            __REGION_LIST_1,
                            is_in=False
                        )

                    if __REJECTED_SKUS_2:
                        result_df = apply_sku_rejection_rules(
                            __REJECTED_SKUS_2,
                            result_df,
                            _REGION_LIST_2,
                            is_in=False
                        )

                    result_df.sort_values(
                        by=["is_po_sku", "SKU"], ascending=[False, True], inplace=True
                    )

                    result_df["RSA Notes"] = ""

                    excel_cols = [
                        "Distributor",
                        "SKU",
                        "Product Name",
                        "Assortment",
                        "Supply Control",
                        "Avg Weekly Sales LM (Qty)",
                        "Total Stock (Qty)",
                        "Current WOI",
                        "PO Qty",
                        "PO Value",
                        "WOI (Stock + PO Ori)",
                        "Remark",
                        "Suggested PO Qty",
                        "Suggested PO Value",
                        "WOI After Buffer (Stock + Suggested Qty)",
                        "Stock + Suggested Qty WOI (Projection at EOM)",
                        "Remaining Allocation (By Region)",
                        "is_po_sku",
                        "RSA Notes",
                    ]

                    result_df = result_df.reindex(columns=excel_cols)

                    excel_dfs[distributor_name] = result_df.copy()

                    # Format columns for display
                    result_df["PO Value"] = result_df["PO Value"].apply(
                        lambda x: f"{x:,.2f}" if pd.notnull(x) else 0
                    )
                    result_df["Suggested PO Value"] = result_df["Suggested PO Value"].apply(
                        lambda x: f"{x:,.2f}" if pd.notnull(x) else 0
                    )
                    result_df["Remaining Allocation (By Region)"] = result_df["Remaining Allocation (By Region)"].apply(
                        lambda x: f"{round(x):,d}" if pd.notnull(x) else 0
                    )
                    result_df["Avg Weekly Sales LM (Qty)"] = result_df["Avg Weekly Sales LM (Qty)"].apply(
                        lambda x: f"{round(x):,d}" if pd.notnull(x) else 0
                    )
                    result_df["WOI (Stock + PO Ori)"] = result_df["WOI (Stock + PO Ori)"].apply(
                        lambda x: f"{x:.2f}" if pd.notnull(x) else ""
                    )
                    result_df["Stock + Suggested Qty WOI (Projection at EOM)"] = result_df[
                        "Stock + Suggested Qty WOI (Projection at EOM)"
                    ].apply(lambda x: f"{x:.2f}" if pd.notnull(x) else "")
                    result_df["Current WOI"] = result_df["Current WOI"].apply(
                        lambda x: f"{x:.2f}" if pd.notnull(x) else ""
                    )
                    result_df["WOI After Buffer (Stock + Suggested Qty)"] = result_df[
                        "WOI After Buffer (Stock + Suggested Qty)"
                    ].apply(lambda x: f"{x:.2f}" if pd.notnull(x) else "")

                    display_dfs.append(result_df)

                progress.progress(1.0, "Processing complete")

                if output_format == "Single Sheet (All Distributors Stacked)":
                    final_excel_df = pd.concat(excel_dfs.values(), ignore_index=True)
                    xlsx_data = to_excel_single_sheet(final_excel_df, all_npd_sku_list)
                    file_name = "po_simulator_result_single_sheet.xlsx"
                else:
                    xlsx_data = to_excel_with_styling(excel_dfs, all_npd_sku_list)
                    file_name = "po_simulator_result_separate_sheets.xlsx"

                st.download_button(
                    label=f"📥 Download PO Simulator Excel ({output_format})",
                    data=xlsx_data,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

                st.subheader("Simulated PO Data")
                final_display_df = pd.concat(display_dfs, ignore_index=True)

                final_cols = [
                    "Distributor",
                    "SKU",
                    "Product Name",
                    "Assortment",
                    "Supply Control",
                    "Avg Weekly Sales LM (Qty)",
                    "Total Stock (Qty)",
                    "Current WOI",
                    "PO Qty",
                    "PO Value",
                    "WOI (Stock + PO Ori)",
                    "Remark",
                    "Suggested PO Qty",
                    "Suggested PO Value",
                    "WOI After Buffer (Stock + Suggested Qty)",
                    "Stock + Suggested Qty WOI (Projection at EOM)",
                    "Remaining Allocation (By Region)",
                ]

                st.dataframe(final_display_df.reindex(columns=final_cols).reset_index(drop=True))

            except Exception as e:
                st.error(f"An error occurred: {e}")
                st.info(
                    "Please ensure the uploaded file is a valid .xlsx or .csv and contains all the required columns."
                )

    st.stop()


# ─── Page: Request PO (For SPV) ──────────────────────────────────────────────

st.markdown("""<div class="hero-wrap">
    <div class="hero-tag">✦ REQUEST PO</div>
    <div class="hero-title">REQUEST PO</div></div>""", unsafe_allow_html=True)
st.divider()

st.markdown("""<div class="pipeline-step active"><span class="step-number">1</span>
<strong>ISI DATA</strong>
<span class="badge badge-info" style="margin-left:.8rem;">Mandatory</span></div>""", unsafe_allow_html=True)

with st.popover("ⓘ Info Tutorial"):
    st.markdown("""
**Tentang PO File:**
1. Klik **Make a Copy** setelah pilih Distributor.
2. Copy SKU dan QTY untuk dimasukkan dalam Spreadsheet.
3. Buat Share File jadi **Anyone with Link - View** - Wajib.
4. Paste link Spreadsheet yang sudah di buat **Make Copy**.
5. Pilih Distributor dan Nama RSA yang akan di assign.
6. Lakukan Preview File terlebih dahulu.
7. Export File bisa dalam bentuk PDF atau Excel.

📌 **Template PO:** [Klik di sini](https://docs.google.com/spreadsheets/d/1_4SFn2_SvGm1on0EJkntYjC2cLvNZyDjX54zcQAWRtQ/copy)
""")

st.markdown("<br>", unsafe_allow_html=True)

# ── Drill-down suggestion ──
with st.container(border=True):
    drill_col1, drill_col2 = st.columns([2,1])
    with drill_col1:
        drill_dist = st.multiselect("Pilih Distributor untuk lihat suggestion SKU",
                                     options=["(Pilih Distributor)"] + CUSTOMER_NAMES, key="drill_distri")
    with drill_col2:
        brand_options = get_brand_list()
        drill_brand = st.selectbox("Filter Brand", options=["All"] + brand_options, key="drill_brand")

    if drill_dist and any(d != "(Pilih Distributor)" for d in drill_dist):
        drill_df = get_distributor_suggestions(drill_dist, drill_brand)
        if drill_df.empty:
            st.info("ℹ️ Tidak ada suggestion SKU.")
        else:
            drill_agg = (drill_df.groupby(["DISTRIBUTOR","SKU"], as_index=False)
                         .agg(SUGGESTION_QTY=("SUGGESTION_QTY","sum"),
                              REMAINING_ALLOCATION=("REMAINING_ALLOCATION","sum"),
                              CURRENT_WOI=("CURRENT_WOI","first"),
                              WOI_AFTER_PO=("WOI_AFTER_PO","first"),
                              STATUS_ALOKASI=("STATUS_ALOKASI","first"))
                         .sort_values(["DISTRIBUTOR","SUGGESTION_QTY"], ascending=[True,False])
                         .reset_index(drop=True))
            drill_skus = drill_agg["SKU"].astype(str).str.upper().tolist()
            drill_names = get_sku_data(tuple(drill_skus))
            if not drill_names.empty:
                drill_names["sku"] = drill_names["sku"].astype(str).str.upper()
                drill_agg["SKU"] = drill_agg["SKU"].astype(str).str.upper()
                drill_agg = drill_agg.merge(
                    drill_names[["sku","product_name"]].rename(columns={"sku":"SKU","product_name":"PRODUCT_NAME"}),
                    on="SKU", how="left")
            else:
                drill_agg["PRODUCT_NAME"] = ""

            display_cols = [c for c in ["DISTRIBUTOR","SKU","PRODUCT_NAME","SUGGESTION_QTY","CURRENT_WOI","WOI_AFTER_PO","REMAINING_ALLOCATION","STATUS_ALOKASI"] if c in drill_agg.columns]
            tbl_df = drill_agg[display_cols].copy()
            for col in ["SUGGESTION_QTY","REMAINING_ALLOCATION"]:
                if col in tbl_df.columns:
                    tbl_df[col] = tbl_df[col].apply(lambda x: int(x) if pd.notna(x) else 0)

            def _highlight_alokasi(val):
                if val == "Reject by Steve": return "background-color:#FFB6C1;color:#8B0000;font-weight:600;"
                elif val == "Terdapat Alokasi": return "background-color:#D6EAF8;color:#1A5490;font-weight:600;"
                elif val == "Alokasi Habis": return "background-color:#FADBD8;color:#922B21;font-weight:600;"
                return ""

            fmt_dict = {c: "{:.2f}" for c in ["CURRENT_WOI","WOI_AFTER_PO"] if c in tbl_df.columns}
            if "STATUS_ALOKASI" in tbl_df.columns:
                styled_tbl = tbl_df.style.map(_highlight_alokasi, subset=["STATUS_ALOKASI"]).format(fmt_dict)
            else:
                styled_tbl = tbl_df.style.format(fmt_dict)
            st.dataframe(styled_tbl, use_container_width=True, hide_index=True)

            # Export
            export_buf = io.BytesIO()
            with pd.ExcelWriter(export_buf, engine='openpyxl') as writer:
                tbl_df.to_excel(writer, index=False, sheet_name='PO Suggestion')
                ws_exp = writer.sheets['PO Suggestion']
                hdr_fill = PatternFill(start_color="8B2040", end_color="8B2040", fill_type="solid")
                hdr_font = Font(bold=True, color="FFFFFF")
                for cell in ws_exp[1]:
                    cell.fill = hdr_fill; cell.font = hdr_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                hdrs_exp = {c.value: c.column for c in ws_exp[1]}
                alok_col = hdrs_exp.get("STATUS_ALOKASI")
                if alok_col:
                    alok_blue = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
                    alok_red = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
                    for r in range(2, ws_exp.max_row+1):
                        val = ws_exp.cell(row=r, column=alok_col).value
                        if val == "Terdapat Alokasi":
                            ws_exp.cell(row=r, column=alok_col).fill = alok_blue
                            ws_exp.cell(row=r, column=alok_col).font = Font(color="1A5490", bold=True)
                        elif val == "Alokasi Habis":
                            ws_exp.cell(row=r, column=alok_col).fill = alok_red
                            ws_exp.cell(row=r, column=alok_col).font = Font(color="922B21", bold=True)
                for col_cells in ws_exp.columns:
                    max_len = max((len(str(c.value)) if c.value else 0) for c in col_cells)
                    ws_exp.column_dimensions[col_cells[0].column_letter].width = min(max_len+4, 40)
                ws_exp.freeze_panes = "A2"

            fname_label = "_".join([d for d in drill_dist if d != "(Pilih Distributor)"])[:50]
            st.download_button(label="Export PO Suggestion (.xlsx)", data=export_buf.getvalue(),
                                file_name=f"PO_Suggestion_{fname_label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True, key="dl_po_suggestion")

            tpl_match = re.search(r'/spreadsheets/d/([a-zA-Z0-9_-]+)', TEMPLATE_PO_URL)
            copy_url = f"https://docs.google.com/spreadsheets/d/{tpl_match.group(1)}/copy" if tpl_match else TEMPLATE_PO_URL
            st.markdown(f"""<div style="margin-top:1rem;padding:.9rem 1.2rem;background:#FFF5F8;border:1px solid #F0C8D6;border-radius:10px;display:flex;align-items:center;justify-content:space-between;">
                <div style="font-size:.88rem;color:#5A1E38;">📄 Gunakan template PO ini, lalu isi dengan data suggestion di atas</div>
                <a href="{copy_url}" target="_blank" style="background:#F49CB6;color:#fff!important;text-decoration:none;padding:.45rem 1.1rem;border-radius:8px;font-size:.85rem;font-weight:700;white-space:nowrap;margin-left:1rem;">
                📝 Make a Copy Template PO</a></div>""", unsafe_allow_html=True)

st.divider()

RSA = ['Aqil','Alfaradi','Erliana','Rizky','Geirda','Rintan','Shaltsa','Daffa']
tabs = st.tabs(["🔗 Google Spreadsheet"])

with tabs[0]:
    gsheet_url = st.text_input("Google Spreadsheet URL",
                                placeholder="https://docs.google.com/spreadsheets/d/...",
                                label_visibility="collapsed")
    if st.button("Load Data"):
        if not gsheet_url.strip():
            st.warning("Masukkan link Google Spreadsheet dulu.")
        else:
            csv_url, sheet_id = gsheet_to_csv_url(gsheet_url)
            if csv_url is None:
                st.error("Link tidak valid.")
            else:
                with st.spinner("🌸 Loading data..."):
                    try:
                        req = urllib.request.Request(csv_url, headers={"User-Agent": "Mozilla/5.0"})
                        with urllib.request.urlopen(req, timeout=30) as resp:
                            data = resp.read()
                        def _read_csv_safe(raw, **kwargs):
                            for enc in ("utf-8","cp1252","latin-1","iso-8859-1"):
                                try:
                                    return pd.read_csv(io.BytesIO(raw), encoding=enc, on_bad_lines='skip', engine='python', **kwargs)
                                except Exception:
                                    continue
                            return pd.read_csv(io.BytesIO(raw), encoding="utf-8", encoding_errors="replace", on_bad_lines='skip', engine='python', **kwargs)
                        df_loaded = _read_csv_safe(data, dtype=str)
                        df_column = _read_csv_safe(data, header=8)
                        df_loaded = numeric_coerce(df_loaded)
                        st.session_state['raw_df'] = df_loaded
                        st.session_state['df'] = df_column
                        st.session_state['gsheet_url'] = gsheet_url
                        st.session_state['data_source'] = f"Google Sheet ({sheet_id[0]}...)"
                        st.session_state['source_type'] = 'GSHEET'
                        st.session_state.pop('export_bytes', None)
                        st.success("✅ Data berhasil dimuat!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"⚠️ Gagal memuat data: {e}")
                        st.info("Pastikan link sudah diset ke 'Anyone with the link'.")

    if 'df' not in st.session_state:
        st.stop()

    df = st.session_state['df'].copy()

    # ── Validation ──
    val_errors = []
    has_product_code = 'PRODUCT CODE' in df.columns
    has_qty = 'QTY' in df.columns
    if not has_product_code: val_errors.append("❌ Kolom **PRODUCT CODE** tidak ditemukan.")
    if not has_qty: val_errors.append("❌ Kolom **QTY** tidak ditemukan.")
    if has_qty:
        df_check = df.copy()
        df_check['_qty_num'] = pd.to_numeric(df_check['QTY'].astype(str).str.replace(',','.').str.strip(), errors='coerce')
        invalid_qty = df_check[df_check['QTY'].notna() & df_check['QTY'].astype(str).str.strip().ne('') & df_check['_qty_num'].isna()]['QTY'].unique().tolist()
        if invalid_qty:
            val_errors.append(f"❌ Kolom **QTY** memiliki nilai bukan angka: `{'`, `'.join([str(v) for v in invalid_qty[:10]])}`")
    if has_product_code:
        df_pc = df['PRODUCT CODE'].dropna().astype(str).str.strip()
        df_pc = df_pc[df_pc != '']
        invalid_pc = df_pc[~df_pc.str.upper().str.match(r'^G2G-\S+$')].unique().tolist()
        if invalid_pc:
            val_errors.append(f"⚠️ **PRODUCT CODE** tidak sesuai format: `{'`, `'.join([str(v) for v in invalid_pc[:10]])}`")
        pc_list = df_pc.str.upper().unique().tolist()
        if pc_list:
            with st.spinner("🔍 Memvalidasi PRODUCT CODE ke BigQuery..."):
                bq_check = get_sku_data(tuple(pc_list))
            if not bq_check.empty:
                found_skus = set(bq_check['sku'].astype(str).str.upper().tolist())
                not_found = [p for p in pc_list if p not in found_skus]
                if not_found:
                    val_errors.append(f"❌ **PRODUCT CODE** tidak ditemukan di sistem: `{'`, `'.join(not_found[:10])}`")
            else:
                val_errors.append("⚠️ Tidak dapat memvalidasi PRODUCT CODE ke BigQuery")

    if val_errors:
        st.warning("**⚠️ Data perlu dicek ulang:**")
        for err in val_errors: st.markdown(f"- {err}")
        st.info("💡 Pastikan PRODUCT CODE sesuai format (contoh: **G2G-223**) dan QTY berisi angka.")
    else:
        st.success("✅ Validasi OK — PRODUCT CODE dan QTY sesuai format.")

    with st.container(border=True):
        col1, col2 = st.columns([1,1])
        with col1:
            st.markdown("**DISTRIBUTOR**")
            pilih = st.selectbox("", options=["(Pilih)"] + CUSTOMER_NAMES, key="distri", label_visibility="collapsed")
        with col2:
            st.markdown("**RSA NAME**")
            rsa_pilih = st.selectbox("", options=["(Pilih)"] + RSA, key="rsa", label_visibility="collapsed")

    if pilih == "(Pilih)":
        st.info("Silakan pilih Distributor terlebih dahulu.")
        st.stop()

    if 'DISTRIBUTOR' not in df.columns:
        df['DISTRIBUTOR'] = pilih
    else:
        df['DISTRIBUTOR'] = df['DISTRIBUTOR'].fillna(pilih)

    df['QTY'] = pd.to_numeric(df['QTY'], errors='coerce').fillna(0)
    df['DPP'] = pd.to_numeric(df['DPP'].astype(str).str.replace(',','.'), errors='coerce').fillna(0)
    df['TOTAL PRICE'] = df['QTY'] * df['DPP']

    discount = 0
    sub_total = df['TOTAL PRICE'].sum()
    tax = sub_total * 0.11
    grand_total = sub_total - discount + tax
    count_sku = df['PRODUCT CODE'].notna().sum()

    with st.expander(f"👁 Lihat isi · {len(df)} baris", expanded=False):
        st.dataframe(df.iloc[:,:6].reset_index(drop=True), use_container_width=True)
        st.success(f"GRAND TOTAL: Rp {grand_total:,.0f} | Total SKU: {count_sku}")

    PRODUCT_LIST_URL = "https://docs.google.com/spreadsheets/d/1_4SFn2_SvGm1on0EJkntYjC2cLvNZyDjX54zcQAWRtQ/export?format=csv&gid=91084545"

    @st.cache_data(ttl=3600, show_spinner=False)
    def load_product_list() -> pd.DataFrame:
        try:
            df_raw = pd.read_csv(PRODUCT_LIST_URL, header=None, nrows=10)
            header_row = 0
            for i, row in df_raw.iterrows():
                row_str = ' '.join(row.astype(str).str.upper().tolist())
                if any(k in row_str for k in ['PRODUCT','SKU','CODE']):
                    header_row = i; break
            df_p = pd.read_csv(PRODUCT_LIST_URL, header=header_row)
            df_p.columns = [str(c).strip().upper() for c in df_p.columns]
            return df_p
        except Exception as e:
            st.warning(f"Gagal load Product List: {e}")
            return pd.DataFrame()

    df_product = load_product_list()
    lifecycle_col = next((c for c in df_product.columns if any(k in c for k in ['LIFECYCLE','LIFE CYCLE','STATUS'])), None)
    product_code_col = next((c for c in df_product.columns if any(k in c for k in ['PRODUCT CODE','SKU','CODE'])), None)
    if lifecycle_col and product_code_col:
        df['PRODUCT CODE'] = df['PRODUCT CODE'].astype(str).str.strip()
        df_product[product_code_col] = df_product[product_code_col].astype(str).str.strip()
        df = df.merge(df_product[[product_code_col, lifecycle_col]].rename(
            columns={product_code_col:'PRODUCT CODE', lifecycle_col:'LIFECYCLE STATUS'}),
            on='PRODUCT CODE', how='left')
        preview_cols = [c for c in ['PRODUCT CODE','DESCRIPTION','LIFECYCLE STATUS'] if c in df.columns]
        df_preview = df[preview_cols].dropna(subset=['PRODUCT CODE']).copy()
        matched = df_preview['LIFECYCLE STATUS'].notna().sum()
        unmatched = df_preview['LIFECYCLE STATUS'].isna().sum()
        with st.expander(f"🔗 Preview Lifecycle Status · {matched} matched · {unmatched} not found", expanded=False):
            st.dataframe(df_preview.reset_index(drop=True), use_container_width=True, hide_index=True)

    st.divider()

    @st.cache_data(show_spinner=False)
    def fetch_template_xlsx(url: str) -> bytes:
        m = re.search(r'/spreadsheets/d/([a-zA-Z0-9_-]+)', url)
        if not m: return None
        gid_match = re.search(r'gid=(\d+)', url)
        gid = gid_match.group(1) if gid_match else '0'
        export_url = f"https://docs.google.com/spreadsheets/d/{m.group(1)}/export?format=xlsx&gid={gid}&single=true"
        req = urllib.request.Request(export_url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            raw = resp.read()
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=False, keep_links=False)
        for sn in wb.sheetnames[2:]: del wb[sn]
        ws = wb.active
        if ws.max_row > 200: ws.delete_rows(201, ws.max_row - 200)
        if ws.max_column > 7: ws.delete_cols(8, ws.max_column - 7)
        buf = io.BytesIO(); wb.save(buf)
        return buf.getvalue()

    def export_to_template(df_data, template_bytes, distributor, rsa_name, discount):
        wb = openpyxl.load_workbook(io.BytesIO(template_bytes), data_only=False, keep_links=False)
        ws = wb.active
        ws['B3'] = distributor
        ws['E3'] = datetime.now().strftime("%d %B %Y")
        ws['E5'] = rsa_name
        COL_MAP = {'DISTRIBUTOR':1,'PRODUCT CODE':2,'DESCRIPTION':3,'QTY':4,'DPP':5,'TOTAL PRICE':6}
        START_ROW = 10
        SUMMARY_LABELS = ['SUB-TOTAL','DISCOUNTS','Tax (11%)','GRAND TOTAL']
        df_export = df_data[df_data['PRODUCT CODE'].notna() & ~df_data['QTY'].astype(str).isin(SUMMARY_LABELS)].copy()
        for r_offset, (_, row) in enumerate(df_export.iterrows()):
            excel_row = START_ROW + r_offset
            for col_name, col_idx in COL_MAP.items():
                if col_name == 'TOTAL PRICE':
                    ws.cell(row=excel_row, column=col_idx).value = f"=D{excel_row}*E{excel_row}"
                else:
                    val = row.get(col_name, "")
                    if pd.isna(val) or str(val).strip() in ('','nan','None'): val = None
                    ws.cell(row=excel_row, column=col_idx, value=val)
        last_data_row = START_ROW + len(df_export) - 1
        summary_start = last_data_row + 2
        sub_row, disc_row, tax_row, grand_row = summary_start, summary_start+1, summary_start+2, summary_start+3
        for row_idx, label, formula in [
            (sub_row, "SUB-TOTAL", f"=SUM(F{START_ROW}:F{last_data_row})"),
            (disc_row, "DISCOUNTS", "=0"),
            (tax_row, "Tax (11%)", f"=F{sub_row}*0.11"),
            (grand_row, "GRAND TOTAL", f"=F{sub_row}-F{disc_row}+F{tax_row}"),
        ]:
            ws.cell(row=row_idx, column=4, value=label)
            ws.cell(row=row_idx, column=6, value=formula)
        try:
            wb.calculation.fullCalcOnLoad = True; wb.calculation.calcMode = 'auto'
        except Exception: pass
        from openpyxl.styles import Border, Side
        sig_row = grand_row + 8
        _green_font = Font(bold=True, color="006400", size=10)
        _red_font = Font(bold=True, color="C00000", size=9)
        _black_font = Font(bold=True, color="000000", size=10)
        _thin_green = Side(border_style="thin", color="006400")
        _box_border = Border(left=_thin_green, right=_thin_green, top=_thin_green, bottom=_thin_green)
        ws.cell(row=sig_row, column=1, value="Initiated by,").font = _black_font
        cell_init = ws.cell(row=sig_row, column=2, value=f"ASM Approval, ({distributor})")
        cell_init.font = _green_font; cell_init.border = _box_border
        ws.cell(row=sig_row, column=6, value="APPROVE").font = _black_font
        ws.cell(row=sig_row+1, column=2, value="(mandatory sign)").font = _red_font
        ws.cell(row=sig_row+1, column=6, value="(SIGN/CAP DISTRIBUTOR)").font = _red_font
        buf = io.BytesIO(); wb.save(buf)
        return buf.getvalue()

    def excel_to_pdf(df_data, distributor, rsa_name, sub_total, tax, grand_total, discount=0):
        df_clean = df_data[df_data['PRODUCT CODE'].notna()].copy()
        df_clean['TOTAL PRICE'] = pd.to_numeric(
            df_clean['TOTAL PRICE'].astype(str).str.replace('.','').str.replace(',','.'), errors='coerce').fillna(0)
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
        elements = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title', parent=styles['Title'], fontName='Trebuchet-Bold',
                                      fontSize=12, textColor=colors.HexColor("#B53473"), alignment=1, spaceAfter=12)
        elements.append(Paragraph("PURCHASE ORDER", title_style))
        info_data = [['Distributor:', distributor, 'Date:', datetime.now().strftime("%d %B %Y")], ['RSA:', rsa_name, '', '']]
        info_tbl = Table(info_data, colWidths=[20*mm, 85*mm, 20*mm, 50*mm])
        info_tbl.setStyle(TableStyle([
            ('FONTNAME',(0,0),(-1,-1),'Trebuchet'), ('FONTSIZE',(0,0),(-1,-1),10),
            ('FONTNAME',(0,0),(0,-1),'Trebuchet-Bold'), ('FONTNAME',(2,0),(2,-1),'Trebuchet-Bold'),
            ('BOTTOMPADDING',(0,0),(-1,-1),6),
        ]))
        elements.append(info_tbl)
        elements.append(Spacer(1, 8*mm))
        header = ['No','PRODUCT CODE','DESCRIPTION','QTY','DPP','TOTAL PRICE']
        data = [header]
        for i, (_, row) in enumerate(df_clean.iterrows(), start=1):
            qty = row.get('QTY',''); dpp = row.get('DPP',''); total = row.get('TOTAL PRICE',0)
            data.append([str(i), str(row.get('PRODUCT CODE','')), str(row.get('DESCRIPTION',''))[:40],
                         f"{qty}" if pd.notna(qty) else "",
                         f"{dpp:,.0f}" if isinstance(dpp,(int,float)) else str(dpp),
                         f"{total:,.0f}" if isinstance(total,(int,float)) else str(total)])
        data += [['','','','','SUB-TOTAL', f"{sub_total:,.0f}"],
                 ['','','','','DISCOUNTS', f"-{discount:,.0f}"],
                 ['','','','','Tax (11%)', f"{tax:,.0f}"],
                 ['','','','','GRAND TOTAL', f"{grand_total:,.0f}"]]
        tbl = Table(data, colWidths=[10*mm,25*mm,70*mm,20*mm,20*mm,30*mm])
        tbl.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#BF3979')),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white), ('FONTSIZE',(0,0),(-1,0),7),
            ('ALIGN',(0,0),(-1,0),'CENTER'), ('FONTNAME',(0,1),(-1,-5),'Trebuchet'),
            ('FONTSIZE',(0,1),(-1,-1),6), ('ALIGN',(3,1),(-1,-1),'RIGHT'),
            ('GRID',(0,0),(-1,-5),0.4,colors.HexColor('#FFB6C1')),
            ('FONTNAME',(4,-4),(-1,-1),'Trebuchet-Bold'),
            ('BACKGROUND',(4,-1),(-1,-1),colors.HexColor('#FFB6C1')),
            ('BOTTOMPADDING',(0,0),(-1,-1),4), ('TOPPADDING',(0,0),(-1,-1),4),
            ('ROWBACKGROUNDS',(0,1),(-1,-5),[colors.white,colors.HexColor('#FAFAFA')]),
        ]))
        elements.append(tbl)
        elements.append(Spacer(1, 15*mm))
        sig_data = [['Initiated by,', f'ASM Approval, ({distributor})', '', 'APPROVE'],
                    ['', '(mandatory sign)', '', '(SIGN/CAP DISTRIBUTOR)']]
        sig_tbl = Table(sig_data, colWidths=[25*mm,70*mm,20*mm,60*mm])
        sig_tbl.setStyle(TableStyle([
            ('FONTNAME',(0,0),(-1,-1),'Trebuchet-Bold'), ('FONTSIZE',(0,0),(-1,0),9),
            ('TEXTCOLOR',(1,0),(1,0),colors.HexColor('#006400')),
            ('BOX',(1,0),(1,0),0.8,colors.HexColor('#006400')),
            ('TEXTCOLOR',(1,1),(1,1),colors.HexColor('#C00000')),
            ('TEXTCOLOR',(3,1),(3,1),colors.HexColor('#C00000')),
        ]))
        elements.append(sig_tbl)
        doc.build(elements)
        return buf.getvalue()

    if st.button("🔄 Generate", use_container_width=True):
        with st.spinner("Prepare file..."):
            try:
                tpl_bytes = fetch_template_xlsx(st.session_state['gsheet_url'])
                export_bytes = export_to_template(df, tpl_bytes, pilih, rsa_pilih, discount)
                st.session_state['export_bytes'] = export_bytes
            except Exception as e:
                st.error(f"Gagal generate file: {e}")
                st.session_state.pop('export_bytes', None)

    if 'export_bytes' in st.session_state:
        prog = st.progress(0)
        for i in range(101):
            prog.progress(i, text="Processing complete" if i == 100 else f"Loading... {i}%")

        dl_col1, dl_col2 = st.columns(2)
        with dl_col1:
            st.download_button(label="Export Excel", data=st.session_state['export_bytes'],
                                file_name=f"PO_{pilih}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True, key="dl_excel")
        with dl_col2:
            try:
                pdf_bytes = excel_to_pdf(df, pilih, rsa_pilih, sub_total, tax, grand_total, discount)
                st.download_button(label="📄 Export PDF", data=pdf_bytes,
                                    file_name=f"PO_{pilih}_{datetime.now().strftime('%Y%m%d')}.pdf",
                                    mime="application/pdf", use_container_width=True, key="dl_pdf")
            except Exception as e:
                st.error(f"Gagal generate PDF: {e}")
