import io
import streamlit as st
import pandas as pd
import numpy as np
from typing import List
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# --- BigQuery Imports ---
from google.cloud import bigquery
from google.api_core.exceptions import NotFound
from google.oauth2 import service_account

# =========================
# BigQuery Configuration
# =========================
# Use Streamlit secrets if available, fallback to local path
try:
    gcp_secrets = st.secrets["connections"]["bigquery"]
    private_key = gcp_secrets["private_key"].replace("\\n", "\n")
    credentials = service_account.Credentials.from_service_account_info(
        {
            "type": gcp_secrets["type"],
            "project_id": gcp_secrets["project_id"],
            "private_key_id": gcp_secrets["private_key_id"],
            "private_key": private_key,
            "client_email": gcp_secrets["client_email"],
            "client_id": gcp_secrets["client_id"],
            "auth_uri": gcp_secrets["auth_uri"],
            "token_uri": gcp_secrets["token_uri"],
            "auth_provider_x509_cert_url": gcp_secrets["auth_provider_x509_cert_url"],
            "client_x509_cert_url": gcp_secrets["client_x509_cert_url"],
        }
    )
    GCP_PROJECT_ID = st.secrets["bigquery"]["project"]
    BQ_DATASET = st.secrets["bigquery"]["dataset"]
    BQ_TABLE = st.secrets["bigquery"]["stock_analysis_table"]
except Exception:
    # Fallback for local testing if secrets are not configured
    GCP_CREDENTIALS_PATH = r"C:\script\skintific-data-warehouse-ea77119e2e7a.json"
    GCP_PROJECT_ID = "skintific-data-warehouse"
    BQ_DATASET = "rsa"
    BQ_TABLE = "stock_analysis"
    credentials = service_account.Credentials.from_service_account_file(
        GCP_CREDENTIALS_PATH
    )


@st.cache_resource(show_spinner=False)
def get_bq_client() -> bigquery.Client:
    """Initializes and returns a BigQuery client."""
    return bigquery.Client(credentials=credentials, project=GCP_PROJECT_ID)


@st.cache_data(ttl=21600, show_spinner="Fetching distributor data from BigQuery...")
def get_distributor_data() -> List[str]:
    """Fetches unique customer and branch codes from BigQuery."""
    client = get_bq_client()
    table_id = f"{GCP_PROJECT_ID}.{BQ_DATASET}.{BQ_TABLE}"
    try:
        query = f"""
        SELECT DISTINCT distributor
        FROM `{table_id}`
        ORDER BY distributor
        """
        rows = client.query(query).result()
        return [r.distributor for r in rows]
    except NotFound:
        st.error(f"Error: BigQuery table '{table_id}' not found.")
        return []
    except Exception as e:
        st.error(f"An error occurred while fetching data from BigQuery: {e}")
        return []


@st.cache_data(ttl=21600, show_spinner="Fetching SKU data from BigQuery...")
def get_sku_data(sku_list: List[str]) -> pd.DataFrame:
    """
    Fetches SKU data (including product_name) for a given list of SKUs from BigQuery.
    Uses UPPER() on the sku column to handle case-insensitive matching.
    """
    client = get_bq_client()
    table_id = f"{GCP_PROJECT_ID}.gt_schema.master_product"

    # Create a string of the SKUs for the IN clause (already uppercased)
    sku_list_str = ", ".join([f"'{sku}'" for sku in sku_list])

    query = f"""
    SELECT
        sku,
        product_name,
        price_for_distri
    FROM `{table_id}`
    WHERE UPPER(sku) IN ({sku_list_str})
    """
    try:
        df_sku_data = client.query(query).to_dataframe()
        return df_sku_data
    except Exception as e:
        st.error(f"Error fetching SKU data from BigQuery: {e}")
        return pd.DataFrame()


@st.cache_data(ttl=21600, show_spinner="Fetching NPD Allocation data from BigQuery...")
def get_npd_data(sku_list: List[str]) -> pd.DataFrame:
    """
    Fetches NPD Allocation data for a given list of SKUs from BigQuery.
    """
    client = get_bq_client()
    table_id = f"{GCP_PROJECT_ID}.gt_schema.npd_allocation"

    # Create a string of the SKUs for the IN clause
    sku_list_str = ", ".join([f"'{sku}'" for sku in sku_list])

    query = f"""
    SELECT
        calendar_date,
        region,
        sku
    FROM `{table_id}`
    WHERE sku IN ({sku_list_str})
    AND calendar_date = '2026-04-01'
    """
    try:
        df_sku_data = client.query(query).to_dataframe()
        return df_sku_data
    except Exception as e:
        st.error(f"Error fetching NPD data from BigQuery: {e}")
        return pd.DataFrame()


@st.cache_data(ttl=21600, show_spinner="Fetching Stock Analysis data from BigQuery...")
def get_stock_data(distributor_name: str, sku_list: List[str]) -> pd.DataFrame:
    """
    Fetches stock and sales data for a given list of SKUs from BigQuery.
    Uses UPPER() on the sku column to handle case-insensitive matching.
    """
    client = get_bq_client()
    table_id = f"{GCP_PROJECT_ID}.{BQ_DATASET}.{BQ_TABLE}"

    # Create a string of the SKUs for the IN clause
    sku_list_str = ", ".join([f"'{sku}'" for sku in sku_list])

    query = f"""
    SELECT
        UPPER(region) AS region,
        UPPER(distributor) AS distributor,
        sku,
        assortment,
        supply_control_status_gt,
        total_stock,
        buffer_plan_by_lm_qty_adj,
        avg_weekly_st_lm_qty,
        buffer_plan_by_lm_val_adj,
        remaining_allocation_qty_region,
        woi_end_of_month_by_lm
    FROM `{table_id}`
    WHERE UPPER(distributor) = '{distributor_name}'
    AND (
        UPPER(sku) IN ({sku_list_str})
        OR buffer_plan_by_lm_qty_adj > 0
    )
    """
    try:
        df_sku_data = client.query(query).to_dataframe()
        return df_sku_data
    except Exception as e:
        st.error(f"Error fetching Stock Analysis data from BigQuery: {e}")
        return pd.DataFrame()


def calculate_woi(stock: pd.Series, po_qty: pd.Series, avg_weekly_sales: pd.Series) -> pd.Series:
    """
    Calculates Weeks of Inventory (WOI) based on the formula:
    (Stock + PO Quantity) / Average Weekly Sales LM
    """
    return np.where(avg_weekly_sales > 0, (stock + po_qty) / avg_weekly_sales, 0)


def apply_sku_rejection_rules(sku_list: List, df: pd.DataFrame, regions: List[str], is_in: bool) -> pd.DataFrame:
    """
    Auto-rejects specific SKUs based on a provided list and region rules.

    Args:
        sku_list: List of SKUs to apply rules to
        df: DataFrame containing the data
        regions: List of allowed regions
        is_in: If False, only allow SKUs in the specified regions (reject all others)
               If True, reject SKUs in the specified regions
    """
    regions_upper = [r.upper() for r in regions]

    if not is_in:
        condition = (df["SKU"].isin(sku_list)) & (~df["region"].str.upper().isin(regions_upper))
    else:
        condition = (df["SKU"].isin(sku_list)) & (df["region"].str.upper().isin(regions_upper))

    df.loc[condition, "Remark"] = "Reject (Stop by Steve)"

    return df


def to_excel_with_styling(dfs: dict, npd_sku_list: List[str] = None) -> bytes:
    """
    Converts a pandas DataFrame to an Excel file with special styling for the first 7 columns.
    Applies specific color to 'Remaining Allocation (By Region)' column if SKU is in npd_sku_list.
    """
    output = io.BytesIO()
    wb = Workbook()

    del wb["Sheet"]

    po_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    suggestion_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    npd_fill = PatternFill(start_color="B1DBF0", end_color="B1DBF0", fill_type="solid")

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    header_alignment = Alignment(horizontal='left', vertical='center')

    proceed_font = Font(bold=True, color="54CE54")
    reject_font = Font(bold=True, color="D73E3E")
    suggest_font = Font(bold=True, color="F3C94C")

    for sheet_name, df in dfs.items():
        ws = wb.create_sheet(title=sheet_name[:31])

        is_po_sku_series = df["is_po_sku"]
        df_no_flag = df.drop("is_po_sku", axis=1)

        rows = dataframe_to_rows(df_no_flag, index=False, header=True)

        headers = list(df_no_flag.columns)
        col_map = {col: i for i, col in enumerate(headers)}

        currency_cols = ["PO Value", "Suggested PO Value"]
        integer_cols = ["Remaining Allocation (By Region)", "Avg Weekly Sales LM (Qty)"]
        decimal_cols = ["WOI (Stock + PO Ori)", "Current WOI", "WOI After Buffer (Stock + Suggested Qty)", "Stock + Suggested Qty WOI (Projection at EOM)"]

        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                if c_idx <= 11 and r_idx > 1:
                    original_row_index = (r_idx - 2)
                    is_po_row = is_po_sku_series.iloc[original_row_index]

                    if is_po_row:
                        cell.fill = po_fill
                    else:
                        cell.fill = suggestion_fill

                if r_idx > 1:
                    col_name = headers[c_idx - 1]

                    if col_name in currency_cols:
                        cell.number_format = "#,##0.00"
                    elif col_name in integer_cols:
                        cell.number_format = "#,##0"
                    elif col_name in decimal_cols:
                        cell.number_format = "0.00"

                    if col_name == "Remark":
                        remark_value = row[c_idx - 1]
                        if "Proceed" in remark_value:
                            cell.font = proceed_font
                        elif "Reject" in remark_value:
                            cell.font = reject_font
                        elif "Additional" in remark_value:
                            cell.font = suggest_font

                    if col_name in ["Remaining Allocation (By Region)", "Suggested PO Qty", "Suggested PO Value"] and npd_sku_list is not None:
                        sku_col_index = col_map.get("SKU")
                        if sku_col_index is not None:
                            sku_value = row[sku_col_index]
                            if sku_value in npd_sku_list:
                                cell.fill = npd_fill

    wb.save(output)
    output.seek(0)
    return output.getvalue()


def to_excel_single_sheet(df: pd.DataFrame, npd_sku_list: List[str] = None) -> bytes:
    """
    Converts a single pandas DataFrame (all distributors stacked)
    to an Excel file with special styling.
    """
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "PO Simulator"

    po_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    suggestion_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    npd_fill = PatternFill(start_color="B1DBF0", end_color="B1DBF0", fill_type="solid")

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    header_alignment = Alignment(horizontal='left', vertical='center')

    proceed_font = Font(bold=True, color="54CE54")
    reject_font = Font(bold=True, color="D73E3E")
    suggest_font = Font(bold=True, color="F3C94C")

    is_po_sku_series = df["is_po_sku"]
    df_no_flag = df.drop("is_po_sku", axis=1)

    rows = dataframe_to_rows(df_no_flag, index=False, header=True)

    headers = list(df_no_flag.columns)
    col_map = {col: i for i, col in enumerate(headers)}

    currency_cols = ["PO Value", "Suggested PO Value"]
    integer_cols = ["Remaining Allocation (By Region)", "Avg Weekly Sales LM (Qty)"]
    decimal_cols = ["WOI (Stock + PO Ori)", "Current WOI", "WOI After Buffer (Stock + Suggested Qty)", "Stock + Suggested Qty WOI (Projection at EOM)"]

    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            if c_idx <= 11 and r_idx > 1:
                original_row_index = (r_idx - 2)
                is_po_row = is_po_sku_series.iloc[original_row_index]

                if is_po_row:
                    cell.fill = po_fill
                else:
                    cell.fill = suggestion_fill

            if r_idx > 1:
                col_name = headers[c_idx - 1]

                if col_name in currency_cols:
                    cell.number_format = "#,##0.00"
                elif col_name in integer_cols:
                    cell.number_format = "#,##0"
                elif col_name in decimal_cols:
                    cell.number_format = "0.00"

                if col_name == "Remark":
                    remark_value = row[c_idx - 1]
                    if "Proceed" in remark_value:
                        cell.font = proceed_font
                    elif "Reject" in remark_value:
                        cell.font = reject_font
                    elif "Additional" in remark_value:
                        cell.font = suggest_font

                if col_name in ["Remaining Allocation (By Region)", "Suggested PO Qty", "Suggested PO Value"] and npd_sku_list is not None:
                    sku_col_index = col_map.get("SKU")
                    if sku_col_index is not None:
                        sku_value = row[sku_col_index]
                        if sku_value in npd_sku_list:
                            cell.fill = npd_fill

    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_po_template_excel() -> bytes:
    """
    Creates a blank Excel template file with the required headers.
    """
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "PO Template"

    ws["A2"] = "PUCHASE ORDER FORM"
    ws["A3"] = "CUSTOMER NAME :"
    ws["A4"] = "NPWP / ID CARD :"
    ws["A5"] = "ADDRESS :"
    ws["D3"] = "DATE :"
    ws["D4"] = "Berlaku Sampai"
    ws["D5"] = "Issued by"

    ws["A2"].font = Font(bold=True, size=16)

    headers = ["DISTRIBUTOR", "PRODUCT CODE", "DESCRIPTION", "QTY", "DPP", "TOTAL PRICE"]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col_idx, value=header)
        cell.font = Font(bold=True)

    ws.freeze_panes = "A9"

    wb.save(output)
    output.seek(0)
    return output.getvalue()


def main():
    st.set_page_config(page_title="PO Simulator", page_icon="🛒", layout="wide")

    st.title("🛒 PO Simulator")
    st.markdown(
        "Use this app to simulate Purchase Order data and decide on whether to Reject / Approve the PO."
    )

    tab1, tab2 = st.tabs(["📖 Guide & Rules", "🔍 Simulation & Results"])

    # Hardcoded Reject List with specific remarks
    MANUAL_REJECT_SKUS_APPROVAL = ["G2G-252", "G2G-253"]
    MANUAL_REJECT_SKUS_NO_TOLERANCE = [
        "G2G-29705",
        "G2G-224",
        "G2G-247",
        "G2G-225",
        "G2G-226",
        "G2G-228",
        "G2G-74",
        "G2G-186",
        "G2G-202",
        "G2G-840",
        "G2G-844",
        "G2G-841",
        "G2G-800",
        "G2G-213",
        "G2G-217",
        "G2G-27305",
        "G2G-30701",
        "G2G-30702",
        "G2G-30703",
        "G2G-30704",
        "G2G-201",
        "G2G-31"
    ]

    MANUAL_REJECT_SKUS = MANUAL_REJECT_SKUS_APPROVAL + MANUAL_REJECT_SKUS_NO_TOLERANCE

    LIMITED_SKUS_QTY = []
    MAX_QTY_LIMIT = 500

    REJECTED_SKUS_1 = ["G2G-29700", "G2G-27300"]
    REGION_LIST_1 = [
        "Central Sumatera",
        "Northern Sumatera",
        "Jakarta (Csa)",
        "West Kalimantan",
        "South Kalimantan",
        "East Kalimantan"
    ]

    REJECTED_SKUS_2 = []
    REGION_LIST_2 = []

    with tab1:
        st.header("How to Use the PO Simulator")
        with st.expander("📋 Step-by-Step Guide"):
            st.markdown("""
            1. **Upload PO Data**: Upload an Excel or CSV file containing the required columns: 'DISTRIBUTOR', 'PRODUCT CODE', 'DESCRIPTION', 'QTY'. The file should start from the 8th row.
            2. **Review Rejection Lists**: Check the manual rejection SKUs and region-based rejections if applicable.
            3. **Simulate and Analyze**: The app will fetch stock and sales data from BigQuery, perform calculations like Weeks of Inventory (WOI), and apply approval/rejection rules.
            4. **View Results**: Review the simulated data in the table, including remarks on whether to proceed, reject, or suggest adjustments.
            5. **Download Excel**: Click the download button to get a single Excel file with a separate sheet for each distributor's results, or a single sheet for all distributors.
            """)

        st.header("Rules & Calculations Logic")
        with st.expander("⚖️ Rules & Calculations Logic Explanation"):
            st.markdown("""
            The following rules are applied in order to determine the remark for each SKU. The rest are the calculations logic behind the results.

            1.  **Reject**: The SKU is rejected if any of these conditions are met:
                * The SKU **does not exist** in the system (not found in BigQuery master product or stock analysis tables).
                * It is on the **regional rejection list** (Stop by Steve), unless the region is allowed to order.
                * It is on the **manual rejection list** (Stop by Steve) with specific remarks:
                    * **Need approval email**: G2G-252, G2G-253
                    * **No tolerance to open**: All other manually rejected SKUs
                * The **Remaining Allocation (By Region)** is **less than 0** (Negative Allocation).
                * The **Current WOI** is too high (Exceeded the WOI Standard).
                * The PO quantity is **greater than** the suggested PO quantity (over-ordering) -> **Reject with Suggestion**.

            2.  **Proceed**: The SKU is approved if any of these conditions are met:
                * The PO quantity is **less than** the suggested PO quantity (under-ordering) -> **Proceed with Suggestion**.
                * The PO quantity **exactly matches** the suggested PO quantity.
                * It's a new product (NPD) with remaining region allocation **greater than 0**, and the PO quantity is **less than or equal to** the remaining allocation by region.
                * It has no historical trend data, and the supply control status is **not** "STOP PO," "DISCONTINUED," or "OOS."

            3.  **Additional Suggestion**: The SKU is marked as an "Additional Suggestion" if:
                * It was not on the original PO but was **suggested by the system**.
                * **Note**: SKUs that would be rejected are automatically filtered out from suggestions.

            4.  **Suggested WOI (OH + IT + Suggested Qty + ST Projection until EOM)**: The suggested WOI is calculated based on the total stock + Suggested Qty + ST Projection until end of month
            """)

        st.header("Manual Rejection SKUs")
        with st.expander("🚫 List of SKUs that are manually rejected by Steve"):
            reject_data = []
            for sku in MANUAL_REJECT_SKUS_APPROVAL:
                reject_data.append({"SKU": sku, "Remark": "Need approval email"})
            for sku in MANUAL_REJECT_SKUS_NO_TOLERANCE:
                reject_data.append({"SKU": sku, "Remark": "No tolerance to open"})

            reject_df = pd.DataFrame(reject_data).sort_values(by="SKU").reset_index(drop=True)
            st.dataframe(reject_df)

        if REJECTED_SKUS_1:
            st.header("Regional Rejection Rules")
            with st.expander("🌍 SKUs with Regional Restrictions"):
                st.markdown(f"""
                **SKUs: {', '.join(REJECTED_SKUS_1)}**

                These SKUs are **ONLY allowed** in the following regions:
                """)
                for region in REGION_LIST_1:
                    st.markdown(f"- {region}")
                st.markdown("**All other regions will be automatically rejected.**")

    with tab2:
        st.header("1. Download PO Template")

        st.download_button(
            label="📥 Download PO Template",
            data=create_po_template_excel(),
            file_name="po_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.header("2. Upload PO Data")
        uploaded_file = st.file_uploader(
            "Upload a PO file (.xlsx/.csv) containing: 'PRODUCT CODE', 'DESCRIPTION', 'QTY', and 'DISTRIBUTOR'.",
            type=["xlsx", "xls", "csv"],
        )

        if uploaded_file:
            try:
                required_cols = ["PRODUCT CODE", "DESCRIPTION", "QTY", "DISTRIBUTOR"]

                if uploaded_file.name.endswith(".xlsx"):
                    po_df = pd.read_excel(uploaded_file, header=7, engine="openpyxl")
                else:
                    po_df = pd.read_csv(uploaded_file, header=7)
                st.success("File uploaded successfully!")

                po_df.dropna(axis=1, how='all', inplace=True)
                po_df.drop(columns=[col for col in po_df.columns if 'Unnamed' in str(col)], inplace=True)

                if not all(col in po_df.columns for col in required_cols):
                    st.error("The uploaded file is missing one or more required columns.")
                    st.write("Please check for these columns:", required_cols)
                    return

                po_df["QTY"] = pd.to_numeric(po_df["QTY"], errors="coerce")
                po_df.dropna(subset=["QTY"], inplace=True)

                po_df["DISTRIBUTOR"] = po_df["DISTRIBUTOR"].str.upper()
                po_df["PRODUCT CODE"] = po_df["PRODUCT CODE"].astype(str).str.strip().str.upper()

                st.write("Preview of uploaded PO data:")
                st.dataframe(po_df.head())

                po_df = po_df[po_df["QTY"] > 0]
                po_df.dropna(subset=["PRODUCT CODE", "QTY", "DISTRIBUTOR"], inplace=True)

                po_df["is_po_sku"] = True

                po_df.rename(
                    columns={
                        "PRODUCT CODE": "Customer SKU Code",
                        "QTY": "PO Qty",
                        "DISTRIBUTOR": "Distributor"
                    },
                    inplace=True,
                )

                po_df = po_df[["Distributor", "Customer SKU Code", "PO Qty", "is_po_sku"]]

                st.header("3. PO Simulation")

                st.subheader("4. Select Output Format and Download")
                output_format = st.radio(
                    "Choose Excel Output Format:",
                    ["Separate Sheets (One per Distributor)", "Single Sheet (All Distributors Stacked)"],
                    index=0
                )

                progress = st.progress(0)
                progress_step = 1.0 / len(po_df["Distributor"].unique())

                all_npd_sku_list = []
                excel_dfs = {}
                display_dfs = []

                uploaded_distributors = po_df["Distributor"].unique().tolist()

                for i, distributor_name in enumerate(uploaded_distributors):
                    st.subheader(f"Processing data for: **{distributor_name}**")
                    progress.progress((i * progress_step), f"Processing {distributor_name}...")

                    current_po_df = po_df[po_df["Distributor"] == distributor_name].copy()
                    sku_list = current_po_df["Customer SKU Code"].unique().tolist()

                    # --- Fetch data from BigQuery ---
                    sku_df = get_sku_data(sku_list)
                    sku_data_df = get_stock_data(distributor_name, sku_list)

                    # Rename sku_df columns — now includes product_name
                    sku_df.rename(
                        columns={
                            "sku": "Customer SKU Code",
                            "price_for_distri": "SIP",
                            "product_name": "Product Name",
                        },
                        inplace=True,
                    )

                    if sku_data_df.empty:
                        st.warning(
                            "Could not find stock and sales data for the uploaded Distributor/SKUs in BigQuery. Please check the Distributor Name or SKU codes."
                        )
                        return

                    # Rename stock data columns
                    if "sku" in sku_data_df.columns:
                        sku_data_df.rename(
                            columns={"sku": "Customer SKU Code", "distributor": "Distributor"},
                            inplace=True,
                        )

                    # Drop product_name from stock data — we use master_product's version
                    if "product_name" in sku_data_df.columns:
                        sku_data_df.drop(columns=["product_name"], inplace=True)

                    # ===== TRACK SKUs NOT FOUND IN BIGQUERY =====
                    skus_in_sku_df = set(sku_df["Customer SKU Code"].tolist()) if not sku_df.empty else set()
                    skus_in_stock_df = set(sku_data_df["Customer SKU Code"].tolist()) if not sku_data_df.empty else set()
                    skus_not_in_bq = set(sku_list) - (skus_in_sku_df | skus_in_stock_df)

                    if skus_not_in_bq:
                        st.warning(
                            f"The following SKUs were **not found** in the system and will be rejected: "
                            f"`{', '.join(sorted(skus_not_in_bq))}`"
                        )
                    # ===== END TRACK SKUs NOT FOUND =====

                    # Merge PO data with SKU price + product_name data
                    result_df = pd.merge(current_po_df, sku_df, on="Customer SKU Code", how="left")

                    result_df["SIP"] = pd.to_numeric(result_df["SIP"], errors="coerce").fillna(0)
                    result_df["PO Value"] = result_df["SIP"] * result_df["PO Qty"]

                    # Merge with stock data (outer to include suggested SKUs)
                    result_df = pd.merge(result_df, sku_data_df, on="Customer SKU Code", how="outer")

                    # After outer merge, fill Product Name for suggested SKUs from master_product
                    # by re-merging just the product names for any new SKUs introduced via outer join
                    missing_product_names = result_df["Product Name"].isna()
                    if missing_product_names.any():
                        suggested_skus = result_df.loc[missing_product_names, "Customer SKU Code"].unique().tolist()
                        if suggested_skus:
                            extra_sku_df = get_sku_data(suggested_skus)
                            if not extra_sku_df.empty:
                                extra_sku_df.rename(
                                    columns={
                                        "sku": "Customer SKU Code",
                                        "price_for_distri": "SIP",
                                        "product_name": "Product Name",
                                    },
                                    inplace=True,
                                )
                                extra_name_map = extra_sku_df.set_index("Customer SKU Code")["Product Name"].to_dict()
                                result_df.loc[missing_product_names, "Product Name"] = (
                                    result_df.loc[missing_product_names, "Customer SKU Code"].map(extra_name_map)
                                )

                    result_sku_list = result_df["Customer SKU Code"].unique().tolist()

                    npd_df = get_npd_data(result_sku_list)
                    current_npd_sku_list = npd_df['sku'].unique().tolist() if not npd_df.empty else []

                    all_npd_sku_list.extend(current_npd_sku_list)
                    all_npd_sku_list = list(set(all_npd_sku_list))

                    result_df["is_po_sku"] = result_df["is_po_sku"].astype("boolean")
                    result_df["is_po_sku"] = result_df["is_po_sku"].fillna(False)

                    result_df[
                        [
                            "PO Qty",
                            "PO Value",
                            "total_stock",
                            "buffer_plan_by_lm_qty_adj",
                            "avg_weekly_st_lm_qty",
                            "buffer_plan_by_lm_val_adj",
                            "remaining_allocation_qty_region",
                            "woi_end_of_month_by_lm",
                        ]
                    ] = result_df[
                        [
                            "PO Qty",
                            "PO Value",
                            "total_stock",
                            "buffer_plan_by_lm_qty_adj",
                            "avg_weekly_st_lm_qty",
                            "buffer_plan_by_lm_val_adj",
                            "remaining_allocation_qty_region",
                            "woi_end_of_month_by_lm",
                        ]
                    ].fillna(0)

                    result_df = result_df[
                        (result_df["PO Qty"] > 0) | (result_df["buffer_plan_by_lm_qty_adj"] > 0)
                    ]

                    # ===== ENHANCED REJECTION LOGIC FOR SUGGESTED SKUs =====
                    suggested_skus_mask = result_df["is_po_sku"] == False

                    exclude_suggested = (
                        (result_df["Customer SKU Code"].isin(skus_not_in_bq)) |
                        (result_df["Customer SKU Code"].isin(MANUAL_REJECT_SKUS_APPROVAL)) |
                        (result_df["Customer SKU Code"].isin(MANUAL_REJECT_SKUS_NO_TOLERANCE)) |
                        (result_df["remaining_allocation_qty_region"] < 0) |
                        (result_df["supply_control_status_gt"].str.upper().isin(["STOP PO", "DISCONTINUED", "OOS", "UNAVAILABLE"])) |
                        (
                            (result_df["Customer SKU Code"].isin(LIMITED_SKUS_QTY)) &
                            (result_df["buffer_plan_by_lm_qty_adj"] > MAX_QTY_LIMIT)
                        ) |
                        (result_df["buffer_plan_by_lm_qty_adj"] == 0)
                    )

                    if REJECTED_SKUS_1:
                        regions_upper_1 = [r.upper() for r in REGION_LIST_1]
                        regional_reject_1 = (
                            (result_df["Customer SKU Code"].isin(REJECTED_SKUS_1)) &
                            (~result_df["region"].str.upper().isin(regions_upper_1))
                        )
                        exclude_suggested = exclude_suggested | regional_reject_1

                    result_df = result_df[~(suggested_skus_mask & exclude_suggested)]
                    # ===== END ENHANCED REJECTION LOGIC =====

                    result_df["distributor_name"] = distributor_name

                    result_df["WOI PO Original"] = calculate_woi(
                        result_df["total_stock"],
                        result_df["PO Qty"],
                        result_df["avg_weekly_st_lm_qty"],
                    )

                    result_df["WOI Suggest"] = calculate_woi(
                        result_df["total_stock"],
                        result_df["buffer_plan_by_lm_qty_adj"],
                        result_df["avg_weekly_st_lm_qty"],
                    )

                    result_df["Current WOI"] = calculate_woi(
                        result_df["total_stock"],
                        0,
                        result_df["avg_weekly_st_lm_qty"],
                    )

                    conditions = [
                        result_df["Customer SKU Code"].isin(skus_not_in_bq),
                        (result_df["Customer SKU Code"].isin(LIMITED_SKUS_QTY)) & (result_df["PO Qty"] > MAX_QTY_LIMIT),
                        (result_df["remaining_allocation_qty_region"] < 0),
                        (result_df["is_po_sku"] == False),
                        result_df["Customer SKU Code"].isin(MANUAL_REJECT_SKUS_APPROVAL),
                        result_df["Customer SKU Code"].isin(MANUAL_REJECT_SKUS_NO_TOLERANCE),
                        (result_df["supply_control_status_gt"].str.upper().isin(["STOP PO", "DISCONTINUED", "OOS", "UNAVAILABLE"])),
                        (
                            (result_df["avg_weekly_st_lm_qty"] == 0) &
                            (result_df["buffer_plan_by_lm_qty_adj"] == 0) &
                            (~result_df["Customer SKU Code"].str.upper().isin(all_npd_sku_list)) &
                            (~result_df["supply_control_status_gt"].str.upper().isin(["STOP PO", "DISCONTINUED", "OOS"]))
                        ),
                        (result_df["buffer_plan_by_lm_qty_adj"] == 0),
                        (result_df["PO Qty"] > result_df["buffer_plan_by_lm_qty_adj"]),
                        (result_df["PO Qty"] < result_df["buffer_plan_by_lm_qty_adj"]),
                        (result_df["PO Qty"] == result_df["buffer_plan_by_lm_qty_adj"]),
                    ]

                    choices = [
                        "Reject (SKU Not Found in System)",
                        f"Reject (Exceeds Qty Limit of {MAX_QTY_LIMIT})",
                        "Reject (Negative Allocation)",
                        "Additional Suggestion",
                        "Reject (Stop by Steve - Need approval email)",
                        "Reject (Stop by Steve - No tolerance to open)",
                        "Reject",
                        "Proceed",
                        "Reject",
                        "Reject with suggestion",
                        "Proceed with suggestion",
                        "Proceed",
                    ]

                    result_df["Remark"] = np.select(conditions, choices, default="N/A (Missing Data)")

                    # Rename columns — note: product_name already renamed to "Product Name" via sku_df merge
                    new_column_names = {
                        "distributor_name": "Distributor",
                        "Customer SKU Code": "SKU",
                        "assortment": "Assortment",
                        "supply_control_status_gt": "Supply Control",
                        "PO Qty": "PO Qty",
                        "PO Value": "PO Value",
                        "total_stock": "Total Stock (Qty)",
                        "avg_weekly_st_lm_qty": "Avg Weekly Sales LM (Qty)",
                        "buffer_plan_by_lm_qty_adj": "Suggested PO Qty",
                        "buffer_plan_by_lm_val_adj": "Suggested PO Value",
                        "WOI PO Original": "WOI (Stock + PO Ori)",
                        "WOI Suggest": "WOI After Buffer (Stock + Suggested Qty)",
                        "woi_end_of_month_by_lm": "Stock + Suggested Qty WOI (Projection at EOM)",
                        "remaining_allocation_qty_region": "Remaining Allocation (By Region)",
                    }

                    result_df.rename(columns=new_column_names, inplace=True)

                    if REJECTED_SKUS_1:
                        result_df = apply_sku_rejection_rules(
                            REJECTED_SKUS_1,
                            result_df,
                            REGION_LIST_1,
                            is_in=False
                        )

                    if REJECTED_SKUS_2:
                        result_df = apply_sku_rejection_rules(
                            REJECTED_SKUS_2,
                            result_df,
                            REGION_LIST_2,
                            is_in=False
                        )

                    result_df.sort_values(
                        by=["is_po_sku", "SKU"], ascending=[False, True], inplace=True
                    )

                    result_df["RSA Notes"] = ""

                    excel_cols = [
                        "Distributor",
                        "SKU",
                        "Product Name",
                        "Assortment",
                        "Supply Control",
                        "Avg Weekly Sales LM (Qty)",
                        "Total Stock (Qty)",
                        "Current WOI",
                        "PO Qty",
                        "PO Value",
                        "WOI (Stock + PO Ori)",
                        "Remark",
                        "Suggested PO Qty",
                        "Suggested PO Value",
                        "WOI After Buffer (Stock + Suggested Qty)",
                        "Stock + Suggested Qty WOI (Projection at EOM)",
                        "Remaining Allocation (By Region)",
                        "is_po_sku",
                        "RSA Notes",
                    ]

                    result_df = result_df.reindex(columns=excel_cols)

                    excel_dfs[distributor_name] = result_df.copy()

                    # Format columns for display
                    result_df["PO Value"] = result_df["PO Value"].apply(
                        lambda x: f"{x:,.2f}" if pd.notnull(x) else 0
                    )
                    result_df["Suggested PO Value"] = result_df["Suggested PO Value"].apply(
                        lambda x: f"{x:,.2f}" if pd.notnull(x) else 0
                    )
                    result_df["Remaining Allocation (By Region)"] = result_df["Remaining Allocation (By Region)"].apply(
                        lambda x: f"{round(x):,d}" if pd.notnull(x) else 0
                    )
                    result_df["Avg Weekly Sales LM (Qty)"] = result_df["Avg Weekly Sales LM (Qty)"].apply(
                        lambda x: f"{round(x):,d}" if pd.notnull(x) else 0
                    )
                    result_df["WOI (Stock + PO Ori)"] = result_df["WOI (Stock + PO Ori)"].apply(
                        lambda x: f"{x:.2f}" if pd.notnull(x) else ""
                    )
                    result_df["Stock + Suggested Qty WOI (Projection at EOM)"] = result_df[
                        "Stock + Suggested Qty WOI (Projection at EOM)"
                    ].apply(lambda x: f"{x:.2f}" if pd.notnull(x) else "")
                    result_df["Current WOI"] = result_df["Current WOI"].apply(
                        lambda x: f"{x:.2f}" if pd.notnull(x) else ""
                    )
                    result_df["WOI After Buffer (Stock + Suggested Qty)"] = result_df[
                        "WOI After Buffer (Stock + Suggested Qty)"
                    ].apply(lambda x: f"{x:.2f}" if pd.notnull(x) else "")

                    display_dfs.append(result_df)

                progress.progress(1.0, "Processing complete")

                if output_format == "Single Sheet (All Distributors Stacked)":
                    final_excel_df = pd.concat(excel_dfs.values(), ignore_index=True)
                    xlsx_data = to_excel_single_sheet(final_excel_df, all_npd_sku_list)
                    file_name = "po_simulator_result_single_sheet.xlsx"
                else:
                    xlsx_data = to_excel_with_styling(excel_dfs, all_npd_sku_list)
                    file_name = "po_simulator_result_separate_sheets.xlsx"

                st.download_button(
                    label=f"📥 Download PO Simulator Excel ({output_format})",
                    data=xlsx_data,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

                st.subheader("Simulated PO Data")
                final_display_df = pd.concat(display_dfs, ignore_index=True)

                final_cols = [
                    "Distributor",
                    "SKU",
                    "Product Name",
                    "Assortment",
                    "Supply Control",
                    "Avg Weekly Sales LM (Qty)",
                    "Total Stock (Qty)",
                    "Current WOI",
                    "PO Qty",
                    "PO Value",
                    "WOI (Stock + PO Ori)",
                    "Remark",
                    "Suggested PO Qty",
                    "Suggested PO Value",
                    "WOI After Buffer (Stock + Suggested Qty)",
                    "Stock + Suggested Qty WOI (Projection at EOM)",
                    "Remaining Allocation (By Region)",
                ]

                st.dataframe(final_display_df.reindex(columns=final_cols).reset_index(drop=True))

            except Exception as e:
                st.error(f"An error occurred: {e}")
                st.info(
                    "Please ensure the uploaded file is a valid .xlsx or .csv and contains all the required columns."
                )


if __name__ == "__main__":
    main()
