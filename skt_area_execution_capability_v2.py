import streamlit as st
import pandas as pd
import uuid
import io
import base64
import re
from pathlib import Path
from datetime import datetime
from pendulum import now as pendulum_now
from google.oauth2 import service_account
from google.cloud import bigquery
from openpyxl.worksheet.datavalidation import DataValidation
from assessment_logic import (
    VALUE_THRESHOLDS, normalize_username, value_to_grade, get_sla_grade,
    bad_stock_grade_for_ytd, validate_allocation_row,
)

st.set_page_config(
    page_title="Distributor Operational Assessment",
    layout="wide",
    page_icon="📋"
)

def _load_logo_base64():
    """Transparent-background SKINTIFIC wordmark, used on the login screen
    instead of a lock emoji. Falls back to None (caller shows nothing) if the
    asset is missing rather than crashing the page."""
    logo_path = Path(__file__).parent / "assets" / "skintific_logo.png"
    if not logo_path.exists():
        return None
    return base64.b64encode(logo_path.read_bytes()).decode()

LOGO_B64 = _load_logo_base64()

@st.cache_data
def _load_guide_html(filename):
    """Reads a generated User_Guide HTML file and inlines its screenshots/*.png
    references as base64 data URIs, so the whole guide is self-contained and can
    be rendered inside an iframe via st.components.v1.html() without needing the
    screenshots folder served separately. Returns None if the file is missing."""
    guide_path = Path(__file__).parent / filename
    if not guide_path.exists():
        return None
    html = guide_path.read_text(encoding="utf-8")
    screenshots_dir = Path(__file__).parent / "screenshots"

    def _embed(match):
        img_path = screenshots_dir / match.group(1)
        if img_path.exists():
            b64 = base64.b64encode(img_path.read_bytes()).decode()
            return f'src="data:image/png;base64,{b64}"'
        return match.group(0)

    return re.sub(r'src="screenshots/([^"]+)"', _embed, html)

GUIDE_HTML = {
    "English": _load_guide_html("User_Guide.html"),
    "Bahasa Indonesia": _load_guide_html("User_Guide_ID.html"),
}

# =====================================================
# SKINTIFIC BRAND CSS (BLUE) — unchanged from the verified mock
# =====================================================
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

.block-container { padding-top: 1.5rem !important; max-width: 1050px !important; }

.brand-header {
    background: linear-gradient(135deg, #14506A 0%, #1E6B8A 55%, #4FA9CB 100%);
    padding: 2rem 2.4rem;
    border-radius: 18px;
    margin-bottom: 1.8rem;
    color: white;
    box-shadow: 0 6px 24px rgba(30, 107, 138, 0.35);
}
.brand-header h1 { margin: 0; font-size: 1.7rem; font-weight: 700; letter-spacing: -0.02em; }
.brand-header p  { margin: 0.3rem 0 0; font-size: 0.88rem; opacity: 0.85; }

.sec-label {
    font-size: 0.72rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.1em;
    color: #1E6B8A;
    margin-bottom: 0.6rem;
}

.cat-divider {
    display: flex;
    align-items: center;
    gap: 0.7rem;
    margin: 2rem 0 1rem;
    font-size: 0.82rem;
    font-weight: 700;
    color: #1E6B8A;
    text-transform: uppercase;
    letter-spacing: 0.06em;
}
.cat-divider::after {
    content: '';
    flex: 1;
    height: 1.5px;
    background: linear-gradient(90deg, #BEE3EE 0%, transparent 100%);
}

.q-title-bar {
    display: flex;
    align-items: center;
    gap: 0.6rem;
    margin-bottom: 0.9rem;
}
.q-badge {
    background: #1E6B8A;
    color: white;
    width: 26px; height: 26px;
    border-radius: 50%;
    display: inline-flex;
    align-items: center;
    justify-content: center;
    font-size: 0.75rem;
    font-weight: 700;
    flex-shrink: 0;
}
.q-name   { font-size: 0.98rem; font-weight: 700; color: #1C1C1C; flex: 1; }
.q-maxpts {
    background: #EAF6FB;
    color: #1E6B8A;
    padding: 0.15rem 0.65rem;
    border-radius: 999px;
    font-size: 0.75rem;
    font-weight: 600;
    white-space: nowrap;
}

.sla-info {
    background: #F4FAFC;
    border: 1px solid #BEE3EE;
    border-radius: 10px;
    padding: 0.9rem 1.1rem;
    font-size: 0.84rem;
    color: #0F4C5C;
    margin-bottom: 0.8rem;
}
.sla-info ul { margin: 0.4rem 0 0 1.1rem; padding: 0; }
.sla-info li { margin-bottom: 0.25rem; }

.sla-sub {
    background: #F9FAFB;
    border: 1px solid #EDEDED;
    border-radius: 10px;
    padding: 0.8rem 1rem 0.4rem;
    margin-bottom: 0.8rem;
}
.sla-sub-label { font-size: 0.8rem; font-weight: 700; color: #374151; margin-bottom: 0.3rem; }

.ytd-box {
    background: linear-gradient(135deg, #EAF6FB, #DCF0F7);
    border: 1.5px solid #BEE3EE;
    border-radius: 12px;
    padding: 1rem 1.3rem;
}
.ytd-box-label { font-size: 0.72rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.07em; color: #1D6E86; }
.ytd-box-value { font-size: 1.25rem; font-weight: 700; color: #14506A; margin-top: 0.15rem; }

[data-testid="stVerticalBlockBorderWrapper"] {
    border: 1.5px solid #BEE3EE !important;
    border-radius: 14px !important;
    background: #FAFDFE !important;
}

.q-card-wrap [data-testid="stVerticalBlockBorderWrapper"] {
    border-left: 5px solid #1E6B8A !important;
    box-shadow: 0 2px 12px rgba(30, 107, 138, 0.07) !important;
}

div[data-testid="stButton"] > button[kind="primary"],
div[data-testid="stFormSubmitButton"] > button[kind="primaryFormSubmit"] {
    background: linear-gradient(135deg, #1E6B8A, #4FA9CB) !important;
    color: white !important;
    border: none !important;
    border-radius: 12px !important;
    padding: 0.75rem 2rem !important;
    font-weight: 700 !important;
    font-size: 1.05rem !important;
    box-shadow: 0 4px 14px rgba(30, 107, 138, 0.38) !important;
}

.result-wrap {
    background: linear-gradient(135deg, #EAF6FB, #DCF0F7);
    border: 2px solid #BEE3EE;
    border-radius: 18px;
    padding: 2.2rem;
    text-align: center;
    margin-bottom: 1.5rem;
}
.result-score { font-size: 4.5rem; font-weight: 800; color: #1E6B8A; line-height: 1; }
.result-pts   { font-size: 0.95rem; color: #1D6E86; margin-top: 0.2rem; }
.result-rating{ font-size: 1.1rem; font-weight: 700; margin-top: 0.7rem; }
.result-dist  { font-size: 0.83rem; color: #1D6E86; margin-top: 0.25rem; }

[data-testid="stSidebar"] { background: #FAFDFE !important; }
[data-testid="stSidebar"] hr { border-color: #BEE3EE !important; }

div[role="radiogroup"] { gap: 0.5rem !important; flex-wrap: wrap !important; }
div[role="radiogroup"] label {
    padding: 0.55rem 0.7rem !important;
    margin-right: 0.4rem !important;
    border-radius: 10px !important;
}
div[role="radiogroup"] label p {
    font-size: 1.05rem !important;
    line-height: 1.45 !important;
}

label[data-baseweb="radio"] > div:has(~ input:checked) {
    background-color: #1E6B8A !important;
}

[data-testid="stProgressBarTrack"] > div {
    background-color: #1E6B8A !important;
}

.stSelectbox label, .stTextInput label, .stNumberInput label, .stRadio > label {
    font-weight: 600 !important; color: #374151 !important; font-size: 0.95rem !important;
}

@media (max-width: 640px) {
    .brand-header h1 { font-size: 1.35rem !important; }
    .q-name { font-size: 1.05rem !important; }
    div[role="radiogroup"] label p { font-size: 1.12rem !important; }
}

.login-card {
    max-width: 440px;
    margin: 40px auto 0;
    background: #fff;
    border: 1.5px solid #BEE3EE;
    border-radius: 16px;
    padding: 40px 44px;
    box-shadow: 0 4px 24px rgba(30, 107, 138, 0.12);
}
.login-logo  { text-align: center; margin-bottom: 20px; }
.login-logo img { max-width: 240px; width: 100%; height: auto; }
.role-pill {
    display: inline-block;
    background: #EAF6FB;
    color: #14506A;
    border: 1px solid #BEE3EE;
    padding: 0.25rem 0.9rem;
    border-radius: 999px;
    font-size: 0.82rem;
    font-weight: 700;
}
</style>
""", unsafe_allow_html=True)

# =====================================================
# BIGQUERY CONNECTION (same pattern as skt_area_execution_capability.py)
# =====================================================
gcp_secrets = dict(st.secrets["connections"]["bigquery"])
gcp_secrets["private_key"] = gcp_secrets["private_key"].replace("\\n", "\n")
credentials = service_account.Credentials.from_service_account_info(gcp_secrets)

PROJECT_ID   = st.secrets["bigquery"]["project"]
DATASET      = st.secrets["bigquery"]["dataset"]
TABLE        = "distributor_assessment"
USERS_TABLE  = "assessment_users"

bq_client = bigquery.Client(credentials=credentials, project=PROJECT_ID)

# =====================================================
# MASTER DATA (real BigQuery — matches skt_area_execution_capability.py)
# =====================================================
@st.cache_data(ttl=600)
def load_master_distributor():
    query = f"""
        SELECT DISTINCT region, distributor_company, distributor, distributor_code
        FROM `{PROJECT_ID}.{DATASET}.master_distributor`
        WHERE status = 'Active' AND brand IN (
        "SKT & G2G & TPH & FR & BB & NP",
        "SKT & G2G & FR & BB & NP",
        "SKT & TPH & FR")
        ORDER BY region, distributor_company
    """
    df = bq_client.query(query).to_dataframe()
    for col in ["region", "distributor_company", "distributor", "distributor_code"]:
        df[col] = df[col].astype(str).str.strip()
    return df

@st.cache_data(ttl=600)
def get_distributors_for_supervisor(full_name, region):
    """Distributors owned by this Area Sales Supervisor specifically — matched
    on master_distributor.spv_skt / spv_tph against the logged-in user's
    full_name, not just region. SELECT DISTINCT means a distributor where both
    spv_skt and spv_tph equal this user still returns exactly one row (avoids
    the duplicate-mapping case). Case-insensitive on full_name since it's a
    free-typed display name that may not match spv_skt/spv_tph casing exactly.
    NOTE: spv_skt/spv_tph column names are user-provided and not independently
    verified against the live schema — confirm a real account resolves to the
    expected distributor list after deploy."""
    query = f"""
        SELECT DISTINCT region, distributor, distributor_code
        FROM `{PROJECT_ID}.{DATASET}.master_distributor`
        WHERE status = 'Active'
          AND region = @region
          AND (UPPER(spv_skt) = UPPER(@full_name) OR UPPER(spv_tph) = UPPER(@full_name))
        ORDER BY distributor
    """
    job_config = bigquery.QueryJobConfig(query_parameters=[
        bigquery.ScalarQueryParameter("region", "STRING", region),
        bigquery.ScalarQueryParameter("full_name", "STRING", full_name),
    ])
    df = bq_client.query(query, job_config=job_config).to_dataframe()
    for col in ["region", "distributor", "distributor_code"]:
        df[col] = df[col].astype(str).str.strip()
    return df

@st.cache_data(ttl=600)
def get_ytd_sell_through(distributor_name, year):
    # Brand filter restricts Bad Stock allowance to Skintific + Timephoria only.
    # Column/value names are user-provided, not independently verified against
    # live fact_sell_through_all rows — sanity-check after deploy.
    query = """
        SELECT SUM(value) AS ytd_value
        FROM `pbi_gt_dataset.fact_sell_through_all`
        WHERE distributor_name = @distributor_name
        AND EXTRACT(YEAR FROM calendar_date) = @selected_year
        AND brand IN ('SKINTIFIC', 'TIMEPHORIA')
    """
    job_config = bigquery.QueryJobConfig(query_parameters=[
        bigquery.ScalarQueryParameter("distributor_name", "STRING", distributor_name),
        bigquery.ScalarQueryParameter("selected_year", "INT64", year),
    ])
    df = bq_client.query(query, job_config=job_config).to_dataframe()
    if df.empty or df["ytd_value"].iloc[0] is None:
        return 0
    return df["ytd_value"].iloc[0]

BAD_STOCK_Q = "BAD STOCK HANDLING PERFORMANCE"

def compute_bad_stock_score(distributor_name, year):
    """Fetches YTD sell-through (the only part that needs BigQuery), then
    delegates the pure compliance-% -> grade math to bad_stock_grade_for_ytd
    (assessment_logic.py) so that rule is unit-testable without a BQ connection."""
    ytd_val = get_ytd_sell_through(distributor_name, year)
    utilization = st.session_state.get(f"util_{BAD_STOCK_Q}", 0)
    grade, bs_allow, utilization_out, compliance_pct = bad_stock_grade_for_ytd(ytd_val, utilization)
    return grade, ytd_val, bs_allow, utilization_out, compliance_pct

# =====================================================
# QUESTIONS CONFIG — all 10 metrics, owned across 4 roles
# =====================================================
questions = {
    "OPERATIONAL LEADER (SPV / OPERATIONAL MANAGER)": {
        "A": ("Exist, exclusive for SKINTIFIC", 5),
        "B": ("Exist, mix with other principle", 3),
        "C": ("Do not exist", 0),
    },
    "SALESMAN": {
        "A": ("Exist, exclusive for SKINTIFIC", 7),
        "B": ("Exist, mix with other principle", 5),
        "C": ("Exist (do not meet qty requirement)", 3),
        "D": ("Do not exist", 0),
    },
    "ADMINISTRATIVE & AR SUPPORT": {
        "A": ("Exist, exclusive for SKINTIFIC", 3),
        "B": ("Exist, mix with other principle", 1),
        "C": ("Do not exist", 0),
    },
    "WAREHOUSE FACILITY STANDARD": {
        "A": ("Dedicated warehouse + temperature control", 3),
        "B": ("Warehouse exists but limited facility", 1),
        "C": ("No dedicated warehouse", 0),
    },
    "DELIVERY SLA COMPLIANCE": {
        "A": ("≥ 95% on-time delivery", 8),
        "B": ("Partial SLA compliance", 4),
        "C": ("Below SLA requirement", 0),
    },
    "INVENTORY CONTROL & STOCK OPNAME": {
        "A": ("Stock opname ≥ 2x/year", 6),
        "B": ("Stock opname 1x/year", 4),
        "C": ("No regular stock opname", 0),
    },
    "DATA REPORTING COMPLIANCE": {
        "A": ("On Time", 8),
        "B": ("+ 1 Days", 4),
        "C": ("+ more than 1 Days", 0),
    },
    "ACCOUNT RECEIVABLE (AR) PERFORMANCE": {
        "A": ("On Time", 4),
        "B": ("+ 2 Days", 2),
        "C": ("+ more than 2 Days", 0),
    },
    BAD_STOCK_Q: {
        "A": ("100% compliance", 2),
        "B": ("≥ 80% compliance", 1),
        "C": ("< 80% compliance", 0),
    },
    "BANK GUARANTEE UPDATE COMPLIANCE": {
        "A": ("100% BG updated within agreed timeline", 4),
        "B": ("BG updated with delay / negotiation", 2),
        "C": ("BG not willing to be used", 0),
    },
}

Q_NUMBER = {name: i + 1 for i, name in enumerate(questions)}

Q_CATEGORIES_MASTER = {
    "👥  People & Roles": [
        "OPERATIONAL LEADER (SPV / OPERATIONAL MANAGER)",
        "SALESMAN",
        "ADMINISTRATIVE & AR SUPPORT",
    ],
    "🏭  Infrastructure & Delivery": [
        "WAREHOUSE FACILITY STANDARD",
        "DELIVERY SLA COMPLIANCE",
    ],
    "📊  Operations & Compliance": [
        "INVENTORY CONTROL & STOCK OPNAME",
        "DATA REPORTING COMPLIANCE",
        "ACCOUNT RECEIVABLE (AR) PERFORMANCE",
        BAD_STOCK_Q,
        "BANK GUARANTEE UPDATE COMPLIANCE",
    ],
}

# =====================================================
# ROLES — same URL, 4 distinct user interfaces
# =====================================================
ROLES = {
    "Area Sales Supervisor": {
        "icon": "🧑‍💼",
        "desc": "Evaluate distributor operational capability across people, infrastructure, delivery & inventory.",
        "questions": [
            "OPERATIONAL LEADER (SPV / OPERATIONAL MANAGER)",
            "SALESMAN",
            "ADMINISTRATIVE & AR SUPPORT",
            "WAREHOUSE FACILITY STANDARD",
            "DELIVERY SLA COMPLIANCE",
            "INVENTORY CONTROL & STOCK OPNAME",
            BAD_STOCK_Q,
        ],
        "bulk": False,
    },
    "Distributor Manager": {
        "icon": "🏢",
        "desc": "Update Bank Guarantee Update Compliance for this distributor.",
        "questions": ["BANK GUARANTEE UPDATE COMPLIANCE"],
        "bulk": True,
    },
    "Admin RSA": {
        "icon": "🗂️",
        "desc": "Update Data Reporting Compliance for this distributor.",
        "questions": ["DATA REPORTING COMPLIANCE"],
        "bulk": True,
    },
    "Account Receivable": {
        "icon": "💰",
        "desc": "Update Account Receivable (AR) Performance for this distributor.",
        "questions": ["ACCOUNT RECEIVABLE (AR) PERFORMANCE"],
        "bulk": True,
    },
}

Q_OWNER = {q: role for role, cfg in ROLES.items() for q in cfg["questions"]}
ROLE_MAX = {role: sum(max(v[1] for v in questions[q].values()) for q in cfg["questions"])
            for role, cfg in ROLES.items()}
TOTAL_MAX_SCORE = sum(max(v[1] for v in g.values()) for g in questions.values())

# =====================================================
# LOGIN — BigQuery-backed (skintific-data-warehouse.gt_schema.assessment_users)
# =====================================================
def check_login(username, password):
    """Returns the user dict {username, full_name, role, region} if valid, else None.
    Username comparison is case-insensitive on both sides (LOWER() on the stored
    column too) so a legacy/mixed-case row still matches a lowercase login attempt."""
    query = f"""
        SELECT username, password, full_name, role, region
        FROM `{PROJECT_ID}.{DATASET}.{USERS_TABLE}`
        WHERE LOWER(username) = @username AND is_active = TRUE
        LIMIT 1
    """
    job_config = bigquery.QueryJobConfig(query_parameters=[
        bigquery.ScalarQueryParameter("username", "STRING", normalize_username(username)),
    ])
    df = bq_client.query(query, job_config=job_config).to_dataframe()
    if df.empty:
        return None
    row = df.iloc[0]
    if password != str(row["password"]):
        return None
    return {
        "username": str(row["username"]),
        "full_name": str(row["full_name"]),
        "role": str(row["role"]),
        "region": None if pd.isna(row["region"]) else str(row["region"]),
    }

def verify_and_change_password(username, old_password, new_password):
    """Available to every role. Verifies old_password against the stored value,
    then updates it. Returns (True, None) on success or (False, error_message).
    No audit-log hook here — this app has no audit logging system to feed."""
    query = f"""
        SELECT password FROM `{PROJECT_ID}.{DATASET}.{USERS_TABLE}`
        WHERE LOWER(username) = @username AND is_active = TRUE
        LIMIT 1
    """
    job_config = bigquery.QueryJobConfig(query_parameters=[
        bigquery.ScalarQueryParameter("username", "STRING", normalize_username(username)),
    ])
    df = bq_client.query(query, job_config=job_config).to_dataframe()
    if df.empty:
        return False, "User not found."
    if old_password != str(df.iloc[0]["password"]):
        return False, "Old password is incorrect."

    update_query = f"""
        UPDATE `{PROJECT_ID}.{DATASET}.{USERS_TABLE}`
        SET password = @new_password
        WHERE LOWER(username) = @username
    """
    job_config = bigquery.QueryJobConfig(query_parameters=[
        bigquery.ScalarQueryParameter("new_password", "STRING", new_password),
        bigquery.ScalarQueryParameter("username", "STRING", normalize_username(username)),
    ])
    bq_client.query(update_query, job_config=job_config).result()
    return True, None

def short_name(q_name):
    return (q_name
             .replace("OPERATIONAL LEADER (SPV / OPERATIONAL MANAGER)", "Operational Leader")
             .replace("ADMINISTRATIVE & AR SUPPORT", "Admin & AR Support")
             .replace("WAREHOUSE FACILITY STANDARD", "Warehouse Facility")
             .replace("DELIVERY SLA COMPLIANCE", "Delivery SLA")
             .replace("INVENTORY CONTROL & STOCK OPNAME", "Inventory Control")
             .replace("DATA REPORTING COMPLIANCE", "Data Reporting")
             .replace("ACCOUNT RECEIVABLE (AR) PERFORMANCE", "AR Performance")
             .replace(BAD_STOCK_Q, "Bad Stock")
             .replace("BANK GUARANTEE UPDATE COMPLIANCE", "Bank Guarantee")
             .title())

def filtered_categories(role):
    role_qs = set(ROLES[role]["questions"])
    result = {}
    for cat, qlist in Q_CATEGORIES_MASTER.items():
        keep = [q for q in qlist if q in role_qs]
        if keep:
            result[cat] = keep
    return result

# =====================================================
# COMBINED PROGRESS / DUPLICATE CHECKS — real BigQuery queries
# (short TTL cache: read often via reruns, but must stay reasonably fresh
# since it drives duplicate-submission prevention)
# =====================================================
@st.cache_data(ttl=30)
def get_combined_progress(distributor_name, period):
    query = f"""
        SELECT metric, ANY_VALUE(point) AS points, ANY_VALUE(submitted_role) AS submitted_role
        FROM `{PROJECT_ID}.{DATASET}.{TABLE}`
        WHERE distributor = @distributor AND assessment_period = @period
        GROUP BY metric
    """
    job_config = bigquery.QueryJobConfig(query_parameters=[
        bigquery.ScalarQueryParameter("distributor", "STRING", distributor_name),
        bigquery.ScalarQueryParameter("period", "STRING", period),
    ])
    df = bq_client.query(query, job_config=job_config).to_dataframe()
    done_points = dict(zip(df["metric"], df["points"])) if not df.empty else {}

    total, items = 0, []
    for q in questions:
        if q in done_points:
            pts = int(done_points[q])
            total += pts
            items.append({"num": Q_NUMBER[q], "name": q, "status": "done", "points": pts})
        else:
            items.append({"num": Q_NUMBER[q], "name": q, "status": "pending", "points": None})
    return total, items

@st.cache_data(ttl=30)
def get_role_bulk_progress(role, period):
    metric = ROLES[role]["questions"][0]
    master_df = load_master_distributor()
    all_distributors = sorted(master_df["distributor"].dropna().unique())

    query = f"""
        SELECT DISTINCT distributor
        FROM `{PROJECT_ID}.{DATASET}.{TABLE}`
        WHERE submitted_role = @role AND assessment_period = @period AND metric = @metric
    """
    job_config = bigquery.QueryJobConfig(query_parameters=[
        bigquery.ScalarQueryParameter("role", "STRING", role),
        bigquery.ScalarQueryParameter("period", "STRING", period),
        bigquery.ScalarQueryParameter("metric", "STRING", metric),
    ])
    df = bq_client.query(query, job_config=job_config).to_dataframe()
    done_set = set(df["distributor"]) if not df.empty else set()

    done    = [d for d in all_distributors if d in done_set]
    pending = [d for d in all_distributors if d not in done_set]
    return done, pending

def check_role_already_submitted(role, distributor_name, period):
    """Live (uncached) check used right before allowing a new submission."""
    query = f"""
        SELECT 1
        FROM `{PROJECT_ID}.{DATASET}.{TABLE}`
        WHERE submitted_role = @role AND distributor = @distributor AND assessment_period = @period
        LIMIT 1
    """
    job_config = bigquery.QueryJobConfig(query_parameters=[
        bigquery.ScalarQueryParameter("role", "STRING", role),
        bigquery.ScalarQueryParameter("distributor", "STRING", distributor_name),
        bigquery.ScalarQueryParameter("period", "STRING", period),
    ])
    df = bq_client.query(query, job_config=job_config).to_dataframe()
    return not df.empty

@st.cache_data(ttl=300)
def get_ass_missing_distributors(period):
    """Active distributors with no Area Sales Supervisor submission for this
    period. ASS submits all 7 of its metrics in one atomic batch, so checking
    for any row with this role+distributor+period is enough — no metric filter
    needed (unlike get_role_bulk_progress, which checks one bulk role's single
    metric)."""
    master_df = load_master_distributor()
    all_distributors = master_df[["region", "distributor"]].drop_duplicates()

    query = f"""
        SELECT DISTINCT distributor
        FROM `{PROJECT_ID}.{DATASET}.{TABLE}`
        WHERE submitted_role = 'Area Sales Supervisor' AND assessment_period = @period
    """
    job_config = bigquery.QueryJobConfig(query_parameters=[
        bigquery.ScalarQueryParameter("period", "STRING", period),
    ])
    df = bq_client.query(query, job_config=job_config).to_dataframe()
    done_set = set(df["distributor"]) if not df.empty else set()

    missing = all_distributors[~all_distributors["distributor"].isin(done_set)]
    return missing.sort_values(["region", "distributor"]).reset_index(drop=True)

@st.cache_data(ttl=300)
def get_ass_users_not_submitted(period):
    """Active Area Sales Supervisor accounts with zero submissions for this
    period — matched on full_name since that's what's stored in
    representative_name, case-insensitively since it's free-typed at login time."""
    query = f"""
        SELECT u.username, u.full_name, u.region
        FROM `{PROJECT_ID}.{DATASET}.{USERS_TABLE}` u
        WHERE u.role = 'Area Sales Supervisor' AND u.is_active = TRUE
          AND NOT EXISTS (
            SELECT 1 FROM `{PROJECT_ID}.{DATASET}.{TABLE}` da
            WHERE da.submitted_role = 'Area Sales Supervisor'
              AND da.assessment_period = @period
              AND UPPER(da.representative_name) = UPPER(u.full_name)
          )
        ORDER BY u.region, u.full_name
    """
    job_config = bigquery.QueryJobConfig(query_parameters=[
        bigquery.ScalarQueryParameter("period", "STRING", period),
    ])
    return bq_client.query(query, job_config=job_config).to_dataframe()

@st.cache_data(ttl=300)
def get_total_ass_users():
    """Count of active Area Sales Supervisor accounts, for the X/Y submitted
    ratio in the Reporting panel."""
    query = f"""
        SELECT COUNT(*) AS total
        FROM `{PROJECT_ID}.{DATASET}.{USERS_TABLE}`
        WHERE role = 'Area Sales Supervisor' AND is_active = TRUE
    """
    df = bq_client.query(query).to_dataframe()
    return int(df["total"].iloc[0]) if not df.empty else 0

def check_bulk_already_submitted(role, distributor_names, period, metric):
    """Live (uncached) batch check — one query covers the whole upload batch."""
    if not distributor_names:
        return set()
    query = f"""
        SELECT DISTINCT distributor
        FROM `{PROJECT_ID}.{DATASET}.{TABLE}`
        WHERE submitted_role = @role AND assessment_period = @period AND metric = @metric
          AND distributor IN UNNEST(@distributor_names)
    """
    job_config = bigquery.QueryJobConfig(query_parameters=[
        bigquery.ScalarQueryParameter("role", "STRING", role),
        bigquery.ScalarQueryParameter("period", "STRING", period),
        bigquery.ScalarQueryParameter("metric", "STRING", metric),
        bigquery.ArrayQueryParameter("distributor_names", "STRING", distributor_names),
    ])
    df = bq_client.query(query, job_config=job_config).to_dataframe()
    return set(df["distributor"]) if not df.empty else set()

def insert_assessment_rows(rows_to_insert):
    """Single batched insert_rows_json call — same pattern as the original
    single-role app. Works for 1 row (Area Sales Supervisor's bad stock row)
    or many rows (a bulk Excel upload) without any chunking needed."""
    table_id = f"{PROJECT_ID}.{DATASET}.{TABLE}"
    return bq_client.insert_rows_json(table_id, rows_to_insert)

# =====================================================
# NPD / SKU FOCUS ALLOCATION UPLOADS (Distributor Manager only)
# Separate from the 10-metric scoring system — sets per-SKU allocation
# targets for a date range, not a grade. Effective Date / End Date are never
# trusted from the uploaded Excel; they're always re-stamped from whatever
# the user picked in the page's date filters at Confirm & Submit time.
# No duplicate-blocking: every upload appends rows to distributor_sku_allocation
# (submission_id/submitted_at let you trace batches and pick the latest
# revision per distributor+SKU yourself if a correction is re-uploaded).
# =====================================================
ALLOCATION_TABLE = "distributor_sku_allocation"

ALLOCATION_PROGRAMS = {
    "NPD": {"icon": "📦", "label": "New Product Development (NPD)"},
    "SKU Focus": {"icon": "🎯", "label": "SKU Focus"},
}

def generate_allocation_template(program_type, effective_date, end_date):
    master_df = load_master_distributor()
    template_df = master_df[["region", "distributor", "distributor_code"]].rename(
        columns={"distributor": "distributor_name"}
    )
    template_df["brand"] = ""
    template_df["sku_code"] = ""
    template_df["allocation_target"] = ""
    template_df["effective_date"] = effective_date
    template_df["end_date"] = end_date
    # Requested column order: Region, Distributor Name, Distributor Code, Brand,
    # SKU Code, Allocation Target, Effective Date, End Date
    template_df = template_df[[
        "region", "distributor_name", "distributor_code",
        "brand", "sku_code", "allocation_target", "effective_date", "end_date",
    ]]

    notes_df = pd.DataFrame([
        {"field": "distributor_code", "example": "Optional - leave blank if not tied to a specific distributor"},
        {"field": "brand", "example": "SKINTIFIC (required)"},
        {"field": "sku_code", "example": "SKINTIFIC-4331 (required)"},
        {"field": "allocation_target", "example": "Optional - whole number, e.g. 500. Leave blank if not yet known"},
        {"field": "effective_date / end_date", "example": "Locked to the dates chosen on the page - edits here are ignored on upload"},
    ])

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        template_df.to_excel(writer, sheet_name="Template", index=False)
        notes_df.to_excel(writer, sheet_name="Notes", index=False)
    buffer.seek(0)
    return buffer

def parse_allocation_upload(df_upload, master_df):
    """Returns (preview_rows, row_errors). brand/sku_code are required to
    consider a row "filled in". distributor_code and allocation_target are
    optional (nullable in distributor_sku_allocation) — left blank, they're
    stored as NULL rather than rejecting the row. An explicitly-typed but
    unrecognized distributor_code is still flagged as an error (likely a
    typo), since blank is the only accepted way to skip it. Effective/End
    Date are deliberately ignored here — the caller re-stamps them from the
    live page filter."""
    code_to_name = dict(zip(master_df["distributor_code"], master_df["distributor"]))
    code_to_region = dict(zip(master_df["distributor_code"], master_df["region"]))

    preview_rows, row_errors = [], []
    for i, r in df_upload.iterrows():
        raw_code = str(r.get("distributor_code", "")).strip()
        code = "" if raw_code.lower() in ("", "nan") else raw_code
        if code and code not in code_to_name:
            row_errors.append(f"Row {i + 2}: unknown distributor_code '{code}' — skipped.")
            continue

        raw_brand = r.get("brand", "")
        raw_sku = r.get("sku_code", "")
        brand = "" if pd.isna(raw_brand) else str(raw_brand).strip()
        sku_code = "" if pd.isna(raw_sku) else str(raw_sku).strip()

        if brand == "" and sku_code == "":
            continue  # entirely blank row, not yet filled in — silently skip
        if brand == "" or sku_code == "":
            row_errors.append(f"Row {i + 2}: both brand and sku_code are required — skipped.")
            continue

        raw_target = r.get("allocation_target", "")
        if pd.isna(raw_target) or str(raw_target).strip() == "":
            allocation_target = None
        else:
            try:
                allocation_target = int(float(raw_target))
            except (TypeError, ValueError):
                row_errors.append(f"Row {i + 2}: invalid allocation_target '{raw_target}' for {brand}/{sku_code} — skipped.")
                continue
            if allocation_target < 0:
                row_errors.append(f"Row {i + 2}: allocation_target for {brand}/{sku_code} cannot be negative — skipped.")
                continue

        if code:
            dist_name = code_to_name[code]
            region_val = code_to_region[code]
        else:
            # No code given — fall back to whatever was typed in the
            # distributor_name / region columns instead of forcing NULL.
            raw_name = str(r.get("distributor_name", "")).strip()
            raw_region = str(r.get("region", "")).strip()
            dist_name = raw_name if raw_name and raw_name.lower() != "nan" else None
            region_val = raw_region if raw_region and raw_region.lower() != "nan" else None

        preview_rows.append({
            "distributor_code": code or None,
            "distributor_name": dist_name,
            "region": region_val,
            "brand": brand,
            "sku_code": sku_code,
            "allocation_target": allocation_target,
        })

    return preview_rows, row_errors

def insert_allocation_rows(rows_to_insert):
    table_id = f"{PROJECT_ID}.{DATASET}.{ALLOCATION_TABLE}"
    return bq_client.insert_rows_json(table_id, rows_to_insert)

# =====================================================
# USER MANAGEMENT — lets Distributor Manager create new logins (any role)
# without touching BigQuery directly. assessment_users has no enforced
# uniqueness constraint, so username_exists() is checked explicitly here.
# =====================================================
def username_exists(username):
    """Case-insensitive existence check — LOWER() on the stored column so a
    legacy/mixed-case row still collides with a new lowercase username."""
    query = f"""
        SELECT 1 FROM `{PROJECT_ID}.{DATASET}.{USERS_TABLE}`
        WHERE LOWER(username) = @username
        LIMIT 1
    """
    job_config = bigquery.QueryJobConfig(query_parameters=[
        bigquery.ScalarQueryParameter("username", "STRING", normalize_username(username)),
    ])
    df = bq_client.query(query, job_config=job_config).to_dataframe()
    return not df.empty

def create_user(username, password, full_name, new_role, region, email):
    table_id = f"{PROJECT_ID}.{DATASET}.{USERS_TABLE}"
    row = {
        "username": normalize_username(username),
        "password": password,
        "full_name": full_name.strip(),
        "role": new_role,
        "region": region,
        "email": email.strip() if email else None,
        "is_active": True,
        "created_at": pendulum_now("Asia/Jakarta").to_datetime_string(),
    }
    return bq_client.insert_rows_json(table_id, [row])

# =====================================================
# BULK EXCEL TEMPLATE — keyed on distributor_code (short, typo-resistant)
# =====================================================
def generate_bulk_template(metric_name):
    master_df = load_master_distributor()
    base_df = master_df[["region", "distributor_code", "distributor"]].rename(columns={"distributor": "distributor_name"})

    grades = questions[metric_name]
    info_df = pd.DataFrame([
        {"grade": g, "description": desc, "points": pts}
        for g, (desc, pts) in grades.items()
    ])

    is_value_based = metric_name in VALUE_THRESHOLDS
    template_df = base_df.copy()
    if is_value_based:
        template_df["value"] = ""
    template_df["grade"] = ""
    template_df["point"] = ""
    template_df["description"] = ""

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        template_df.to_excel(writer, sheet_name="Template", index=False)
        info_df.to_excel(writer, sheet_name="Info", index=False)

        ws = writer.sheets["Template"]
        last_row = len(template_df) + 1
        # Column order: region(A), distributor_code(B), distributor_name(C), [value(D)], grade, point, description
        grade_col = "E" if is_value_based else "D"
        point_col = "F" if is_value_based else "E"
        desc_col  = "G" if is_value_based else "F"
        value_col = "D" if is_value_based else None

        if is_value_based:
            threshold = VALUE_THRESHOLDS[metric_name]
            for row in range(2, last_row + 1):
                ws[f"{grade_col}{row}"] = f'=IF({value_col}{row}="","",IF({value_col}{row}<=0,"A",IF({value_col}{row}<={threshold},"B","C")))'
                ws[f"{point_col}{row}"] = f'=IF({grade_col}{row}="","",IFERROR(VLOOKUP({grade_col}{row},Info!$A:$C,3,FALSE),""))'
                ws[f"{desc_col}{row}"]  = f'=IF({grade_col}{row}="","",IFERROR(VLOOKUP({grade_col}{row},Info!$A:$C,2,FALSE),""))'
            dv = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="-999", allow_blank=True)
            dv.add(f"{value_col}2:{value_col}{last_row}")
            ws.add_data_validation(dv)
        else:
            for row in range(2, last_row + 1):
                ws[f"{point_col}{row}"] = f'=IFERROR(VLOOKUP({grade_col}{row},Info!$A:$C,3,FALSE),"")'
                ws[f"{desc_col}{row}"]  = f'=IFERROR(VLOOKUP({grade_col}{row},Info!$A:$C,2,FALSE),"")'
            dv = DataValidation(type="list", formula1=f'"{",".join(grades.keys())}"', allow_blank=True)
            dv.add(f"{grade_col}2:{grade_col}{last_row}")
            ws.add_data_validation(dv)
    buffer.seek(0)
    return buffer

# =====================================================
# BRAND HEADER
# =====================================================
st.markdown(f"""
<div class="brand-header">
    <div style="display:flex; align-items:center; gap:1.2rem;">
        <div style="font-size:3rem; line-height:1;">📋</div>
        <div>
            <h1>Distributor Operational Assessment</h1>
            <p>Multi-role assessment across {len(questions)} operational metrics &nbsp;·&nbsp; Total Score: {TOTAL_MAX_SCORE} pts</p>
        </div>
    </div>
</div>
""", unsafe_allow_html=True)

# =====================================================
# LOGIN GATE
# =====================================================
if "user" not in st.session_state:
    st.session_state.user = None

if st.session_state.user is None:
    logo_html = (
        f'<img src="data:image/png;base64,{LOGO_B64}" alt="SKINTIFIC" />'
        if LOGO_B64 else "🔐"
    )
    st.markdown(f"""
    <div class="login-card">
        <div class="login-logo">{logo_html}</div>
    </div>
    """, unsafe_allow_html=True)

    _, col_form, _ = st.columns([1, 1.4, 1])
    with col_form:
        with st.form("login_form", border=False):
            username = st.text_input("Username", placeholder="e.g. budi")
            password = st.text_input("Password", type="password", placeholder="••••••••")
            login_clicked = st.form_submit_button("🔑  Log In", use_container_width=True, type="primary")

        if login_clicked:
            user = check_login(username, password)
            if user:
                st.session_state.user = user
                st.rerun()
            else:
                st.error("❌ Invalid username or password.")
    st.stop()

role = st.session_state.user["role"]
representative_name = st.session_state.user["full_name"]

with st.sidebar:
    region_line = f"<br><span class='role-pill'>📍 {st.session_state.user['region']}</span>" if st.session_state.user.get("region") else ""
    st.markdown(
        f"**{ROLES[role]['icon']} {representative_name}**<br>"
        f"<span class='role-pill'>{role}</span>{region_line}",
        unsafe_allow_html=True,
    )
    if st.button("🚪 Logout", use_container_width=True):
        st.session_state.user = None
        st.rerun()

    with st.expander("🔑 Change Password"):
        old_pw = st.text_input("Old Password", type="password", key="cp_old")
        new_pw = st.text_input("New Password", type="password", key="cp_new")
        confirm_pw = st.text_input("Confirm New Password", type="password", key="cp_confirm")
        if st.button("Update Password", use_container_width=True, key="cp_submit"):
            if not old_pw or not new_pw or not confirm_pw:
                st.error("All fields are required.")
            elif new_pw != confirm_pw:
                st.error("New Password and Confirm New Password don't match.")
            else:
                ok, err = verify_and_change_password(st.session_state.user["username"], old_pw, new_pw)
                if ok:
                    st.success("✅ Password changed. Please log in again.")
                    st.session_state.user = None
                    st.rerun()
                else:
                    st.error(err)

    with st.expander("📖 User Guide"):
        st.caption("Download the user guide (opens in your browser):")
        if GUIDE_HTML.get("English"):
            st.download_button(
                "⬇️ English", data=GUIDE_HTML["English"],
                file_name="User_Guide.html", mime="text/html",
                use_container_width=True, key="dl_guide_en",
            )
        if GUIDE_HTML.get("Bahasa Indonesia"):
            st.download_button(
                "⬇️ Bahasa Indonesia", data=GUIDE_HTML["Bahasa Indonesia"],
                file_name="User_Guide_ID.html", mime="text/html",
                use_container_width=True, key="dl_guide_id",
            )
        if not GUIDE_HTML.get("English") and not GUIDE_HTML.get("Bahasa Indonesia"):
            st.caption("Guide not available in this deployment.")

with st.container(border=True):
    st.markdown('<div class="sec-label">👤 Logged In As</div>', unsafe_allow_html=True)
    region_suffix = f"  ·  📍 {st.session_state.user['region']} region only" if st.session_state.user.get("region") else ""
    st.caption(f"{ROLES[role]['icon']}  **{representative_name}**  ·  {ROLES[role]['desc']}{region_suffix}")

# =====================================================
# METADATA — PERIOD
# Area Sales Supervisor: current month + 2 previous (3 total).
# Other roles: 6 months before + current + 6 months after (13 total).
# =====================================================
months = ["January","February","March","April","May","June",
          "July","August","September","October","November","December"]

def build_period_options(months_before, months_after):
    now = datetime.now()
    base_index = now.year * 12 + (now.month - 1)
    opts = []
    for offset in range(-months_before, months_after + 1):
        idx = base_index + offset
        year = idx // 12
        month_num = idx % 12 + 1
        opts.append((months[month_num - 1], year))
    return opts

period_options = build_period_options(2, 0) if role == "Area Sales Supervisor" else build_period_options(6, 6)
period_labels  = [f"{m} {y}" for m, y in period_options]
default_index  = next(i for i, (m, y) in enumerate(period_options)
                       if m == months[datetime.now().month - 1] and y == datetime.now().year)

with st.container(border=True):
    st.markdown('<div class="sec-label">📅 Assessment Period</div>', unsafe_allow_html=True)
    assessment_period = st.selectbox("Assessment Period", period_labels, index=default_index, label_visibility="visible")

selected_month, _sep_year = assessment_period.rsplit(" ", 1)
selected_year = int(_sep_year)

# =====================================================
# SESSION STATE INIT (UI-flow only — submission truth lives in BigQuery)
# =====================================================
for key, default in [
    ("show_confirm", False), ("pending_submission", None), ("submission_result", None), ("just_submitted", False),
    ("show_bulk_confirm", False), ("pending_bulk_submission", None), ("bulk_submission_result", None), ("just_submitted_bulk", False),
    ("show_alloc_confirm", False), ("pending_allocation", None), ("just_submitted_alloc", None), ("last_alloc_count", 0),
]:
    if key not in st.session_state:
        st.session_state[key] = default

is_bulk_role = ROLES[role]["bulk"]

# =====================================================================================
# ===========================  AREA SALES SUPERVISOR (single distributor)  ============
# =====================================================================================
if not is_bulk_role:

    user_region = st.session_state.user.get("region")

    if not user_region:
        st.error("⚠️ Your account has no region assigned. Contact an admin to fix your user record.")
        st.stop()

    supervisor_df = get_distributors_for_supervisor(representative_name, user_region)

    with st.container(border=True):
        st.markdown('<div class="sec-label">📍 Location</div>', unsafe_allow_html=True)
        c4, c5 = st.columns(2)
        with c4:
            st.selectbox("Region", [user_region], disabled=True,
                         help="Your account is scoped to this region only.")
            region = user_region
        with c5:
            dist_opts = sorted(supervisor_df["distributor"].dropna().unique())
            distributor = st.selectbox("Distributor", ["— Select Distributor —"] + dist_opts)
            if not dist_opts:
                st.caption("⚠️ No distributors are mapped to you (spv_skt/spv_tph) in this region yet. Contact an admin if this looks wrong.")

    metadata_ok = distributor != "— Select Distributor —"
    if not metadata_ok:
        st.info("✏️  Select a distributor above to unlock the assessment form.")
        st.stop()

    st.success(f"✅  **{representative_name.strip()}**  ·  {ROLES[role]['icon']} **{role}**  ·  {region}  ·  {distributor}  ·  {assessment_period}")

    role_key = (role, distributor, assessment_period)

    pending = st.session_state.pending_submission
    if pending and (pending["role"], pending["distributor"], pending["assessment_period"]) != role_key:
        st.session_state.pending_submission = None
        st.session_state.show_confirm = False

    already_submitted_by_role = check_role_already_submitted(role, distributor, assessment_period)

    if already_submitted_by_role:
        st.warning(
            f"⚠️ **{role}** has already submitted for **{distributor}** in **{assessment_period}**. "
            f"Only one submission is allowed per role, per distributor, per assessment period."
        )

    # ── SIDEBAR — YOUR LIVE SCORE + COMBINED ASSESSMENT PROGRESS ──
    with st.sidebar:
        st.markdown("---")
        st.markdown(f"### {ROLES[role]['icon']} {role}")

        role_qs  = ROLES[role]["questions"]
        role_max = ROLE_MAX[role]
        live_score = 0
        for q_name in role_qs:
            grades = questions[q_name]
            if q_name == "DELIVERY SLA COMPLIANCE":
                inner = st.session_state.get("inner_city_sla", "100%")
                outer = st.session_state.get("outer_city_sla", "100%")
                _, pts = get_sla_grade(inner, outer)
            elif q_name == BAD_STOCK_Q:
                g, *_ = compute_bad_stock_score(distributor, selected_year)
                pts = grades[g][1]
            else:
                g = st.session_state.get(f"grade_{q_name}", list(grades.keys())[0])
                pts = grades.get(g, ("", 0))[1]
            live_score += pts

        pct_live = live_score / role_max if role_max else 0
        st.metric("Your Live Score", f"{live_score} / {role_max}", f"{int(pct_live*100)}%")
        st.progress(pct_live)

        for q_name in role_qs:
            max_pts = max(v[1] for v in questions[q_name].values())
            grades = questions[q_name]
            if q_name == "DELIVERY SLA COMPLIANCE":
                inner = st.session_state.get("inner_city_sla", "100%")
                outer = st.session_state.get("outer_city_sla", "100%")
                _, pts = get_sla_grade(inner, outer)
            elif q_name == BAD_STOCK_Q:
                g, *_ = compute_bad_stock_score(distributor, selected_year)
                pts = grades[g][1]
            else:
                g = st.session_state.get(f"grade_{q_name}", list(grades.keys())[0])
                pts = grades.get(g, ("", 0))[1]
            icon = "🟢" if pts == max_pts else ("🟡" if pts > 0 else "🔴")
            st.caption(f"{icon} **{Q_NUMBER[q_name]}. {short_name(q_name)}** — {pts}/{max_pts} pts")

        st.markdown("---")
        st.markdown("### 📋 Assessment Progress")
        st.caption(f"{distributor} · {assessment_period}")
        combined_total, items = get_combined_progress(distributor, assessment_period)
        for it in items:
            owner = Q_OWNER[it["name"]]
            if it["status"] == "done":
                st.caption(f"✅ **{it['num']}. {short_name(it['name'])}** — {it['points']} pts _({owner})_")
            else:
                st.caption(f"⏳ **{it['num']}. {short_name(it['name'])}** — pending _({owner})_")

        st.markdown("---")
        st.metric("Combined Score So Far", f"{combined_total} / {TOTAL_MAX_SCORE}")
        st.caption("Score updates as each role submits their part (cached up to 30s).")

    # ── ASSESSMENT FORM (role-scoped) ──
    answers = {}

    for category, q_list in filtered_categories(role).items():
        st.markdown(f'<div class="cat-divider">{category}</div>', unsafe_allow_html=True)

        for q_name in q_list:
            grades  = questions[q_name]
            q_num   = Q_NUMBER[q_name]
            max_pts = max(v[1] for v in grades.values())

            if q_name == BAD_STOCK_Q:
                ytd_check = get_ytd_sell_through(distributor, selected_year)
                if not ytd_check:
                    # No sell-through data to assess against — hide this
                    # metric's card entirely and auto-award max score, no
                    # error/warning shown (per spec).
                    answers[q_name] = {
                        "grade": "A", "ytd_value": ytd_check, "bs_allowance": 0,
                        "bs_utilization": 0, "compliance_pct": 100.0,
                    }
                    continue

            with st.container(border=True):
                st.markdown(f"""
                <div class="q-title-bar">
                    <span class="q-badge">{q_num}</span>
                    <span class="q-name">{q_name}</span>
                    <span class="q-maxpts">max {max_pts} pts</span>
                </div>
                """, unsafe_allow_html=True)

                if q_name == "DELIVERY SLA COMPLIANCE":
                    st.markdown("""
                    <div class="sla-info">
                        📌 <strong>Scoring Logic</strong>
                        <ul>
                            <li>Either Inner <b>or</b> Outer &lt; 80% → 0 pts</li>
                            <li>Either is 99%–80% (none below 80%) → 4 pts</li>
                            <li>Both Inner <b>and</b> Outer at 100% → 8 pts</li>
                        </ul>
                    </div>
                    """, unsafe_allow_html=True)

                    col_a, col_b = st.columns(2)
                    with col_a:
                        st.markdown('<div class="sla-sub"><div class="sla-sub-label">🏙️ Inner City (2 × 24 Hours)</div></div>', unsafe_allow_html=True)
                        inner_city = st.radio("Inner City SLA", ["100%", "99%-80%", "<80%"],
                                              key="inner_city_sla", horizontal=True,
                                              label_visibility="collapsed")
                    with col_b:
                        st.markdown('<div class="sla-sub"><div class="sla-sub-label">🗺️ Outer City (3 × 24 Hours)</div></div>', unsafe_allow_html=True)
                        outer_city = st.radio("Outer City SLA", ["100%", "99%-80%", "<80%"],
                                              key="outer_city_sla", horizontal=True,
                                              label_visibility="collapsed")

                    answers[q_name] = {"inner": inner_city, "outer": outer_city}
                    continue

                if q_name == BAD_STOCK_Q:
                    ytd_val_preview = get_ytd_sell_through(distributor, selected_year)
                    bs_allow_preview = ytd_val_preview * 0.005

                    c_ytd, c_allow = st.columns(2)
                    with c_ytd:
                        st.markdown(f"""
                        <div class="ytd-box">
                            <div class="ytd-box-label">📈 YTD Sell Through</div>
                            <div class="ytd-box-value">Rp {ytd_val_preview:,.0f}</div>
                        </div>""", unsafe_allow_html=True)
                    with c_allow:
                        st.markdown(f"""
                        <div class="ytd-box">
                            <div class="ytd-box-label">💰 Bad Stock Allowance (0.5%)</div>
                            <div class="ytd-box-value">Rp {bs_allow_preview:,.0f}</div>
                        </div>""", unsafe_allow_html=True)

                    st.markdown("""
                    <div class="sla-info">
                        📌 <strong>Scoring Logic</strong>
                        <ul>
                            <li>Compliance % = Utilization / Allowance × 100</li>
                            <li>Compliance ≥ 100% (utilization meets/exceeds allowance) → 2 pts</li>
                            <li>Compliance ≥ 80% → 1 pt</li>
                            <li>Compliance &lt; 80% → 0 pts</li>
                        </ul>
                    </div>
                    """, unsafe_allow_html=True)

                    st.number_input(
                        "Bad Stock Utilization (Rp)",
                        min_value=0, max_value=int(ytd_val_preview) if ytd_val_preview else 0,
                        value=0, step=10_000,
                        key=f"util_{BAD_STOCK_Q}", format="%d",
                        help="Actual Rupiah value of bad stock claimed/utilized by the distributor this period.",
                    )

                    grade_opt, ytd_val, bs_allow, utilization, compliance_pct = compute_bad_stock_score(distributor, selected_year)
                    badge_color = "#059669" if grade_opt == "A" else ("#D97706" if grade_opt == "B" else "#DC2626")
                    st.markdown(
                        f"<div style='margin-top:0.6rem; font-weight:700; color:{badge_color}; font-size:1rem;'>"
                        f"Compliance: {compliance_pct:.1f}%  →  Grade {grade_opt}  ({grades[grade_opt][1]} pts)"
                        f"</div>", unsafe_allow_html=True
                    )

                    answers[q_name] = {
                        "grade": grade_opt, "ytd_value": ytd_val, "bs_allowance": bs_allow,
                        "bs_utilization": utilization, "compliance_pct": compliance_pct,
                    }
                    continue

                grade_opt = st.radio(
                    "Select Grade",
                    list(grades.keys()),
                    format_func=lambda x: f"{x}  —  {grades[x][0]}  ({grades[x][1]} pts)",
                    key=f"grade_{q_name}",
                    horizontal=True,
                    label_visibility="collapsed",
                )

                if q_name == "SALESMAN":
                    st.caption("💡 If fewer than 5 salesmen, enter  **-**  for empty slots.")
                    cols5 = st.columns(5)
                    names5 = []
                    for i, col in enumerate(cols5, 1):
                        with col:
                            n = st.text_input(f"Name {i}", key=f"name_{q_name}_{i}",
                                              placeholder=f"Salesman {i}")
                            names5.append(n)
                    answers[q_name] = {"grade": grade_opt, "person_name": names5}

                elif q_name in ["OPERATIONAL LEADER (SPV / OPERATIONAL MANAGER)",
                                 "ADMINISTRATIVE & AR SUPPORT"]:
                    person = st.text_input(
                        "Name (required if role exists)",
                        key=f"name_{q_name}",
                        placeholder="Full name",
                    )
                    answers[q_name] = {"grade": grade_opt, "person_name": person}

                else:
                    answers[q_name] = {"grade": grade_opt}

    st.write("")
    review_clicked = st.button(
        f"🔍  Review & Submit ({role})", use_container_width=True, type="primary",
        disabled=already_submitted_by_role,
    )

    # ── VALIDATION → BUILD PENDING SUBMISSION → OPEN CONFIRM DIALOG ──
    if review_clicked:
        errors = []

        name_qs = [q for q in [
            "OPERATIONAL LEADER (SPV / OPERATIONAL MANAGER)",
            "SALESMAN",
            "ADMINISTRATIVE & AR SUPPORT",
        ] if q in answers]

        for q in name_qs:
            grade = answers[q]["grade"]
            dne   = [k for k, v in questions[q].items() if "Do not exist" in v[0]]

            if q == "SALESMAN":
                names = [st.session_state.get(f"name_{q}_{i}", "").strip() for i in range(1, 6)]
                if grade in dne and any(n != "" for n in names):
                    errors.append("SALESMAN: Selected 'Do not exist' but names were entered.")
                elif grade not in dne and sum(1 for n in names if n != "") != 5:
                    errors.append("SALESMAN: Exactly 5 names required (use  -  for empty slots).")
            else:
                person = st.session_state.get(f"name_{q}", "").strip()
                if grade in dne and person != "":
                    errors.append(f"{q}: Selected 'Do not exist' but a name was entered.")
                if grade not in dne and person == "":
                    errors.append(f"{q}: Name is required because the role exists.")

        if errors:
            st.error("**Please fix the following before submitting:**")
            for e in errors:
                st.markdown(f"- {e}")
            st.stop()

        sid = str(uuid.uuid4())
        submitted_at = pendulum_now("Asia/Jakarta").to_datetime_string()
        rows = []

        role_total = 0
        for q, val in answers.items():
            if q == "DELIVERY SLA COMPLIANCE":
                i, o = val["inner"], val["outer"]
                role_total += 0 if (i == "<80%" or o == "<80%") else (4 if (i == "99%-80%" or o == "99%-80%") else 8)
            else:
                role_total += questions[q][val["grade"]][1]

        for q, val in answers.items():
            dne  = [k for k, v in questions[q].items() if "Do not exist" in v[0]] if q != "DELIVERY SLA COMPLIANCE" else []
            base = dict(
                submission_id=sid, submitted_at=submitted_at,
                representative_name=representative_name.strip().upper(),
                submitted_role=role, submitted_by_username=st.session_state.user["username"],
                region=region, distributor=distributor, assessment_period=assessment_period,
                inner_city_sla=None, outer_city_sla=None,
                bs_ytd_value=None, bs_allowance=None, bs_utilization=None, bs_compliance_pct=None,
                days_late=None,
            )

            if q == "DELIVERY SLA COMPLIANCE":
                g, p = get_sla_grade(val["inner"], val["outer"])
                rows.append({**base, "metric": q, "grade": g, "person_name": None, "point": p,
                             "inner_city_sla": val["inner"], "outer_city_sla": val["outer"]})

            elif q == "SALESMAN":
                g = val["grade"]
                ns = [st.session_state.get(f"name_{q}_{i}", "").strip() for i in range(1, 6)]
                for n in ns:
                    rows.append({**base, "metric": q, "grade": g,
                                  "person_name": None if g in dne else n.upper(),
                                  "point": questions[q][g][1]})
            elif q == BAD_STOCK_Q:
                g = val["grade"]
                rows.append({**base, "metric": q, "grade": g, "person_name": None,
                              "point": questions[q][g][1],
                              "bs_ytd_value": val["ytd_value"], "bs_allowance": val["bs_allowance"],
                              "bs_utilization": val["bs_utilization"], "bs_compliance_pct": val["compliance_pct"]})
            else:
                g = val["grade"]
                raw = st.session_state.get(f"name_{q}", "").strip()
                rows.append({**base, "metric": q, "grade": g,
                              "person_name": None if g in dne else (raw.upper() or None),
                              "point": questions[q][g][1]})

        st.session_state.pending_submission = {
            "sid": sid, "rows": rows, "role_total": role_total, "role": role,
            "metrics": list(answers.keys()),
            "distributor": distributor, "assessment_period": assessment_period,
            "representative_name": representative_name.strip(),
        }
        st.session_state.show_confirm = True
        st.rerun()

    # ── PRE-SUBMIT CONFIRMATION POPUP ──
    def collapse_salesman_for_display(rows):
        display_rows = []
        salesman_added = False
        for r in rows:
            if r["metric"] == "SALESMAN":
                if not salesman_added:
                    names = [x["person_name"] for x in rows if x["metric"] == "SALESMAN" and x["person_name"]]
                    display_rows.append({**r, "person_name": ", ".join(names) if names else None})
                    salesman_added = True
            else:
                display_rows.append(r)
        return display_rows

    @st.dialog("📝 Confirm Submission", width="large")
    def show_confirm_dialog():
        data = st.session_state.pending_submission
        role_total = data["role_total"]
        role_max   = ROLE_MAX[data["role"]]

        st.markdown(f"**{data['distributor']}**  ·  {data['assessment_period']}")
        st.caption(f"Role: {ROLES[data['role']]['icon']} {data['role']}  ·  Representative: {data['representative_name'].upper()}")

        display_rows = collapse_salesman_for_display(data["rows"])
        df_res = pd.DataFrame(display_rows)[["metric", "grade", "point", "person_name"]]
        st.dataframe(df_res, use_container_width=True, hide_index=True)

        st.metric(f"{data['role']} Subtotal", f"{role_total} / {role_max}")
        st.caption("Please double-check the answers above. Once confirmed, this cannot be submitted again for this role on this distributor & period.")

        st.write("")
        b1, b2 = st.columns(2)
        with b1:
            if st.button("✏️  Back to Edit", use_container_width=True):
                st.session_state.show_confirm = False
                st.rerun()
        with b2:
            if st.button("✅  Confirm & Submit", type="primary", use_container_width=True):
                # Final live race-condition guard right before writing
                if check_role_already_submitted(data["role"], data["distributor"], data["assessment_period"]):
                    st.error("Someone already submitted this in the time you were reviewing. Refresh and try again.")
                else:
                    errors = insert_assessment_rows(data["rows"])
                    if errors:
                        st.error("❌ Failed to insert into BigQuery")
                        st.json(errors)
                    else:
                        get_combined_progress.clear()
                        st.session_state.submission_result = data
                        st.session_state.just_submitted = True
                        st.session_state.show_confirm = False
                        st.session_state.pending_submission = None
                        st.rerun()

    if st.session_state.show_confirm and st.session_state.pending_submission:
        show_confirm_dialog()

    # ── FINAL RESULT DISPLAY ──
    result = st.session_state.submission_result
    if result and (result["role"], result["distributor"], result["assessment_period"]) == role_key:
        st.success(
            f"✅ **{result['role']}** submission saved for **{result['distributor']}** · "
            f"{result['assessment_period']}  (+{result['role_total']} pts)  |  ID: `{result['sid']}`"
        )

        if st.session_state.just_submitted:
            st.balloons()
            st.session_state.just_submitted = False

        combined_total, items = get_combined_progress(distributor, assessment_period)
        all_done = all(it["status"] == "done" for it in items)

        if all_done:
            pct = int((combined_total / TOTAL_MAX_SCORE) * 100)
            if   pct >= 80: rating, rc = "Excellent ⭐",       "#059669"
            elif pct >= 60: rating, rc = "Good 👍",            "#D97706"
            elif pct >= 40: rating, rc = "Fair ⚠️",            "#EA580C"
            else:           rating, rc = "Needs Improvement ❌","#DC2626"

            st.markdown(f"""
            <div class="result-wrap">
                <div class="result-score">{combined_total}</div>
                <div class="result-pts">out of {TOTAL_MAX_SCORE} pts &nbsp;·&nbsp; all 4 roles submitted 🎉</div>
                <div class="result-rating" style="color:{rc};">{rating} &nbsp;·&nbsp; {pct}%</div>
                <div class="result-dist">{distributor} &nbsp;·&nbsp; {assessment_period}</div>
            </div>
            """, unsafe_allow_html=True)
        else:
            pending_roles = sorted({Q_OWNER[it["name"]] for it in items if it["status"] == "pending"})
            st.info(f"⏳ Waiting on: {', '.join(pending_roles)} before this assessment is complete.")

        with st.expander("📦 Rows inserted into BigQuery (this submission)", expanded=False):
            df_res = pd.DataFrame(result["rows"])[["metric", "grade", "point", "person_name", "submitted_role", "assessment_period"]]
            st.dataframe(df_res, use_container_width=True, hide_index=True)

# =====================================================================================
# ===========================  BULK ROLES (Distributor Manager / Admin RSA / AR)  ======
# =====================================================================================
else:
    metric_name = ROLES[role]["questions"][0]
    grades = questions[metric_name]

    st.success(f"✅  **{representative_name.strip()}**  ·  {ROLES[role]['icon']} **{role}**  ·  {assessment_period}  ·  bulk mode (no per-distributor selection)")

    pending_bulk = st.session_state.pending_bulk_submission
    if pending_bulk and (pending_bulk["role"], pending_bulk["assessment_period"]) != (role, assessment_period):
        st.session_state.pending_bulk_submission = None
        st.session_state.show_bulk_confirm = False

    # ── SIDEBAR — bulk submission status across all distributors ──
    with st.sidebar:
        st.markdown("---")
        st.markdown(f"### {ROLES[role]['icon']} {role}")
        st.caption(f"Bulk mode · {assessment_period}")

        done_list, pending_list = get_role_bulk_progress(role, assessment_period)
        n_total = len(done_list) + len(pending_list)
        pct = len(done_list) / n_total if n_total else 0
        st.metric("Distributors Submitted", f"{len(done_list)} / {n_total}", f"{int(pct*100)}%")
        st.progress(pct)

        st.markdown("---")
        if done_list:
            st.caption("✅ **Submitted**")
            for d in done_list:
                st.caption(f"&nbsp;&nbsp;• {d}")
        if pending_list:
            st.caption("⏳ **Pending**")
            for d in pending_list:
                st.caption(f"&nbsp;&nbsp;• {d}")

    is_value_based = metric_name in VALUE_THRESHOLDS

    # ── ADMIN: Create New User (Distributor Manager only) ──
    if role == "Distributor Manager":
        with st.expander("👤  Create New User (Admin)", expanded=False):
            st.caption("Create a login for any of the 4 roles — no BigQuery access needed.")

            region_opts = sorted(load_master_distributor()["region"].dropna().unique())

            cu1, cu2 = st.columns(2)
            with cu1:
                new_username = st.text_input("Username", placeholder="e.g. fajar", key="new_user_username")
                new_full_name = st.text_input("Full Name", placeholder="e.g. Fajar Pratama", key="new_user_full_name")
                new_email = st.text_input("Email (optional)", placeholder="e.g. fajar@skintific.com", key="new_user_email")
            with cu2:
                new_password = st.text_input("Password", type="password", placeholder="••••••••", key="new_user_password")
                new_user_role = st.selectbox("Role", list(ROLES.keys()), key="new_user_role")
                if new_user_role == "Area Sales Supervisor":
                    new_region = st.selectbox("Region", region_opts, key="new_user_region")
                else:
                    new_region = None

            create_clicked = st.button("➕  Create User", type="primary", use_container_width=True, key="create_user_btn")

            if create_clicked:
                errors = []
                if new_username.strip() == "":
                    errors.append("Username is required.")
                if new_password == "":
                    errors.append("Password is required.")
                if new_full_name.strip() == "":
                    errors.append("Full Name is required.")
                if new_user_role == "Area Sales Supervisor" and not new_region:
                    errors.append("Region is required for Area Sales Supervisor.")
                if not errors and username_exists(new_username):
                    errors.append(f"Username '{normalize_username(new_username)}' already exists.")

                if errors:
                    for e in errors:
                        st.error(e)
                else:
                    insert_errors = create_user(new_username, new_password, new_full_name, new_user_role, new_region, new_email)
                    if insert_errors:
                        st.error("❌ Failed to create user")
                        st.json(insert_errors)
                    else:
                        st.success(f"✅ User '{normalize_username(new_username)}' created with role **{new_user_role}**.")

        with st.expander("📊  Reporting — Pending ASS Assessments", expanded=False):
            report_period = st.selectbox("Period", period_labels, index=default_index, key="report_period")

            missing_df = get_ass_missing_distributors(report_period)
            not_submitted_df = get_ass_users_not_submitted(report_period)
            total_distributors = load_master_distributor()["distributor"].nunique()
            total_ass_users = get_total_ass_users()
            submitted_distributors = total_distributors - len(missing_df)
            submitted_ass_users = total_ass_users - len(not_submitted_df)

            rc1, rc2 = st.columns(2)
            with rc1:
                st.metric("Distributors Submitted (ASS)", f"{submitted_distributors}/{total_distributors}")
            with rc2:
                st.metric("ASS Users Submitted", f"{submitted_ass_users}/{total_ass_users}")

            st.markdown(f"**Distributors without an ASS assessment — {report_period}**")
            if missing_df.empty:
                st.success("All distributors have an ASS assessment for this period.")
            else:
                st.dataframe(missing_df, use_container_width=True, hide_index=True)

            st.markdown(f"**ASS users who haven't submitted yet — {report_period}**")
            if not_submitted_df.empty:
                st.success("All ASS users have submitted for this period.")
            else:
                st.dataframe(not_submitted_df, use_container_width=True, hide_index=True)

    # ── ADMIN: NPD & SKU Focus allocation uploads (Distributor Manager only) ──
    @st.dialog("📝 Confirm Allocation Submission", width="large")
    def show_alloc_confirm_dialog():
        data = st.session_state.pending_allocation
        cfg = ALLOCATION_PROGRAMS[data["program_type"]]

        st.markdown(f"**{cfg['icon']} {cfg['label']}**  ·  {data['eff_date']} → {data['end_date']}")
        df_res = pd.DataFrame(data["rows"])[["distributor_code", "distributor", "brand", "sku_code", "allocation_target"]]
        st.dataframe(df_res, use_container_width=True, hide_index=True)
        st.metric("Rows in this batch", f"{len(data['rows'])}")
        st.caption("Please double-check before confirming. Effective/End Date are locked to the filter you chose, not whatever was in the Excel file.")

        st.write("")
        b1, b2 = st.columns(2)
        with b1:
            if st.button("✏️  Back to Edit", use_container_width=True, key="alloc_back"):
                st.session_state.show_alloc_confirm = False
                st.rerun()
        with b2:
            if st.button("✅  Confirm & Submit", type="primary", use_container_width=True, key="alloc_confirm"):
                insert_errors = insert_allocation_rows(data["rows"])
                if insert_errors:
                    st.error("❌ Failed to insert into BigQuery")
                    st.json(insert_errors)
                else:
                    st.session_state.just_submitted_alloc = data["program_type"]
                    st.session_state.last_alloc_count = len(data["rows"])
                    st.session_state.show_alloc_confirm = False
                    st.session_state.pending_allocation = None
                    st.rerun()

    if st.session_state.show_alloc_confirm and st.session_state.pending_allocation:
        show_alloc_confirm_dialog()

    def render_allocation_section(program_type):
        cfg = ALLOCATION_PROGRAMS[program_type]
        with st.expander(f"{cfg['icon']}  {cfg['label']} Allocation Upload", expanded=False):
            st.caption(f"Set {cfg['label']} allocation targets per distributor + SKU for a date range.")

            dc1, dc2 = st.columns(2)
            with dc1:
                eff_date = st.date_input("Effective Date", key=f"alloc_eff_{program_type}")
            with dc2:
                end_date = st.date_input("End Date", key=f"alloc_end_{program_type}")

            if end_date < eff_date:
                st.error("End Date must be on or after Effective Date.")
                return

            st.caption("📌 Choose the date range above **first** — it's locked into every row of the template, and the final submission always uses these dates regardless of what's edited in the uploaded file.")

            master_df = load_master_distributor()
            template_buffer = generate_allocation_template(program_type, eff_date, end_date)
            st.download_button(
                f"⬇️  Download {program_type} Template",
                data=template_buffer,
                file_name=f"{program_type.replace(' ', '_')}_template_{eff_date}_{end_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key=f"dl_alloc_{program_type}",
            )

            uploaded_file = st.file_uploader(
                f"Upload filled {program_type} template (.xlsx)", type=["xlsx"], key=f"upload_alloc_{program_type}"
            )

            ready_rows = []
            if uploaded_file is not None:
                try:
                    df_upload = pd.read_excel(uploaded_file, sheet_name="Template")
                except Exception as e:
                    st.error(f"Could not read the uploaded file: {e}")
                    return

                df_upload.columns = [str(c).strip().lower() for c in df_upload.columns]
                required_cols = {"region", "distributor_name", "distributor_code", "brand", "sku_code", "allocation_target"}
                if not required_cols.issubset(set(df_upload.columns)):
                    st.error(f"Template must contain columns: {', '.join(sorted(required_cols))}")
                    return

                preview_rows, row_errors = parse_allocation_upload(df_upload, master_df)

                if row_errors:
                    with st.expander(f"⚠️ {len(row_errors)} row(s) skipped", expanded=True):
                        for e in row_errors:
                            st.caption(e)

                if preview_rows:
                    st.markdown(f"**Preview** — Effective/End Date will be set to **{eff_date} → {end_date}** for every row")
                    st.dataframe(pd.DataFrame(preview_rows), use_container_width=True, hide_index=True)
                    ready_rows = preview_rows
                    st.caption(f"{len(ready_rows)} row(s) ready to submit.")
                else:
                    st.info("No valid rows found — fill in brand / sku_code / allocation_target for at least one distributor and re-upload.")

            st.write("")
            review_clicked = st.button(
                f"🔍  Review & Submit {program_type} ({len(ready_rows)} row{'s' if len(ready_rows) != 1 else ''})",
                use_container_width=True, type="primary",
                disabled=len(ready_rows) == 0, key=f"alloc_review_{program_type}",
            )

            if review_clicked:
                sid = str(uuid.uuid4())
                submitted_at = pendulum_now("Asia/Jakarta").to_datetime_string()
                rows = [{
                    "submission_id": sid, "submitted_at": submitted_at,
                    "submitted_by_username": st.session_state.user["username"],
                    "program_type": program_type,
                    "region": r["region"], "distributor_code": r["distributor_code"], "distributor": r["distributor_name"],
                    "brand": r["brand"], "sku_code": r["sku_code"], "allocation_target": r["allocation_target"],
                    "effective_date": eff_date.isoformat(), "end_date": end_date.isoformat(),
                } for r in ready_rows]

                st.session_state.pending_allocation = {
                    "sid": sid, "rows": rows, "program_type": program_type,
                    "eff_date": eff_date, "end_date": end_date,
                }
                st.session_state.show_alloc_confirm = True
                st.rerun()

            if st.session_state.just_submitted_alloc == program_type:
                st.success(f"✅ {program_type} allocation submitted — {st.session_state.last_alloc_count} row(s) inserted into BigQuery.")
                st.balloons()
                st.session_state.just_submitted_alloc = None

    if role == "Distributor Manager":
        render_allocation_section("NPD")
        render_allocation_section("SKU Focus")

    with st.container(border=True):
        st.markdown(f'<div class="sec-label">📖 Grade Reference — {metric_name}</div>', unsafe_allow_html=True)
        if is_value_based:
            st.caption(f"Enter the **number of days late** as the value — grade is derived automatically (0 → A, 1–{VALUE_THRESHOLDS[metric_name]} → B, more → C):")
        for g, (desc, pts) in grades.items():
            st.caption(f"**{g}** — {desc}  ({pts} pts)")

    with st.container(border=True):
        st.markdown('<div class="sec-label">📥 Step 1 — Download Template</div>', unsafe_allow_html=True)
        if is_value_based:
            st.caption("Download the Excel template, fill in the **value** column (number of days late) for each distributor — the **grade**, **point** and **description** columns auto-fill — then upload it below.")
        else:
            st.caption("Download the Excel template, fill in the **grade** column (A/B/C) for each distributor — the **point** and **description** columns auto-fill — then upload it below.")
        template_buffer = generate_bulk_template(metric_name)
        st.download_button(
            "⬇️  Download Excel Template",
            data=template_buffer,
            file_name=f"{role.replace(' ', '_')}_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    with st.container(border=True):
        st.markdown('<div class="sec-label">📤 Step 2 — Upload Completed Template</div>', unsafe_allow_html=True)
        uploaded_file = st.file_uploader("Upload filled template (.xlsx)", type=["xlsx"], key=f"upload_{role}")

    bulk_ready_rows = []
    if uploaded_file is not None:
        try:
            df_upload = pd.read_excel(uploaded_file, sheet_name="Template")
        except Exception as e:
            st.error(f"Could not read the uploaded file: {e}")
            st.stop()

        df_upload.columns = [str(c).strip().lower() for c in df_upload.columns]
        required_cols = (
            {"region", "distributor_code", "distributor_name", "value"} if is_value_based
            else {"region", "distributor_code", "distributor_name", "grade"}
        )
        if not required_cols.issubset(set(df_upload.columns)):
            st.error(f"Template must contain columns: {', '.join(sorted(required_cols))}")
            st.stop()

        master_df    = load_master_distributor()
        code_to_name = dict(zip(master_df["distributor_code"], master_df["distributor"]))
        code_to_region = dict(zip(master_df["distributor_code"], master_df["region"]))
        valid_grades = set(grades.keys())

        # Resolve codes -> names first, so the batch dup-check (which queries
        # distributor_assessment.distributor, a name column) covers the WHOLE
        # uploaded file in one query.
        candidate_codes = [str(r.get("distributor_code", "")).strip() for _, r in df_upload.iterrows()]
        candidate_names = [code_to_name[c] for c in candidate_codes if c in code_to_name]
        already_set = check_bulk_already_submitted(role, candidate_names, assessment_period, metric_name)

        preview_rows = []
        row_errors = []
        for i, r in df_upload.iterrows():
            code = str(r.get("distributor_code", "")).strip()
            if code not in code_to_name:
                row_errors.append(f"Row {i + 2}: unknown distributor_code '{code}' — skipped.")
                continue
            dist_name = code_to_name[code]

            if is_value_based:
                raw_value = r.get("value", "")
                if pd.isna(raw_value) or str(raw_value).strip() == "":
                    continue
                try:
                    num_value = float(raw_value)
                except (TypeError, ValueError):
                    row_errors.append(f"Row {i + 2}: invalid value '{raw_value}' for {dist_name} (must be a number of days) — skipped.")
                    continue
                if num_value < 0:
                    num_value = 0
                grade_val = value_to_grade(metric_name, num_value)
            else:
                raw_grade = r.get("grade", "")
                grade_val = "" if pd.isna(raw_grade) else str(raw_grade).strip().upper()
                if grade_val == "":
                    continue
                if grade_val not in valid_grades:
                    row_errors.append(
                        f"Row {i + 2}: invalid grade '{grade_val}' for {dist_name} "
                        f"(must be one of {'/'.join(sorted(valid_grades))}) — skipped."
                    )
                    continue

            already = dist_name in already_set
            preview_row = {"distributor_code": code, "distributor_name": dist_name, "region": code_to_region.get(code)}
            if is_value_based:
                preview_row["value"] = num_value
            preview_row.update({
                "grade": grade_val,
                "description": grades[grade_val][0],
                "points": grades[grade_val][1],
                "status": "already submitted" if already else "ready",
            })
            preview_rows.append(preview_row)

        if row_errors:
            with st.expander(f"⚠️ {len(row_errors)} row(s) skipped", expanded=True):
                for e in row_errors:
                    st.caption(e)

        if preview_rows:
            st.markdown("**Preview**")
            st.dataframe(pd.DataFrame(preview_rows), use_container_width=True, hide_index=True)
            bulk_ready_rows = [r for r in preview_rows if r["status"] == "ready"]
            skipped_dup = len(preview_rows) - len(bulk_ready_rows)
            msg = f"{len(bulk_ready_rows)} distributor(s) ready to submit"
            if skipped_dup:
                msg += f", {skipped_dup} already submitted (will be skipped)"
            st.caption(msg + ".")
        else:
            col_hint = "value" if is_value_based else "grade"
            st.info(f"No valid rows found — fill in the `{col_hint}` column for at least one distributor and re-upload.")

    st.write("")
    bulk_review_clicked = st.button(
        f"🔍  Review & Submit Bulk ({len(bulk_ready_rows)} distributor{'s' if len(bulk_ready_rows) != 1 else ''})",
        use_container_width=True, type="primary",
        disabled=len(bulk_ready_rows) == 0,
    )

    if bulk_review_clicked:
        sid = str(uuid.uuid4())
        submitted_at = pendulum_now("Asia/Jakarta").to_datetime_string()
        rows = [{
            "submission_id": sid, "submitted_at": submitted_at,
            "representative_name": representative_name.strip().upper(),
            "submitted_role": role, "submitted_by_username": st.session_state.user["username"],
            "region": r.get("region"), "distributor": r["distributor_name"], "assessment_period": assessment_period,
            "metric": metric_name, "grade": r["grade"], "person_name": None, "point": r["points"],
            "inner_city_sla": None, "outer_city_sla": None,
            "bs_ytd_value": None, "bs_allowance": None, "bs_utilization": None, "bs_compliance_pct": None,
            "days_late": r.get("value") if is_value_based else None,
        } for r in bulk_ready_rows]

        st.session_state.pending_bulk_submission = {
            "sid": sid, "rows": rows, "role": role, "metric": metric_name,
            "assessment_period": assessment_period,
            "representative_name": representative_name.strip(),
        }
        st.session_state.show_bulk_confirm = True
        st.rerun()

    @st.dialog("📝 Confirm Bulk Submission", width="large")
    def show_bulk_confirm_dialog():
        data = st.session_state.pending_bulk_submission
        total_pts = sum(r["point"] for r in data["rows"])

        st.markdown(f"**{data['role']}** · {data['metric']}  ·  {data['assessment_period']}")
        st.caption(f"Representative: {data['representative_name'].upper()}")

        df_res = pd.DataFrame(data["rows"])[["distributor", "grade", "point"]]
        st.dataframe(df_res, use_container_width=True, hide_index=True)

        cm1, cm2 = st.columns(2)
        with cm1:
            st.metric("Distributors in this batch", f"{len(data['rows'])}")
        with cm2:
            st.metric("Total Points (this batch)", f"{total_pts}")

        st.caption("Please double-check before confirming. Distributors already submitted for this role/period were excluded automatically.")

        st.write("")
        b1, b2 = st.columns(2)
        with b1:
            if st.button("✏️  Back to Edit", use_container_width=True, key="bulk_back"):
                st.session_state.show_bulk_confirm = False
                st.rerun()
        with b2:
            if st.button("✅  Confirm & Submit Bulk", type="primary", use_container_width=True, key="bulk_confirm"):
                # Final live race-condition guard — re-check right before insert
                names = [r["distributor"] for r in data["rows"]]
                still_already = check_bulk_already_submitted(data["role"], names, data["assessment_period"], data["metric"])
                final_rows = [r for r in data["rows"] if r["distributor"] not in still_already]

                if not final_rows:
                    st.warning("All distributors in this batch were already submitted by someone else. Nothing to insert.")
                else:
                    errors = insert_assessment_rows(final_rows)
                    if errors:
                        st.error("❌ Failed to insert into BigQuery")
                        st.json(errors)
                    else:
                        get_role_bulk_progress.clear()
                        get_combined_progress.clear()
                        st.session_state.bulk_submission_result = {**data, "rows": final_rows}
                        st.session_state.just_submitted_bulk = True
                        st.session_state.show_bulk_confirm = False
                        st.session_state.pending_bulk_submission = None
                        st.rerun()

    if st.session_state.show_bulk_confirm and st.session_state.pending_bulk_submission:
        show_bulk_confirm_dialog()

    # ── FINAL RESULT DISPLAY (bulk) ──
    bulk_result = st.session_state.bulk_submission_result
    if bulk_result and bulk_result["role"] == role and bulk_result["assessment_period"] == assessment_period:
        st.success(
            f"✅ Bulk submission saved — **{len(bulk_result['rows'])} distributor(s)** updated for "
            f"**{role}** · {bulk_result['assessment_period']}  |  ID: `{bulk_result['sid']}`"
        )

        if st.session_state.just_submitted_bulk:
            st.balloons()
            st.session_state.just_submitted_bulk = False

        with st.expander("📦 Rows inserted into BigQuery (this batch)", expanded=False):
            df_res = pd.DataFrame(bulk_result["rows"])[
                ["distributor", "metric", "grade", "point", "submitted_role", "assessment_period"]
            ]
            st.dataframe(df_res, use_container_width=True, hide_index=True)
