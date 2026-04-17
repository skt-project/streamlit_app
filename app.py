import streamlit as st
import pandas as pd
from datetime import datetime
import io
import zipfile
import math
import re
import base64
import urllib.request
from pathlib import Path
import openpyxl

# --- BigQuery Imports ---
from google.cloud import bigquery
from google.api_core.exceptions import NotFound
from google.oauth2 import service_account

try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    MATPLOTLIB_OK = True
except ImportError:
    MATPLOTLIB_OK = False

st.set_page_config(
    page_title="DataFlow — Glad2Glow",
    layout="wide",
    page_icon='📁',
    initial_sidebar_state="expanded",
)

DASHBOARD_URL_DEFAULT = "https://po-simulator.streamlit.app/"

# Google Drive share link untuk po_template.xlsx
TEMPLATE_DRIVE_URL = "https://docs.google.com/spreadsheets/d/1FD2WN8PutkwzXXRYSj1jpA4EyxqzAfStyg2KC3grR30/edit?usp=sharing"

# =========================
# BigQuery Configuration
# =========================
try:
    gcp_secrets = st.secrets["connections"]["bigquery"]
    private_key = gcp_secrets["private_key"].replace("\\n", "\n")
    _bq_credentials = service_account.Credentials.from_service_account_info(
        {
            "type": gcp_secrets["type"],
            "project_id": gcp_secrets["project_id"],
            "private_key_id": gcp_secrets["private_key_id"],
            "private_key": private_key,
            "client_email": gcp_secrets["client_email"],
            "client_id": gcp_secrets["client_id"],
            "auth_uri": gcp_secrets["auth_uri"],
            "token_uri": gcp_secrets["token_uri"],
            "auth_provider_x509_cert_url": gcp_secrets["auth_provider_x509_cert_url"],
            "client_x509_cert_url": gcp_secrets["client_x509_cert_url"],
        }
    )
    GCP_PROJECT_ID = st.secrets["bigquery"]["project"]
except Exception:
    _bq_credentials = None
    GCP_PROJECT_ID = "skintific-data-warehouse"


@st.cache_resource(show_spinner=False)
def get_bq_client() -> bigquery.Client:
    """Initializes and returns a BigQuery client."""
    return bigquery.Client(credentials=_bq_credentials, project=GCP_PROJECT_ID)


@st.cache_data(ttl=3600, show_spinner=False)
def fetch_customer_names() -> list:
    """Fetch distributor list from BigQuery master_distributor table."""
    try:
        client = get_bq_client()
        query = """
            SELECT distributor
            FROM `skintific-data-warehouse.gt_schema.master_distributor`
            ORDER BY distributor
        """
        rows = client.query(query).result()
        return [r.distributor for r in rows if r.distributor]
    except Exception as e:
        st.warning(f"⚠️ Gagal memuat daftar distributor dari BigQuery: {e}")
        return []


# Load customer names dynamically from BigQuery
CUSTOMER_NAMES = fetch_customer_names()


def _drive_to_direct(url: str) -> str:
    """Convert Google Drive / Google Sheets share URL → direct download URL."""
    m = re.search(r'/spreadsheets/d/([a-zA-Z0-9_-]+)', url)
    if m:
        return f"https://docs.google.com/spreadsheets/d/{m.group(1)}/export?format=xlsx"
    m = re.search(r'/file/d/([a-zA-Z0-9_-]+)', url)
    if m:
        return f"https://drive.usercontent.google.com/download?id={m.group(1)}&export=download&authuser=0"
    return url


@st.cache_data(show_spinner=False)
def _fetch_template_bytes(url: str) -> bytes:
    """Download template xlsx dari Google Drive/Sheets, cache hasilnya."""
    if not url or not url.strip():
        raise ValueError("TEMPLATE_DRIVE_URL belum diset. Isi konstanta di bagian atas app.py.")
    direct = _drive_to_direct(url)
    headers = {"User-Agent": "Mozilla/5.0"}
    req = urllib.request.Request(direct, headers=headers)
    with urllib.request.urlopen(req, timeout=30) as resp:
        data = resp.read()
    if not data or data[:4] != b'PK\x03\x04':
        raise ValueError(
            "File yang didownload bukan xlsx yang valid. "
            "Pastikan file bisa diakses publik ('Anyone with the link')."
        )
    return data


def _logo_src() -> str:
    local = Path(__file__).parent / "logo.png"
    if local.exists():
        b64 = base64.b64encode(local.read_bytes()).decode()
        return f"data:image/png;base64,{b64}"
    return "https://glad2glow.com/cdn/shop/files/logo.png?height=628&pad_color=ffffff&v=1745724802&width=1200"


LOGO_URL = _logo_src()

PO_TEMPLATE_COLS = [
    'Distributor', 'SKU', 'Product Name', 'Assortment', 'Supply Control',
    'Avg Weekly Sales LM (Qty)', 'Total Stock (Qty)', 'Current WOI',
    'PO Qty', 'PO Value', 'WOI (Stock + PO Ori)', 'Remark',
    'Suggested PO Qty', 'Suggested PO Value',
    'WOI After Buffer (Stock + Suggested Qty)',
    'Stock + Suggested Qty WOI (Projection at EOM)',
    'Remaining Allocation (By Region)', 'RSA Notes',
]

PO_IMG_COLS = PO_TEMPLATE_COLS[6:13]
PO_COLS_copy = PO_TEMPLATE_COLS[:13]


def check_password():

    def login_form():
        st.markdown("""
        <style>
        body, p, div, span, label, input, button, h1, h2, h3, h4, h5, h6 { font-family: 'Trebuchet MS', sans-serif; }
        html, body, [data-testid="stAppViewContainer"] {
            background: linear-gradient(145deg, #F13E93 40%,#FFA6A6 50%, #F26076 60%) !important;
            min-height: 50vh;
        }
        [data-testid="stSidebar"] { display: none !important; }
        .login-card {
            background: rgba(255,255,255,0.92);
            border-radius: 24px;
            padding: 2.5rem 2rem 2rem;
            box-shadow: 0 20px 60px rgba(202,97,128,0.25);
            border: 1px solid rgba(252,183,199,0.5);
            backdrop-filter: blur(12px);
        }
        .login-petal {
            position: fixed; pointer-events: none; font-size: 1.8rem; opacity: 0.18;
        }
        </style>
        <div class="login-petal" style="top:8%;left:5%;">🌸</div>
        <div class="login-petal" style="top:15%;right:8%;">🌺</div>
        <div class="login-petal" style="bottom:20%;left:10%;">🌸</div>
        <div class="login-petal" style="bottom:10%;right:6%;">💐</div>
        """, unsafe_allow_html=True)

        _, col, _ = st.columns([1, 1.5, 1])
        with col:
            st.markdown(
                f'<div style="text-align:center;padding:0.5rem 0 0.2rem;">'
                f'<img src="{LOGO_URL}" style="max-width:200px;height:auto;display:inline-block;background:transparent;" />'
                f'</div>',
                unsafe_allow_html=True
            )
            st.markdown("""
            <div style="text-align:center; margin: 1rem 0 1.5rem;">
                <div style="font-size:1.5rem; font-weight:700; color:#6E253A; letter-spacing:0.5px;">
                    DataFlow Automator
                </div>
                <div style="color:#000000; font-size:0.85rem; margin-top:0.3rem;">
                    Masukkan passwordmu
                </div>
            </div>
            """, unsafe_allow_html=True)

            with st.form("login_form"):
                password = st.text_input(
                    "🔒 Password",
                    type="password",
                    placeholder="Masukkan password...",
                )
                submitted = st.form_submit_button(
                    "Login", use_container_width=True
                )
                if submitted:
                    if password == st.secrets["glowithyou"]:
                        st.session_state["authenticated"] = True
                        st.rerun()
                    else:
                        st.error("❌ Password salah. Silakan coba lagi.")

            st.markdown("""
            <div style="text-align:center; margin-top:1.5rem; color:#6E253A; font-size:0.72rem;">
                🌸 Secured with Streamlit Secrets &nbsp;·&nbsp; Glad2Glow
            </div>
            </div>
            """, unsafe_allow_html=True)

    if st.session_state.get("authenticated"):
        return True

    login_form()
    return False


if not check_password():
    st.stop()

CUSTOM_CSS = """
<style>

body, p, div, span, label, input, textarea, select, button,
h1, h2, h3, h4, h5, h6, li, td, th, caption, small, strong, em {
    font-family: 'Trebuchet MS', sans-serif;
    box-sizing: border-box;
}

[data-testid="stSidebarCollapseButton"] { display: none !important; }

:root {
    --rose:      #CA6180;
    --rose-dark: #A84D6A;
    --blush:     #FCB7C7;
    --dark:      #BF3979;
    --dark2:     #751F58;
    --border:    rgba(0,0,0,0.08);
    --shadow:    0 2px 12px rgba(0,0,0,0.07);
    --shadow-lg: 0 8px 32px rgba(0,0,0,0.12);
    --g-rose:    linear-gradient(135deg, #751F58, #A84D6A);
}

h1,h2,h3,h4,h5,h6 { color: #FFFFFF !important; }

[data-testid="stSidebar"] {
    background: var(--dark) !important;
    border-right: 1px solid rgba(255,255,255,0.06) !important;
    box-shadow: 4px 0 24px rgba(0,0,0,0.35);
}
[data-testid="stSidebar"], [data-testid="stSidebar"] * { color: #FFFFFF !important; }
[data-testid="stSidebar"] input {
    background: rgba(255,255,255,0.1) !important;
    border: 1px solid rgba(255,255,255,0.2) !important;
    color: #FFFFFF !important; border-radius: 8px !important;
}
[data-testid="stSidebar"] input::placeholder { color: rgba(255,255,255,0.4) !important; }
[data-testid="stSidebar"] [data-baseweb="select"] > div {
    background: rgba(255,255,255,0.1) !important;
    border: 1px solid rgba(255,255,255,0.2) !important;
    color: #FFFFFF !important;
}
[data-testid="stSidebar"] hr { border-color: rgba(255,255,255,0.1) !important; }

[data-testid="stMain"] input, [data-testid="stMain"] textarea {
    color: #FFFFFF !important; background: rgba(255,255,255,0.15) !important;
    border: 1.5px solid rgba(255,255,255,0.35) !important; border-radius: 10px !important;
}
[data-testid="stMain"] input:focus {
    border-color: #FFFFFF !important;
    box-shadow: 0 0 0 3px rgba(255,255,255,0.15) !important;
}
[data-testid="stMain"] .stSelectbox > div > div {
    color: #FFFFFF !important; background: rgba(255,255,255,0.15) !important;
    border: 1.5px solid rgba(255,255,255,0.35) !important; border-radius: 10px !important;
}
[data-testid="stMain"] [data-baseweb="select"] span,
[data-testid="stMain"] [data-baseweb="select"] div { color: #FFFFFF !important; }

.hero-wrap { padding: 0.5rem 0 1.5rem; }
.hero-tag {
    display: inline-block;
    background: rgba(168,77,106,0.6);
    color: #FFFFFF !important;
    font-size: 0.72rem; font-weight: 700;
    letter-spacing: 1.5px; text-transform: uppercase;
    padding: 0.25rem 0.8rem; border-radius: 20px;
    border: 1px solid rgba(255,255,255,0.25);
    margin-bottom: 0.75rem;
}
.hero-title {
    font-size: 2rem; font-weight: 700; color: #FFFFFF !important;
    margin-bottom: 0.5rem; line-height: 1.25;
}
.hero-sub { color: rgba(255,255,255,0.85) !important; font-size: 0.92rem; line-height: 1.7; max-width: 600px; }

.pipeline-step {
    background: #A84D6A; border: none;
    border-radius: 12px; padding: 0.9rem 1.3rem;
    margin-bottom: 1rem;
    box-shadow: 0 2px 12px rgba(0,0,0,0.15);
}
.pipeline-step *, .pipeline-step strong, .pipeline-step code { color: #FFFFFF !important; }
.pipeline-step.active   { background: #8B2040; }
.pipeline-step.completed { background: #3E7D6A; }
.step-number {
    display: inline-flex; align-items: center; justify-content: center;
    width: 24px; height: 24px; border-radius: 50%;
    background: var(--g-rose); color: #FFF !important;
    font-weight: 700; font-size: 0.75rem;
    margin-right: 0.6rem; vertical-align: middle;
}

.metric-card {
    background: rgba(168,77,106,0.35); border: 1px solid rgba(255,255,255,0.2);
    border-radius: 12px; padding: 1.1rem 1.3rem;
    box-shadow: var(--shadow);
    transition: transform 0.2s, box-shadow 0.2s;
    position: relative; overflow: hidden;
}
.metric-card::before {
    content: ''; position: absolute; top: 0; left: 0;
    width: 3px; height: 100%; background: rgba(255,255,255,0.5);
    border-radius: 3px 0 0 3px;
}
.metric-card:hover { transform: translateY(-2px); box-shadow: var(--shadow-lg); }
.metric-label {
    color: rgba(255,255,255,0.7) !important; font-size: 0.68rem;
    text-transform: uppercase; letter-spacing: 1.1px; font-weight: 700; margin-bottom: 0.3rem;
}
.metric-value { font-size: 1.5rem; font-weight: 700; color: #FFFFFF !important; }
.metric-rose  .metric-value { color: var(--rose) !important; }
.metric-pink  .metric-value { color: #C96080 !important; }
.metric-muted .metric-value { color: var(--rose-dark) !important; }
.metric-blush .metric-value { color: #B05A77 !important; }

.feat-card {
    background: rgba(168,77,106,0.35); border: 1px solid rgba(255,255,255,0.2);
    border-radius: 16px; padding: 2rem 1.5rem;
    text-align: center; box-shadow: var(--shadow);
    transition: transform 0.2s, box-shadow 0.2s;
}
.feat-card:hover { transform: translateY(-3px); box-shadow: var(--shadow-lg); }
.feat-icon { font-size: 2.2rem; margin-bottom: 0.8rem; }
.feat-title { font-size: 0.95rem; font-weight: 700; color: #FFFFFF !important; margin-bottom: 0.4rem; }
.feat-desc  { font-size: 0.8rem; color: rgba(255,255,255,0.75) !important; line-height: 1.6; }

.badge {
    display: inline-block; padding: 0.18rem 0.65rem;
    border-radius: 20px; font-size: 0.65rem; font-weight: 700;
    text-transform: uppercase; letter-spacing: 0.7px;
}
.badge-success { background: rgba(5,150,105,0.25); color: #FFFFFF !important; border: 1px solid rgba(255,255,255,0.2); }
.badge-warning { background: rgba(168,77,106,0.4); color: #FFFFFF !important; border: 1px solid rgba(255,255,255,0.2); }
.badge-info    { background: rgba(255,255,255,0.2); color: #FFFFFF !important; border: 1px solid rgba(255,255,255,0.3); }

.stDownloadButton > button, .stButton > button {
    background: var(--g-rose) !important; color: #FFF !important;
    border: none !important; border-radius: 10px !important;
    font-weight: 700 !important; padding: 0.55rem 1.4rem !important;
    box-shadow: 0 2px 10px rgba(202,97,128,0.28) !important;
    transition: all 0.2s !important;
}
.stDownloadButton > button:hover, .stButton > button:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 5px 18px rgba(202,97,128,0.4) !important;
}

[data-testid="stDataFrame"] { border-radius: 10px !important; overflow: hidden !important; }
hr { border-color: #E8EAED !important; margin: 1.2rem 0 !important; }
.stAlert { border-radius: 10px !important; }
.stAlert * { color: inherit !important; }
.streamlit-expanderHeader {
    background: rgba(255,255,255,0.15) !important; border-radius: 10px !important;
    border: 1px solid rgba(255,255,255,0.25) !important; color: #FFFFFF !important;
}
.streamlit-expanderHeader * { color: #FFFFFF !important; }
.gsheet-card {
    background: rgba(255,255,255,0.1); border: 1.5px dashed rgba(255,255,255,0.3);
    border-radius: 14px; padding: 1.4rem;
}
.gsheet-card * { color: #FFFFFF !important; }
.split-info {
    background: rgba(168,77,106,0.3); border: 1px solid rgba(255,255,255,0.2);
    border-radius: 10px; padding: 0.85rem 1.1rem; margin: 0.5rem 0;
}
.split-info * { color: #FFFFFF !important; }
[data-testid="stMain"] [data-testid="stDateInput"] input {
    color: #FFFFFF !important; background: rgba(255,255,255,0.15) !important;
    border: 1.5px solid rgba(255,255,255,0.35) !important; border-radius: 10px !important;
}

.product-banner {
    border-radius: 20px; overflow: hidden;
    background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
    padding: 3rem 2.5rem; margin-top: 2rem;
    display: flex; align-items: center; justify-content: space-between;
    position: relative; min-height: 180px;
}
.product-banner::before {
    content: ''; position: absolute; inset: 0;
    background: radial-gradient(ellipse 400px 300px at 80% 50%,
        rgba(202,97,128,0.18) 0%, transparent 70%);
    pointer-events: none;
}
.banner-left { flex: 1; z-index: 1; }
.banner-label {
    font-size: 0.7rem; font-weight: 700; letter-spacing: 2px;
    text-transform: uppercase; color: var(--blush) !important;
    margin-bottom: 0.6rem;
}
.banner-title {
    font-size: 1.6rem; font-weight: 700; color: #FFFFFF !important;
    line-height: 1.3; margin-bottom: 0.5rem;
}
.banner-sub { font-size: 0.85rem; color: rgba(255,255,255,0.6) !important; }
.banner-right { z-index: 1; text-align: right; }
.banner-stat { font-size: 3.5rem; font-weight: 700; color: #FFFFFF !important; line-height: 1; }
.banner-stat-label { font-size: 0.78rem; color: rgba(255,255,255,0.55) !important; margin-top: 0.2rem; }
</style>
"""
st.markdown(CUSTOM_CSS, unsafe_allow_html=True)


def detect_date_columns(df: pd.DataFrame) -> list:
    date_cols = []
    for col in df.columns:
        if df[col].dtype in ['datetime64[ns]', 'datetime64[ns, UTC]']:
            date_cols.append(col)
            continue
        if df[col].dtype == object:
            sample = df[col].dropna().head(50)
            if len(sample) == 0:
                continue
            try:
                parsed = pd.to_datetime(sample, errors='coerce')
                if parsed.notna().sum() / len(sample) > 0.6:
                    date_cols.append(col)
            except Exception:
                pass
    return date_cols


def safe_to_datetime(series: pd.Series) -> pd.Series:
    try:
        return pd.to_datetime(series, errors='coerce')
    except Exception:
        return series


def _convert_to_xlsx(fname: str, fbytes: bytes) -> tuple[str, bytes]:
    """Convert xls/csv to xlsx bytes. Returns (new_fname, xlsx_bytes). Preserves sheet visibility."""
    ext = fname.rsplit(".", 1)[-1].lower()
    base = fname.rsplit(".", 1)[0]
    if ext == "xlsx":
        return fname, fbytes
    elif ext == "xls":
        import xlrd as _xlrd
        _book = _xlrd.open_workbook(file_contents=fbytes)
        _wb = openpyxl.Workbook()
        _wb.remove(_wb.active)
        _vis_map = {0: 'visible', 1: 'hidden', 2: 'veryHidden'}
        for _si in range(_book.nsheets):
            _ws_old = _book.sheet_by_index(_si)
            _ws_new = _wb.create_sheet(title=_ws_old.name)
            _vis = getattr(_book, 'sheet_visibility', None)
            _ws_new.sheet_state = _vis_map.get(_vis[_si] if _vis else 0, 'visible')
            for _r in range(_ws_old.nrows):
                for _c in range(_ws_old.ncols):
                    _ws_new.cell(row=_r + 1, column=_c + 1, value=_ws_old.cell_value(_r, _c))
        _buf = io.BytesIO()
        _wb.save(_buf)
        return base + ".xlsx", _buf.getvalue()
    elif ext == "csv":
        _df = pd.read_csv(io.BytesIO(fbytes), dtype=str, encoding_errors="replace")
        _buf = io.BytesIO()
        _df.to_excel(_buf, index=False, engine="openpyxl")
        return base + ".xlsx", _buf.getvalue()
    return fname, fbytes


def _append_to_template(df: pd.DataFrame, template_bytes: bytes, start_row: int = 9) -> bytes:
    """Tulis df ke po_template.xlsx mulai start_row, concat ke bawah row demi row."""
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    ws = wb.active

    header_row_idx = start_row - 1
    tpl_headers = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row_idx, column=col_idx).value
        if val is not None:
            tpl_headers[str(val).strip().lower()] = col_idx

    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            cell.value = None

    df_cols = [str(c).strip() for c in df.columns]
    for r_offset, row_data in enumerate(df.itertuples(index=False)):
        cur_row = start_row + r_offset
        for c_idx, col_name in enumerate(df_cols):
            tpl_col = tpl_headers.get(col_name.lower())
            if tpl_col is None:
                continue
            val = row_data[c_idx]
            ws.cell(row=cur_row, column=tpl_col).value = (
                None if (val is None or str(val).strip() in ("", "nan")) else val
            )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Sheet1") -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl', datetime_format='YYYY-MM-DD HH:mm:SS') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        for col_cells in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col_cells)
            header_len = len(str(col_cells[0].value or ""))
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(max_len, header_len) + 3, 50)
    return buf.getvalue()


def create_zip_of_files(file_dict: dict) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for fname, data in file_dict.items():
            zf.writestr(fname, data)
    return buf.getvalue()


def split_dataframe(df: pd.DataFrame, max_rows: int = 7500) -> list:
    if len(df) <= max_rows:
        return [df]
    n_chunks = math.ceil(len(df) / max_rows)
    return [df.iloc[i * max_rows:(i + 1) * max_rows].reset_index(drop=True) for i in range(n_chunks)]


def split_by_po_groups(df: pd.DataFrame, po_column: str | None, max_rows: int = 7500) -> list:
    if po_column is None or po_column not in df.columns:
        return split_dataframe(df, max_rows)
    if len(df) <= max_rows:
        return [df.reset_index(drop=True)]

    po_order = [v for v in df[po_column].drop_duplicates().tolist() if v == v and v is not None]
    po_sizes = df.groupby(po_column, sort=False).size()

    chunks = []
    current_pos: list = []
    current_size = 0

    for po_val in po_order:
        po_size = int(po_sizes.get(po_val, 0))
        if po_size == 0:
            continue
        if current_pos and current_size + po_size > max_rows:
            mask = df[po_column].isin(current_pos)
            chunks.append(df[mask].reset_index(drop=True))
            current_pos = []
            current_size = 0
        current_pos.append(po_val)
        current_size += po_size

    if current_pos:
        mask = df[po_column].isin(current_pos)
        chunks.append(df[mask].reset_index(drop=True))

    return chunks if chunks else [df.reset_index(drop=True)]


def gsheet_to_csv_url(url: str):
    match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', url)
    if not match:
        return None, None
    sheet_id = match.group(1)
    gid_match = re.search(r'gid=(\d+)', url)
    gid = gid_match.group(1) if gid_match else '0'
    csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
    return csv_url, sheet_id


def attach_po_counts(df: pd.DataFrame, po_column: str) -> tuple:
    grouped = df.groupby(po_column).size().reset_index(name='count')
    grouped = grouped.sort_values('count', ascending=False).reset_index(drop=True)
    return df.merge(grouped, on=po_column, how='left'), grouped


def numeric_coerce(df: pd.DataFrame) -> pd.DataFrame:
    for col in df.columns:
        try:
            numeric_vals = pd.to_numeric(df[col], errors='coerce')
            if numeric_vals.notna().sum() / max(df[col].notna().sum(), 1) > 0.8:
                df[col] = numeric_vals
        except Exception:
            pass
    return df


_REMARK_STYLES = [
    ('reject with suggestion', '#FFF3CD', '#856404'),
    ('reject (stop by steve',  '#F8D7DA', '#721C24'),
    ('reject',                 '#F8D7DA', '#721C24'),
    ('proceed with suggestion','#D4EDDA', '#155724'),
    ('proceed',                '#D4EDDA', '#155724'),
    ('additional suggestion',  '#D1ECF1', '#0C5460'),
]


def df_to_image_bytes(df: pd.DataFrame, title: str = "") -> bytes:
    if not MATPLOTLIB_OK:
        raise RuntimeError("matplotlib tidak tersedia.")
    n_rows, n_cols = df.shape
    col_width = 2.0
    fig_w = max(16, n_cols * col_width)
    fig_h = max(2.5, n_rows * 0.38 + 1.6)

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.axis('off')
    if title:
        ax.set_title(title, fontsize=11, fontweight='bold', color='#1a1a2e', pad=10)

    cell_text = [[str(v) if pd.notna(v) else '' for v in row] for row in df.values]
    tbl = ax.table(
        cellText=cell_text,
        colLabels=df.columns.tolist(),
        loc='center',
        cellLoc='left',
    )
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(7.5)
    tbl.auto_set_column_width(col=list(range(n_cols)))

    for j in range(n_cols):
        tbl[0, j].set_facecolor('#1a1a2e')
        tbl[0, j].set_text_props(color='white', fontweight='bold')

    remark_idx = df.columns.tolist().index('Remark') if 'Remark' in df.columns else None

    for i in range(1, n_rows + 1):
        row_bg = '#EBF5FB' if i % 2 == 0 else '#FFFFFF'
        for j in range(n_cols):
            tbl[i, j].set_facecolor(row_bg)
            tbl[i, j].set_text_props(color='#1a1a2e')
        if remark_idx is not None:
            val = str(df.iloc[i - 1, remark_idx]).lower().strip()
            for keyword, bg, fg in _REMARK_STYLES:
                if keyword in val:
                    tbl[i, remark_idx].set_facecolor(bg)
                    tbl[i, remark_idx].set_text_props(color=fg, fontweight='bold')
                    break

    fig.patch.set_facecolor('white')
    buf = io.BytesIO()
    plt.savefig(buf, format='jpg', bbox_inches='tight', dpi=150, facecolor='white')
    buf.seek(0)
    plt.close(fig)
    return buf.getvalue()


def validate_po_template(df: pd.DataFrame) -> tuple[bool, list]:
    missing = [c for c in PO_TEMPLATE_COLS if c not in df.columns]
    return len(missing) == 0, missing


def _get_sheet_names(file_bytes: bytes, engine: str) -> list[str]:
    """Kembalikan daftar sheet yang visible/unhide saja."""
    try:
        if engine == 'xlrd':
            import xlrd
            book = xlrd.open_workbook(file_contents=file_bytes)
            if hasattr(book, 'sheet_visibility'):
                return [book.sheet_name(i) for i in range(book.nsheets)
                        if book.sheet_visibility[i] == 0]
            else:
                return [book.sheet_name(i) for i in range(book.nsheets)
                        if getattr(book.sheet_by_index(i), 'visibility', 0) == 0]
        else:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            sheets = [ws.title for ws in wb.worksheets if ws.sheet_state == 'visible']
            wb.close()
            return sheets
    except Exception:
        return []


def _excel_engine(fname: str) -> str:
    if fname.lower().endswith('.xls'):
        import importlib.util
        if importlib.util.find_spec('xlrd') is None:
            raise ImportError(
                "Library `xlrd` tidak ditemukan di environment ini. "
                "Jalankan `pip install xlrd>=2.0.1` di terminal yang sama dengan Streamlit, "
                "atau konversi file .xls ke .xlsx terlebih dahulu."
            )
        return 'xlrd'
    return 'openpyxl'


def detect_header_row(file_bytes: bytes, fname: str = "", max_scan: int = 15, sheet_name=0) -> int:
    engine = _excel_engine(fname)
    df_raw = pd.read_excel(
        io.BytesIO(file_bytes), sheet_name=sheet_name, header=None,
        engine=engine, dtype=str, nrows=max_scan
    )
    best_row, best_score = 0, -1
    for i in range(len(df_raw)):
        vals = [str(v).strip() for v in df_raw.iloc[i].values if pd.notna(v) and str(v).strip()]
        text_count = sum(
            1 for v in vals
            if not v.replace('.', '', 1).replace(',', '', 1).lstrip('-').isdigit()
        )
        score = text_count * 10 + len(vals)
        if score > best_score:
            best_score = score
            best_row = i
    return best_row


# =========================
# Sidebar
# =========================
with st.sidebar:
    st.markdown(
        f'<div style="text-align:center;padding:0.5rem 0 0.2rem;">'
        f'<img src="{LOGO_URL}" style="max-width:200px;height:auto;display:inline-block;background:transparent;" />'
        f'</div>',
        unsafe_allow_html=True
    )
    if 'page' not in st.session_state:
        st.session_state['page'] = 'extractor'

    st.markdown(
        "<div style='padding:0 0.4rem 0.3rem;;text-align:center;font-size:1rem;font-weight:700;"
        "letter-spacing:2.5px;color:rgba(255,255,255,0.35);text-transform:uppercase;'>MENU</div>",
        unsafe_allow_html=True
    )

    if st.button("Data Extractor", use_container_width=True, key="nav_extractor"):
        st.session_state['page'] = 'extractor'
        st.rerun()
    if st.button("PO Changer", use_container_width=True, key="nav_po"):
        st.session_state['page'] = 'po_changer'
        st.rerun()

    st.link_button(
        "PO SIMULATOR",
        DASHBOARD_URL_DEFAULT,
        use_container_width=True)

    st.divider()
    st.markdown(
        "<div style='padding:0 0.6rem;font-size:0.6rem;font-weight:700;letter-spacing:2px;"
        "color:rgba(255,255,255,0.35);text-transform:uppercase;margin-bottom:0.6rem;'>CONFIGURATION</div>",
        unsafe_allow_html=True
    )
    max_rows_per_file = st.number_input(
        "Max rows per file",
        min_value=100, max_value=100000, value=7500, step=500,
        help="Batas baris per file output"
    )
    po_col_override = st.text_input(
        "Nama kolom PO (opsional)",
        placeholder="Auto-detect",
        help="Kosongkan untuk auto-detect kolom PO"
    )

    st.divider()
    st.divider()
    if st.button("Logout", use_container_width=True):
        st.session_state["authenticated"] = False
        for k in ("raw_df", "data_source", "source_type"):
            st.session_state.pop(k, None)
        st.rerun()

    st.markdown(
        "<div style='text-align:center;color:rgba(255,255,255,0.25);font-size:0.62rem;margin-top:1rem;'>DataFlow v1.0 · Glad2Glow"
        "<div style='text-align:center;color:rgba(255,255,255,0.25);font-size:0.62rem;margin-top:0.1rem;'> Created by shaltsa </div>",
        unsafe_allow_html=True
    )

# =========================
# PO Changer Page
# =========================
if st.session_state.get('page') == 'po_changer':
    st.markdown("""
    <div class="hero-wrap">
        <div class="hero-tag">✦ PO Management</div>
        <div class="hero-title">PO Changer</div>
    </div>
    """, unsafe_allow_html=True)

    st.divider()

    _INVALID_QTY = {"-", "null", "none", "", "0", "0.0"}

    def _read_one(fname: str, fbytes: bytes, sheet_name=0):
        ext = fname.rsplit(".", 1)[-1].lower()
        if ext == "csv":
            raw_preview = pd.read_csv(io.BytesIO(fbytes), header=None, dtype=str, nrows=15, encoding_errors="replace")
            best_row, best_score = 0, -1
            for i in range(len(raw_preview)):
                vals = [str(v).strip() for v in raw_preview.iloc[i].values if pd.notna(v) and str(v).strip()]
                text_count = sum(1 for v in vals if not v.replace('.','',1).replace(',','',1).lstrip('-').isdigit())
                score = text_count * 10 + len(vals)
                if score > best_score:
                    best_score, best_row = score, i
            df = pd.read_csv(io.BytesIO(fbytes), header=best_row, dtype=str, encoding_errors="replace")
            return df, best_row
        else:
            engine = _excel_engine(fname)
            hrow = detect_header_row(fbytes, fname, sheet_name=sheet_name)
            df = pd.read_excel(io.BytesIO(fbytes), sheet_name=sheet_name, header=hrow, engine=engine, dtype=str)
            return df, hrow

    def _parse_idx(rng: str):
        rng = rng.strip()
        if not rng or rng == ":":
            return None, None
        if ":" not in rng:
            n = int(rng)
            return n, n + 1
        left, right = rng.split(":", 1)
        start = int(left.strip()) if left.strip() else None
        end   = int(right.strip()) if right.strip() else None
        stop  = (end + 1) if end is not None else None
        return start, stop

    def _apply_range_r(df: pd.DataFrame, rng_r: str, rng_c: str) -> pd.DataFrame:
        if rng_r.strip():
            try:
                rs, re = _parse_idx(rng_r)
                df = df.iloc[rs:re]
            except Exception:
                pass
        if rng_c.strip():
            try:
                cs, ce = _parse_idx(rng_c)
                df = df.iloc[:, cs:ce]
            except Exception:
                pass
        return df

    folder_files = st.file_uploader(
        "📁 Upload File PO (.xlsx / .xls / .csv / .zip)",
        type=["xlsx", "xls", "csv", "zip"],
        accept_multiple_files=True,
        key="po_folder",
    )

    if folder_files:
        raw_entries = []
        _converted_names = []
        for uf in folder_files:
            fb = uf.read()
            if uf.name.lower().endswith(".zip"):
                with zipfile.ZipFile(io.BytesIO(fb)) as zf:
                    for zname in zf.namelist():
                        ext = zname.rsplit(".", 1)[-1].lower() if "." in zname else ""
                        if ext in ("xlsx", "xls", "csv"):
                            _new_name, _new_bytes = _convert_to_xlsx(zname, zf.read(zname))
                            raw_entries.append((_new_name, _new_bytes))
                            if ext != "xlsx":
                                _converted_names.append(f"{zname} → {_new_name}")
            else:
                _new_name, _new_bytes = _convert_to_xlsx(uf.name, fb)
                raw_entries.append((_new_name, _new_bytes))
                if uf.name.rsplit(".", 1)[-1].lower() != "xlsx":
                    _converted_names.append(f"{uf.name} → {_new_name}")

        if _converted_names:
            st.caption("🔄 Auto-convert: " + "  ·  ".join(_converted_names))

        st.markdown("""
        <div class="pipeline-step active">
            <span class="step-number" style="background:#3B2B5E;">🗂</span>
            <strong>Konfigurasi per File</strong>
        </div>
        """, unsafe_allow_html=True)

        parsed = []
        for idx, (fname, fbytes) in enumerate(raw_entries):
            with st.container(border=True):
                hc1, hc2 = st.columns([2, 1])
                with hc1:
                    st.markdown(f"**#{idx+1} &nbsp; {fname}**")
                with hc2:
                    try:
                        _wb = openpyxl.load_workbook(io.BytesIO(fbytes), data_only=True)
                        sheets = [ws.title for ws in _wb.worksheets if ws.sheet_state == 'visible']
                        _wb.close()
                    except Exception:
                        sheets = []
                    if len(sheets) > 1:
                        sheet_sel = st.selectbox(
                            "Sheet:", options=sheets,
                            key=f"fs_{idx}_{fname}",
                            label_visibility="collapsed",
                        )
                    elif sheets:
                        sheet_sel = sheets[0]
                        st.caption(f"📄 `{sheets[0]}`")
                    else:
                        sheet_sel = 0
                        st.caption("⚠️ Tidak ada sheet visible")

                try:
                    df_f, hrow = _read_one(fname, fbytes, sheet_sel)
                    df_f = df_f.loc[:, ~df_f.columns.astype(str).str.startswith('Unnamed')].dropna(how='all')
                    df_f = df_f[df_f.apply(lambda r: r.astype(str).str.strip().ne('').any(), axis=1)]
                    parse_err = None
                except Exception as e:
                    df_f, hrow, parse_err = None, "-", str(e)

                if df_f is not None:
                    df_f.columns = [str(c).strip().upper() for c in df_f.columns]

                    _qty_col = next((c for c in df_f.columns if c.strip().upper() in ("QTY", "QUANTITY")), None)
                    _before_qty = len(df_f)
                    if _qty_col:
                        def _qty_valid(v):
                            s = str(v).strip().lower()
                            if s in _INVALID_QTY:
                                return False
                            try:
                                return float(s.replace(",", ".")) > 0
                            except ValueError:
                                return False
                        df_f = df_f[df_f[_qty_col].apply(_qty_valid)].reset_index(drop=True)
                        _qty_removed = _before_qty - len(df_f)
                        if _qty_removed:
                            st.caption(f"🗑 {_qty_removed:,} baris dibuang (QTY tidak valid — kolom **{_qty_col}**)")

                    _dist_cols = [c for c in df_f.columns if "DISTRIBUTOR" in c.upper()]
                    _has_dist = len(_dist_cols) > 0

                    dc1, dc2 = st.columns([1, 2])
                    with dc1:
                        st.caption("Distributor" + (" *(sudah ada di kolom)*" if _has_dist else ""))
                    with dc2:
                        dist_val = st.selectbox(
                            "Distributor",
                            options=["(Pilih)"] + CUSTOMER_NAMES,
                            key=f"dist_{idx}_{fname}",
                            label_visibility="collapsed",
                        )

                    _df_preview = df_f.copy()
                    if not _has_dist and dist_val not in ("", "(Pilih)"):
                        _df_preview.insert(0, "DISTRIBUTOR", dist_val)
                    with st.expander(f"👁 Lihat isi  ·  header row {hrow}  ·  {len(_df_preview)} baris  ·  {_df_preview.shape[1]} kolom  ·  indeks 0–{_df_preview.shape[1]-1}", expanded=False):
                        st.dataframe(_df_preview.iloc[:,:6].reset_index(drop=True), use_container_width=True)

                    rc1, rc2 = st.columns(2)
                    with rc1:
                        row_rng = st.text_input("Row Range", value="",
                                                key=f"row_{idx}_{fname}",
                                                placeholder="5:10 | 5: | :10 | 5")
                    with rc2:
                        col_rng = st.text_input("Column Range", value="",
                                                key=f"col_{idx}_{fname}",
                                                placeholder="0:3 | 2: | :5 | 2")

                    if row_rng.strip() or col_rng.strip():
                        _pv = _apply_range_r(df_f.copy(), row_rng, col_rng)
                        st.caption(f"📍 Preview rentang: {len(_pv)} baris × {_pv.shape[1]} kolom")
                        st.dataframe(_pv, use_container_width=True, hide_index=True)

                    parsed.append({
                        "name": fname, "df": df_f,
                        "row_rng": row_rng, "col_rng": col_rng,
                        "dist_val": dist_val, "has_dist": _has_dist,
                        "error": None,
                    })
                else:
                    st.error(f"❌ {parse_err}")
                    parsed.append({"name": fname, "df": None, "row_rng": "", "col_rng": "",
                                   "dist_val": "", "has_dist": False, "error": parse_err})

        ready = [p for p in parsed if p["df"] is not None]
        st.divider()
        if st.button("➕ Gabungkan Semua", disabled=not ready,
                     use_container_width=True, key="folder_concat_btn"):
            _frames = []
            for p in ready:
                _df = _apply_range_r(p["df"].copy(), p["row_rng"], p["col_rng"])
                if not p["has_dist"] and p["dist_val"] not in ("", "(Pilih)"):
                    _df.insert(0, "DISTRIBUTOR", p["dist_val"])
                _frames.append(_df)

            try:
                _tpl_bytes = _fetch_template_bytes(TEMPLATE_DRIVE_URL)
                _twb = openpyxl.load_workbook(io.BytesIO(_tpl_bytes), data_only=True)
                _tws = _twb.active
                _ref_cols = [
                    str(_tws.cell(row=8, column=c).value).strip().upper()
                    for c in range(1, _tws.max_column + 1)
                    if _tws.cell(row=8, column=c).value is not None
                ]
                _twb.close()
            except Exception:
                _tpl_bytes = None
                _ref_cols = _frames[0].columns.tolist()

            _frames = [f.reindex(columns=_ref_cols) for f in _frames]
            combined_df = pd.concat(_frames, ignore_index=True)
            st.session_state["folder_result"] = {"df": combined_df, "tpl_bytes": _tpl_bytes}
            st.rerun()

        _res = st.session_state.get("folder_result")
        if _res is not None:
            combined_df = _res["df"]
            _dist_label = st.session_state.get("po_distributor") or ""
            if not _dist_label or _dist_label == "(Pilih Distributor)":
                _dist_label = "Combined"

            st.success(f"✅ **{len(combined_df):,}** baris · {combined_df.shape[1]} kolom dari {len(ready)} file")
            for col in ['QTY', 'DPP', 'TOTAL PRICE']:
                if col in combined_df.columns:
                    combined_df[col] = pd.to_numeric(combined_df[col], errors='coerce')
            st.dataframe(combined_df.head(10), use_container_width=True, hide_index=True)

            _ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            _default_fname = f"PO_{_dist_label}_{_ts}"
            _fname_input = st.text_input(
                "Nama file (opsional)",
                value=_default_fname,
                placeholder=_default_fname,
                key="folder_dl_fname",
            )
            _fname_final = (_fname_input.strip() or _default_fname).removesuffix(".xlsx") + ".xlsx"

            _tpl_bytes_dl = _res.get("tpl_bytes")
            if _tpl_bytes_dl is None:
                try:
                    _tpl_bytes_dl = _fetch_template_bytes(TEMPLATE_DRIVE_URL)
                except Exception as _e:
                    st.error(f"❌ Gagal mengambil template dari Drive: {_e}")
                    _tpl_bytes_dl = None
            if _tpl_bytes_dl is not None:
                st.download_button(
                    label="Download Template PO",
                    data=_append_to_template(combined_df, _tpl_bytes_dl, start_row=9),
                    file_name=_fname_final,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

    st.markdown("""
    <div class="pipeline-step active">
        <span class="step-number">1</span>
        <strong>Upload File PO</strong>
        <span class="badge badge-info" style="margin-left:0.8rem;">Harus sesuai template</span>
    </div>
    """, unsafe_allow_html=True)
    po_file = st.file_uploader("Upload file PO (.xlsx / .xls)", type=["xlsx", "xls"])

    if po_file is not None:
        file_bytes = po_file.read()
        _po_engine = _excel_engine(po_file.name)

        po_sheets = _get_sheet_names(file_bytes, _po_engine)
        if len(po_sheets) > 1:
            po_selected_sheet = st.selectbox(
                f"Sheet tersedia ({len(po_sheets)} sheet):",
                options=po_sheets,
                key="po_sheet_sel",
            )
        elif po_sheets:
            po_selected_sheet = po_sheets[0]
            st.caption(f"Sheet: **{po_selected_sheet}**")
        else:
            po_selected_sheet = 0

        try:
            raw_preview = pd.read_excel(
                io.BytesIO(file_bytes), sheet_name=po_selected_sheet,
                header=None, engine=_po_engine, dtype=str, nrows=5
            )
        except Exception as e:
            st.error(f"❌ Gagal membaca file: {e}")
            st.stop()

        raw_preview.index = [f"Row {i}" for i in raw_preview.index]

        try:
            po_df = pd.read_excel(
                io.BytesIO(file_bytes), sheet_name=po_selected_sheet,
                header=int(0), engine=_po_engine, dtype=str
            )
            po_df = po_df.loc[:, ~po_df.columns.str.startswith('Unnamed')]
            po_df = po_df.dropna(how='all')
        except Exception as e:
            st.error(f"❌ Gagal membaca file: {e}")
            st.stop()

        st.success("✅ Format file valid — sesuai template PO.")

        next_step = 2
        st.markdown(f"""
        <div class="pipeline-step active">
            <span class="step-number">{next_step}</span>
            <strong>Preview Data — Top 10 WOI</strong>
        </div>
        """, unsafe_allow_html=True)

        woi_col = next((c for c in po_df.columns if 'woi' in c.lower() and 'stock' in c.lower()), None)
        if woi_col is None:
            woi_col = next((c for c in po_df.columns if 'woi' in c.lower()), 'Current WOI')

        preview_df = po_df.copy()
        preview_df = preview_df[preview_df['Remark'].str.contains('Reject', na=False)]
        preview_df[woi_col] = pd.to_numeric(preview_df[woi_col], errors='coerce')
        top10 = preview_df.nlargest(10, woi_col)[PO_IMG_COLS].reset_index(drop=True)

        styled = top10.style.set_properties(**{
            'background-color': '#D6EAF8',
            'color': '#1a1a2e',
            'border': '1px solid #AED6F1',
        }).format(na_rep='-')
        st.dataframe(styled, use_container_width=True, hide_index=True)
        st.caption(f"Menampilkan 10 baris dengan nilai **{woi_col}** terbesar · kolom Distributor s/d Remark")

        alloc_col = next((c for c in po_df.columns if 'allocation' in c.lower()), None)
        if alloc_col:
            st.markdown(f"""
            <div class="pipeline-step active">
                <span class="step-number">{next_step + 1}</span>
                <strong>Remaining Allocation (tidak kosong &amp; tidak negatif)</strong>
            </div>
            """, unsafe_allow_html=True)
            _alloc_df = po_df.copy()
            _alloc_df[alloc_col] = pd.to_numeric(_alloc_df[alloc_col], errors='coerce')
            _alloc_df = _alloc_df[_alloc_df[alloc_col].notna() & (_alloc_df[alloc_col] > 0)]
            _show_cols = [c for c in PO_IMG_COLS + [alloc_col] if c in _alloc_df.columns]
            _show_cols = list(dict.fromkeys(_show_cols))
            if _alloc_df.empty:
                st.info("Tidak ada baris dengan allocation tersedia.")
            else:
                st.caption(f"**{len(_alloc_df):,}** baris · kolom **{alloc_col}** ≥ 0")
                st.dataframe(
                    _alloc_df[_show_cols].reset_index(drop=True),
                    use_container_width=True, hide_index=True,
                )

        final_step = next_step + (2 if alloc_col else 1)
        st.markdown(f"""
        <div class="pipeline-step active">
            <span class="step-number">{final_step}</span>
            <strong>Auto Kategorisasi &amp; Download Gambar</strong>
        </div>
        """, unsafe_allow_html=True)

        if not MATPLOTLIB_OK:
            st.error("❌ Library matplotlib tidak tersedia. Tambahkan ke requirements.txt.")
            st.stop()

        img_df = po_df[PO_COLS_copy].copy()
        sc_col     = 'Supply Control'
        remark_col = 'Remark'
        stop_keywords   = ['STOP PO', 'OOS', 'DISCONTINUE', 'UNAVAILABLE']
        sc_col = next(
            (c for c in img_df.columns if 'supply' in c.lower() and 'control' in c.lower()),
            None)
        stop_mask = img_df[sc_col].str.strip().str.upper().isin([k.upper() for k in stop_keywords])

        stop_df = img_df[stop_mask].reset_index(drop=True)
        cat1_col, cat2_col, cat3_col = st.columns(3)

        with cat1_col:
            st.markdown("""
            <div class="metric-card" style="text-align:center;border-left:4px solid #CA6180;margin-bottom:0.6rem;">
                <div style="font-size:1.6rem;">🚫</div>
                <div style="font-weight:700;color:#CA6180;">Product Stop PO</div>
                <div style="color:#A8849A;font-size:0.78rem;">Supply Control: STOP PO / OOS / DISCONTINUE</div>
            </div>""", unsafe_allow_html=True)
            if stop_df.empty:
                st.info("Tidak ada data kategori ini.")
            else:
                st.caption(f"{len(stop_df)} baris ditemukan")
                try:
                    img_bytes = df_to_image_bytes(stop_df, title="Product Stop PO")
                    st.download_button(
                        label="Download product stop po (.png)",
                        data=img_bytes,
                        file_name=f"product_stop_po_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png",
                        mime="image/png",
                        use_container_width=True,
                    )
                except Exception as e:
                    st.error(f"Gagal generate gambar: {e}")

        non_stop_df = img_df[~stop_mask].copy()
        col = non_stop_df[remark_col].str.lower()
        steve_mask = (col.str.contains(
            "reject (stop by steve", na=False, regex=False) |
            col.str.contains("reject (negative allocation)", na=False, regex=False))

        steve_df = non_stop_df[steve_mask].reset_index(drop=True)

        with cat2_col:
            st.markdown("""
            <div class="metric-card" style="text-align:center;border-left:4px solid #FCB7C7;margin-bottom:0.6rem;">
                <div style="font-size:1.6rem;">❌</div>
                <div style="font-weight:700;color:#CA6180;">Reject by Steve</div>
                <div style="color:#A8849A;font-size:0.78rem;">Remark: reject by steve</div>
            </div>""", unsafe_allow_html=True)
            if steve_df.empty:
                st.info("Tidak ada data kategori ini.")
            else:
                st.caption(f"{len(steve_df)} baris ditemukan")
                try:
                    img_bytes = df_to_image_bytes(steve_df, title="Reject by Steve")
                    st.download_button(
                        label="⬇ Download reject by steve (.png)",
                        data=img_bytes,
                        file_name=f"reject_by_steve_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png",
                        mime="image/png",
                        use_container_width=True,
                    )
                except Exception as e:
                    st.error(f"Gagal generate gambar: {e}")

        approval_keywords = ['Reject', 'Reject with suggestion']
        approval_mask = (
            non_stop_df[remark_col]
            .str.strip()
            .str.lower()
            .isin([k.lower() for k in approval_keywords])
            & ~steve_mask)
        approval_df = non_stop_df[approval_mask].reset_index(drop=True)

        with cat3_col:
            st.markdown("""
            <div class="metric-card" style="text-align:center;border-left:4px solid #A84D6A;margin-bottom:0.6rem;">
                <div style="font-size:1.6rem;">⚠️</div>
                <div style="font-weight:700;color:#CA6180;">Products Need Approval</div>
                <div style="color:#A8849A;font-size:0.78rem;">Remark: reject / reject by suggestion</div>
            </div>""", unsafe_allow_html=True)
            if approval_df.empty:
                st.info("Tidak ada data kategori ini.")
            else:
                st.caption(f"{len(approval_df)} baris ditemukan")
                try:
                    img_bytes = df_to_image_bytes(approval_df, title="Products Need Approval")
                    st.download_button(
                        label="Download products need approval (.png)",
                        data=img_bytes,
                        file_name=f"products_need_approval_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png",
                        mime="image/png",
                        use_container_width=True,
                    )
                except Exception as e:
                    st.error(f"Gagal generate gambar: {e}")

        st.markdown("""
        <div class="pipeline-step active">
            <span class="step-number">4</span>
            <strong>Product Code yang Harus Dihapus</strong>
        </div>
        """, unsafe_allow_html=True)

        _sku_col = 'SKU'
        _sku_series = []
        for _sdf in (stop_df, steve_df):
            if _sku_col in _sdf.columns:
                _sku_series.append(_sdf[_sku_col].dropna().astype(str).str.strip())
        _remark_series = []
        if 'Supply Control' in stop_df.columns:
            _remark_series.append(
                stop_df['Supply Control'].dropna().astype(str).str.strip()
            )
        if 'Remark' in steve_df.columns:
            _remark_series.append(
                steve_df['Remark'].dropna().astype(str).str.strip()
            )
        _sku_all = (
            pd.concat(_sku_series).pipe(lambda s: s[s != ''])
            .drop_duplicates().sort_values().reset_index(drop=True)
            if _sku_series else pd.Series(dtype=str)
        )
        _remark_all = (
            pd.concat(_remark_series, ignore_index=True)
            .astype(str)
            .str.strip()
            .loc[lambda s: s.ne('')]
            .drop_duplicates()
            .sort_values()
            .reset_index(drop=True)
            if _remark_series else pd.Series(dtype="string")
        )

        if _sku_all.empty:
            st.info("Tidak ada product code yang perlu dihapus.")
        else:
            _pc1, _pc2 = st.columns([1, 1])
            with _pc1:
                st.dataframe(
                    pd.DataFrame({"Remark/Supply Control": _remark_all}),
                    use_container_width=True, hide_index=True,
                )
            with _pc2:
                with st.expander("📋 Copy SKU"):
                    st.code("\n".join(_sku_all.tolist()), language=None)

    st.markdown("""
     <div class="hero-wrap">
         <div class="hero-tag">✦ Template PO</div>
         <div class="hero-title">Upload Template PO</div>
     </div>
     """, unsafe_allow_html=True)
    st.markdown("""
     <div class="pipeline-step active">
         <span class="step-number">5</span>
         <strong>Upload File Template</strong>
     </div>
     """, unsafe_allow_html=True)
    tpl_files = st.file_uploader("Upload template (.xlsx / .xls)",
                                 type=["xlsx", "xls"],
                                 accept_multiple_files=True,
                                 key="tpl_uploader")

    for _fi, tpl_file in enumerate(tpl_files or []):
        with st.container(border=True):
            st.markdown(f"**#{_fi+1} &nbsp; {tpl_file.name}**")

            tpl_bytes = tpl_file.read()
            _tpl_name, tpl_bytes = _convert_to_xlsx(tpl_file.name, tpl_bytes)
            if tpl_file.name.rsplit(".", 1)[-1].lower() != "xlsx":
                st.caption(f"🔄 Auto-convert: {tpl_file.name} → {_tpl_name}")
            _tpl_engine = "openpyxl"

            tpl_sheets = _get_sheet_names(tpl_bytes, _tpl_engine)
            if not tpl_sheets:
                st.warning("⚠️ Tidak ada sheet visible.")
                continue
            sc1, sc2 = st.columns([2, 1])
            with sc1:
                if len(tpl_sheets) > 1:
                    tpl_selected_sheet = st.selectbox(
                        f"Sheet ({len(tpl_sheets)} visible):",
                        options=tpl_sheets,
                        key=f"tpl_sheet_sel_{_fi}",
                    )
                else:
                    tpl_selected_sheet = tpl_sheets[0]
                    st.caption(f"📄 Sheet: **{tpl_selected_sheet}**")
            with sc2:
                _auto_hrow = detect_header_row(tpl_bytes, _tpl_name, sheet_name=tpl_selected_sheet)
                _hrow_input = st.number_input(
                    "Header row (baris ke-)", min_value=1,
                    value=int(_auto_hrow) + 1, step=1,
                    key=f"tpl_hrow_{_fi}",
                )
            _tpl_hrow = int(_hrow_input) - 1

            try:
                tpl_df = pd.read_excel(io.BytesIO(tpl_bytes), sheet_name=tpl_selected_sheet,
                                       header=_tpl_hrow, engine=_tpl_engine, dtype=str)
                tpl_df = tpl_df.loc[:, ~tpl_df.columns.str.startswith('Unnamed')]
                tpl_df = tpl_df.dropna(how='all').reset_index(drop=True)
            except Exception as e:
                st.error(f"❌ Gagal membaca file: {e}")
                continue

            st.caption(f"**{len(tpl_df):,} baris · {len(tpl_df.columns)} kolom**")
            with st.expander("👁 Preview data", expanded=False):
                st.dataframe(tpl_df, use_container_width=True, hide_index=True)

            qty_col_t = next((c for c in tpl_df.columns
                              if any(k in c.lower() for k in ['qty', 'quantity'])), None)
            sku_col_t = next((c for c in tpl_df.columns
                              if any(k in c.lower() for k in ['sku', 'product code', 'kode', 'code'])), None)

            if qty_col_t and sku_col_t:
                st.markdown("""
                <div class="pipeline-step active">
                    <span class="step-number">2</span>
                    <strong>Kosongkan Quantity per Product Code</strong>
                </div>
                """, unsafe_allow_html=True)

                with st.container(border=True):
                    st.caption(f"SKU: **{sku_col_t}** · Quantity: **{qty_col_t}**")
                    mc1, mc2 = st.columns([2, 1])
                    with mc1:
                        tpl_codes = st.text_area(
                            "Daftar Product Code (satu per baris)",
                            placeholder="SKU001\nSKU-ABC\nPROD123",
                            height=150, key=f"tpl_codes_{_fi}"
                        )
                    with mc2:
                        st.markdown("**Format:**")
                        st.code("SKU001\nSKU-ABC", language=None)
                        tpl_apply = st.button("🗑 Kosongkan Quantity",
                                              use_container_width=True, key=f"tpl_apply_{_fi}")

                    if tpl_apply and tpl_codes.strip():
                        targets = {c.strip() for c in tpl_codes.strip().splitlines() if c.strip()}
                        mask_t = tpl_df[sku_col_t].astype(str).str.strip().isin(targets)

                        if mask_t.sum() == 0:
                            st.warning("⚠️ Tidak ada Product Code yang cocok.")
                        else:
                            wb = openpyxl.load_workbook(io.BytesIO(tpl_bytes))
                            ws = next((s for s in wb.worksheets
                                       if s.title == tpl_selected_sheet), wb.active)
                            header_excel_row = _tpl_hrow + 1
                            headers_ws = {
                                ws.cell(row=header_excel_row, column=c).value: c
                                for c in range(1, ws.max_column + 1)
                            }
                            sku_col_idx = headers_ws.get(sku_col_t)
                            qty_col_idx = headers_ws.get(qty_col_t)

                            if sku_col_idx and qty_col_idx:
                                cleared = 0
                                for row in ws.iter_rows(min_row=header_excel_row + 1,
                                                        max_row=ws.max_row):
                                    if str(row[sku_col_idx - 1].value or "").strip() in targets:
                                        row[qty_col_idx - 1].value = None
                                        cleared += 1
                                out_buf = io.BytesIO()
                                wb.save(out_buf)
                                st.session_state[f"tpl_out_{_fi}"] = {
                                    "buf": out_buf.getvalue(), "mask": mask_t,
                                    "sku": sku_col_t, "qty": qty_col_t,
                                    "cleared": cleared, "df": tpl_df,
                                }
                            else:
                                st.warning("⚠️ Kolom tidak ditemukan di worksheet asli.")

                _res_t = st.session_state.get(f"tpl_out_{_fi}")
                if _res_t:
                    st.success(f"✅ Quantity dikosongkan untuk **{_res_t['cleared']}** baris.")
                    with st.expander("👁 Lihat baris yang diubah", expanded=False):
                        st.dataframe(
                            _res_t["df"][_res_t["mask"]][[_res_t["sku"], _res_t["qty"]]].reset_index(drop=True),
                            use_container_width=True, hide_index=True,
                        )
                    customer_name = st.selectbox(
                        "Distributor",
                        options=["(Pilih)"] + CUSTOMER_NAMES,
                        key=f"tpl_cust_{_fi}",
                        label_visibility="collapsed",
                    )
                    file_label = re.sub(r'[\\/*?:"<>|]', "", (customer_name or "").strip()) or "Unnamed_Customer"
                    st.download_button(
                        label="Download Hasil Modifikasi (.xlsx)",
                        data=_res_t["buf"],
                        file_name=f"Form PO {file_label}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key=f"tpl_dl_{_fi}",
                    )
            else:
                st.info("Kolom SKU / QTY tidak terdeteksi. Periksa header row.")

    st.stop()

# =========================
# Data Extractor Page
# =========================
st.markdown("""
    <div class="hero-wrap">
        <div class="hero-tag">✦ Raw Data Sales</div>
        <div class="hero-title">Data Extractor</div>
        <div class="hero-sub"></div>
    </div>
    """, unsafe_allow_html=True)
st.divider()

st.markdown("""
<div class="pipeline-step active">
    <span class="step-number">1</span>
    <strong>Pilih Sumber Data</strong>
    <span class="badge badge-info" style="margin-left:0.8rem;">REQUIRED</span>
</div>
""", unsafe_allow_html=True)
if st.session_state.get('page') == 'extractor':
    tab1, tab2 = st.tabs(["🔗 Google Spreadsheet", "📂 Upload File"])

    with tab1:
        st.session_state["extractor"] = "google_sheet"

        gsheet_url = st.text_input(
            "Google Spreadsheet URL",
            placeholder="https://docs.google.com/spreadsheets/d/...",
            help="Pastikan sheet sudah public (view access)",
            label_visibility="collapsed"
        )

        load_gsheet = st.button("Load Data")

        if load_gsheet:
            if not gsheet_url.strip():
                st.warning("Masukkan link Google Spreadsheet dulu.")
            else:
                csv_url, sheet_id = gsheet_to_csv_url(gsheet_url)

                if csv_url is None:
                    st.error("Link tidak valid.")
                else:
                    with st.spinner("Loading data..."):
                        df_loaded = pd.read_csv(csv_url, dtype=str)
                        df_loaded = numeric_coerce(df_loaded)

                        st.session_state['raw_df'] = df_loaded
                        st.session_state['data_source'] = f"Google Sheet ({sheet_id[:10]}...)"
                        st.session_state['source_type'] = 'GSHEET'

                        st.success(f"Loaded {len(df_loaded):,} rows")

    with tab2:
        st.session_state["extractor"] = "upload_file"

        uploaded_file = st.file_uploader(
            "Upload file data",
            type=["csv", "xlsx", "xls"],
            help="Format yang didukung: CSV, Excel (.xlsx, .xls)",
            label_visibility="collapsed"
        )

        if uploaded_file is not None:
            try:
                file_name = uploaded_file.name.lower()

                with st.spinner("🌸 Memproses file..."):
                    if file_name.endswith(".csv"):
                        df_loaded = pd.read_csv(uploaded_file, dtype=str)
                    elif file_name.endswith(".xlsx") or file_name.endswith(".xls"):
                        df_loaded = pd.read_excel(uploaded_file, dtype=str)
                    else:
                        st.error("❌ Format file tidak didukung.")
                        df_loaded = None

                    if df_loaded is not None:
                        df_loaded = numeric_coerce(df_loaded)

                        st.session_state['raw_df'] = df_loaded
                        st.session_state['data_source'] = f"Upload File ({uploaded_file.name})"
                        st.session_state['source_type'] = 'UPLOAD'

                        st.success(f"✅ File berhasil diproses — {len(df_loaded):,} baris.")

            except Exception as e:
                st.error(f"❌ Gagal membaca file: {e}")

if 'raw_df' not in st.session_state or st.session_state['raw_df'] is None:
    st.markdown("<br>", unsafe_allow_html=True)
    ec1, ec2, ec3 = st.columns(3)
    with ec1:
        st.markdown("""
        <div class="feat-card">
            <div class="feat-icon">🕐</div>
            <div class="feat-title">Auto Timestamp</div>
            <div class="feat-desc">
                Kolom <code style="background:#000000;padding:1px 5px;border-radius:4px;
                font-size:0.8rem;">waktu_extract</code> otomatis ditambahkan ke setiap dataset
            </div>
        </div>""", unsafe_allow_html=True)
    with ec2:
        st.markdown("""
        <div class="feat-card">
            <div class="feat-icon">✨</div>
            <div class="feat-title">Smart Filter</div>
            <div class="feat-desc">
                Auto-detect kolom tanggal dan filter berdasarkan date range yang dipilih
            </div>
        </div>""", unsafe_allow_html=True)
    with ec3:
        st.markdown("""
        <div class="feat-card">
            <div class="feat-icon">📦</div>
            <div class="feat-title">Group &amp; Split</div>
            <div class="feat-desc">
                Group by PO Number, agregat count, dan split otomatis maks 7.500 baris per file
            </div>
        </div>""", unsafe_allow_html=True)

    st.stop()

raw_df = st.session_state['raw_df'].copy()
source_type = st.session_state.get('source_type', 'FILE')

extract_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

col1, col2, col3, col4 = st.columns(4)
with col1:
    st.markdown(f"""
    <div class="metric-card metric-rose">
        <div class="metric-label">Total Rows</div>
        <div class="metric-value">{len(raw_df):,}</div>
    </div>""", unsafe_allow_html=True)
with col2:
    raw_df['NETTO'] = pd.to_numeric(raw_df['NETTO'], errors='coerce')
    total_netto = raw_df['NETTO'].sum()
    st.markdown(f"""
    <div class="metric-card metric-pink">
        <div class="metric-label">Total NETTO Distributor</div>
        <div class="metric-value">{
            f"{total_netto:,.0f}".replace(",", ".")
            }</div>
    </div>""", unsafe_allow_html=True)
with col3:
    st.markdown(f"""
    <div class="metric-card metric-muted">
        <div class="metric-label">Source Type</div>
        <div class="metric-value" style="font-size:1.1rem;">{source_type}</div>
    </div>""", unsafe_allow_html=True)
with col4:
    st.markdown(f"""
    <div class="metric-card metric-blush">
        <div class="metric-label">Extract Time</div>
        <div class="metric-value" style="font-size:0.88rem;">{extract_time}</div>
    </div>""", unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

with st.expander('📊 Raw Data Preview', expanded=False):
    st.dataframe(raw_df.head(50), use_container_width=True, height=280)

st.divider()

st.markdown("""
<div class="pipeline-step active">
    <span class="step-number">3</span>
    <strong>Schema Mapping — Format ST Data Template</strong>
</div>
""", unsafe_allow_html=True)

OUTPUT_MAPPING = [
    ("*Customer Code",         ["brand", "customer code", "kode customer"],      True),
    ("Customer Name",          ["customer name", "nama customer"],                False),
    ("*Customer Branch Code",  ["branch code", "customer branch code", "kdsls"],  True),
    ("Customer Branch Name",   ["branch name", "customer branch name", "branch"], False),
    ("Customer Address",       ["address", "alamat", "customer address"],         False),
    ("*PO Date",               ["po date", "tanggal", "tgl", "date"],             True),
    ("*PO Number",             ["po number", "nofaktur", "no faktur", "po no"],   True),
    ("*Customer Store Code",   ["store code", "customer store code", "kdoutlet"], True),
    ("Customer Store Name",    ["store name", "customer store name", "nama outlet", "outlet"], False),
    ("*Customer SKU Code",     ["sku code", "customer sku code", "pcode"],        True),
    ("Customer SKU Name",      ["sku name", "customer sku name", "nama barang"],  False),
    ("*Qty",                   ["qty", "quantity", "jumlah"],                     True),
]

src_options = ["(Kosongkan)"] + raw_df.columns.tolist()


def _auto_detect(hints, cols):
    clean_cols = [c.lower().lstrip("*").strip() for c in cols]
    for hint in hints:
        for i, cc in enumerate(clean_cols):
            if hint in cc or cc in hint:
                return cols[i]
    return None


with st.container(border=True):
    st.caption("Petakan kolom sumber ke kolom output ST Data Template. 🔴 = wajib")

    b1, b2 = st.columns([1, 1])
    with b1:
        _a = _auto_detect(OUTPUT_MAPPING[0][1], raw_df.columns.tolist())
        _i = src_options.index(_a) if _a in src_options else 0
        st.selectbox("🔴 *Customer Code ← sumber", src_options, index=_i, key="sm_0")
    with b2:
        cust_filter = st.multiselect(
            "🔍 Filter Customer (opsional — filter baris berdasarkan nilai kolom ini)",
            options=CUSTOMER_NAMES,
            placeholder="Ketik nama customer untuk mencari...",
            key="sm_cust_filter",
        )

    mc1, mc2 = st.columns(2)
    for _i, (_out, _hints, _req) in enumerate(OUTPUT_MAPPING[1:], start=1):
        _a = _auto_detect(_hints, raw_df.columns.tolist())
        _idx = src_options.index(_a) if _a in src_options else 0
        _lbl = ("🔴 " if _req else "") + _out + " ← sumber"
        with (mc1 if _i % 2 == 1 else mc2):
            st.selectbox(_lbl, src_options, index=_idx, key=f"sm_{_i}")

    apply_map = st.button("▶ Terapkan Mapping", use_container_width=True, key="apply_schema_map")

if apply_map:
    _rows = {}
    for _i, (_out, _hints, _req) in enumerate(OUTPUT_MAPPING):
        _src = st.session_state.get(f"sm_{_i}")
        if _src and _src != "(Kosongkan)" and _src in raw_df.columns:
            _rows[_out] = raw_df[_src].values
        else:
            _rows[_out] = [""] * len(raw_df)
    _mdf = pd.DataFrame(_rows)
    _cf = st.session_state.get("sm_cust_filter") or []
    if _cf:
        _src0 = st.session_state.get("sm_0")
        if _src0 and _src0 != "(Kosongkan)" and _src0 in raw_df.columns:
            _mdf = _mdf[_mdf["*Customer Code"].astype(str).isin(_cf)].reset_index(drop=True)
    _dist = st.session_state.get("sm_distributor") or ""
    if _dist and _dist != "(Pilih Distributor)":
        _mdf.insert(0, "Distributor", _dist)
    st.session_state["mapped_df"] = _mdf

if "mapped_df" in st.session_state:
    work_df = st.session_state["mapped_df"]
    st.success(f"✅ Mapping diterapkan — {work_df.shape[0]:,} baris · {work_df.shape[1]} kolom output")
    st.dataframe(work_df.head(20), use_container_width=True, hide_index=True)
else:
    work_df = raw_df

st.divider()

st.markdown("""
<div class="pipeline-step active">
    <span class="step-number">5</span>
    <strong>Group by PO Number &amp; Count</strong>
</div>
""", unsafe_allow_html=True)

po_column = None
if po_col_override and po_col_override.strip():
    if po_col_override.strip() in work_df.columns:
        po_column = po_col_override.strip()
    else:
        st.error(f"❌ Kolom '{po_col_override}' tidak ditemukan dalam data!")

if po_column is None:
    po_candidates = [c for c in work_df.columns
                     if any(k in c.lower().lstrip("*").strip()
                            for k in ["po number", "nofaktur", "po num", "po no"])]
    if not po_candidates:
        po_candidates = [c for c in work_df.columns if "po" in c.lower()]
    if po_candidates:
        po_column = po_candidates[0]

if po_column:
    st.success(f"🌸 Kolom PO terdeteksi: **{po_column}**")
    processed_df, po_grouped = attach_po_counts(work_df, po_column)

    gcol1, gcol2, gcol3 = st.columns(3)
    with gcol1:
        st.markdown(f"""
        <div class="metric-card metric-rose">
            <div class="metric-label">Unique PO Numbers</div>
            <div class="metric-value">{len(po_grouped):,}</div>
        </div>""", unsafe_allow_html=True)
    with gcol2:
        st.markdown(f"""
        <div class="metric-card metric-pink">
            <div class="metric-label">Total Records</div>
            <div class="metric-value">{len(processed_df):,}</div>
        </div>""", unsafe_allow_html=True)
    with gcol3:
        avg_count = po_grouped['count'].mean()
        st.markdown(f"""
        <div class="metric-card metric-muted">
            <div class="metric-label">Avg Count / PO</div>
            <div class="metric-value">{avg_count:.1f}</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    tab_summary, tab_detail = st.tabs(["📈 PO Summary", "📋 Full Data"])
    with tab_summary:
        st.dataframe(po_grouped.head(100), use_container_width=True, height=300)
    with tab_detail:
        st.dataframe(processed_df.drop(columns=['count'], errors='ignore').head(100),
                     use_container_width=True, height=300)
else:
    st.warning("⚠️ Kolom PO Number tidak ditemukan. Pilih secara manual atau pastikan mapping *PO Number sudah diisi.")
    manual_po = st.selectbox("Pilih kolom PO secara manual:", options=["(Lewati)"] + work_df.columns.tolist())
    if manual_po != "(Lewati)":
        po_column = manual_po
        processed_df, po_grouped = attach_po_counts(work_df, po_column)
        st.dataframe(po_grouped.head(50), use_container_width=True)
    else:
        processed_df = work_df.copy()

st.divider()

st.markdown(f"""
<div class="pipeline-step active">
    <span class="step-number">6</span>
    <strong>Split Processed Data</strong>
    <span class="badge badge-info" style="margin-left:0.8rem;">MAX {max_rows_per_file:,} ROWS/FILE</span>
</div>
""", unsafe_allow_html=True)

_out_df = processed_df.drop(columns=['count'], errors='ignore')
split_dfs = split_by_po_groups(_out_df, po_column, max_rows=max_rows_per_file)
n_files = len(split_dfs)

po_note = " · PO Number tidak dipisah antar file" if po_column else ""
st.markdown(f"""
<div class="split-info">
    🌸 <strong>Split Result:</strong> {len(_out_df):,} rows → <strong>{n_files} file(s)</strong>
    &nbsp;|&nbsp; Maks {max_rows_per_file:,} rows/file{po_note}
</div>
""", unsafe_allow_html=True)

if n_files > 1:
    split_cols = st.columns(min(n_files, 4))
    for i, sdf in enumerate(split_dfs):
        with split_cols[i % 4]:
            n_po_in_file = sdf[po_column].nunique() if po_column and po_column in sdf.columns else "-"
            st.markdown(f"""
            <div class="metric-card metric-rose" style="text-align:center;margin-bottom:0.5rem;">
                <div class="metric-label">File {i + 1}</div>
                <div class="metric-value" style="font-size:1.2rem;">{len(sdf):,}</div>
                <div style="color:#888888;font-size:0.72rem;">rows · {n_po_in_file} PO</div>
            </div>
            """, unsafe_allow_html=True)

st.divider()

st.markdown("""
<div class="pipeline-step active">
    <span class="step-number">7</span>
    <strong>Download Outputs</strong>
    <span class="badge badge-success" style="margin-left:0.8rem;">READY</span>
</div>
""", unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

cust_code_col = "*Customer Code" if "*Customer Code" in _out_df.columns else None
if cust_code_col is None:
    cust_code_col = next((c for c in _out_df.columns if "customer" in c.lower() and "code" in c.lower()), None)

dl1, dl2 = st.columns(2)

with dl1:
    if cust_code_col:
        cust_grouped = _out_df.groupby(cust_code_col).size().reset_index(name='Total Rows')
        cust_grouped = cust_grouped.sort_values('Total Rows', ascending=False).reset_index(drop=True)
        cust_excel = to_excel_bytes(cust_grouped, "Customer Grouping")
        st.markdown(f"""
        <div class="metric-card" style="text-align:center;border-left:4px solid #CA6180;margin-bottom:0.6rem;">
            <div style="font-size:1.8rem;margin-bottom:0.3rem;">📊</div>
            <div style="font-weight:700;color:#CA6180;margin-bottom:0.2rem;">Customer Grouping</div>
            <div style="color:#A8849A;font-size:0.78rem;">Group by {cust_code_col}</div>
        </div>
        """, unsafe_allow_html=True)
        st.download_button(
            label="⬇ Download Customer Grouping (.xlsx)",
            data=cust_excel,
            file_name=f"customer_grouping_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    else:
        st.info("ℹ️ Kolom *Customer Code belum dimapping.")

with dl2:
    st.markdown("""
    <div class="metric-card" style="text-align:center;border-left:4px solid #CA6180;margin-bottom:0.6rem;">
        <div style="font-size:1.8rem;margin-bottom:0.3rem;">📦</div>
        <div style="font-weight:700;color:#CA6180;margin-bottom:0.2rem;">ST Data Output</div>
        <div style="color:#A8849A;font-size:0.78rem;">Format ST Data Template — siap import</div>
    </div>
    """, unsafe_allow_html=True)
    if n_files == 1:
        proc_excel = to_excel_bytes(split_dfs[0], "ST Data")
        st.download_button(
            label="⬇ Download ST Data (.xlsx)",
            data=proc_excel,
            file_name=f"st_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    else:
        file_dict = {f"st_data_part_{i+1}.xlsx": to_excel_bytes(sdf, f"Part {i+1}") for i, sdf in enumerate(split_dfs)}
        zip_data = create_zip_of_files(file_dict)
        st.download_button(
            label=f"⬇ Download All ({n_files} files .zip)",
            data=zip_data,
            file_name=f"st_data_split_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
            mime="application/zip",
            use_container_width=True,
        )

st.markdown("<br>", unsafe_allow_html=True)

all_files = {}
if cust_code_col:
    all_files["01_customer_grouping.xlsx"] = to_excel_bytes(cust_grouped, "Customer Grouping")
for i, sdf in enumerate(split_dfs):
    all_files[f"0{i+2}_st_data_part_{i+1}.xlsx"] = to_excel_bytes(sdf, f"Part {i+1}")

all_zip = create_zip_of_files(all_files)
st.download_button(
    label="⬇ Download Complete Bundle (.zip)",
    data=all_zip,
    file_name=f"dataflow_complete_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
    mime="application/zip",
    use_container_width=True,
)

st.divider()
st.markdown("""
<div style="background:linear-gradient(135deg,rgba(202,97,128,0.08) 0%,rgba(252,183,199,0.12) 100%);
            border:1px solid rgba(202,97,128,0.2); border-radius:14px;
            padding:1.2rem 1.5rem; text-align:center;">
    <div style="font-size:1.1rem;font-weight:700;color:#CA6180;margin-bottom:0.3rem;">
        🌸 Pipeline Complete
    </div>
    <div style="color:#A8849A;font-size:0.85rem;">
        Output dalam format ST Data Template — siap diimport ke sistem.
    </div>
</div>
""", unsafe_allow_html=True)