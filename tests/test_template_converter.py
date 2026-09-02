"""Regression tests: customer/PO identity fields must stay text (leading zeros).

Run: pytest tests/test_template_converter.py -v
"""
from __future__ import annotations

import ast
import io
import sys
from pathlib import Path
from unittest.mock import MagicMock

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

# Streamlit/starlette may be uninstallable in this environment; the converter
# only needs the decorator + secrets fallback at import time.
def _passthrough(*args, **kwargs):
    def deco(fn):
        return fn
    if args and callable(args[0]) and not kwargs:
        return args[0]
    return deco

_mock_st = MagicMock()
_mock_st.cache_resource = _passthrough
_mock_st.cache_data = _passthrough
_mock_st.secrets = MagicMock()
_mock_st.secrets.__getitem__.side_effect = KeyError('connections')
sys.modules.setdefault('streamlit', _mock_st)

import template_converter as tc

REPO_ROOT = Path(__file__).resolve().parent.parent

TEXT_COLS = [
    "Customer SKU Code",
    "Customer SKU Name",
    "Customer Store Code",
    "Customer Store Name",
    "PO Number",
]


class FakeUpload(io.BytesIO):
    def __init__(self, data: bytes, name: str):
        super().__init__(data)
        self.name = name


def _static_fields(**overrides):
    base = {
        "Customer Code": "9999",
        "Customer Name": "PT Test",
        "Customer Branch Code": "",
        "Customer Branch Name": "Jakarta",
        "Customer Address": "Jl. Test",
    }
    base.update(overrides)
    return base


def _mapping():
    return {
        "PO Date": "po_date",
        "PO Number": "po_number",
        "Customer Store Code": "customer_store_code",
        "Customer Store Name": "customer_store_name",
        "Customer SKU Code": "customer_sku_code",
        "Customer SKU Name": "customer_sku_name",
        "Qty": "qty",
    }


def _source_frame(rows):
    return pd.DataFrame(rows)
# =====================================================================
# UNIT — as_text
# =====================================================================
class TestAsText:
    @pytest.mark.sanity
    @pytest.mark.parametrize("raw,expected", [
        ("00123", "00123"),
        ("00001", "00001"),
        ("012345", "012345"),
        ("ABC00123", "ABC00123"),
        ("SKU-001", "SKU-001"),
        ("123", "123"),
        (123, "123"),
        (123.0, "123"),
        ("123.0", "123"),
        ("00123.0", "00123"),
        ("", ""),
        (None, ""),
        (float("nan"), ""),
        ("nan", ""),
        ("NaN", ""),
        ("  00123  ", "00123"),
        ("STORE001", "STORE001"),
        ("PO00123", "PO00123"),
    ])
    def test_preserves_text_and_leading_zeros(self, raw, expected):
        assert tc.as_text(raw) == expected

    def test_does_not_int_convert_leading_zero_strings(self):
        assert tc.as_text("000789") == "000789"
        assert tc.as_text("000789") != "789"


# =====================================================================
# UNIT — CSV upload
# =====================================================================
class TestCsvUpload:
    @pytest.mark.sanity
    def test_csv_preserves_leading_zeros(self):
        csv = (
            "customer_sku_code,customer_store_code,po_number,"
            "customer_sku_name,customer_store_name\n"
            "00123,00045,000789,SKU-001,TOKO 001\n"
        )
        uploaded = FakeUpload(csv.encode("utf-8"), "sample.csv")
        df = tc.read_any_table(uploaded)
        assert list(df["customer_sku_code"]) == ["00123"]
        assert list(df["customer_store_code"]) == ["00045"]
        assert list(df["po_number"]) == ["000789"]
        assert list(df["customer_sku_name"]) == ["SKU-001"]
        assert list(df["customer_store_name"]) == ["TOKO 001"]

    def test_csv_blank_is_empty_string_not_nan(self):
        csv = "po_number,customer_sku_code\n,00123\n"
        uploaded = FakeUpload(csv.encode("utf-8"), "blank.csv")
        df = tc.read_any_table(uploaded)
        val = df["po_number"].iloc[0]
        assert val == ""
        assert str(val).lower() != "nan"

    def test_csv_plain_numeric_looking_stays_string(self):
        csv = "po_number,customer_sku_code,customer_store_code\n123,456,789\n"
        uploaded = FakeUpload(csv.encode("utf-8"), "plain.csv")
        df = tc.read_any_table(uploaded)
        assert list(df["po_number"]) == ["123"]
        assert list(df["customer_sku_code"]) == ["456"]
        assert list(df["customer_store_code"]) == ["789"]

# =====================================================================
# UNIT — Excel upload
# =====================================================================
def _xlsx_bytes(headers, rows, text_cells=True):
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
        if text_cells:
            r = ws.max_row
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell.value, str):
                    cell.number_format = "@"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestExcelUpload:
    @pytest.mark.sanity
    def test_excel_preserves_leading_zeros(self):
        data = _xlsx_bytes(
            ["po_number", "customer_sku_code", "customer_store_code",
             "customer_sku_name", "customer_store_name"],
            [("00123", "00123", "00045", "SKU-001", "TOKO 001")],
        )
        uploaded = FakeUpload(data, "sample.xlsx")
        df = tc.read_any_table(uploaded)
        assert list(df["po_number"]) == ["00123"]
        assert list(df["customer_sku_code"]) == ["00123"]
        assert list(df["customer_store_code"]) == ["00045"]

    def test_excel_multiple_leading_zero_po_numbers(self):
        data = _xlsx_bytes(
            ["PO Number"],
            [("00123",), ("00045",), ("000789",)],
        )
        uploaded = FakeUpload(data, "po.xlsx")
        df = tc.read_any_table(uploaded)
        assert list(df["PO Number"]) == ["00123", "00045", "000789"]


# =====================================================================
# UNIT — mapping / conversion pipeline
# =====================================================================
class TestIntelligentMapping:
    @pytest.mark.sanity
    def test_leading_zeros_survive_mapping(self):
        df = _source_frame([{
            "po_date": "2026-01-15",
            "po_number": "000789",
            "customer_store_code": "00045",
            "customer_store_name": "TOKO 001",
            "customer_sku_code": "00123",
            "customer_sku_name": "SKU-001",
            "qty": "2",
        }])
        mapped, _, failed = tc.intelligent_mapping(
            df, _static_fields(), _mapping(), "11", "PT TEST", enable_fuzzy=False
        )
        assert failed == []
        assert mapped["Customer SKU Code"].iloc[0] == "00123"
        assert mapped["Customer Store Code"].iloc[0] == "00045"
        assert mapped["PO Number"].iloc[0] == "000789"

    def test_alphanumeric_unchanged(self):
        df = _source_frame([{
            "po_date": "2026-01-15",
            "po_number": "PO00123",
            "customer_store_code": "STORE001",
            "customer_store_name": "TOKO ABC",
            "customer_sku_code": "SKU00123",
            "customer_sku_name": "Product A",
            "qty": "1",
        }])
        mapped, _, _ = tc.intelligent_mapping(
            df, _static_fields(), _mapping(), "11", "PT TEST", enable_fuzzy=False
        )
        assert mapped["Customer SKU Code"].iloc[0] == "SKU00123"
        assert mapped["Customer Store Code"].iloc[0] == "STORE001"
        assert mapped["PO Number"].iloc[0] == "PO00123"

    def test_plain_numeric_looking_stays_string(self):
        df = _source_frame([{
            "po_date": "2026-01-15",
            "po_number": 789,
            "customer_store_code": 456,
            "customer_store_name": "TOKO",
            "customer_sku_code": 123,
            "customer_sku_name": "NAME",
            "qty": 1,
        }])
        mapped, _, _ = tc.intelligent_mapping(
            df, _static_fields(), _mapping(), "11", "PT TEST", enable_fuzzy=False
        )
        assert mapped["Customer SKU Code"].iloc[0] == "123"
        assert mapped["Customer Store Code"].iloc[0] == "456"
        assert mapped["PO Number"].iloc[0] == "789"
        assert mapped["Customer SKU Code"].dtype == "string"
        assert mapped["PO Number"].dtype == "string"

    def test_blank_does_not_become_nan_string(self):
        df = _source_frame([{
            "po_date": "2026-01-15",
            "po_number": None,
            "customer_store_code": "",
            "customer_store_name": None,
            "customer_sku_code": float("nan"),
            "customer_sku_name": "",
            "qty": 1,
        }])
        mapped, _, _ = tc.intelligent_mapping(
            df, _static_fields(), _mapping(), "11", "PT TEST", enable_fuzzy=False
        )
        for col in TEXT_COLS:
            val = mapped[col].iloc[0]
            assert val == ""
            assert str(val).lower() not in ("nan", "none", "<na>")
            assert val != "123.0"

    def test_sinar_sakti_prefix_keeps_leading_zeros(self):
        df = _source_frame([{
            "po_date": "2026-01-15",
            "po_number": "000789",
            "customer_store_code": "00045",
            "customer_store_name": "TOKO",
            "customer_sku_code": "00123",
            "customer_sku_name": "NAME",
            "qty": 1,
        }])
        mapped, _, _ = tc.intelligent_mapping(
            df, _static_fields(), _mapping(), "11", "CV SINAR SAKTI", enable_fuzzy=False
        )
        assert mapped["PO Number"].iloc[0] == "SS000789"

    def test_store_code_prefix_keeps_leading_zeros(self):
        df = _source_frame([{
            "po_date": "2026-01-15",
            "po_number": "000789",
            "customer_store_code": "00045",
            "customer_store_name": "TOKO",
            "customer_sku_code": "00123",
            "customer_sku_name": "NAME",
            "qty": 1,
        }])
        mapped, _, _ = tc.intelligent_mapping(
            df,
            _static_fields(**{"Customer Branch Code": "11"}),
            _mapping(),
            "11",
            "PT TEST",
            enable_fuzzy=False,
        )
        assert mapped["Customer Store Code"].iloc[0] == "1100045"

class TestDedupAndExport:
    @pytest.mark.sanity
    def test_dedup_does_not_strip_leading_zeros(self):
        df = pd.DataFrame({
            "Customer Code": ["119999", "119999"],
            "Customer Name": ["PT Test", "PT Test"],
            "Customer Branch Code": ["", ""],
            "Customer Branch Name": ["Jakarta", "Jakarta"],
            "Customer Address": ["Jl. Test", "Jl. Test"],
            "PO Date": ["2026-01-15", "2026-01-15"],
            "PO Number": ["000789", "000789"],
            "Customer Store Code": ["00045", "00045"],
            "Customer Store Name": ["TOKO 001", "TOKO 001"],
            "Customer SKU Code": ["00123", "00123"],
            "Customer SKU Name": ["SKU-001", "SKU-001"],
            "Qty": [1, 2],
        })
        out = tc.deduplicate_and_sum_qty(df)
        assert len(out) == 1
        assert out["PO Number"].iloc[0] == "000789"
        assert out["Customer SKU Code"].iloc[0] == "00123"
        assert out["Customer Store Code"].iloc[0] == "00045"
        assert out["Qty"].iloc[0] == 3

    def test_excel_export_cells_are_text_with_leading_zeros(self):
        df = pd.DataFrame({
            "Customer Code": ["119999"],
            "Customer Name": ["PT Test"],
            "Customer Branch Code": [""],
            "Customer Branch Name": ["Jakarta"],
            "Customer Address": ["Jl. Test"],
            "PO Date": ["2026-01-15"],
            "PO Number": ["000789"],
            "Customer Store Code": ["00045"],
            "Customer Store Name": ["TOKO 001"],
            "Customer SKU Code": ["00123"],
            "Customer SKU Name": ["SKU-001"],
            "Qty": [1],
        })
        raw = tc.to_excel_bytes(df, "MappedData")
        wb = load_workbook(io.BytesIO(raw))
        ws = wb["MappedData"]
        headers = [c.value for c in ws[2]]  # startrow=1 -> header on Excel row 2
        data_row = ws[3]
        values = {headers[i]: data_row[i].value for i in range(len(headers))}
        assert values["PO Number"] == "000789"
        assert values["Customer SKU Code"] == "00123"
        assert values["Customer Store Code"] == "00045"
        assert values["Customer SKU Name"] == "SKU-001"
        assert values["Customer Store Name"] == "TOKO 001"
        idx = {headers[i]: i for i in range(len(headers))}
        for col in TEXT_COLS:
            cell = data_row[idx[col]]
            assert cell.data_type in ("s", "str", "inlineStr")
            assert not isinstance(cell.value, (int, float))


class TestEndToEndCsvAndExcel:
    @pytest.mark.sanity
    def test_csv_upload_through_mapping_and_export(self):
        csv = (
            "po_date,po_number,customer_store_code,customer_store_name,"
            "customer_sku_code,customer_sku_name,qty\n"
            "2026-01-15,000789,00045,TOKO 001,00123,SKU-001,1\n"
        )
        uploaded = FakeUpload(csv.encode("utf-8"), "full.csv")
        df = tc.read_any_table(uploaded)
        mapped, _, _ = tc.intelligent_mapping(
            df, _static_fields(), _mapping(), "11", "PT TEST", enable_fuzzy=False
        )
        mapped = tc.deduplicate_and_sum_qty(mapped)
        assert mapped["PO Number"].iloc[0] == "000789"
        assert mapped["Customer SKU Code"].iloc[0] == "00123"
        assert mapped["Customer Store Code"].iloc[0] == "00045"
        raw = tc.to_excel_bytes(mapped, "MappedData")
        wb = load_workbook(io.BytesIO(raw))
        headers = [c.value for c in wb["MappedData"][2]]
        data = [c.value for c in wb["MappedData"][3]]
        values = dict(zip(headers, data))
        assert values["PO Number"] == "000789"
        assert values["Customer SKU Code"] == "00123"
        assert values["Customer Store Code"] == "00045"

    def test_excel_upload_through_mapping_and_export(self):
        data = _xlsx_bytes(
            ["po_date", "po_number", "customer_store_code", "customer_store_name",
             "customer_sku_code", "customer_sku_name", "qty"],
            [("2026-01-15", "000789", "00045", "TOKO 001", "00123", "SKU-001", "1")],
        )
        uploaded = FakeUpload(data, "full.xlsx")
        df = tc.read_any_table(uploaded)
        mapped, _, _ = tc.intelligent_mapping(
            df, _static_fields(), _mapping(), "11", "PT TEST", enable_fuzzy=False
        )
        mapped = tc.deduplicate_and_sum_qty(mapped)
        for col, expected in [
            ("PO Number", "000789"),
            ("Customer SKU Code", "00123"),
            ("Customer Store Code", "00045"),
            ("Customer SKU Name", "SKU-001"),
            ("Customer Store Name", "TOKO 001"),
        ]:
            assert mapped[col].iloc[0] == expected


class Test3MProductCode:
    def test_barcode_with_leading_zeros_not_int_converted(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "TEMPLATE"
        ws.cell(row=1, column=1, value="No. Trans : JL/M3-000789 [ 09-02-2026 ] - ONE MART")
        ws.cell(row=1, column=8, value="00045")
        ws.cell(row=1, column=8).number_format = "@"
        ws.cell(row=2, column=1, value="001234567890")
        ws.cell(row=2, column=1).number_format = "@"
        ws.cell(row=2, column=2, value="Product Zero")
        ws.cell(row=2, column=3, value=2)
        buf = io.BytesIO()
        wb.save(buf)
        uploaded = FakeUpload(buf.getvalue(), "3m.xlsx")
        cleaned = tc.clean_3m_daily_st(uploaded)
        assert not cleaned.empty
        assert cleaned["Product Code"].iloc[0] == "001234567890"
        assert cleaned["No. TRANSAKSI"].iloc[0] == "JL/M3-000789"
        assert cleaned["ID CUST DISTRIBUTOR"].iloc[0] == "00045"


class TestSmokeSyntax:
    @pytest.mark.sanity
    def test_template_converter_parses(self):
        source = (REPO_ROOT / "template_converter.py").read_text(encoding="utf-8")
        ast.parse(source, filename="template_converter.py")