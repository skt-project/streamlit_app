"""SKU_MAPPING_TEMPLATE 2.0.xlsx — the new official SKU template (2026-09-03).

Audited structure (see noo_sku/config.py and the implementation report for the
full inspection):

  Sheet 'SKU TEMPLATE FOR STREAMLIT'   the upload sheet -- header on row 3,
                                       NOTHING below it: the CONTOH example
                                       that used to live inline (rows 4-5) has
                                       been moved OUT.
  Sheet 'Contoh Pengisian'             the example, now on its own sheet, with
                                       the SAME 4-column header text on row 1
                                       -- a real risk of being picked as the
                                       "upload" sheet by header-signature alone.
  Sheet 'GUIDELINE'                    unchanged, prose only.

These tests run against the actual bundled file, not a synthetic fixture,
because the risk being tested (sheet-selection ambiguity) only exists in the
real file's specific structure.
"""
from __future__ import annotations

import io

import openpyxl
import pytest

from noo_sku import config, parsers, sources


@pytest.fixture()
def template_bytes():
    return sources.load_local_sku_template()


# ─── Test 7 — the app serves THIS file ────────────────────────────────────────
@pytest.mark.sanity
def test_template_asset_is_bundled_and_readable(template_bytes):
    assert template_bytes[:2] == b"PK", "must be a real .xlsx (zip) file"
    assert len(template_bytes) > 1000


@pytest.mark.sanity
def test_missing_asset_fails_loudly_not_silently(monkeypatch, tmp_path):
    """Serving the wrong (stale Drive) template silently would be worse than
    a clear error -- confirm the loader refuses rather than substituting."""
    from pathlib import Path

    monkeypatch.setattr(config, "SKU_TEMPLATE_LOCAL_PATH",
                        Path(tmp_path) / "does_not_exist.xlsx")
    with pytest.raises(FileNotFoundError):
        sources.load_local_sku_template()


def test_prepare_sku_template_is_a_harmless_no_op_on_the_new_file(template_bytes):
    """No gramasi column exists to strip; must return the file unchanged."""
    assert sources.prepare_sku_template(template_bytes) == template_bytes


# ─── Structural audit ─────────────────────────────────────────────────────────
@pytest.mark.sanity
def test_template_has_the_expected_sheets_and_no_gramasi_column(template_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    assert set(wb.sheetnames) >= {
        "SKU TEMPLATE FOR STREAMLIT", "Contoh Pengisian", "GUIDELINE"}
    ws = wb["SKU TEMPLATE FOR STREAMLIT"]
    header = [c.value for c in ws[config.SKU_HEADER_ROW] if c.value]
    assert header == config.SKU_COLUMNS
    assert not any("size" in h.lower() for h in header)


def test_upload_sheet_has_no_data_below_the_header_in_the_blank_template(
        template_bytes):
    """The example used to live inline right under the header; confirming
    it is gone here is what makes 'start reading immediately below the
    header, no hardcoded skip' a safe rule for this template."""
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    ws = wb["SKU TEMPLATE FOR STREAMLIT"]
    for row in ws.iter_rows(min_row=config.SKU_HEADER_ROW + 1, max_row=10,
                            values_only=True):
        assert not any(row), "no example rows should remain under the header"


# ─── Test 8 — dynamic header detection, no hardcoded skiprows ─────────────────
@pytest.mark.sanity
def test_data_begins_immediately_below_the_header_no_row_hardcoded(template_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    ws = wb["SKU TEMPLATE FOR STREAMLIT"]
    rows = [["SKINTIFIC-296", "SKINTIFIC TEST PRODUCT", "SKC-296", "S296 TEST"],
           ["SKINTIFIC-41101", "SKINTIFIC 4D HYALURONIC", "SKC-411", "S411"]]
    for i, values in enumerate(rows, start=config.SKU_HEADER_ROW + 1):
        for j, v in enumerate(values, start=1):
            ws.cell(row=i, column=j, value=v)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    parsed = parsers.parse_upload(buf)
    assert parsed.sheet_name == "SKU TEMPLATE FOR STREAMLIT"
    assert parsed.header_row == config.SKU_HEADER_ROW
    assert parsed.row_numbers == [config.SKU_HEADER_ROW + 1,
                                  config.SKU_HEADER_ROW + 2]
    assert [r["Principal Product Code"] for r in parsed.rows] == [
        "SKINTIFIC-296", "SKINTIFIC-41101"]


@pytest.mark.sanity
def test_header_detection_is_not_hardcoded_to_a_fixed_row_number(template_bytes):
    """If BD Support inserts or removes a banner row and the header shifts,
    parsing must still find it by signature, not by a pinned row index."""
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    ws = wb["SKU TEMPLATE FOR STREAMLIT"]
    ws.insert_rows(1, amount=2)  # header (and everything else) shifts down by 2
    new_header_row = config.SKU_HEADER_ROW + 2
    ws.cell(row=new_header_row + 1, column=1, value="SKINTIFIC-296")
    ws.cell(row=new_header_row + 1, column=2, value="SKINTIFIC TEST PRODUCT")
    ws.cell(row=new_header_row + 1, column=3, value="SKC-296")
    ws.cell(row=new_header_row + 1, column=4, value="S296 TEST")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    parsed = parsers.parse_upload(buf)
    assert parsed.header_row == config.SKU_HEADER_ROW + 2
    assert len(parsed.rows) == 1


# ─── Test 9 — the "Contoh Pengisian" example sheet is never read as data ──────
@pytest.mark.sanity
def test_example_sheet_is_never_selected_as_the_upload_sheet(template_bytes):
    """Contoh Pengisian carries the identical 4-column header text on row 1 --
    the single easiest way for this template to be mis-parsed."""
    parsed = parsers.parse_upload(io.BytesIO(template_bytes))
    assert parsed.sheet_name == "SKU TEMPLATE FOR STREAMLIT"
    assert parsed.sheet_name != "Contoh Pengisian"


@pytest.mark.sanity
def test_example_sheet_content_never_appears_in_parsed_rows(template_bytes):
    """The example's own SKU code (TYY114002) must never surface as a row."""
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    ws = wb["SKU TEMPLATE FOR STREAMLIT"]
    ws.cell(row=config.SKU_HEADER_ROW + 1, column=1, value="SKINTIFIC-296")
    ws.cell(row=config.SKU_HEADER_ROW + 1, column=2, value="TEST")
    ws.cell(row=config.SKU_HEADER_ROW + 1, column=3, value="SKC-1")
    ws.cell(row=config.SKU_HEADER_ROW + 1, column=4, value="TEST DB")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    parsed = parsers.parse_upload(buf)
    codes = [r["Principal Product Code"] for r in parsed.rows]
    assert "TYY114002" not in codes
    assert codes == ["SKINTIFIC-296"]


def test_reference_only_sheets_are_excluded_from_header_scanning():
    assert "contoh pengisian" in parsers._REFERENCE_ONLY_SHEETS
    assert "guideline" in parsers._REFERENCE_ONLY_SHEETS


# ─── Wrong-template detection still works with the new file ─────────────────
def test_new_sku_template_uploaded_into_noo_section_is_still_rejected(
        template_bytes):
    parsed = parsers.parse_upload(io.BytesIO(template_bytes))
    message = parsers.check_template_kind(parsed, parsers.UPLOAD_NOO)
    assert message and "SKU Mapping" in message
