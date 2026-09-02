import io
import re
from typing import Dict, List, Tuple, Optional
import streamlit as st
import pandas as pd
from difflib import get_close_matches
from google.cloud import bigquery
from google.api_core.exceptions import NotFound
from google.oauth2 import service_account

# =========================
# Environment / Secrets
# =========================
try:
    gcp_secrets = st.secrets["connections"]["bigquery"]
    private_key = gcp_secrets["private_key"].replace("\\n", "\n")
    credentials = service_account.Credentials.from_service_account_info({
        "type": gcp_secrets["type"],
        "project_id": gcp_secrets["project_id"],
        "private_key_id": gcp_secrets["private_key_id"],
        "private_key": private_key,
        "client_email": gcp_secrets["client_email"],
        "client_id": gcp_secrets["client_id"],
        "auth_uri": gcp_secrets["auth_uri"],
        "token_uri": gcp_secrets["token_uri"],
        "auth_provider_x509_cert_url": gcp_secrets["auth_provider_x509_cert_url"],
        "client_x509_cert_url": gcp_secrets["client_x509_cert_url"],
    })
    GCP_PROJECT_ID = st.secrets["bigquery"]["project"]
    BQ_DATASET = st.secrets["bigquery"]["dataset"]
    BQ_CONFIGS_TABLE = st.secrets["bigquery"]["config_table"]
except Exception:
    GCP_CREDENTIALS_PATH = r"C:\script\skintific-data-warehouse-ea77119e2e7a.json"
    GCP_PROJECT_ID = "skintific-data-warehouse"
    BQ_DATASET = "gt_schema"
    BQ_CONFIGS_TABLE = "distributor_configs"
    try:
        credentials = service_account.Credentials.from_service_account_file(
            GCP_CREDENTIALS_PATH
        )
    except Exception:
        credentials = None

# =========================
# Master Schema
# =========================
MASTER_SCHEMA: List[str] = [
    "Customer Code",
    "Customer Name",
    "Customer Branch Code",
    "Customer Branch Name",
    "Customer Address",
    "PO Date",
    "PO Number",
    "Customer Store Code",
    "Customer Store Name",
    "Customer SKU Code",
    "Customer SKU Name",
    "Qty",
]

FIXED_FIRST_5 = MASTER_SCHEMA[:5]

BRAND_PREFIXES = {
    "SKINTIFIC": "11",
    "G2G": "12",
    "TIMEPHORIA": "13",
    "FACERINNA": "1A",
    "BODIBREZE": "17",
    "NEXTPRIME": "17"
}
BRAND_OPTIONS = list(BRAND_PREFIXES.keys())

# Distributors whose raw files require the 3M cleaning pipeline
# (matched via case-insensitive prefix)
M3_DISTRIBUTOR_PREFIX = "pt mitra makmur mandiri sejahtera"

# Master distributor table for BQ lookups
BQ_MASTER_DISTRIBUTOR_TABLE = "skintific-data-warehouse.gt_schema.master_distributor"

# Columns that must remain text end-to-end (leading zeros, codes, names).
# Never coerce these through int/float/pandas numeric inference.
TEXT_FIELDS: Tuple[str, ...] = (
    "Customer SKU Code",
    "Customer SKU Name",
    "Customer Store Code",
    "Customer Store Name",
    "PO Number",
)

_TEXT_FIELD_ALIASES = {
    "customer sku code",
    "customer_sku_code",
    "customer sku name",
    "customer_sku_name",
    "customer store code",
    "customer_store_code",
    "customer store name",
    "customer_store_name",
    "po number",
    "po_number",
    # 3M intermediate columns that feed the master text fields
    "product code",
    "product name",
    "no. transaksi",
    "id cust distributor",
}

_BLANK_TOKENS = {"", "nan", "none", "<na>", "<nat>", "nat", "null"}
_FLOAT_ARTIFACT_RE = re.compile(r"^-?\d+\.0+$")


def is_text_field(col_name: object) -> bool:
    """True when a column is one of the five string-only identity fields (or an alias)."""
    if col_name is None:
        return False
    raw = str(col_name).strip()
    if raw in TEXT_FIELDS:
        return True
    lowered = raw.lower()
    snake = lowered.replace(" ", "_")
    spaced = lowered.replace("_", " ")
    return (
        lowered in _TEXT_FIELD_ALIASES
        or snake in _TEXT_FIELD_ALIASES
        or spaced in _TEXT_FIELD_ALIASES
    )


def as_text(value) -> str:
    """Coerce a cell to text without converting numeric-looking strings to numbers.

    Preserves leading zeros (``00123`` stays ``00123``). Excel float artifacts
    such as ``123.0`` become ``123`` without calling ``int()`` on a string, so
    ``00123.0`` becomes ``00123`` rather than ``123``. Blanks / NaN become
    ``""`` rather than ``"nan"``.
    """
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (ValueError, TypeError):
        pass

    if isinstance(value, bool):
        return str(value)

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)

    if isinstance(value, str):
        s = value.strip()
        if s.lower() in _BLANK_TOKENS:
            return ""
        # Strip a trailing .0 on an all-digit string without going through int(),
        # so leading zeros survive (``00123.0`` -> ``00123``).
        if _FLOAT_ARTIFACT_RE.match(s):
            return s.split(".")[0]
        return s

    s = str(value).strip()
    if s.lower() in _BLANK_TOKENS:
        return ""
    return s


def series_as_text(series: pd.Series) -> pd.Series:
    """Return a pandas ``string`` dtype series with blanks as empty strings."""
    return series.map(as_text, na_action=None).astype("string")


def uniquify_column_names(columns) -> List[str]:
    """Make headers unique the way pandas mangles duplicates (name, name.1, name.2).

    Distributor files often repeat labels such as TYPE. Duplicate names later
    crash Streamlit preview via PyArrow: 'Duplicate column names found'.
    """
    seen: Dict[str, int] = {}
    unique: List[str] = []
    for raw in columns:
        name = "" if raw is None else str(raw)
        count = seen.get(name, 0)
        unique.append(name if count == 0 else f"{name}.{count}")
        seen[name] = count + 1
    return unique


def force_text_columns(
    df: pd.DataFrame, columns: Optional[List[str]] = None
) -> pd.DataFrame:
    """Force identity fields (or an explicit column list) to string dtype."""
    if df is None:
        return df
    df = df.copy()
    cols = columns if columns is not None else [c for c in df.columns if is_text_field(c)]
    for col in cols:
        if col in df.columns:
            df[col] = series_as_text(df[col])
    return df


def _excel_cell_to_text(cell) -> str:
    """Read an openpyxl cell as text, preserving string values and leading zeros."""
    value = cell.value
    if value is None:
        return ""

    data_type = getattr(cell, "data_type", None)
    if data_type in ("s", "str", "inlineStr"):
        return as_text(value)

    fmt = (getattr(cell, "number_format", None) or "").strip()
    fmt_clean = fmt.replace("\\", "").split(";")[0]
    if (
        data_type == "n"
        and isinstance(value, (int, float))
        and re.fullmatch(r"0+", fmt_clean)
    ):
        try:
            if float(value).is_integer():
                return f"{int(value):0{len(fmt_clean)}d}"
        except (ValueError, OverflowError, OSError):
            pass

    return as_text(value)


def _read_excel_cells_as_text(
    uploaded_file,
    sheet_name=0,
    header: Optional[int] = 0,
) -> pd.DataFrame:
    """Load an .xlsx workbook cell-by-cell as text (no pandas type inference)."""
    from openpyxl import load_workbook

    if hasattr(uploaded_file, "seek"):
        uploaded_file.seek(0)

    wb = load_workbook(uploaded_file, data_only=True, read_only=True)
    try:
        if isinstance(sheet_name, str):
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet {sheet_name!r} not found")
            ws = wb[sheet_name]
        else:
            ws = wb.worksheets[int(sheet_name)]

        all_rows: List[List[str]] = []
        for row in ws.iter_rows():
            all_rows.append([_excel_cell_to_text(c) for c in row])
    finally:
        wb.close()

    if not all_rows:
        return pd.DataFrame()

    if header is None:
        return pd.DataFrame(all_rows)

    header_idx = int(header)
    if header_idx >= len(all_rows):
        return pd.DataFrame()

    columns = uniquify_column_names(
        [
            cell if cell != "" else f"Unnamed: {i}"
            for i, cell in enumerate(all_rows[header_idx])
        ]
    )
    data = all_rows[header_idx + 1 :]
    return pd.DataFrame(data, columns=columns)


def _read_excel_as_text(
    uploaded_file,
    sheet_name=0,
    header: Optional[int] = 0,
) -> pd.DataFrame:
    """Read Excel as text; fall back to pandas dtype=str for .xls / engine issues."""
    name = str(getattr(uploaded_file, "name", "") or "").lower()
    use_openpyxl = (not name.endswith(".xls")) or name.endswith(".xlsx")
    if use_openpyxl:
        try:
            return _read_excel_cells_as_text(
                uploaded_file, sheet_name=sheet_name, header=header
            )
        except ValueError:
            raise
        except Exception:
            pass

    if hasattr(uploaded_file, "seek"):
        uploaded_file.seek(0)

    kwargs = {"dtype": str, "keep_default_na": False, "header": header}
    if sheet_name is not None:
        kwargs["sheet_name"] = sheet_name
    df = pd.read_excel(uploaded_file, **kwargs)
    return df.apply(lambda col: col.map(as_text))



# =========================
# BigQuery Client
# =========================
@st.cache_resource(show_spinner=False)
def get_bq_client() -> bigquery.Client:
    return bigquery.Client(credentials=credentials, project=credentials.project_id)


# =========================
# BigQuery Bootstrap
# =========================
def ensure_bq_objects():
    client = get_bq_client()
    dataset_ref = bigquery.Dataset(f"{client.project}.{BQ_DATASET}")
    try:
        client.get_dataset(dataset_ref)
    except NotFound:
        client.create_dataset(dataset_ref)

    schema_configs = [
        bigquery.SchemaField("distributor", "STRING", mode="REQUIRED"),
        bigquery.SchemaField("static_fields", "JSON", mode="REQUIRED"),
        bigquery.SchemaField("mapping", "JSON", mode="REQUIRED"),
        bigquery.SchemaField("updated_at", "TIMESTAMP", mode="REQUIRED"),
        bigquery.SchemaField("updated_by", "STRING", mode="REQUIRED"),
    ]
    table_configs_id = f"{client.project}.{BQ_DATASET}.{BQ_CONFIGS_TABLE}"
    try:
        client.get_table(table_configs_id)
    except NotFound:
        table = bigquery.Table(table_configs_id, schema=schema_configs)
        client.create_table(table)


# =========================
# BigQuery Helpers
# =========================
def list_distributors() -> List[str]:
    client = get_bq_client()
    sql = f"""
    SELECT distributor
    FROM `{client.project}.{BQ_DATASET}.{BQ_CONFIGS_TABLE}`
    ORDER BY distributor
    """
    rows = client.query(sql).result()
    return [r.distributor for r in rows]


def get_config(distributor: str) -> Optional[Dict]:
    client = get_bq_client()
    sql = f"""
    SELECT static_fields, mapping
    FROM `{client.project}.{BQ_DATASET}.{BQ_CONFIGS_TABLE}`
    WHERE distributor = @distributor
    LIMIT 1
    """
    job_config = bigquery.QueryJobConfig(
        query_parameters=[
            bigquery.ScalarQueryParameter("distributor", "STRING", distributor)
        ]
    )
    rows = list(client.query(sql, job_config=job_config).result())
    if not rows:
        return None
    return {
        "static_fields": rows[0].static_fields,
        "mapping": rows[0].mapping,
    }


@st.cache_data(show_spinner=False)
def lookup_branch_info_by_store_prefix(store_code_prefix: str) -> Optional[Dict]:
    """
    Looks up Customer Branch Code and Customer Branch Name from master_distributor
    using the first 6 digits of the Customer Store Code.

    Column mapping in master_distributor:
        distributor_code  → Customer Branch Code
        distributor       → Customer Branch Name

    Results are cached so repeated calls with the same prefix won't
    re-hit BigQuery (most rows in a file share the same branch).
    """
    if not store_code_prefix or store_code_prefix.strip() in ("", "nan"):
        return None

    client = get_bq_client()
    sql = f"""
    SELECT
        distributor_code,
        distributor
    FROM `{BQ_MASTER_DISTRIBUTOR_TABLE}`
    WHERE distributor_code = @store_prefix
    LIMIT 1
    """
    job_config = bigquery.QueryJobConfig(
        query_parameters=[
            bigquery.ScalarQueryParameter("store_prefix", "STRING", store_code_prefix)
        ]
    )
    try:
        rows = list(client.query(sql, job_config=job_config).result())
    except Exception:
        return None

    if not rows:
        return None

    return {
        "Customer Branch Code": rows[0].distributor_code or "",
        "Customer Branch Name": rows[0].distributor or "",
    }


# =========================
# 3M Daily ST Cleaning
# =========================
def is_3m_distributor(distributor_name: str) -> bool:
    """Returns True if the distributor should use the 3M cleaning pipeline."""
    return distributor_name.lower().startswith(M3_DISTRIBUTOR_PREFIX)


def clean_3m_daily_st(uploaded_file) -> pd.DataFrame:
    """
    Parse and flatten the 3M Daily ST Upload Template (.xlsx).

    The raw file has a report-style layout where each transaction block starts
    with a header line like:
        "No. Trans : JL/M3-26020183 [ 09-02-2026 ] - ONE MART"
    followed by product rows. Column 7 of that header row carries the
    distributor store ID ("Store Code Suggestion").

    Returns a flat intermediate DataFrame with columns:
        Product Code | Product Name | Kuantitas | No. TRANSAKSI | PO Date |
        ID CUST DISTRIBUTOR | Customer Store Name
    """
    df = _read_excel_as_text(uploaded_file, sheet_name="TEMPLATE", header=None)

    records = []
    current_trans = None
    current_po_date = None
    current_store_id = None
    current_store_name = None

    for _, row in df.iterrows():
        cell0 = as_text(row[0])

        # ── Transaction header ────────────────────────────────────────────────
        # Pattern: "No. Trans : JL/M3-26020183 [ 09-02-2026 ] - ONE MART"
        if "No. Trans :" in cell0:
            match = re.match(
                r"No\.\s*Trans\s*:\s*(\S+)\s*\[\s*(\d{2}-\d{2}-\d{4})\s*\]\s*-\s*(.+)",
                cell0,
            )
            if match:
                current_trans = as_text(match.group(1))
                # Convert DD-MM-YYYY → YYYY-MM-DD to align with master schema
                current_po_date = pd.to_datetime(
                    match.group(2), format="%d-%m-%Y"
                ).strftime("%Y-%m-%d")
                current_store_name = as_text(match.group(3))

            col7 = as_text(row[7]) if 7 in row.index else ""
            # Leave blank for unregistered stores
            current_store_id = "" if col7 in ("Not Registered", "") else col7

        # ── Product row (col 0 is a numeric barcode ≥ 10 digits) ─────────────
        elif cell0.isdigit() and len(cell0) >= 10:
            records.append(
                {
                    "Product Code": cell0,  # keep as text; do not int() (strips leading zeros)
                    "Product Name": as_text(row[1]),
                    "Kuantitas": row[2],
                    "No. TRANSAKSI": as_text(current_trans),
                    "PO Date": current_po_date,
                    "ID CUST DISTRIBUTOR": as_text(current_store_id),
                    "Customer Store Name": as_text(current_store_name),
                }
            )

    result = pd.DataFrame(records)
    if not result.empty:
        result["Kuantitas"] = pd.to_numeric(result["Kuantitas"], errors="coerce")
        result = force_text_columns(
            result,
            [
                "Product Code",
                "Product Name",
                "No. TRANSAKSI",
                "ID CUST DISTRIBUTOR",
                "Customer Store Name",
            ],
        )
    return result


def map_3m_to_master(
    cleaned: pd.DataFrame,
    static_fields: Dict[str, str],
    brand_prefix: str,
) -> Tuple[pd.DataFrame, List[str], List[str]]:
    """
    Maps the intermediate 3M cleaned DataFrame to MASTER_SCHEMA.

    Customer Name and Customer Branch Code are resolved by looking up the
    first 6 digits of each row's Customer Store Code in master_distributor.
    Falls back to static_fields values when no BQ match is found.

    Returns (mapped_df, unregistered_stores, bq_lookup_misses).
    """
    out = pd.DataFrame(index=cleaned.index)

    # ── Fixed columns from static_fields (overridden below where BQ lookup wins)
    for col in FIXED_FIRST_5:
        out[col] = static_fields.get(col, "")

    # Apply brand prefix to Customer Code
    out["Customer Code"] = brand_prefix + static_fields.get("Customer Code", "")

    # ── Dynamic columns ───────────────────────────────────────────────────────
    out["PO Date"] = cleaned["PO Date"]
    out["PO Number"] = series_as_text(cleaned["No. TRANSAKSI"])
    out["Customer Store Code"] = series_as_text(cleaned["ID CUST DISTRIBUTOR"])
    out["Customer Store Name"] = series_as_text(cleaned["Customer Store Name"])
    out["Customer SKU Code"] = series_as_text(cleaned["Product Code"])
    out["Customer SKU Name"] = series_as_text(cleaned["Product Name"])
    out["Qty"] = cleaned["Kuantitas"]

    # ── BQ lookup: Customer Name & Customer Branch Code per store prefix ───────
    bq_lookup_misses: List[str] = []

    def enrich_from_bq(store_code: str) -> pd.Series:
        prefix = store_code[:6] if len(store_code) >= 6 else store_code
        # Blank store code → blank Name, Branch Code, and Branch Name
        if not prefix or prefix in ("", "nan"):
            return pd.Series({
                "Customer Name": "",
                "Customer Branch Code": "",
                "Customer Branch Name": "",
            })
        result = lookup_branch_info_by_store_prefix(prefix)
        if result:
            # BQ hit – distributor_code → Branch Code, distributor → Branch Name
            return pd.Series({
                "Customer Name": static_fields.get("Customer Name", ""),
                "Customer Branch Code": result.get("Customer Branch Code", ""),
                "Customer Branch Name": result.get("Customer Branch Name", ""),
            })
        # BQ miss – record prefix for warning and fall back to static_fields
        bq_lookup_misses.append(prefix)
        return pd.Series({
            "Customer Name": static_fields.get("Customer Name", ""),
            "Customer Branch Code": static_fields.get("Customer Branch Code", ""),
            "Customer Branch Name": static_fields.get("Customer Branch Name", ""),
        })

    enriched = out["Customer Store Code"].apply(enrich_from_bq)
    out["Customer Name"] = enriched["Customer Name"]
    out["Customer Branch Code"] = enriched["Customer Branch Code"]
    out["Customer Branch Name"] = enriched["Customer Branch Name"]

    # Deduplicate miss list
    bq_lookup_misses = list(dict.fromkeys(bq_lookup_misses))

    # ── Collect unregistered stores ───────────────────────────────────────────
    unregistered = (
        cleaned.loc[cleaned["ID CUST DISTRIBUTOR"] == "", "Customer Store Name"]
        .unique()
        .tolist()
    )

    out = force_text_columns(out[MASTER_SCHEMA])
    return out, unregistered, bq_lookup_misses


# =========================
# Deduplication / Qty Sum
# =========================
def deduplicate_and_sum_qty(df: pd.DataFrame) -> pd.DataFrame:
    """
    Groups by all MASTER_SCHEMA columns except Qty and sums Qty for
    duplicate rows (i.e. rows where every non-Qty column is identical).

    Returns a DataFrame with the same column order as MASTER_SCHEMA.
    """
    group_cols = [col for col in MASTER_SCHEMA if col != "Qty"]

    df = df.copy()
    df["Qty"] = pd.to_numeric(df["Qty"], errors="coerce").fillna(0)

    # Identity fields must stay text so groupby does not coerce "00123" -> 123
    df = force_text_columns(df)

    result = (
        df.groupby(group_cols, as_index=False, dropna=False)["Qty"]
        .sum()
    )

    return force_text_columns(result[MASTER_SCHEMA])


# =========================
# Intelligent Mapping
# =========================
def intelligent_mapping(
    df: pd.DataFrame,
    static_fields: Dict[str, str],
    mapping: Dict[str, str],
    brand_prefix: str,
    distributor: str,   # ✅ required
    enable_fuzzy: bool = True,
    fuzzy_cutoff: float = 0.6,
) -> Tuple[pd.DataFrame, Dict[str, str], List[str]]:

    out = pd.DataFrame()
    effective_mapping = {}
    failed_columns = []

    df = df.copy()
    df.columns = uniquify_column_names([col.lower() for col in df.columns])
    df = force_text_columns(df)
    mapping_lower = {k: v.lower() for k, v in mapping.items()}

    for col in FIXED_FIRST_5:
        out[col] = [static_fields.get(col, "")] * len(df)

    customer_code_static = static_fields.get("Customer Code", "")
    out["Customer Code"] = brand_prefix + customer_code_static

    needed = [c for c in MASTER_SCHEMA if c not in FIXED_FIRST_5]

    for target in needed:
        src = mapping_lower.get(target, "")
        if src and src in df.columns:
            if target == "PO Date":
                out[target] = pd.to_datetime(df[src], errors="coerce").dt.strftime("%Y-%m-%d")
            else:
                out[target] = df[src]
            effective_mapping[target] = src
        else:
            out[target] = None

    if enable_fuzzy:
        for target in needed:
            if out[target].isna().all():
                guesses = get_close_matches(
                    target.lower(), df.columns.tolist(), n=1, cutoff=fuzzy_cutoff
                )
                if guesses:
                    src = guesses[0]
                    out[target] = df[src]
                    effective_mapping[target] = src
                else:
                    failed_columns.append(target)

    # Coerce identity fields AFTER fuzzy matching so missing columns (all-NA)
    # can still be filled by fuzzy match. as_text turns NA into "" (not "nan").
    out = force_text_columns(out)

    # ✅ PREFIX LOGIC (SAFE)
    if distributor.upper() == "CV SINAR SAKTI":
        po = series_as_text(out["PO Number"])
        out["PO Number"] = po.map(
            lambda x: x if (not x or x.startswith("SS")) else "SS" + x
        )

    branch_code_prefix = static_fields.get("Customer Branch Code", "")
    original_store_code_col = effective_mapping.get("Customer Store Code")

    if branch_code_prefix and original_store_code_col:
        store = series_as_text(out["Customer Store Code"])
        out["Customer Store Code"] = store.map(lambda x: branch_code_prefix + x)
        effective_mapping["Customer Store Code"] = (
            f"PREFIXED({branch_code_prefix}){original_store_code_col}"
        )

    out = force_text_columns(out[MASTER_SCHEMA])
    return out, effective_mapping, failed_columns


# =========================
# Utilities
# =========================
def read_any_table(uploaded_file) -> pd.DataFrame:
    """Read CSV/Excel with every column as text so leading zeros are not inferred away."""
    name = uploaded_file.name.lower()
    if hasattr(uploaded_file, "seek"):
        uploaded_file.seek(0)

    if name.endswith(".csv"):
        df = pd.read_csv(
            uploaded_file,
            dtype=str,
            keep_default_na=False,
            na_filter=False,
        )
    elif name.endswith(".xlsx"):
        df = _read_excel_as_text(uploaded_file, header=0)
    elif name.endswith(".xls"):
        df = pd.read_excel(
            uploaded_file,
            dtype=str,
            keep_default_na=False,
            engine="xlrd",
        )
    else:
        st.error("Unsupported file type. Please upload a .csv, .xls, or .xlsx file.")
        return pd.DataFrame()

    # Sanitize every column: preserve leading-zero strings, turn NaN into "".
    df.columns = uniquify_column_names(df.columns)
    df = df.apply(lambda col: col.map(as_text))
    df = force_text_columns(df)
    return df


def to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Data") -> bytes:
    """Serialize a DataFrame to Excel, writing identity fields as true Excel text."""
    buf = io.BytesIO()
    df = force_text_columns(df)
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, header=True, startrow=1, sheet_name=sheet_name)
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        text_fmt = workbook.add_format({"num_format": "@"})
        # pandas startrow=1 -> header at row 1, data from row 2 (0-indexed)
        data_start_row = 2
        for col_idx, col_name in enumerate(df.columns):
            if not is_text_field(col_name):
                continue
            worksheet.set_column(col_idx, col_idx, None, text_fmt)
            for row_idx, value in enumerate(df[col_name].tolist()):
                worksheet.write_string(
                    data_start_row + row_idx,
                    col_idx,
                    as_text(value),
                    text_fmt,
                )
    buf.seek(0)
    return buf.getvalue()


# =========================
# UI – 3M pipeline section
# =========================
def render_3m_pipeline(dist: str, brand: str, brand_prefix: str):
    """Handles upload + conversion for CV MITRA MAKMUR MANDIRI distributors."""
    st.info(
        "ℹ️ This distributor uses the **3M Daily ST** format. "
        "Please upload the raw *3M Daily ST Upload Template* Excel file."
    )

    uploaded = st.file_uploader(
        "Upload 3M Daily ST File (.xlsx)", type=["xlsx", "xls"], key="m3_uploader"
    )
    if not uploaded:
        return

    # ── Fetch distributor config ──────────────────────────────────────────────
    cfg = get_config(dist)
    if not cfg:
        st.error(
            "Configuration for the selected distributor was not found. "
            "Please contact an administrator."
        )
        return

    # ── Parse raw report layout ───────────────────────────────────────────────
    with st.spinner("Parsing 3M report layout…"):
        try:
            cleaned = clean_3m_daily_st(uploaded)
        except Exception as e:
            st.error(f"Error parsing 3M file: {e}")
            return

    if cleaned.empty:
        st.warning("No product rows were found in the uploaded file.")
        return

    # ── Map to MASTER_SCHEMA (includes BQ lookup for Name & Branch Code) ──────
    with st.spinner("Looking up Customer Name & Branch Code from master_distributor…"):
        try:
            mapped, unregistered, bq_misses = map_3m_to_master(
                cleaned, cfg["static_fields"], brand_prefix
            )
        except Exception as e:
            st.error(f"Error mapping to master schema: {e}")
            return

    # ── Deduplicate rows and sum Qty ──────────────────────────────────────────
    rows_before = len(mapped)
    mapped = deduplicate_and_sum_qty(mapped)
    mapped = force_text_columns(mapped)
    rows_after = len(mapped)
    rows_merged = rows_before - rows_after
    if rows_merged > 0:
        st.info(
            f"ℹ️ **{rows_merged} duplicate row(s)** were merged and their Qty summed "
            f"({rows_before} → {rows_after} rows)."
        )

    st.success("Mapped to master schema.")
    st.write("Converted sample:")
    st.dataframe(mapped.head())

    # ── Mapping log ───────────────────────────────────────────────────────────
    st.subheader("Mapping Log")
    st.write("✅ **Successful Mappings:**")
    mapping_log = [
        {
            "Target Column": "Customer Code",
            "Source Column": f"PREFIXED({brand_prefix}){cfg['static_fields'].get('Customer Code', '')}",
            "Status": "Mapped",
        },
        {
            "Target Column": "Customer Name",
            "Source Column": "BQ lookup → master_distributor (first 6 digits of Store Code)",
            "Status": "Mapped",
        },
        {
            "Target Column": "Customer Branch Code",
            "Source Column": "BQ lookup → master_distributor WHERE distributor_code = first 6 chars of Store Code",
            "Status": "Mapped",
        },
        {
            "Target Column": "Customer Branch Name",
            "Source Column": "BQ lookup → master_distributor WHERE distributor_code = first 6 chars of Store Code",
            "Status": "Mapped",
        },
        {
            "Target Column": "Customer Address",
            "Source Column": cfg["static_fields"].get("Customer Address", ""),
            "Status": "Mapped",
        },
        {
            "Target Column": "PO Date",
            "Source Column": "Parsed from transaction header [DD-MM-YYYY]",
            "Status": "Mapped",
        },
        {
            "Target Column": "PO Number",
            "Source Column": "No. TRANSAKSI (transaction header)",
            "Status": "Mapped",
        },
        {
            "Target Column": "Customer Store Code",
            "Source Column": "Store Code Suggestion (col 7)",
            "Status": "Mapped",
        },
        {
            "Target Column": "Customer Store Name",
            "Source Column": "Parsed from transaction header",
            "Status": "Mapped",
        },
        {
            "Target Column": "Customer SKU Code",
            "Source Column": "BARCODE (col 0)",
            "Status": "Mapped",
        },
        {
            "Target Column": "Customer SKU Name",
            "Source Column": "NAMA PRODUK (col 1)",
            "Status": "Mapped",
        },
        {
            "Target Column": "Qty",
            "Source Column": "QTY (col 2) — summed for duplicate rows",
            "Status": "Mapped",
        },
    ]
    st.table(pd.DataFrame(mapping_log))

    # ── BQ lookup miss warning ────────────────────────────────────────────────
    if bq_misses:
        st.warning(
            f"⚠️ **{len(bq_misses)} store prefix(es)** were not found in "
            "`master_distributor`. Fell back to static config values for "
            "Customer Name & Customer Branch Code:"
        )
        st.write(", ".join(sorted(bq_misses)))

    # ── Unregistered stores warning ───────────────────────────────────────────
    if unregistered:
        st.warning(
            f"⚠️ **{len(unregistered)} store(s)** are marked *Not Registered* "
            "and will have an empty Customer Store Code:"
        )
        st.write(", ".join(sorted(unregistered)))

    # ── Download ──────────────────────────────────────────────────────────────
    try:
        xlsx = to_excel_bytes(mapped, "MappedData")
        st.download_button(
            label="📥 Download Converted Excel",
            data=xlsx,
            file_name=f"{dist}_converted.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        st.error(f"Error generating download file: {e}")


# =========================
# UI – Standard pipeline
# =========================
def render_standard_pipeline(dist: str, brand: str, brand_prefix: str):
    """Handles upload + conversion for standard distributor formats."""
    uploaded = st.file_uploader(
        "Upload Distributor File (.xlsx/.csv)", type=["xlsx", "xls", "csv"]
    )
    if not uploaded:
        return

    st.write("Preview of uploaded data:")
    try:
        df = read_any_table(uploaded)
        df.columns = uniquify_column_names([col.lower() for col in df.columns])
        st.dataframe(df.head())
    except Exception as e:
        st.error(f"Error reading the uploaded file: {e}")
        return

    cfg = get_config(dist)
    if not cfg:
        st.error(
            "Configuration for the selected distributor was not found. "
            "Please contact an administrator."
        )
        return

    try:
        mapped, effective_map, failed_columns = intelligent_mapping(
            df,
            cfg["static_fields"],
            cfg["mapping"],
            brand_prefix,
            distributor=dist
        )
    except Exception as e:
        st.error(f"Error during data mapping: {e}")
        return

    # ── Deduplicate rows and sum Qty ──────────────────────────────────────────
    rows_before = len(mapped)
    mapped = deduplicate_and_sum_qty(mapped)
    mapped = force_text_columns(mapped)
    rows_after = len(mapped)
    rows_merged = rows_before - rows_after
    if rows_merged > 0:
        st.info(
            f"ℹ️ **{rows_merged} duplicate row(s)** were merged and their Qty summed "
            f"({rows_before} → {rows_after} rows)."
        )

    st.success("Mapped to master schema.")
    st.write("Converted sample:")
    st.dataframe(mapped.head())

    st.subheader("Mapping Log")

    if effective_map:
        st.write("✅ **Successful Mappings:**")
        successful_log = [
            {
                "Target Column": target,
                "Source Column": source,
                "Status": "Mapped" if target in cfg["mapping"] else "Fuzzy Match",
            }
            for target, source in effective_map.items()
        ]
        st.table(pd.DataFrame(successful_log))

    if failed_columns:
        st.write("❌ **Failed Mappings (Columns not found in the uploaded file):**")
        failed_log = [
            {
                "Target Column": target,
                "Expected Source Column (from config)": cfg["mapping"].get(target, "N/A"),
            }
            for target in failed_columns
        ]
        st.table(pd.DataFrame(failed_log))

    try:
        xlsx = to_excel_bytes(mapped, "MappedData")
        st.download_button(
            "📥 Download Converted Excel",
            data=xlsx,
            file_name=f"{dist}_converted.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        st.error(f"Error generating download file: {e}")


# =========================
# Main
# =========================
def main():
    st.set_page_config(
        page_title="Distributor Converter", page_icon="📦", layout="wide"
    )
    st.header("📂 Distributor Data Converter")
    st.markdown(
        "Upload your distributor's data file and it will be converted to the standard template."
    )

    try:
        ensure_bq_objects()
    except Exception as e:
        st.error(f"BigQuery setup error: {e}")
        st.stop()

    distributors = list_distributors()
    if not distributors:
        st.info(
            "No distributors configured yet. "
            "Please configure at least one in BigQuery directly."
        )
        return

    dist = st.selectbox("Select Distributor", distributors)
    brand = st.selectbox("Select Brand", BRAND_OPTIONS)
    brand_prefix = BRAND_PREFIXES.get(brand, "")

    # ── Route to the correct pipeline ────────────────────────────────────────
    if is_3m_distributor(dist):
        render_3m_pipeline(dist, brand, brand_prefix)
    else:
        render_standard_pipeline(dist, brand, brand_prefix)


if __name__ == "__main__":
    main()
